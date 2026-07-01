from __future__ import annotations

import csv
import gzip
import io
from datetime import date, datetime

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from metacritic_calendar_app.main import app
from metacritic_calendar_app.models import (
    BoxOfficeMojoReleaseItem,
    BoxOfficeMojoReleaseWindowSnapshot,
    MetacriticCalendarItem,
    MetacriticCalendarSnapshot,
    MetacriticTvClassificationItem,
    MetacriticTvClassificationSnapshot,
    TvImdbEpisodeCountItem,
    TvImdbEpisodeCountSnapshot,
)
from metacritic_calendar_app.routes import calendar_service
from metacritic_calendar_app.services.calendar import MetacriticCalendarService
from metacritic_calendar_app.services.imdb_episode_counts import (
    ImdbEpisodeCountLookup,
    ImdbEpisodeCountService,
    TvImdbEpisodeCountService,
    parse_imdb_episode_dates,
)
from metacritic_calendar_app.services.box_office_mojo import BoxOfficeMojoCalendarService
from metacritic_calendar_app.services.tv_classification import (
    DAYPART_OTHER,
    DAYPART_PRIME_TIME,
    LANGUAGE_TYPE_ENGLISH,
    LANGUAGE_TYPE_OTHER,
    PROGRAM_TYPE_MINI_SERIES,
    PROGRAM_TYPE_MOVIE,
    PROGRAM_TYPE_OTHER,
    PROGRAM_TYPE_SERIES,
    TV_CLASSIFICATION_EXPORT_HEADERS,
    MetacriticTvClassificationReportService,
)
from title_url_lookup_app.services.imdb_dataset import ImdbTitleMatch, ImdbDatasetLookupService
import pytest

@pytest.fixture(autouse=True)
def mock_imdb_lookup(monkeypatch):
    monkeypatch.setattr(ImdbDatasetLookupService, "lookup_title", lambda self, query: [])


def test_parse_movie_calendar_payload():
    payload = {
        "data": {
            "item": {
                "headline": "Movie Release Calendar",
                "dateUpdated": {"date": "2026-04-20 18:00:15"},
                "body": """
                    <h3>MON / April 20</h3>
                    <table width="100%">
                      <tbody>
                        <tr>
                          <td><shortcode api="{&quot;metaScore&quot;:64,&quot;title&quot;:&quot;Tow&quot;,&quot;url&quot;:&quot;/movie/tow/&quot;}"></shortcode></td>
                          <td><a href="https://www.metacritic.com/movie/tow/">Tow</a><br>(Roadside Attractions/Vertical)<br>Drama</td>
                          <td>RENT/BUY</td>
                        </tr>
                        <tr>
                          <td><shortcode api="{&quot;metaScore&quot;:55,&quot;title&quot;:&quot;Bonus Feature&quot;,&quot;url&quot;:&quot;/movie/bonus-feature/&quot;}"></shortcode></td>
                          <td><a href="https://www.metacritic.com/movie/bonus-feature/">Bonus Feature</a><br>(Sample Studio)<br>Comedy</td>
                          <td>Special Event</td>
                        </tr>
                      </tbody>
                    </table>
                    <h2>2027 releases</h2>
                    <h3>FRI / January 15, 2027</h3>
                    <table width="100%">
                      <tbody>
                        <tr>
                          <td>&nbsp;</td>
                          <td><strong>The Beekeeper 2</strong><br>(Amazon MGM Studios)<br>Action/Thriller</td>
                          <td><strong>WIDE</strong></td>
                        </tr>
                      </tbody>
                    </table>
                """,
            }
        }
    }

    items = calendar_service.parse_payload(calendar_service.TARGETS["movies"], payload)

    assert len(items) == 3
    assert items[0].title == "Tow"
    assert items[0].availability == "RENT/BUY"
    assert items[1].title == "Bonus Feature"
    assert items[1].availability == "Special Event"
    assert items[2].title == "The Beekeeper 2"
    assert items[2].release_date == "2027-01-15"


def test_parse_games_calendar_payload():
    payload = {
        "data": {
            "item": {
                "headline": "Notable Video Game Releases: New and Upcoming",
                "datePublished": {"date": "2026-04-20 07:00:00"},
                "body": """
                    <h2>Just released</h2>
                    <h3>This week's new games</h3>
                    <table width="100%">
                      <tbody>
                        <tr>
                          <td><shortcode api="{&quot;metaScore&quot;:80,&quot;title&quot;:&quot;Vampire Crawlers&quot;,&quot;url&quot;:&quot;/game/vampire-crawlers/&quot;}"></shortcode></td>
                          <td><a href="https://www.metacritic.com/game/vampire-crawlers/">Vampire Crawlers</a> PC PS5 XBX NS<br>Roguelite/Deckbuilder - April 21</td>
                          <td>&nbsp;<br/></td>
                        </tr>
                      </tbody>
                    </table>
                    <h2>Coming soon</h2>
                    <h3>Week of May 4-10</h3>
                    <table width="100%">
                      <tbody>
                        <tr>
                          <td>&nbsp;</td>
                          <td><a href="https://www.metacritic.com/game/mixtape/">Mixtape</a> PC PS5 XBX NS2<br>Adventure - May 7</td>
                          <td>&nbsp;</td>
                        </tr>
                      </tbody>
                    </table>
                """,
            }
        }
    }

    items = calendar_service.parse_payload(calendar_service.TARGETS["games"], payload)

    assert len(items) == 2
    assert items[0].title == "Vampire Crawlers"
    assert items[0].provider == "PC PS5 XBX NS"
    assert items[0].details == "Roguelite/Deckbuilder"
    assert items[0].release_date == "2026-04-21"
    assert items[0].metascore == 80
    assert items[1].title == "Mixtape"
    assert items[1].release_date == "2026-05-07"


def test_calendar_parser_repairs_mojibake_accents():
    payload = {
        "data": {
            "item": {
                "headline": "Movie Release Calendar",
                "dateUpdated": {"date": "2026-05-07 12:00:00"},
                "body": """
                    <h3>THU / May 28</h3>
                    <table width="100%">
                      <tbody>
                        <tr>
                          <td>&nbsp;</td>
                          <td><strong>Emi Mart\u00c3\u00adnez: The Kid Who Stops Time</strong><br>Foreign/Documentary/Sports</td>
                          <td>Streaming (Netflix)</td>
                        </tr>
                      </tbody>
                    </table>
                """,
            }
        }
    }

    items = calendar_service.parse_payload(calendar_service.TARGETS["movies"], payload)

    assert items[0].title == "Emi Mart\u00ednez: The Kid Who Stops Time"


def test_parse_tv_calendar_excludes_rent_buy_availability_and_special_titles():
    payload = {
        "data": {
            "item": {
                "headline": "TV Premiere Calendar",
                "dateUpdated": {"date": "2026-04-20 18:00:15"},
                "body": """
                    <h3>MON / April 20</h3>
                    <table width="100%">
                      <tbody>
                        <tr>
                          <td>&nbsp;</td>
                          <td><a href="https://www.metacritic.com/movie/weekend-at-the-end-of-the-world/">Weekend at the End of the World</a><br>Horror/Comedy</td>
                          <td>RENT/BUY</td>
                        </tr>
                        <tr>
                          <td>&nbsp;</td>
                          <td><strong>Streaming Movie</strong> <img alt="movie" src="badge.png"><br>Drama</td>
                          <td>Netflix</td>
                        </tr>
                        <tr>
                          <td>&nbsp;</td>
                          <td><strong>All American: The Final Season</strong><br>Drama</td>
                          <td>CW 9p</td>
                        </tr>
                        <tr>
                          <td>&nbsp;</td>
                          <td><strong>Example Comedy Special</strong><br>Comedy</td>
                          <td>Netflix</td>
                        </tr>
                        <tr>
                          <td>&nbsp;</td>
                          <td><a href="https://www.metacritic.com/tv/example-show/">Example Show</a><br>Drama</td>
                          <td>Netflix</td>
                        </tr>
                      </tbody>
                    </table>
                """,
            }
        }
    }

    items = calendar_service.parse_payload(calendar_service.TARGETS["tv"], payload)

    assert len(items) == 4
    assert items[0].title == "Weekend at the End of the World"
    assert items[0].section == "movies"
    assert items[0].availability == "RENT/BUY"
    assert items[1].title == "Streaming Movie"
    assert items[1].section == "movies"
    assert items[1].availability == "Netflix"
    assert items[2].title == "All American: The Final Season"
    assert items[2].section == "tv"
    assert items[2].availability == "CW 9p"
    assert items[3].title == "Example Show"
    assert items[3].section == "tv"
    assert items[3].availability == "Netflix"


def test_tv_classification_report_parses_requested_columns():
    payload = {
        "data": {
            "item": {
                "headline": "TV Premiere Calendar",
                "dateUpdated": {"date": "2026-05-05 12:00:00"},
                "body": """
                    <h3>TUE / May 5</h3>
                    <table width="100%">
                      <tbody>
                        <tr>
                          <td><shortcode api="{&quot;title&quot;:&quot;The Drama&quot;,&quot;url&quot;:&quot;/movie/the-drama/&quot;}"></shortcode></td>
                          <td><strong>($)</strong> <a href="https://www.metacritic.com/movie/the-drama/">The Drama</a><br>Rom-Com</td>
                          <td>RENT/BUY</td>
                        </tr>
                        <tr>
                          <td>&nbsp;</td>
                          <td><strong>Gary</strong><br>Comedy/Drama Special: This links to <a href="https://www.metacritic.com/tv/the-bear/">The Bear</a>.</td>
                          <td>Hulu</td>
                        </tr>
                        <tr>
                          <td>&nbsp;</td>
                          <td><strong>Soccer's American Dream</strong> <img alt="limited series" src="badge.png"><br>Reality/Sports</td>
                          <td>Vice 10p</td>
                        </tr>
                      </tbody>
                    </table>
                    <h3>WED / May 6</h3>
                    <table width="100%">
                      <tbody>
                        <tr>
                          <td>&nbsp;</td>
                          <td><strong>Busted on Bodycam</strong> <img alt="new series" src="badge.png"><br>Reality</td>
                          <td>Tubi</td>
                        </tr>
                        <tr>
                          <td><shortcode api="{&quot;title&quot;:&quot;Season 2&quot;,&quot;url&quot;:&quot;/tv/citadel/season-2/&quot;}"></shortcode></td>
                          <td><a href="https://www.metacritic.com/tv/citadel/season-2/">Citadel</a> - <a href="https://www.youtube.com/watch?v=abc">Trailer</a><br>Drama: Season two description.</td>
                          <td>Prime Video</td>
                        </tr>
                        <tr>
                          <td>&nbsp;</td>
                          <td><strong>Love Is Blind: Poland</strong> <img alt="new series" src="badge.png"><br>Foreign/Reality</td>
                          <td>Netflix</td>
                        </tr>
                      </tbody>
                    </table>
                """,
            }
        }
    }
    service = MetacriticTvClassificationReportService(calendar_service=calendar_service)

    items = service.parse_payload(payload)

    assert [item.title for item in items] == [
        "Gary",
        "Soccer's American Dream",
        "The Drama",
        "Busted on Bodycam",
        "Citadel",
        "Love Is Blind: Poland",
    ]
    assert items[0].release_date == "2026-05-05"
    assert items[0].network == "Hulu"
    assert items[0].daypart == DAYPART_OTHER
    assert items[0].program_type == PROGRAM_TYPE_OTHER
    assert items[0].language_type == LANGUAGE_TYPE_ENGLISH
    assert items[0].genre_1 == "Comedy"
    assert items[0].genre_2 == "Drama Special"
    assert items[0].metacritic_url == ""

    assert items[1].network == "Vice"
    assert items[1].daypart == DAYPART_PRIME_TIME
    assert items[1].program_type == PROGRAM_TYPE_MINI_SERIES
    assert items[1].genre_1 == "Reality"
    assert items[1].genre_2 == "Sports"

    assert items[2].network == "RENT/BUY"
    assert items[2].program_type == PROGRAM_TYPE_MOVIE
    assert items[2].genre_1 == "Rom Com"
    assert items[3].program_type == PROGRAM_TYPE_SERIES
    assert items[4].title == "Citadel"
    assert items[4].metacritic_url == "https://www.metacritic.com/tv/citadel/season-2/"
    assert items[4].genre_1 == "Drama"
    assert items[5].language_type == LANGUAGE_TYPE_OTHER
    assert items[5].genre_1 == "Reality"


def test_tv_classification_report_repairs_mojibake_titles():
    payload = {
        "data": {
            "item": {
                "headline": "TV Premiere Calendar",
                "dateUpdated": {"date": "2026-05-07 12:00:00"},
                "body": """
                    <h3>THU / May 28</h3>
                    <table width="100%">
                      <tbody>
                        <tr>
                          <td>&nbsp;</td>
                          <td><strong>Emi Mart\u00c3\u00adnez: The Kid Who Stops Time</strong> <img alt="movie" src="badge.png"><br>Foreign/Documentary/Sports</td>
                          <td>Netflix</td>
                        </tr>
                      </tbody>
                    </table>
                """,
            }
        }
    }
    service = MetacriticTvClassificationReportService(calendar_service=calendar_service)

    items = service.parse_payload(payload)

    assert items[0].title == "Emi Mart\u00ednez: The Kid Who Stops Time"
    assert items[0].program_type == PROGRAM_TYPE_MOVIE


def test_tv_classification_report_fetches_present_day_to_next_three_months(monkeypatch):
    payload = {
        "data": {
            "item": {
                "headline": "TV Premiere Calendar",
                "dateUpdated": {"date": "2026-05-05 12:00:00"},
                "body": """
                    <h3>MON / May 4</h3>
                    <table><tbody>
                      <tr>
                        <td>&nbsp;</td>
                        <td><strong>Yesterday Show</strong><br>Drama</td>
                        <td>Netflix</td>
                      </tr>
                    </tbody></table>
                    <h3>TUE / May 5</h3>
                    <table><tbody>
                      <tr>
                        <td>&nbsp;</td>
                        <td><a href="https://www.metacritic.com/movie/today-movie/">Today Movie</a> <img alt="movie" src="movie.png"><br>Drama</td>
                        <td>RENT/BUY</td>
                      </tr>
                      <tr>
                        <td>&nbsp;</td>
                        <td><strong>Today Show</strong><br>Comedy</td>
                        <td>Hulu</td>
                      </tr>
                    </tbody></table>
                    <h3>WED / August 5</h3>
                    <table><tbody>
                      <tr>
                        <td>&nbsp;</td>
                        <td><strong>Window Edge</strong><br>Reality</td>
                        <td>Peacock</td>
                      </tr>
                    </tbody></table>
                    <h3>THU / August 6</h3>
                    <table><tbody>
                      <tr>
                        <td>&nbsp;</td>
                        <td><strong>Too Far Show</strong><br>Drama</td>
                        <td>Prime Video</td>
                      </tr>
                    </tbody></table>
                """,
            }
        }
    }
    class FakeImdbDatasetLookup:
        def __init__(self):
            self.queries = []

        def lookup_title(self, query):
            self.queries.append(query)
            imdb_ids = {
                "Today Movie": "tt1000000",
                "Today Show": "tt1000001",
                "Window Edge": "tt1000002",
            }
            imdb_id = imdb_ids.get(query.title)
            if not imdb_id:
                return []
            return [
                ImdbTitleMatch(
                    imdb_id=imdb_id,
                    url=f"https://www.imdb.com/title/{imdb_id}/",
                    display_title=query.title,
                    original_title=query.title,
                    title_type="movie" if query.title_type == "movie" else "tvSeries",
                    start_year=query.year,
                    end_year="",
                    score=180,
                    matched_on=["primary_title"],
                )
            ]

    fake_imdb_lookup = FakeImdbDatasetLookup()
    service = MetacriticTvClassificationReportService(
        calendar_service=calendar_service,
        imdb_dataset_lookup=fake_imdb_lookup,
    )
    monkeypatch.setattr(calendar_service, "_fetch_calendar_payload", lambda client, target: payload)

    snapshot = service.fetch_snapshot(today=date(2026, 5, 5))

    assert snapshot.window_start == date(2026, 5, 5)
    assert snapshot.window_end == date(2026, 8, 5)
    assert [item.title for item in snapshot.items] == ["Today Show", "Window Edge"]
    assert [item.imdb_ttcode for item in snapshot.items] == ["tt1000001", "tt1000002"]
    assert [(query.title, query.title_type) for query in fake_imdb_lookup.queries] == [
        ("Today Show", "tv"),
        ("Window Edge", "tv"),
    ]
    assert "2 parsed TV row(s) were outside the 3-month date window." in snapshot.notes
    assert "1 parsed TV row(s) were skipped because the availability is Rent/Buy." in snapshot.notes


def test_index_page_renders():
    client = TestClient(app)
    response = client.get("/")

    assert response.status_code == 200
    assert "Asana LF Automation Tasks" in response.text
    assert "1. Billboard New Entries Brand Discovery" in response.text
    assert "4. TV Season & Episode Metadata Adding" in response.text


def test_calendar_search_page_accepts_direct_browser_visit():
    client = TestClient(app)
    response = client.get("/calendar/search")

    assert response.status_code == 200
    assert "Asana LF Automation Tasks" in response.text
    assert 'name="task"' in response.text


def test_calendar_search_accepts_multiple_sections(monkeypatch):
    captured = {}
    snapshot = MetacriticCalendarSnapshot(
        calendar_type="games_tv",
        generated_at=datetime(2026, 4, 21, 9, 0, 0),
        items=[
            MetacriticCalendarItem(
                section="tv",
                section_label="TV Shows",
                source_title="TV Premiere Calendar",
                source_url="https://www.metacritic.com/news/tv-premiere-dates/",
                group_label="MON / April 20",
                release_date="2026-04-20",
                title="Example Show",
                url="https://www.metacritic.com/tv/example-show/",
                provider="Netflix",
            ),
            MetacriticCalendarItem(
                section="tv",
                section_label="TV Shows",
                source_title="TV Premiere Calendar",
                source_url="https://www.metacritic.com/news/tv-premiere-dates/",
                group_label="MON / June 1",
                release_date="2026-06-01",
                title="Outside Show",
                url="https://www.metacritic.com/tv/outside-show/",
                provider="Hulu",
            )
        ],
    )

    def fake_fetch_snapshot(calendar_type):
        captured["calendar_type"] = calendar_type
        return snapshot

    monkeypatch.setattr("metacritic_calendar_app.routes.calendar_service.fetch_snapshot", fake_fetch_snapshot)

    client = TestClient(app)
    response = client.post(
        "/calendar/search",
        data={
            "calendar_type": ["games", "tv"],
            "custom_start_date": "2026-04-01",
            "custom_end_date": "2026-04-30",
        },
    )

    assert response.status_code == 200
    assert captured["calendar_type"] == ["games", "tv"]
    assert "Example Show" in response.text
    assert "Outside Show" not in response.text
    assert "1 row(s) were outside the selected date range." in response.text


def test_calendar_service_resolves_multiple_calendar_types():
    keys, snapshot_type = calendar_service.resolve_calendar_types("games,tv")

    assert keys == ["games", "tv"]
    assert snapshot_type == "games_tv"


def test_movies_calendar_selection_includes_tv_feed_movie_rows(monkeypatch):
    service = MetacriticCalendarService()
    movie_payload = {
        "data": {
            "item": {
                "headline": "Movie Release Calendar",
                "dateUpdated": {"date": "2026-04-20 18:00:15"},
                "body": "<h3>FRI / May 8</h3><table><tbody></tbody></table>",
            }
        }
    }
    tv_payload = {
        "data": {
            "item": {
                "headline": "TV Premiere Calendar",
                "dateUpdated": {"date": "2026-04-20 18:00:15"},
                "body": """
                    <h3>TUE / May 5</h3>
                    <table width="100%">
                      <tbody>
                        <tr>
                          <td>&nbsp;</td>
                          <td><a href="https://www.metacritic.com/movie/the-drama/">The Drama</a><br>Rom-Com</td>
                          <td>RENT/BUY</td>
                        </tr>
                        <tr>
                          <td>&nbsp;</td>
                          <td><strong>Example Show</strong><br>Drama</td>
                          <td>Netflix</td>
                        </tr>
                      </tbody>
                    </table>
                """,
            }
        }
    }

    def fake_fetch_calendar_payload(client, target):
        if target.key == "movies":
            return movie_payload
        if target.key == "tv":
            return tv_payload
        raise AssertionError(f"Unexpected target {target.key}")

    monkeypatch.setattr(service, "_fetch_calendar_payload", fake_fetch_calendar_payload)

    snapshot = service.fetch_snapshot("movies")

    assert snapshot.calendar_type == "movies"
    assert [item.title for item in snapshot.items] == ["The Drama"]
    assert snapshot.items[0].section == "movies"
    assert snapshot.items[0].availability == "RENT/BUY"


def test_api_calendar_route(monkeypatch):
    snapshot = MetacriticCalendarSnapshot(
        calendar_type="games",
        generated_at=datetime(2026, 4, 21, 9, 0, 0),
        items=[
            MetacriticCalendarItem(
                section="games",
                section_label="Games",
                source_title="Notable Video Game Releases: New and Upcoming",
                source_url="https://www.metacritic.com/news/major-new-and-upcoming-video-games-ps5-xbox-switch-pc/",
                group_label="This week's new games",
                release_date="2026-04-21",
                title="Vampire Crawlers",
                url="https://www.metacritic.com/game/vampire-crawlers/",
                provider="PC PS5 XBX NS",
                availability="",
                details="Roguelite/Deckbuilder",
                metascore=80,
            ),
            MetacriticCalendarItem(
                section="games",
                section_label="Games",
                source_title="Notable Video Game Releases: New and Upcoming",
                source_url="https://www.metacritic.com/news/major-new-and-upcoming-video-games-ps5-xbox-switch-pc/",
                group_label="Week of May 4-10",
                release_date="2026-05-07",
                title="Mixtape",
                url="https://www.metacritic.com/game/mixtape/",
                provider="PC PS5",
                availability="",
                details="Adventure",
            )
        ],
    )

    monkeypatch.setattr("metacritic_calendar_app.routes.calendar_service.fetch_snapshot", lambda calendar_type: snapshot)

    client = TestClient(app)
    response = client.get("/api/calendar?calendar_type=games&start_date=2026-04-01&end_date=2026-04-30")

    assert response.status_code == 200
    payload = response.json()
    assert payload["calendar_type"] == "games"
    assert [item["title"] for item in payload["items"]] == ["Vampire Crawlers"]
    assert "Date filter: 2026-04-01 to 2026-04-30." in payload["notes"]


def test_imdb_episode_count_service_counts_seasons_and_episodes(tmp_path, monkeypatch):
    source_dir = tmp_path / "source"
    source_dir.mkdir()
    source_path = source_dir / "title.episode.tsv.gz"
    with gzip.open(source_path, "wt", encoding="utf-8", newline="") as file_handle:
        writer = csv.DictWriter(
            file_handle,
            fieldnames=["tconst", "parentTconst", "seasonNumber", "episodeNumber"],
            delimiter="\t",
        )
        writer.writeheader()
        writer.writerows(
            [
                {"tconst": "tt1000001", "parentTconst": "tt1234567", "seasonNumber": "1", "episodeNumber": "1"},
                {"tconst": "tt1000002", "parentTconst": "tt1234567", "seasonNumber": "1", "episodeNumber": "2"},
                {"tconst": "tt1000003", "parentTconst": "tt1234567", "seasonNumber": "2", "episodeNumber": "1"},
                {"tconst": "tt2000001", "parentTconst": "tt7654321", "seasonNumber": "1", "episodeNumber": "1"},
            ]
        )

    class FakeImdbDatasetLookup:
        def lookup_title(self, query):
            return [
                ImdbTitleMatch(
                    imdb_id="tt1234567",
                    url="https://www.imdb.com/title/tt1234567/",
                    display_title="Example Show",
                    original_title="Example Show",
                    title_type="tvSeries",
                    start_year="2026",
                    end_year="",
                    score=185,
                    matched_on=["primary_title"],
                )
            ]

    service = ImdbEpisodeCountService(imdb_dataset_lookup=FakeImdbDatasetLookup())
    monkeypatch.setattr(service, "_imdb_dataset_dir", lambda: tmp_path)
    monkeypatch.setattr(service, "_episode_dataset_source", lambda: str(source_path))
    monkeypatch.setattr(
        service,
        "_fetch_latest_season_episode_page",
        lambda imdb_id, season_number: _fake_imdb_response(
            """
            <script type="application/ld+json">
              {"@type":"TVEpisode","datePublished":"2026-05-03"}
            </script>
            """
        ),
    )

    lookup = service.lookup_show("Example Show")

    assert lookup.imdb_match_status == "found"
    assert lookup.imdb_id == "tt1234567"
    assert lookup.season_count == 2
    assert lookup.latest_season_number == 2
    assert lookup.latest_season_episode_count == 1
    assert lookup.latest_season_start_date == "2026-05-03"
    assert lookup.latest_season_end_date == "2026-06-02"
    assert lookup.latest_season_date_source == "imdb"
    assert lookup.episode_count == 3


def test_parse_imdb_episode_dates_supports_json_and_visible_dates():
    html = """
        <script type="application/ld+json">
          {"@type":"TVEpisode","datePublished":"2026-05-01"}
          {"@type":"TVEpisode","datePublished":"2026-05-08"}
        </script>
    """

    dates = parse_imdb_episode_dates(html)

    assert [item.isoformat() for item in dates] == ["2026-05-01", "2026-05-08"]


def test_tv_imdb_episode_count_service_filters_selected_date_window():
    class FakeCalendarService:
        def fetch_snapshot(self, calendar_type):
            assert calendar_type == "tv"
            return MetacriticCalendarSnapshot(
                calendar_type="tv",
                generated_at=datetime(2026, 4, 29, 9, 0, 0),
                items=[
                    MetacriticCalendarItem(
                        section="tv",
                        section_label="TV Shows",
                        source_title="TV Premiere Calendar",
                        source_url="https://www.metacritic.com/news/tv-premiere-dates/",
                        group_label="WED / April 29",
                        release_date="2026-04-29",
                        title="Today Show",
                        availability="Netflix",
                    ),
                    MetacriticCalendarItem(
                        section="tv",
                        section_label="TV Shows",
                        source_title="TV Premiere Calendar",
                        source_url="https://www.metacritic.com/news/tv-premiere-dates/",
                        group_label="FRI / April 24",
                        release_date="2026-04-24",
                        title="Recent Show",
                        availability="Peacock",
                    ),
                    MetacriticCalendarItem(
                        section="tv",
                        section_label="TV Shows",
                        source_title="TV Premiere Calendar",
                        source_url="https://www.metacritic.com/news/tv-premiere-dates/",
                        group_label="WED / May 6",
                        release_date="2026-05-06",
                        title="Week Show",
                        availability="Disney+",
                    ),
                    MetacriticCalendarItem(
                        section="movies",
                        section_label="Movies",
                        source_title="TV Premiere Calendar",
                        source_url="https://www.metacritic.com/news/tv-premiere-dates/",
                        group_label="TUE / May 5",
                        release_date="2026-05-05",
                        title="Rent Buy Movie",
                        availability="RENT/BUY",
                    ),
                    MetacriticCalendarItem(
                        section="movies",
                        section_label="Movies",
                        source_title="TV Premiere Calendar",
                        source_url="https://www.metacritic.com/news/tv-premiere-dates/",
                        group_label="MON / May 4",
                        release_date="2026-05-04",
                        title="Streaming Movie",
                        url="https://www.metacritic.com/movie/streaming-movie/",
                        availability="Netflix",
                    ),
                    MetacriticCalendarItem(
                        section="tv",
                        section_label="TV Shows",
                        source_title="TV Premiere Calendar",
                        source_url="https://www.metacritic.com/news/tv-premiere-dates/",
                        group_label="TUE / May 5",
                        release_date="2026-05-05",
                        title="Movie Url Row",
                        url="https://www.metacritic.com/movie/movie-url-row/",
                        availability="Hulu",
                    ),
                    MetacriticCalendarItem(
                        section="tv",
                        section_label="TV Shows",
                        source_title="TV Premiere Calendar",
                        source_url="https://www.metacritic.com/news/tv-premiere-dates/",
                        group_label="THU / May 7",
                        release_date="2026-05-07",
                        title="Outside Show",
                        availability="Hulu",
                    ),
                    MetacriticCalendarItem(
                        section="tv",
                        section_label="TV Shows",
                        source_title="TV Premiere Calendar",
                        source_url="https://www.metacritic.com/news/tv-premiere-dates/",
                        group_label="TBD",
                        release_date="",
                        title="Undated Show",
                        availability="Max",
                    ),
                ],
            )

    class FakeImdbEpisodeCountService:
        def __init__(self):
            self.looked_up_titles = []

        def lookup_show(self, title):
            self.looked_up_titles.append(title)
            return ImdbEpisodeCountLookup(
                imdb_id="tt1234567",
                imdb_url="https://www.imdb.com/title/tt1234567/",
                imdb_title=title,
                imdb_start_year="2026",
                imdb_title_type="tvSeries",
                imdb_match_status="found",
                imdb_match_score=185,
                season_count=1,
                latest_season_number=1,
                latest_season_episode_count=1,
                episode_count=1,
            )

    imdb_service = FakeImdbEpisodeCountService()
    service = TvImdbEpisodeCountService(
        calendar_service=FakeCalendarService(),
        imdb_episode_count_service=imdb_service,
    )

    snapshot = service.fetch_snapshot("week", today=date(2026, 4, 29))

    assert snapshot.date_window_key == "week"
    assert snapshot.window_start == date(2026, 4, 29)
    assert snapshot.window_end == date(2026, 5, 6)
    assert [item.title for item in snapshot.items] == ["Today Show", "Week Show"]
    assert imdb_service.looked_up_titles == ["Today Show", "Week Show"]
    assert snapshot.items[0].latest_season_start_date == "2026-04-29"
    assert snapshot.items[0].latest_season_end_date == "2026-05-29"
    assert "2 TV rows were outside the selected date window." in snapshot.notes
    assert "1 TV row was skipped because the release date was unavailable." in snapshot.notes
    assert "1 TV row was skipped because the availability is Rent/Buy." in snapshot.notes
    assert "2 TV rows were skipped because the title is tagged as a movie." in snapshot.notes

    last_7_days_snapshot = service.fetch_snapshot("last_7_days", today=date(2026, 4, 29))

    assert last_7_days_snapshot.date_window_key == "last_7_days"
    assert last_7_days_snapshot.window_start == date(2026, 4, 23)
    assert last_7_days_snapshot.window_end == date(2026, 4, 29)
    assert [item.title for item in last_7_days_snapshot.items] == ["Today Show", "Recent Show"]

    custom_snapshot = service.fetch_snapshot(
        "custom",
        start_date="2026-05-06",
        end_date="2026-05-07",
        today=date(2026, 4, 29),
    )

    assert custom_snapshot.date_window_key == "custom"
    assert custom_snapshot.date_window_label == "Custom Date Range"
    assert custom_snapshot.window_start == date(2026, 5, 6)
    assert custom_snapshot.window_end == date(2026, 5, 7)
    assert [item.title for item in custom_snapshot.items] == ["Week Show", "Outside Show"]


def test_tv_imdb_episode_count_route_api_and_downloads(monkeypatch):
    from metacritic_calendar_app import routes

    snapshot = TvImdbEpisodeCountSnapshot(
        generated_at=datetime(2026, 4, 29, 10, 0, 0),
        source_url="https://www.metacritic.com/news/tv-premiere-dates/",
        date_window_key="month",
        date_window_label="Upcoming Month (30 days)",
        window_start=date(2026, 4, 29),
        window_end=date(2026, 5, 29),
        items=[
            TvImdbEpisodeCountItem(
                release_date="2026-05-01",
                title="Example Show",
                metacritic_url="https://www.metacritic.com/tv/example-show/",
                network_distributor="Netflix",
                provider="Netflix",
                imdb_id="tt1234567",
                imdb_url="https://www.imdb.com/title/tt1234567/",
                imdb_title="Example Show",
                imdb_start_year="2026",
                imdb_title_type="tvSeries",
                imdb_match_status="found",
                imdb_match_score=185,
                season_count=2,
                latest_season_number=2,
                latest_season_episode_count=4,
                latest_season_start_date="2026-05-01",
                latest_season_end_date="2026-05-22",
                latest_season_date_source="imdb",
                episode_count=12,
            )
        ],
    )

    captured_fetches = []

    def fake_fetch_snapshot(date_window="year", start_date=None, end_date=None):
        captured_fetches.append((date_window, start_date, end_date))
        update = {"date_window_key": date_window}
        if start_date and end_date:
            update["window_start"] = date.fromisoformat(start_date)
            update["window_end"] = date.fromisoformat(end_date)
            update["date_window_label"] = "Custom Date Range"
        return snapshot.model_copy(update=update)

    monkeypatch.setattr(routes.tv_imdb_episode_count_service, "fetch_snapshot", fake_fetch_snapshot)

    client = TestClient(app)

    page_response = client.get("/tv/imdb-episode-counts")
    assert page_response.status_code == 200
    assert "Asana LF Automation Tasks" in page_response.text
    assert 'select name="date_window"' in page_response.text
    assert '<option value="today"' in page_response.text
    assert '<option value="last_7_days"' in page_response.text
    assert '<option value="week"' in page_response.text
    assert '<option value="month"' in page_response.text
    assert '<option value="daily_segment" selected' in page_response.text
    assert '<option value="custom"' in page_response.text
    assert 'type="date" name="custom_start_date"' in page_response.text
    assert 'type="date" name="custom_end_date"' in page_response.text

    form_response = client.post("/tv/imdb-episode-counts/search", data={"date_window": "month"})
    assert form_response.status_code == 200
    assert captured_fetches[-1] == ("month", None, None)
    assert "Example Show" in form_response.text
    assert "tt1234567" in form_response.text
    assert "4" in form_response.text
    assert "Upcoming Month (30 days)" in form_response.text
    assert 'value="month" selected' in form_response.text
    assert "network_distributor" in form_response.text
    assert "latest_season_episode_count" in form_response.text
    assert "IMDb Match" not in form_response.text
    assert "Genre / Details" not in form_response.text

    custom_form_response = client.post(
        "/tv/imdb-episode-counts/search",
        data={
            "date_window": "custom",
            "custom_start_date": "2026-04-01",
            "custom_end_date": "2026-04-07",
        },
    )
    assert custom_form_response.status_code == 200
    assert captured_fetches[-1] == ("custom", "2026-04-01", "2026-04-07")
    assert 'value="custom" selected' in custom_form_response.text
    assert 'name="custom_start_date" value="2026-04-01"' in custom_form_response.text
    assert 'name="custom_end_date" value="2026-04-07"' in custom_form_response.text

    export_id = _extract_tv_imdb_export_id(form_response.text)
    csv_response = client.get(f"/tv/imdb-episode-counts/export/{export_id}/csv")
    assert csv_response.status_code == 200
    assert 'filename="tv_imdb_episode_counts_month.csv"' in csv_response.headers["content-disposition"]
    csv_rows = list(csv.DictReader(io.StringIO(csv_response.content.decode("utf-8-sig"))))
    assert list(csv_rows[0].keys()) == [
        "release_date",
        "title",
        "network_distributor",
        "imdb_id",
        "metacritic_url",
        "latest_season_number",
        "latest_season_episode_count",
        "latest_season_start_date",
        "latest_season_end_date",
    ]
    assert csv_rows[0]["release_date"] == "01-05-2026"
    assert csv_rows[0]["title"] == "Example Show"
    assert csv_rows[0]["network_distributor"] == "Netflix"
    assert csv_rows[0]["imdb_id"] == "tt1234567"
    assert csv_rows[0]["metacritic_url"] == "https://www.metacritic.com/tv/example-show/"
    assert csv_rows[0]["latest_season_number"] == "2"
    assert csv_rows[0]["latest_season_episode_count"] == "4"
    assert csv_rows[0]["latest_season_start_date"] == "01-05-2026"
    assert csv_rows[0]["latest_season_end_date"] == "22-05-2026"

    xlsx_response = client.get(f"/tv/imdb-episode-counts/export/{export_id}/xlsx")
    assert xlsx_response.status_code == 200
    assert 'filename="tv_imdb_episode_counts_month.xlsx"' in xlsx_response.headers["content-disposition"]
    workbook = load_workbook(io.BytesIO(xlsx_response.content))
    assert workbook.sheetnames == ["Export"]
    sheet = workbook["Export"]
    assert [cell.value for cell in sheet[1]] == [
        "release_date",
        "title",
        "network_distributor",
        "imdb_id",
        "metacritic_url",
        "latest_season_number",
        "latest_season_episode_count",
        "latest_season_start_date",
        "latest_season_end_date",
    ]
    assert sheet["A2"].value == "01-05-2026"
    assert sheet["B2"].value == "Example Show"
    assert sheet["C2"].value == "Netflix"
    assert sheet["D2"].value == "tt1234567"
    assert sheet["E2"].value == "https://www.metacritic.com/tv/example-show/"
    assert sheet["F2"].value == 2
    assert sheet["G2"].value == 4
    assert sheet["H2"].value == "01-05-2026"
    assert sheet["I2"].value == "22-05-2026"

    api_response = client.get("/api/tv/imdb-episode-counts?date_window=week")
    assert api_response.status_code == 200
    assert captured_fetches[-1] == ("week", None, None)
    api_payload = api_response.json()
    assert api_payload["date_window_key"] == "week"
    assert set(api_payload["items"][0]) == {
        "release_date",
        "title",
        "network_distributor",
        "imdb_id",
        "metacritic_url",
        "latest_season_number",
        "latest_season_episode_count",
        "latest_season_start_date",
        "latest_season_end_date",
    }
    assert api_payload["items"][0]["imdb_id"] == "tt1234567"
    assert api_payload["items"][0]["latest_season_episode_count"] == 4
    assert api_payload["items"][0]["network_distributor"] == "Netflix"
    assert api_payload["items"][0]["latest_season_end_date"] == "2026-05-22"


    custom_api_response = client.get(
        "/api/tv/imdb-episode-counts?date_window=custom&start_date=2026-04-01&end_date=2026-04-07"
    )
    assert custom_api_response.status_code == 200
    assert captured_fetches[-1] == ("custom", "2026-04-01", "2026-04-07")
    assert custom_api_response.json()["window_start"] == "2026-04-01"


def test_tv_classification_report_route_api_and_downloads(monkeypatch):
    from metacritic_calendar_app import routes

    snapshot = MetacriticTvClassificationSnapshot(
        generated_at=datetime(2026, 5, 5, 12, 0, 0),
        source_url="https://www.metacritic.com/news/tv-premiere-dates/",
        window_start=date(2026, 5, 5),
        window_end=date(2026, 8, 5),
        items=[
            MetacriticTvClassificationItem(
                release_date="2026-05-05",
                title="Soccer's American Dream",
                imdb_ttcode="tt1234567",
                network="Vice",
                daypart=DAYPART_PRIME_TIME,
                program_type=PROGRAM_TYPE_MINI_SERIES,
                language_type=LANGUAGE_TYPE_ENGLISH,
                genre_1="Reality",
                genre_2="Sports",
            )
        ],
        notes=["Source: Metacritic TV premiere calendar."],
    )

    monkeypatch.setattr(routes.tv_classification_report_service, "fetch_snapshot", lambda: snapshot)

    client = TestClient(app)

    page_response = client.get("/tv/classification-report")
    assert page_response.status_code == 200
    assert "Asana LF Automation Tasks" in page_response.text

    form_response = client.post("/tv/classification-report/search")
    assert form_response.status_code == 200
    assert "Soccer&#39;s American Dream" in form_response.text
    assert "Download CSV" in form_response.text
    assert "Download Excel" in form_response.text
    assert "TV Classification Results" in form_response.text
    assert "2026-05-05" in form_response.text
    assert "2026-08-05" in form_response.text

    export_id = _extract_tv_classification_export_id(form_response.text, "csv")
    csv_response = client.get(f"/tv/classification-report/export/{export_id}/csv")
    assert csv_response.status_code == 200
    assert 'filename="metacritic_tv_classification_report.csv"' in csv_response.headers["content-disposition"]
    csv_rows = list(csv.DictReader(io.StringIO(csv_response.content.decode("utf-8-sig"))))
    assert list(csv_rows[0].keys()) == TV_CLASSIFICATION_EXPORT_HEADERS
    assert csv_rows[0]["Release Date"] == "2026-05-05"
    assert csv_rows[0]["IMDb ttcode"] == "tt1234567"
    assert csv_rows[0]["Network"] == "Vice"
    assert csv_rows[0]["Daypart"] == DAYPART_PRIME_TIME
    assert csv_rows[0]["Program Type"] == PROGRAM_TYPE_MINI_SERIES
    assert csv_rows[0]["Genre 2"] == "Sports"

    xlsx_response = client.get(f"/tv/classification-report/export/{export_id}/xlsx")
    assert xlsx_response.status_code == 200
    assert 'filename="metacritic_tv_classification_report.xlsx"' in xlsx_response.headers["content-disposition"]
    workbook = load_workbook(io.BytesIO(xlsx_response.content))
    assert workbook.sheetnames == ["Sheet1"]
    sheet = workbook["Sheet1"]
    assert [cell.value for cell in sheet[1]] == TV_CLASSIFICATION_EXPORT_HEADERS
    assert sheet["A2"].value == "2026-05-05"
    assert sheet["A2"].number_format == "@"
    assert sheet["B2"].value == "Soccer's American Dream"
    assert sheet["C2"].value == "tt1234567"
    assert sheet["D2"].value == "Vice"

    api_response = client.get("/api/tv/classification-report")
    assert api_response.status_code == 200
    api_payload = api_response.json()
    assert api_payload["window_start"] == "2026-05-05"
    assert api_payload["window_end"] == "2026-08-05"
    assert api_payload["items"][0] == {
        "release_date": "2026-05-05",
        "title": "Soccer's American Dream",
        "imdb_ttcode": "tt1234567",
        "network": "Vice",
        "daypart": DAYPART_PRIME_TIME,
        "program_type": PROGRAM_TYPE_MINI_SERIES,
        "language_type": LANGUAGE_TYPE_ENGLISH,
        "genre_1": "Reality",
        "genre_2": "Sports",
        "genre_3": "",
    }


def test_parse_box_office_mojo_release_page():
    page_html = """
        <div id="table">
          <table>
            <tr class="mojo-group-label">
              <th colspan="3" class="a-size-large mojo-table-header">April 17, 2026</th>
            </tr>
            <tr>
              <td class="mojo-field-type-release">
                <a href="/release/rl123/?ref_=bo_rs_table_1"><img alt="" src="poster.jpg" /></a>
                <div class="mojo-schedule-release-details">
                  <a class="a-link-normal" href="/release/rl123/?ref_=bo_rs_table_1"><h3>Lee Cronin's The Mummy</h3></a>
                  <div class="a-section a-spacing-none mojo-schedule-genres">Horror</div>
                  <div class="a-section a-spacing-none">With: Jack Reynor, Laia Costa</div>
                  <div class="a-section a-spacing-none">2 hr 14 min</div>
                </div>
              </td>
              <td class="mojo-field-type-release_studios">Warner Bros.</td>
              <td class="mojo-field-type-release_scale">Wide</td>
            </tr>
            <tr>
              <td class="mojo-field-type-release">
                <div class="mojo-schedule-release-details">
                  <a class="a-link-normal" href="/release/rl456/?ref_=bo_rs_table_2"><h3>Cave of Forgotten Dreams</h3></a>
                  <div class="a-section a-spacing-none"><span class="a-size-base a-color-secondary">2026 Re-release</span></div>
                  <div class="a-section a-spacing-none mojo-schedule-genres">Documentary History 3D</div>
                  <div class="a-section a-spacing-none">With: Werner Herzog</div>
                  <div class="a-section a-spacing-none">1 hr 30 min</div>
                </div>
              </td>
              <td class="mojo-field-type-release_studios">IFC Films</td>
              <td class="mojo-field-type-release_scale">Limited</td>
            </tr>
            <tr class="mojo-group-label">
              <th colspan="3" class="a-size-large mojo-table-header">April 20, 2026</th>
            </tr>
            <tr>
              <td class="mojo-field-type-release">
                <div class="mojo-schedule-release-details">
                  <a class="a-link-normal" href="/release/rl789/?ref_=bo_rs_table_3"><h3>That Time I Got Reincarnated as a Slime the Movie: Scarlet Bond</h3></a>
                  <div class="a-section a-spacing-none"><span class="a-size-base a-color-secondary">Re-Release 2026 Anime Nights Program</span></div>
                  <div class="a-section a-spacing-none mojo-schedule-genres">Action Adventure Animation Fantasy</div>
                  <div class="a-section a-spacing-none">With: Miho Okasaki, Yuma Uchida</div>
                  <div class="a-section a-spacing-none">1 hr 54 min</div>
                </div>
              </td>
              <td class="mojo-field-type-release_studios">Sony Pictures Releasing</td>
              <td class="mojo-field-type-release_scale">Limited</td>
            </tr>
          </table>
        </div>
    """

    service = BoxOfficeMojoCalendarService()
    items = service.parse_release_page(page_html)

    assert len(items) == 3
    assert items[0].release_date == "2026-04-17"
    assert items[0].title == "Lee Cronin's The Mummy"
    assert items[0].url == "https://www.boxofficemojo.com/release/rl123/"
    assert items[0].genres == "Horror"
    assert items[0].cast == "Jack Reynor, Laia Costa"
    assert items[0].runtime == "2 hr 14 min"
    assert items[1].release_notes == "2026 Re-release"
    assert items[1].distributor == "IFC Films"
    assert items[2].release_notes == "Re-Release 2026 Anime Nights Program"
    assert items[2].scale == "Limited"


def test_fetch_last_7_days_snapshot_uses_yesterday_as_window_end(monkeypatch):
    service = BoxOfficeMojoCalendarService()

    def fake_fetch_release_items_for_window(window_start, window_end, anchor_dates):
        assert window_start == date(2026, 4, 15)
        assert window_end == date(2026, 4, 21)
        assert anchor_dates[0] == date(2026, 4, 15)
        return [
            BoxOfficeMojoReleaseItem(
                release_date="2026-04-18",
                title="Recent Movie",
                url="https://www.boxofficemojo.com/release/rl555/",
                release_notes="",
                genres="Thriller",
                cast="Example Cast",
                runtime="1 hr 42 min",
                distributor="Sample Distributor",
                scale="Wide",
            )
        ]

    monkeypatch.setattr(service, "_fetch_release_items_for_window", fake_fetch_release_items_for_window)

    snapshot = service.fetch_last_7_days_snapshot(date(2026, 4, 22))

    assert snapshot.report_key == "usa_last_7_days"
    assert snapshot.report_label == "USA last 7 days"
    assert snapshot.window_start == date(2026, 4, 15)
    assert snapshot.window_end == date(2026, 4, 21)
    assert snapshot.items[0].title == "Recent Movie"


def test_fetch_upcoming_12_months_snapshot_uses_today_based_window(monkeypatch):
    service = BoxOfficeMojoCalendarService()

    def fake_fetch_release_items_for_window(window_start, window_end, anchor_dates):
        assert window_start == date(2026, 4, 22)
        assert window_end == date(2027, 4, 22)
        assert anchor_dates[0] == date(2026, 4, 22)
        assert anchor_dates[1] == date(2026, 5, 1)
        assert anchor_dates[-1] == date(2027, 4, 1)
        return [
            BoxOfficeMojoReleaseItem(
                release_date="2026-04-25",
                title="Future Movie",
                url="https://www.boxofficemojo.com/release/rl999/",
                release_notes="",
                genres="Drama",
                cast="Example Cast",
                runtime="1 hr 50 min",
                distributor="Sample Distributor",
                scale="Wide",
            )
        ]

    monkeypatch.setattr(service, "_fetch_release_items_for_window", fake_fetch_release_items_for_window)

    snapshot = service.fetch_upcoming_12_months_snapshot(date(2026, 4, 22))

    assert snapshot.report_key == "usa_upcoming_12_months"
    assert snapshot.report_label == "USA upcoming 12 months"
    assert snapshot.window_start == date(2026, 4, 22)
    assert snapshot.window_end == date(2027, 4, 22)
    assert snapshot.items[0].title == "Future Movie"


def test_box_office_mojo_route_api_and_downloads(monkeypatch):
    from metacritic_calendar_app import routes

    snapshot = BoxOfficeMojoReleaseWindowSnapshot(
        report_key="usa_last_7_days",
        report_label="USA last 7 days",
        generated_at=datetime(2026, 4, 22, 11, 30, 0),
        window_start=datetime(2026, 4, 15).date(),
        window_end=datetime(2026, 4, 21).date(),
        source_url="https://www.boxofficemojo.com/calendar/2026-04-15/",
        items=[
            BoxOfficeMojoReleaseItem(
                release_date="2026-04-17",
                title="Lee Cronin's The Mummy",
                url="https://www.boxofficemojo.com/release/rl123/",
                release_notes="",
                genres="Horror",
                cast="Jack Reynor, Laia Costa",
                runtime="2 hr 14 min",
                distributor="Warner Bros.",
                scale="Wide",
            )
        ],
        notes=["Source: Box Office Mojo domestic release schedule, which reflects U.S. theatrical release dates."],
    )
    upcoming_snapshot = BoxOfficeMojoReleaseWindowSnapshot(
        report_key="usa_upcoming_12_months",
        report_label="USA upcoming 12 months",
        generated_at=datetime(2026, 4, 22, 11, 30, 0),
        window_start=datetime(2026, 4, 22).date(),
        window_end=datetime(2027, 4, 22).date(),
        source_url="https://www.boxofficemojo.com/calendar/2026-04-22/",
        items=[
            BoxOfficeMojoReleaseItem(
                release_date="2026-05-01",
                title="Future Movie",
                url="https://www.boxofficemojo.com/release/rl999/",
                release_notes="",
                genres="Drama",
                cast="Example Cast",
                runtime="1 hr 50 min",
                distributor="Future Studio",
                scale="Wide",
            )
        ],
        notes=["Source: Box Office Mojo domestic release schedule, which reflects U.S. theatrical release dates."],
    )

    monkeypatch.setattr(routes, "box_office_mojo_service", BoxOfficeMojoCalendarService())
    monkeypatch.setattr(routes.box_office_mojo_service, "fetch_last_7_days_snapshot", lambda: snapshot)
    monkeypatch.setattr(routes.box_office_mojo_service, "fetch_upcoming_12_months_snapshot", lambda: upcoming_snapshot)

    client = TestClient(app)

    form_response = client.post("/box-office-mojo/search")
    assert form_response.status_code == 200
    assert "Download CSV" in form_response.text
    assert "Download Excel" in form_response.text

    export_id = _extract_export_id(form_response.text, "csv")

    csv_response = client.get(f"/box-office-mojo/export/{export_id}/csv")
    assert csv_response.status_code == 200
    csv_rows = list(csv.DictReader(io.StringIO(csv_response.content.decode("utf-8-sig"))))
    assert csv_rows[0]["title"] == "Lee Cronin's The Mummy"
    assert csv_rows[0]["distributor"] == "Warner Bros."

    xlsx_response = client.get(f"/box-office-mojo/export/{export_id}/xlsx")
    assert xlsx_response.status_code == 200
    workbook = load_workbook(io.BytesIO(xlsx_response.content))
    assert "Summary" in workbook.sheetnames
    assert "Releases" in workbook.sheetnames

    api_response = client.get("/api/box-office-mojo/last-7-days")
    assert api_response.status_code == 200
    assert api_response.json()["items"][0]["title"] == "Lee Cronin's The Mummy"

    upcoming_form_response = client.post("/box-office-mojo/upcoming-12-months/search")
    assert upcoming_form_response.status_code == 200
    assert "USA upcoming 12 months" in upcoming_form_response.text
    assert "Download CSV" in upcoming_form_response.text

    upcoming_api_response = client.get("/api/box-office-mojo/upcoming-12-months")
    assert upcoming_api_response.status_code == 200
    assert upcoming_api_response.json()["items"][0]["title"] == "Future Movie"


def _extract_export_id(html: str, fmt: str) -> str:
    marker = "/box-office-mojo/export/"
    start = html.index(marker) + len(marker)
    end = html.index(f"/{fmt}", start)
    return html[start:end]


def _extract_tv_imdb_export_id(html: str) -> str:
    marker = "/tv/imdb-episode-counts/export/"
    start = html.index(marker) + len(marker)
    end = html.index("/csv", start)
    return html[start:end]


def _extract_tv_classification_export_id(html: str, fmt: str) -> str:
    marker = "/tv/classification-report/export/"
    start = html.index(marker) + len(marker)
    end = html.index(f"/{fmt}", start)
    return html[start:end]


class _FakeImdbResponse:
    def __init__(self, text: str, status_code: int = 200):
        self.text = text
        self.status_code = status_code
        self.headers = {}


def _fake_imdb_response(text: str, status_code: int = 200) -> _FakeImdbResponse:
    return _FakeImdbResponse(text, status_code)


def test_calendar_export_includes_imdb_id():
    snapshot = MetacriticCalendarSnapshot(
        calendar_type="movies",
        generated_at=datetime(2026, 4, 21, 9, 0, 0),
        items=[
            MetacriticCalendarItem(
                section="movies",
                section_label="Movies",
                source_title="Movie Release Calendar",
                source_url="https://www.metacritic.com/news/upcoming-movie-release-dates-schedule/",
                group_label="MON / April 20",
                release_date="2026-04-20",
                title="Example Movie",
                url="https://www.metacritic.com/movie/example-movie/",
                imdb_id="tt9876543",
            )
        ],
    )
    csv_bytes = calendar_service.snapshot_to_csv_bytes(snapshot)
    csv_rows = list(csv.DictReader(io.StringIO(csv_bytes.decode("utf-8-sig"))))
    assert "imdb_id" in csv_rows[0]
    assert csv_rows[0]["imdb_id"] == "tt9876543"


def test_box_office_mojo_export_includes_imdb_id():
    snapshot = BoxOfficeMojoReleaseWindowSnapshot(
        report_key="usa_last_7_days",
        report_label="USA last 7 days",
        generated_at=datetime(2026, 4, 22, 11, 30, 0),
        window_start=date(2026, 4, 15),
        window_end=date(2026, 4, 21),
        source_url="https://www.boxofficemojo.com/calendar/2026-04-15/",
        items=[
            BoxOfficeMojoReleaseItem(
                release_date="2026-04-17",
                title="Lee Cronin's The Mummy",
                url="https://www.boxofficemojo.com/release/rl123/",
                imdb_id="tt1122334",
            )
        ],
    )
    service = BoxOfficeMojoCalendarService()
    
    # Check CSV export
    csv_bytes = service.snapshot_to_csv_bytes(snapshot)
    csv_rows = list(csv.DictReader(io.StringIO(csv_bytes.decode("utf-8-sig"))))
    assert "imdb_id" in csv_rows[0]
    assert csv_rows[0]["imdb_id"] == "tt1122334"
    
    # Check Excel export
    xlsx_bytes = service.snapshot_to_xlsx_bytes(snapshot)
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    releases_sheet = workbook["Releases"]
    headers = [cell.value for cell in releases_sheet[1]]
    assert "IMDb ID" in headers
    imdb_index = headers.index("IMDb ID")
    assert releases_sheet.cell(row=2, column=imdb_index + 1).value == "tt1122334"

