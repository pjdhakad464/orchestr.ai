import io
from datetime import date, timedelta
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app import routes
from app.config import settings
from app.main import app


def test_index_page_renders():
    client = TestClient(app)
    response = client.get("/")
    assert response.status_code == 200
    assert "Welcome to OrchestrAI" in response.text


def test_excel_validator_page_renders():
    client = TestClient(app)
    response = client.get("/excel-validator")
    assert response.status_code == 200
    assert "Data Ops" in response.text
    assert "Project" in response.text
    assert "Validation" in response.text
    assert "Workbook Sheet Validator" in response.text
    assert "Workbook source file (.xlsx or .csv)" in response.text
    assert "Google Sheet URL" in response.text
    assert "Run by" in response.text
    assert "/static/htmx-fallback.js" in response.text
    assert "Validation progress" in response.text
    assert "validator-progress-track" in response.text


def test_excel_validator_guide_page_renders():
    client = TestClient(app)
    response = client.get("/excel-validator/guide")
    assert response.status_code == 200
    assert "Data Ops Validator Guide" in response.text
    assert "Excel, CSV, and Google Sheets" in response.text
    assert "What This Engine Does" in response.text


def test_validate_excel_route_returns_downloadable_result(monkeypatch, tmp_path):
    _configure_validation_storage(monkeypatch, tmp_path)
    client = TestClient(app)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tasks"
    sheet.append(["Task ID", "Task Name", "Status", "Completion %", "Due Date"])
    sheet.append(["TASK-001", "", "Blocked", 120, date.today() - timedelta(days=1)])
    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = """
    {
      "rules": [
        {"sheet": "Tasks", "column": "Task Name", "check": "required"},
        {"sheet": "Tasks", "column": "Status", "check": "in", "values": ["Open", "Done"]},
        {"sheet": "Tasks", "column": "Completion %", "check": "between", "min": 0, "max": 100},
        {"sheet": "Tasks", "column": "Due Date", "check": "date_not_past"}
      ]
    }
    """

    response = client.post(
        "/validate-excel",
        data={"rules_json": rules, "run_by": "QA Team"},
        files={"workbook": ("tasks.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    assert "Download Validated File" in response.text
    assert "4 issues" in response.text
    assert "tasks_validated.xlsx" in response.text

    validation_id = _extract_validation_id(response.text)
    routes.cache._items.clear()
    download_response = client.get("/validate-excel/download/" + validation_id)
    assert download_response.status_code == 200

    validated = load_workbook(io.BytesIO(download_response.content))
    assert "Validation Summary" in validated.sheetnames
    assert any(path.name == "tasks_validated.xlsx" for path in (tmp_path / "runs").rglob("*.xlsx"))


def test_validate_csv_route_returns_downloadable_result(monkeypatch, tmp_path):
    _configure_validation_storage(monkeypatch, tmp_path)
    client = TestClient(app)

    csv_content = "Task ID,Task Name,Status\nTASK-001,,Blocked\n"
    rules = """
    {
      "rules": [
        {"sheet": "*", "column": "Task Name", "check": "required"}
      ]
    }
    """

    response = client.post(
        "/validate-excel",
        data={"rules_json": rules, "run_by": "CSV User"},
        files={"workbook": ("tasks.csv", csv_content.encode("utf-8"), "text/csv")},
    )

    assert response.status_code == 200
    assert "1 issue" in response.text
    assert "tasks_validated.xlsx" in response.text

    download_response = client.get("/validate-excel/download/" + _extract_validation_id(response.text))
    assert download_response.status_code == 200
    validated = load_workbook(io.BytesIO(download_response.content))
    assert "Validation Summary" in validated.sheetnames


def test_validate_google_sheet_route_returns_downloadable_result(monkeypatch, tmp_path):
    _configure_validation_storage(monkeypatch, tmp_path)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tasks"
    sheet.append(["Task ID", "Task Name"])
    sheet.append(["TASK-001", ""])

    monkeypatch.setattr(routes, "load_google_sheet_workbook", lambda _: (workbook, "team-tracker.gsheet"))

    client = TestClient(app)
    rules = """
    {
      "rules": [
        {"sheet": "Tasks", "column": "Task Name", "check": "required"}
      ]
    }
    """

    response = client.post(
        "/validate-excel",
        data={
            "rules_json": rules,
            "run_by": "Sheets User",
            "google_sheet_url": "https://docs.google.com/spreadsheets/d/test-sheet-id/edit#gid=0",
        },
    )

    assert response.status_code == 200
    assert "1 issue" in response.text
    assert "team-tracker_validated.xlsx" in response.text

    download_response = client.get("/validate-excel/download/" + _extract_validation_id(response.text))
    assert download_response.status_code == 200
    validated = load_workbook(io.BytesIO(download_response.content))
    assert "Validation Summary" in validated.sheetnames


def test_validate_excel_requires_run_by(monkeypatch, tmp_path):
    _configure_validation_storage(monkeypatch, tmp_path)
    client = TestClient(app)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tasks"
    sheet.append(["Task ID", "Task Name"])
    sheet.append(["TASK-001", "Example"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = """
    {
      "rules": [
        {"sheet": "Tasks", "column": "Task Name", "check": "required"}
      ]
    }
    """

    response = client.post(
        "/validate-excel",
        data={"rules_json": rules, "run_by": "   "},
        files={"workbook": ("tasks.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    assert "Enter a name in the Run by field before validating." in response.text


def _extract_validation_id(html: str) -> str:
    marker = '/validate-excel/download/'
    start = html.index(marker) + len(marker)
    end = html.index('"', start)
    return html[start:end]


def _configure_validation_storage(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(settings, "validation_history_db", str(tmp_path / "history.sqlite3"))
    monkeypatch.setattr(settings, "validation_output_dir", str(tmp_path / "runs"))
    monkeypatch.setattr(settings, "validation_history_limit", 10)
