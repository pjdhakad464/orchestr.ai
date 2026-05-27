import gzip
import io
import json
import os
import sqlite3
from pathlib import Path
from datetime import date, timedelta

import httpx
from openpyxl import Workbook, load_workbook

import app.services.workbook_validator as workbook_validator
from app.services.workbook_validator import (
    _matches_social_reference_format,
    _normalize_social_reference,
    _lookup_imdb_record,
    _lookup_wikidata_record,
    _lookup_wikipedia_record,
    build_sample_rules_json,
    parse_validation_rules,
    validate_workbook,
)


def _write_gzipped_tsv(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    with gzip.open(path, "wt", encoding="utf-8", newline="") as file_handle:
        file_handle.write("\t".join(headers) + "\n")
        for row in rows:
            file_handle.write("\t".join(row) + "\n")


def _build_rottentomatoes_search_html(media_type: str, rows: list[dict[str, str]]) -> str:
    rendered_rows: list[str] = []
    for row in rows:
        attribute_pairs = [f'{key}="{value}"' for key, value in row.items() if key not in {"title", "url"}]
        attributes = f" {' '.join(attribute_pairs)}" if attribute_pairs else ""
        rendered_rows.append(
            f"""
            <search-page-media-row{attributes}>
                <a href="{row['url']}" class="unset" data-qa="thumbnail-link" slot="thumbnail">
                    <img alt="{row['title']}" src="https://example.com/poster.jpg">
                </a>
                <a href="{row['url']}" class="unset" data-qa="info-name" slot="title">
                    {row['title']}
                </a>
            </search-page-media-row>
            """
        )

    return f'<search-page-result type="{media_type}">{"".join(rendered_rows)}</search-page-result>'


def _build_rottentomatoes_movie_page_html(title: str, url: str, date_published: str) -> str:
    payload = {
        "@context": "https://schema.org",
        "@type": "Movie",
        "name": title,
        "url": url,
        "datePublished": date_published,
    }
    return f"""
    <html>
      <head>
        <title>{title} | Rotten Tomatoes</title>
        <script type="application/ld+json">{json.dumps(payload)}</script>
      </head>
      <body></body>
    </html>
    """


def _build_rules_with_rottentomatoes_check():
    payload = json.loads(build_sample_rules_json())
    payload["rules"].append(
        {
            "sheet": "*",
            "column": "rottentomatoes",
            "check": "rottentomatoes_url_match",
        }
    )
    return parse_validation_rules(json.dumps(payload))


def test_parse_validation_rules_accepts_root_list():
    rules = parse_validation_rules(
        """
        [
          {"sheet": "Tasks", "column": "Task Name", "check": "required"}
        ]
        """
    )

    assert len(rules) == 1
    assert rules[0].sheet == "Tasks"
    assert rules[0].column == "Task Name"


def test_normalize_social_reference_builds_expected_urls():
    assert _normalize_social_reference("twitter", "@openai") == "https://x.com/openai"
    assert _normalize_social_reference("instagram", "natgeo") == "https://www.instagram.com/natgeo/"
    assert _normalize_social_reference("youtube", "OpenAI") == "https://www.youtube.com/@OpenAI"
    assert _normalize_social_reference("youtube", "@OpenAI") == "https://www.youtube.com/@OpenAI"
    assert _normalize_social_reference("tiktok", "@creator") == "https://www.tiktok.com/@creator"
    assert _normalize_social_reference("wikipedia", "Taylor Swift") == "https://en.wikipedia.org/wiki/Taylor_Swift"
    assert _normalize_social_reference("wikidata", "q42") == "https://www.wikidata.org/wiki/Q42"
    assert _normalize_social_reference("imdb", "tt1234567") == "https://www.imdb.com/title/tt1234567/"


def test_social_reference_format_rejects_wrong_resource_types():
    assert _matches_social_reference_format("facebook", "https://www.facebook.com/p/openai")[0] is False
    assert _matches_social_reference_format("twitter", "https://x.com/openai/status/123")[0] is False
    assert _matches_social_reference_format("instagram", "https://www.instagram.com/p/abc123/")[0] is False
    assert _matches_social_reference_format("youtube", "https://www.youtube.com/watch?v=123")[0] is False
    assert _matches_social_reference_format("youtube", "https://www.youtube.com/c/OpenAI")[0] is False
    assert _matches_social_reference_format("tiktok", "https://www.tiktok.com/@creator/video/123")[0] is False
    assert _matches_social_reference_format("wikidata", "https://www.wikidata.org/wiki/Q42")[0] is True
    assert _matches_social_reference_format("wikidata", "https://www.wikidata.org/wiki/Taylor_Swift")[0] is False
    assert _matches_social_reference_format("imdb", "abc123")[0] is False
    assert _matches_social_reference_format("youtube", "https://www.youtube.com/@OpenAI")[0] is True
    assert _matches_social_reference_format("youtube", "http://www.youtube.com/@youdotcom")[0] is True
    assert _matches_social_reference_format("imdb", "nm1234567")[0] is True


def test_facebook_bad_urls_are_flagged():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "companies",
            "brand_set",
            "facebook_page",
        ]
    )
    sheet.append(
        [
            "Regular Title",
            "Media",
            "Feature",
            "Drama",
            "Drama",
            "Company",
            "Competitive View",
            "https://www.facebook.com/p/Netflix-61550012345678",
        ]
    )
    sheet.append(
        [
            "Second Title",
            "Media",
            "Feature",
            "Drama",
            "Drama",
            "Company",
            "Competitive View",
            "https://www.facebook.com/pages/Robot-and-Monster/151138821613437",
        ]
    )
    sheet.append(
        [
            "Third Title",
            "Media",
            "Feature",
            "Drama",
            "Drama",
            "Company",
            "Competitive View",
            "https://www.facebook.com/php/example",
        ]
    )
    sheet.append(
        [
            "Fourth Title",
            "Media",
            "Feature",
            "Drama",
            "Drama",
            "Company",
            "Competitive View",
            "https://www.facebook.com/page/OpenAI",
        ]
    )
    sheet.append(
        [
            "Fifth Title",
            "Media",
            "Feature",
            "Drama",
            "Drama",
            "Company",
            "Competitive View",
            "https://www.facebook.com/profile.php?id=12345",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "facebook-pages.xlsx", rules)

    bad_cells = {issue.cell for issue in artifact.issues if issue.rule == "url_not_contains_if_present"}
    assert {"H2", "H3", "H4", "H5", "H6"}.issubset(bad_cells)


def test_youtube_urls_with_percent_20_and_percent_7_are_flagged(monkeypatch):
    monkeypatch.setattr(workbook_validator, "_fetch_social_reference", lambda *args, **kwargs: (True, ""))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "companies",
            "brand_set",
            "youtube_channel_username",
            "youtube_channel_company",
        ]
    )
    sheet.append(
        [
            "Regular Title",
            "Media",
            "Feature",
            "Drama",
            "Drama",
            "Company",
            "Competitive View",
            "https://www.youtube.com/@Open%20AI",
            "https://www.youtube.com/@Company%20Channel",
        ]
    )
    sheet.append(
        [
            "Second Title",
            "Media",
            "Feature",
            "Drama",
            "Drama",
            "Company",
            "Competitive View",
            "https://www.youtube.com/@Open%7AI",
            "",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "youtube-percent-20.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    bad_cells = {issue.cell for issue in artifact.issues if issue.rule == "url_not_contains_if_present"}
    assert {"H2", "I2", "H3"}.issubset(bad_cells)
    assert "removing '%20' and replacing it with a space" in catalog["H2"].comment.text
    assert "removing '%20' and replacing it with a space" in catalog["I2"].comment.text
    assert "Remove %7 and add filter that we called pipe." in catalog["H3"].comment.text
    assert catalog["H2"].fill.start_color.rgb == "FFFCE4D6"
    assert catalog["I2"].fill.start_color.rgb == "FFFCE4D6"
    assert catalog["H3"].fill.start_color.rgb == "FFFCE4D6"


def test_youtube_multi_url_cells_check_each_entry(monkeypatch):
    checked_urls: list[str] = []

    def fake_fetch(url: str, platform: str, client=None):
        checked_urls.append(url)
        return True, ""

    monkeypatch.setattr(workbook_validator, "_fetch_social_reference", fake_fetch)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "companies",
            "brand_set",
            "youtube_channel_username",
        ]
    )
    sheet.append(
        [
            "OpenAI",
            "Media",
            "Feature",
            "Drama",
            "Drama",
            "Company",
            "Competitive View",
            "https://www.youtube.com/@OpenAI | https://www.youtube.com/watch?v=123",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "youtube-multi-url-format.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert any(issue.cell == "H2" and issue.rule == "social_reference_format" for issue in artifact.issues)
    assert "entry 2: video, shorts, and playlist URLs are not allowed" in catalog["H2"].comment.text
    assert checked_urls == [
        "https://www.youtube.com/@OpenAI",
        "https://www.youtube.com/watch?v=123",
    ]


def test_youtube_multi_url_cells_flag_unrelated_channels(monkeypatch):
    monkeypatch.setattr(workbook_validator, "_fetch_social_reference", lambda *args, **kwargs: (True, ""))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "companies",
            "brand_set",
            "youtube_channel_username",
        ]
    )
    sheet.append(
        [
            "OpenAI",
            "Media",
            "Feature",
            "Drama",
            "Drama",
            "Company",
            "Competitive View",
            "https://www.youtube.com/@OpenAI | https://www.youtube.com/@DifferentChannel",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "youtube-multi-url-title-match.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert any(issue.cell == "H2" and issue.rule == "social_reference_format" for issue in artifact.issues)
    assert "entry 2: does not appear related to title 'OpenAI'" in catalog["H2"].comment.text


def test_twitter_search_term_keywords_flags_value_errors_in_red():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "companies",
            "brand_set",
            "twitter_search_term_keywords",
        ]
    )
    sheet.append(
        [
            "Example Title",
            "Media",
            "Studio",
            "Competitive View",
            "#VALUE!",
        ]
    )
    sheet.append(
        [
            "Example Title 2",
            "Media",
            "Studio",
            "Competitive View",
            "openai ai",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "twitter-search-term-keywords.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    bad_cells = {issue.cell for issue in artifact.issues if issue.rule == "not_equals"}
    assert "E2" in bad_cells
    assert "E3" not in bad_cells
    assert "#VALUE!" in catalog["E2"].comment.text
    assert catalog["E2"].fill.start_color.rgb == workbook_validator.ERROR_FILL.start_color.rgb


def test_twitter_search_term_fields_are_required_for_movies_and_tv_shows():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "companies",
            "brand_set",
            "twitter_search_terms",
            "twitter_search_term_keywords",
        ]
    )
    sheet.append(
        [
            "Movie One",
            "Movies",
            "Feature",
            "Drama",
            "Drama",
            "Studio",
            "Competitive View",
            "",
            "",
        ]
    )
    sheet.append(
        [
            "Show One",
            "TV Shows",
            "Program Type - Scripted",
            "Drama",
            "Drama",
            "Studio",
            "Competitive View",
            "show one",
            "",
        ]
    )
    sheet.append(
        [
            "Media One",
            "Media",
            "Feature",
            "",
            "",
            "Studio",
            "Competitive View",
            "",
            "",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "twitter-search-terms-required.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    required_issues = {issue.cell for issue in artifact.issues if issue.rule == "not_blank_and_not_in"}
    assert {"H2", "I2", "I3"}.issubset(required_issues)
    assert "H3" not in required_issues
    assert "H4" not in required_issues
    assert "I4" not in required_issues
    assert "twitter_search_terms cannot be blank for Movies and TV Shows." in catalog["H2"].comment.text
    assert "twitter_search_term_keywords cannot be blank for Movies and TV Shows." in catalog["I2"].comment.text
    assert "twitter_search_term_keywords cannot be blank for Movies and TV Shows." in catalog["I3"].comment.text


def test_network_is_required_for_movies_and_tv_shows():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "companies",
            "brand_set",
            "network",
        ]
    )
    sheet.append(
        [
            "Movie One",
            "Movies",
            "Feature",
            "Drama",
            "Drama",
            "Studio",
            "Competitive View",
            "",
        ]
    )
    sheet.append(
        [
            "Show One",
            "TV Shows",
            "Program Type - Scripted",
            "Drama",
            "Drama",
            "Studio",
            "Competitive View",
            "#NA",
        ]
    )
    sheet.append(
        [
            "Movie Two",
            "Movies",
            "Feature",
            "Drama",
            "Drama",
            "Studio",
            "Competitive View",
            "Warner Bros.",
        ]
    )
    sheet.append(
        [
            "Media One",
            "Media",
            "Feature",
            "",
            "",
            "Studio",
            "Competitive View",
            "",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "network-required.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    required_issues = {issue.cell for issue in artifact.issues if issue.rule == "not_blank_and_not_in"}
    assert {"H2", "H3"}.issubset(required_issues)
    assert "H4" not in required_issues
    assert "H5" not in required_issues
    assert "Network cannot be blank for Movies and TV Shows." in catalog["H2"].comment.text
    assert "Network cannot be blank for Movies and TV Shows." in catalog["H3"].comment.text


def test_brand_set_requires_wide_release_data_feed_value_when_title_sub_category_marks_wide_release():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "companies",
            "brand_set",
        ]
    )
    sheet.append(
        [
            "Wide Title Missing Feed",
            "Movies",
            "Release - Wide\nStudio - Independent",
            "Studio",
            "Competitive View",
        ]
    )
    sheet.append(
        [
            "Wide Title Correct Feed",
            "Movies",
            "Release Type - Wide\nStudio - Independent",
            "Studio",
            "Competitive View, [Data Feed] Film - Wide Release + Custom Requests",
        ]
    )
    sheet.append(
        [
            "Wide Title - DAR",
            "Movies",
            "Release - Wide\nStudio - Independent",
            "Studio",
            "Pristine DAR Brands",
        ]
    )
    sheet.append(
        [
            "Wide Title Correct - DAR",
            "Movies",
            "Release Type - Wide\nStudio - Independent",
            "Studio",
            "Pristine DAR Brands, LF // Film - Majors + Independents",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "wide-brand-set.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    contains_issues = [issue for issue in artifact.issues if issue.rule == "contains" and issue.column == "E"]
    bad_cells = {issue.cell for issue in contains_issues}
    assert "E2" in bad_cells
    assert "E3" not in bad_cells
    assert "E4" in bad_cells
    assert "E5" not in bad_cells
    assert "[Data Feed] Film - Wide Release + Custom Requests" in catalog["E2"].comment.text
    assert "LF // Film - Majors + Independents" in catalog["E4"].comment.text
    assert catalog["E2"].fill.start_color.rgb == workbook_validator.ERROR_FILL.start_color.rgb


def test_brand_set_rules_cover_limited_movie_releases_for_dar_and_non_dar_titles():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "companies",
            "brand_set",
        ]
    )
    sheet.append(
        [
            "Limited Title Extra Brand",
            "Movies",
            "Release Type - Limited\nStudio - Independent",
            "Studio",
            "Competitive View, Extra Feed",
        ]
    )
    sheet.append(
        [
            "Limited Title Correct",
            "Movies",
            "Release - Limited\nStudio - Independent",
            "Studio",
            "Competitive View",
        ]
    )
    sheet.append(
        [
            "Limited Title - DAR",
            "Movies",
            "Release Type - Limited\nStudio - Independent",
            "Studio",
            "Pristine DAR Brands",
        ]
    )
    sheet.append(
        [
            "Limited Title Correct - DAR",
            "Movies",
            "Release - Limited\nStudio - Independent",
            "Studio",
            "Pristine DAR Brands, LF // Film - Majors + Independents",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "limited-brand-set.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    regex_issues = {issue.cell for issue in artifact.issues if issue.rule == "regex" and issue.column == "E"}
    contains_issues = {issue.cell for issue in artifact.issues if issue.rule == "contains" and issue.column == "E"}
    assert "E2" in regex_issues
    assert "E3" not in regex_issues
    assert "E4" in contains_issues
    assert "E5" not in contains_issues
    assert "only Competitive View" in catalog["E2"].comment.text
    assert "LF // Film - Majors + Independents" in catalog["E4"].comment.text


def test_movie_released_on_must_match_tmdb_us_release_date(monkeypatch):
    monkeypatch.setattr(
        workbook_validator,
        "_lookup_movie_metadata",
        lambda title: (
            True,
            {"us_release_date": "2025-07-04", "release_type": "Wide", "genres": ["Action", "Adventure"]},
            "",
        ),
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "released_on",
            "companies",
            "brand_set",
            "facebook_page",
        ]
    )
    sheet.append(
        [
            "Example Movie",
            "Movies",
            "Feature",
            "Drama",
            "Drama",
            "July-03-2025",
            "Studio",
            "Competitive View",
            "",
        ]
    )
    sheet.append(
        [
            "Example Movie 2",
            "Movies",
            "Feature",
            "Drama",
            "Drama",
            "July-04-2025",
            "Studio",
            "Competitive View",
            "",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "movies.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    bad_cells = {issue.cell for issue in artifact.issues if issue.rule == "movie_us_release_date_match"}
    assert "F2" in bad_cells
    assert "F3" not in bad_cells
    assert "TMDB USA release date: 2025-07-04" in catalog["F2"].comment.text


def test_movie_released_on_accepts_common_non_month_day_year_formats(monkeypatch):
    monkeypatch.setattr(
        workbook_validator,
        "_lookup_movie_metadata",
        lambda title: (
            True,
            {"us_release_date": "2025-07-04", "release_type": "Wide", "genres": ["Action", "Adventure"]},
            "",
        ),
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "released_on",
            "companies",
            "brand_set",
        ]
    )
    sheet.append(
        [
            "Slash Format Movie",
            "Movies",
            "Feature",
            "Drama",
            "Drama",
            "07/04/2025",
            "Studio",
            "Competitive View",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "movie-slash-release-date.xlsx", rules)

    assert not any(issue.cell == "F2" and issue.rule == "movie_us_release_date_match" for issue in artifact.issues)


def test_movie_release_type_and_genre_add_tmdb_recommendations_to_comments(monkeypatch):
    monkeypatch.setattr(
        workbook_validator,
        "_lookup_movie_metadata",
        lambda title: (
            True,
            {"us_release_date": "2025-07-04", "release_type": "Wide", "genres": ["Action", "Adventure"]},
            "",
        ),
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "released_on",
            "release_type",
            "companies",
            "brand_set",
            "facebook_page",
        ]
    )
    sheet.append(
        [
            "Example Movie",
            "Movies",
            "Feature",
            "Comedy",
            "Comedy",
            "July-04-2025",
            "Limited",
            "Studio",
            "Competitive View",
            "",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "movie-metadata.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert any(issue.cell == "D2" and issue.rule == "movie_genre_match" for issue in artifact.issues)
    assert any(issue.cell == "G2" and issue.rule == "movie_release_type_match" for issue in artifact.issues)
    assert catalog["D2"].fill.fill_type == "solid"
    assert catalog["G2"].fill.fill_type == "solid"
    assert "TMDB genres: Action, Adventure" in catalog["D2"].comment.text
    assert "Wide" in catalog["G2"].comment.text


def test_movie_genre_partial_overlap_still_flags_tmdb_recommendation(monkeypatch):
    monkeypatch.setattr(
        workbook_validator,
        "_lookup_movie_metadata",
        lambda title: (
            True,
            {"us_release_date": "2025-07-04", "release_type": "Wide", "genres": ["Action", "Adventure"]},
            "",
        ),
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "released_on",
            "release_type",
            "companies",
            "brand_set",
        ]
    )
    sheet.append(
        [
            "Example Movie",
            "Movies",
            "Feature",
            "Action, Comedy",
            "Action",
            "July-04-2025",
            "Wide",
            "Studio",
            "Competitive View",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "movie-genre-partial-overlap.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert any(issue.cell == "D2" and issue.rule == "movie_genre_match" for issue in artifact.issues)
    assert "TMDB genres: Action, Adventure" in catalog["D2"].comment.text


def test_rottentomatoes_url_mismatch_adds_correct_movie_url(monkeypatch):
    def fake_network_get(url, client=None, params=None, headers=None):
        if "search" in url:
            html = _build_rottentomatoes_search_html(
                "movie",
                [
                    {
                        "title": "Inception",
                        "url": "https://www.rottentomatoes.com/m/inception",
                        "release-year": "2010",
                    },
                    {
                        "title": "Inception: The Cobol Job",
                        "url": "https://www.rottentomatoes.com/m/inception_the_cobol_job",
                        "release-year": "2010",
                    },
                ],
            )
            return httpx.Response(200, request=httpx.Request("GET", url), text=html)
        return httpx.Response(
            200,
            request=httpx.Request("GET", url),
            text=_build_rottentomatoes_movie_page_html("Inception", "https://www.rottentomatoes.com/m/inception", "2010-07-16"),
        )

    monkeypatch.setattr(workbook_validator, "_network_get", fake_network_get)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "released_on",
            "release_type",
            "companies",
            "brand_set",
            "rottentomatoes",
        ]
    )
    sheet.append(
        [
            "Inception",
            "Movies",
            "Feature",
            "Sci-Fi",
            "Sci-Fi",
            "July-16-2010",
            "Wide",
            "Studio",
            "Competitive View",
            "https://www.rottentomatoes.com/m/not_inception",
        ]
    )
    sheet.append(
        [
            "Inception",
            "Movies",
            "Feature",
            "Sci-Fi",
            "Sci-Fi",
            "July-16-2010",
            "Wide",
            "Studio",
            "Competitive View",
            "https://www.rottentomatoes.com/m/inception",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = _build_rules_with_rottentomatoes_check()
    artifact = validate_workbook(buffer.getvalue(), "rottentomatoes-movies.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert any(issue.cell == "J2" and issue.rule == "rottentomatoes_url_match" for issue in artifact.issues)
    assert not any(issue.cell == "J3" and issue.rule == "rottentomatoes_url_match" for issue in artifact.issues)
    assert "Rotten Tomatoes URL: https://www.rottentomatoes.com/m/inception" in catalog["J2"].comment.text
    assert catalog["J2"].fill.fill_type == "solid"
    assert catalog["J3"].fill.fill_type != "solid"


def test_rottentomatoes_lookup_uses_release_year_to_choose_best_match(monkeypatch):
    def fake_network_get(url, client=None, params=None, headers=None):
        if "search" in url:
            html = _build_rottentomatoes_search_html(
                "movie",
                [
                    {
                        "title": "Gladiator",
                        "url": "https://www.rottentomatoes.com/m/gladiator",
                        "release-year": "2000",
                    },
                    {
                        "title": "Gladiator",
                        "url": "https://www.rottentomatoes.com/m/gladiator_2024",
                        "release-year": "2024",
                    },
                ],
            )
            return httpx.Response(200, request=httpx.Request("GET", url), text=html)
        return httpx.Response(
            200,
            request=httpx.Request("GET", url),
            text=_build_rottentomatoes_movie_page_html("Gladiator", "https://www.rottentomatoes.com/m/gladiator_2024", "2024-11-22"),
        )

    monkeypatch.setattr(workbook_validator, "_network_get", fake_network_get)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "released_on",
            "release_type",
            "companies",
            "brand_set",
            "rottentomatoes",
        ]
    )
    sheet.append(
        [
            "Gladiator",
            "Movies",
            "Feature",
            "Action",
            "Action",
            "November-22-2024",
            "Wide",
            "Studio",
            "Competitive View",
            "",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = _build_rules_with_rottentomatoes_check()
    artifact = validate_workbook(buffer.getvalue(), "rottentomatoes-year-match.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert any(issue.cell == "J2" and issue.rule == "rottentomatoes_url_match" for issue in artifact.issues)
    assert "https://www.rottentomatoes.com/m/gladiator_2024" in catalog["J2"].comment.text


def test_rottentomatoes_ignores_tv_urls_and_only_uses_movie_urls(monkeypatch):
    def fake_network_get(url, client=None, params=None, headers=None):
        if "search" in url:
            html = _build_rottentomatoes_search_html(
                "movie",
                [
                    {
                        "title": "Rock the Block",
                        "url": "https://www.rottentomatoes.com/tv/rock_the_block",
                        "release-year": "2026",
                    },
                    {
                        "title": "Rock the Block",
                        "url": "https://www.rottentomatoes.com/m/rock_the_block_2026",
                        "release-year": "2026",
                    },
                ],
            )
            return httpx.Response(200, request=httpx.Request("GET", url), text=html)
        return httpx.Response(
            200,
            request=httpx.Request("GET", url),
            text=_build_rottentomatoes_movie_page_html(
                "Rock the Block",
                "https://www.rottentomatoes.com/m/rock_the_block_2026",
                "2026-01-10",
            ),
        )

    monkeypatch.setattr(workbook_validator, "_network_get", fake_network_get)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "released_on",
            "release_type",
            "companies",
            "brand_set",
            "rottentomatoes",
        ]
    )
    sheet.append(
        [
            "Rock the Block",
            "Movies",
            "Feature",
            "Documentary",
            "Documentary",
            "January-10-2026",
            "Wide",
            "Studio",
            "Competitive View",
            "https://www.rottentomatoes.com/tv/rock_the_block",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = _build_rules_with_rottentomatoes_check()
    artifact = validate_workbook(buffer.getvalue(), "rottentomatoes-ignore-tv.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert any(issue.cell == "J2" and issue.rule == "rottentomatoes_url_match" for issue in artifact.issues)
    assert "Rotten Tomatoes URL: https://www.rottentomatoes.com/m/rock_the_block_2026" in catalog["J2"].comment.text
    assert "/tv/rock_the_block" not in catalog["J2"].comment.text


def test_rottentomatoes_does_not_fall_back_to_nearest_year(monkeypatch):
    def fake_network_get(url, client=None, params=None, headers=None):
        if "search" in url:
            html = _build_rottentomatoes_search_html(
                "movie",
                [
                    {
                        "title": "Gladiator",
                        "url": "https://www.rottentomatoes.com/m/gladiator",
                        "release-year": "2000",
                    },
                    {
                        "title": "Gladiator",
                        "url": "https://www.rottentomatoes.com/m/gladiator_2024",
                        "release-year": "2024",
                    },
                ],
            )
            return httpx.Response(200, request=httpx.Request("GET", url), text=html)
        raise AssertionError("movie detail page should not be fetched when no matching release year exists")

    monkeypatch.setattr(workbook_validator, "_network_get", fake_network_get)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "released_on",
            "release_type",
            "companies",
            "brand_set",
            "rottentomatoes",
        ]
    )
    sheet.append(
        [
            "Gladiator",
            "Movies",
            "Feature",
            "Action",
            "Action",
            "November-22-2025",
            "Wide",
            "Studio",
            "Competitive View",
            "",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = _build_rules_with_rottentomatoes_check()
    artifact = validate_workbook(buffer.getvalue(), "rottentomatoes-no-nearest-year.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert any(issue.cell == "J2" and issue.rule == "rottentomatoes_url_match" for issue in artifact.issues)
    assert "Rotten Tomatoes title not found for release year 2025" in catalog["J2"].comment.text


def test_rottentomatoes_requires_release_year_to_verify_url(monkeypatch):
    def fail_network_get(*args, **kwargs):
        raise AssertionError("Rotten Tomatoes lookup should not run without a valid release year")

    monkeypatch.setattr(workbook_validator, "_network_get", fail_network_get)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "released_on",
            "release_type",
            "companies",
            "brand_set",
            "rottentomatoes",
        ]
    )
    sheet.append(
        [
            "Inception",
            "Movies",
            "Feature",
            "Sci-Fi",
            "Sci-Fi",
            "",
            "Wide",
            "Studio",
            "Competitive View",
            "https://www.rottentomatoes.com/m/inception",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = _build_rules_with_rottentomatoes_check()
    artifact = validate_workbook(buffer.getvalue(), "rottentomatoes-missing-year.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert any(issue.cell == "J2" and issue.rule == "rottentomatoes_url_match" for issue in artifact.issues)
    assert "released_on must include a valid year" in catalog["J2"].comment.text
    assert catalog["J2"].fill.fill_type == "solid"


def test_movie_tmdb_connectivity_errors_do_not_false_flag_cells(monkeypatch):
    monkeypatch.setattr(
        workbook_validator,
        "_lookup_movie_metadata",
        lambda title: (False, None, "TMDB lookup failed: ConnectError"),
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "released_on",
            "release_type",
            "companies",
            "brand_set",
        ]
    )
    sheet.append(
        [
            "Example Movie - DAR",
            "Movies",
            "Feature",
            "Action",
            "Action",
            "July-04-2025",
            "Wide",
            "Studio",
            "Pristine DAR Brands",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "movie-connectivity.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert not any(issue.rule == "movie_us_release_date_match" for issue in artifact.issues)
    assert not any(issue.rule == "movie_release_type_match" for issue in artifact.issues)
    assert not any(issue.rule == "movie_genre_match" for issue in artifact.issues)
    assert catalog["D2"].fill.fill_type != "solid"
    assert catalog["F2"].fill.fill_type != "solid"
    assert catalog["G2"].fill.fill_type != "solid"


def test_rottentomatoes_connectivity_errors_do_not_false_flag_cells(monkeypatch):
    def fake_network_get(url, client=None, params=None, headers=None):
        request = httpx.Request("GET", url)
        raise httpx.ConnectError("boom", request=request)

    monkeypatch.setattr(workbook_validator, "_network_get", fake_network_get)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "released_on",
            "release_type",
            "companies",
            "brand_set",
            "rottentomatoes",
        ]
    )
    sheet.append(
        [
            "Inception",
            "Movies",
            "Feature",
            "Sci-Fi",
            "Sci-Fi",
            "July-16-2010",
            "Wide",
            "Studio",
            "Competitive View",
            "https://www.rottentomatoes.com/m/not_inception",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = _build_rules_with_rottentomatoes_check()
    artifact = validate_workbook(buffer.getvalue(), "rottentomatoes-connectivity.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert not any(issue.rule == "rottentomatoes_url_match" for issue in artifact.issues)
    assert catalog["J2"].fill.fill_type != "solid"


def test_rottentomatoes_rejects_loose_title_match_after_detail_page_verification(monkeypatch):
    def fake_network_get(url, client=None, params=None, headers=None):
        if "search" in url:
            html = _build_rottentomatoes_search_html(
                "movie",
                [
                    {
                        "title": "Avatar: The Last Airbender",
                        "url": "https://www.rottentomatoes.com/m/avatar_the_last_airbender",
                        "release-year": "2024",
                    }
                ],
            )
            return httpx.Response(200, request=httpx.Request("GET", url), text=html)
        return httpx.Response(
            200,
            request=httpx.Request("GET", url),
            text=_build_rottentomatoes_movie_page_html(
                "Avatar: The Last Airbender",
                "https://www.rottentomatoes.com/m/avatar_the_last_airbender",
                "2024-01-01",
            ),
        )

    monkeypatch.setattr(workbook_validator, "_network_get", fake_network_get)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "released_on",
            "release_type",
            "companies",
            "brand_set",
            "rottentomatoes",
        ]
    )
    sheet.append(
        [
            "Avatar",
            "Movies",
            "Feature",
            "Fantasy",
            "Fantasy",
            "January-01-2024",
            "Wide",
            "Studio",
            "Competitive View",
            "",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = _build_rules_with_rottentomatoes_check()
    artifact = validate_workbook(buffer.getvalue(), "rottentomatoes-title-verification.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert any(issue.cell == "J2" and issue.rule == "rottentomatoes_url_match" for issue in artifact.issues)
    assert "Rotten Tomatoes page title mismatch: Avatar: The Last Airbender" in catalog["J2"].comment.text


def test_rottentomatoes_rejects_detail_page_year_mismatch(monkeypatch):
    def fake_network_get(url, client=None, params=None, headers=None):
        if "search" in url:
            html = _build_rottentomatoes_search_html(
                "movie",
                [
                    {
                        "title": "Inception",
                        "url": "https://www.rottentomatoes.com/m/inception",
                        "release-year": "2010",
                    }
                ],
            )
            return httpx.Response(200, request=httpx.Request("GET", url), text=html)
        return httpx.Response(
            200,
            request=httpx.Request("GET", url),
            text=_build_rottentomatoes_movie_page_html("Inception", "https://www.rottentomatoes.com/m/inception", "2011-07-16"),
        )

    monkeypatch.setattr(workbook_validator, "_network_get", fake_network_get)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "released_on",
            "release_type",
            "companies",
            "brand_set",
            "rottentomatoes",
        ]
    )
    sheet.append(
        [
            "Inception",
            "Movies",
            "Feature",
            "Sci-Fi",
            "Sci-Fi",
            "July-16-2010",
            "Wide",
            "Studio",
            "Competitive View",
            "",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = _build_rules_with_rottentomatoes_check()
    artifact = validate_workbook(buffer.getvalue(), "rottentomatoes-year-verification.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert any(issue.cell == "J2" and issue.rule == "rottentomatoes_url_match" for issue in artifact.issues)
    assert "Rotten Tomatoes page year mismatch: 2011" in catalog["J2"].comment.text


def test_movie_metadata_lookup_strips_dar_suffix_before_tmdb_search(monkeypatch):
    looked_up_titles: list[str] = []

    def fake_lookup(title: str):
        looked_up_titles.append(title)
        return True, {"us_release_date": "2025-07-04", "release_type": "Wide", "genres": ["Action"]}, ""

    monkeypatch.setattr(workbook_validator, "_lookup_movie_metadata", fake_lookup)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "released_on",
            "release_type",
            "companies",
            "brand_set",
        ]
    )
    sheet.append(
        [
            "Example Movie - DAR",
            "Movies",
            "Feature",
            "Action",
            "Action",
            "July-04-2025",
            "Wide",
            "Studio",
            "Pristine DAR Brands",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "movie-dar-title.xlsx", rules)

    assert looked_up_titles == ["Example Movie"]
    assert not any(issue.rule == "movie_us_release_date_match" for issue in artifact.issues)


def test_lookup_movie_metadata_prefers_matching_tmdb_result(monkeypatch):
    class DummyTmdbClient:
        headers: dict[str, str] = {}

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return None

    def fake_tmdb_get(client, path, params=None):
        if path == "/search/movie":
            return {
                "results": [
                    {
                        "id": 11,
                        "title": "Completely Different Movie",
                        "original_title": "Completely Different Movie",
                        "release_date": "2025-07-04",
                        "popularity": 80,
                    },
                    {
                        "id": 22,
                        "title": "Example Movie",
                        "original_title": "Example Movie",
                        "release_date": "2025-07-04",
                        "popularity": 10,
                    },
                ]
            }
        if path == "/movie/22":
            return {"release_date": "2025-07-04", "genres": [{"name": "Action"}]}
        if path == "/movie/22/release_dates":
            return {
                "results": [
                    {
                        "iso_3166_1": "US",
                        "release_dates": [{"release_date": "2025-07-04T00:00:00.000Z", "type": 3}],
                    }
                ]
            }
        raise AssertionError(f"unexpected TMDB path: {path}")

    monkeypatch.setattr(workbook_validator.settings, "tmdb_api_key", "test-key")
    monkeypatch.setattr(workbook_validator.settings, "tmdb_read_access_token", "")
    monkeypatch.setattr(workbook_validator, "_build_tmdb_http_client", lambda: DummyTmdbClient())
    monkeypatch.setattr(workbook_validator, "_tmdb_get", fake_tmdb_get)

    success, metadata, detail = workbook_validator._lookup_movie_metadata("Example Movie", release_year=2025)

    assert success is True
    assert detail == ""
    assert metadata is not None
    assert metadata["release_date"] == "2025-07-04"
    assert metadata["release_type"] == "Wide"


def test_movie_released_on_uses_tmdb_general_release_date_when_us_date_missing(monkeypatch):
    monkeypatch.setattr(
        workbook_validator,
        "_lookup_movie_metadata",
        lambda title, release_year=None: (
            True,
            {
                "release_date": "2025-07-04",
                "release_date_source": "global",
                "release_type": "",
                "genres": ["Action", "Adventure"],
            },
            "",
        ),
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "released_on",
            "companies",
            "brand_set",
        ]
    )
    sheet.append(
        [
            "Example Movie",
            "Movies",
            "Feature",
            "Drama",
            "Drama",
            "July-03-2025",
            "Studio",
            "Competitive View",
        ]
    )
    sheet.append(
        [
            "Example Movie 2",
            "Movies",
            "Feature",
            "Drama",
            "Drama",
            "July-04-2025",
            "Studio",
            "Competitive View",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "movie-global-release-date.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    bad_cells = {issue.cell for issue in artifact.issues if issue.rule == "movie_us_release_date_match"}
    assert "F2" in bad_cells
    assert "F3" not in bad_cells
    assert "TMDB release date: 2025-07-04" in catalog["F2"].comment.text


def test_title_category_combined_values_are_allowed_and_media_sub_category_can_be_blank():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "companies",
            "brand_set",
            "facebook_page",
        ]
    )
    sheet.append(
        [
            "Venue Title",
            "Venues, Events & Attractions",
            "",
            "",
            "",
            "Company",
            "Competitive View",
            "",
        ]
    )
    sheet.append(
        [
            "Fitness Title",
            "Health, Wellness, Fitness",
            "",
            "",
            "",
            "Company",
            "Competitive View",
            "",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "combined-categories.xlsx", rules)

    title_category_issues = {issue.cell for issue in artifact.issues if issue.rule == "in"}
    title_subcategory_issues = {issue.cell for issue in artifact.issues if issue.rule == "not_blank_and_not_in"}

    assert "B2" not in title_category_issues
    assert "B3" not in title_category_issues
    assert "C2" not in title_subcategory_issues
    assert "C3" not in title_subcategory_issues


def test_default_rules_validate_wikidata_and_imdb(monkeypatch):
    lookup_calls: list[str] = []

    def fake_lookup(platform, raw_value, client):
        lookup_calls.append(platform)
        if platform == "wikidata":
            return (
                True,
                {
                    "title": "Correct Title",
                    "alternate_titles": ["Correct Title Film"],
                    "wikidata_id": "Q123",
                    "type": "movie",
                    "url": "https://www.wikidata.org/wiki/Q123",
                    "wikipedia_url": "https://en.wikipedia.org/wiki/Correct_Title",
                    "wikidata_url": "https://www.wikidata.org/wiki/Q123",
                },
                "",
            )
        if platform == "imdb":
            return True, {"title": "Wrong IMDb Title", "type": "movie", "id": "tt1234567"}, ""
        raise AssertionError("unexpected platform")

    monkeypatch.setattr(workbook_validator, "_lookup_reference_record", fake_lookup)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "imdb_id",
            "wikidata_id",
            "companies",
            "brand_set",
        ]
    )
    sheet.append(
        [
            "Correct Title",
            "Movies",
            "Feature",
            "Drama",
            "Drama",
            "tt1234567",
            "Q123",
            "Studio",
            "Competitive View",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "references.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert not any(issue.cell == "G2" and issue.rule == "reference_lookup_match" for issue in artifact.issues)
    assert any(issue.cell == "F2" and issue.rule == "reference_lookup_match" for issue in artifact.issues)
    assert "IMDb title: Wrong IMDb Title" in catalog["F2"].comment.text
    assert lookup_calls == ["wikidata", "imdb"]


def test_legacy_wikipedia_rules_are_skipped(monkeypatch):
    def fail_lookup(*args, **kwargs):
        raise AssertionError("legacy Wikipedia rules should not trigger live lookups")

    monkeypatch.setattr(workbook_validator, "_lookup_reference_record", fail_lookup)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(["title", "wikipedia_page", "wikidata_id"])
    sheet.append(["Correct Title", "https://en.wikipedia.org/wiki/Correct_Title", "Q123"])

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(
        """
        {
          "rules": [
            {"sheet": "*", "column": "wikipedia_page", "check": "social_reference_format", "platform": "wikipedia"},
            {"sheet": "*", "column": "wikipedia_page", "check": "reference_lookup_match", "platform": "wikipedia"}
          ]
        }
        """
    )
    artifact = validate_workbook(buffer.getvalue(), "legacy-wikipedia.xlsx", rules)

    assert not any(issue.cell == "B2" and issue.rule == "reference_lookup_match" for issue in artifact.issues)
    assert not any(issue.cell == "B2" and issue.rule == "social_reference_format" for issue in artifact.issues)


def test_imdb_lookup_reads_official_title_and_name_datasets(tmp_path, monkeypatch):
    title_dataset = tmp_path / "title.basics.tsv.gz"
    name_dataset = tmp_path / "name.basics.tsv.gz"
    dataset_cache_dir = tmp_path / "cache"

    _write_gzipped_tsv(
        title_dataset,
        ["tconst", "titleType", "primaryTitle", "originalTitle", "isAdult", "startYear", "endYear", "runtimeMinutes", "genres"],
        [["tt1234567", "tvSeries", "Localized Title", "Original Show", "0", "2024", r"\N", "45", "Drama"]],
    )
    _write_gzipped_tsv(
        name_dataset,
        ["nconst", "primaryName", "birthYear", "deathYear", "primaryProfession", "knownForTitles"],
        [["nm7654321", "Jane Example", "1980", r"\N", "actor", "tt1234567"]],
    )

    monkeypatch.setattr(workbook_validator.settings, "imdb_title_basics_url", str(title_dataset))
    monkeypatch.setattr(workbook_validator.settings, "imdb_name_basics_url", str(name_dataset))
    monkeypatch.setattr(workbook_validator.settings, "imdb_dataset_dir", str(dataset_cache_dir))
    monkeypatch.setattr(workbook_validator.settings, "omdb_api_key", "")

    def fail_network(*args, **kwargs):
        raise AssertionError("network fallback should not be used when dataset lookup succeeds")

    monkeypatch.setattr(workbook_validator, "_network_get", fail_network)

    title_success, title_metadata, title_detail = _lookup_imdb_record("tt1234567", client=None)
    name_success, name_metadata, name_detail = _lookup_imdb_record("nm7654321", client=None)

    assert title_success is True
    assert title_detail == ""
    assert title_metadata is not None
    assert title_metadata["title"] == "Localized Title"
    assert title_metadata["alternate_titles"] == ["Original Show"]
    assert title_metadata["type"] == "series"

    assert name_success is True
    assert name_detail == ""
    assert name_metadata is not None
    assert name_metadata["title"] == "Jane Example"
    assert name_metadata["type"] == "person"


def test_validate_workbook_uses_imdb_dataset_for_reference_matching(tmp_path, monkeypatch):
    title_dataset = tmp_path / "title.basics.tsv.gz"
    name_dataset = tmp_path / "name.basics.tsv.gz"
    dataset_cache_dir = tmp_path / "cache"

    _write_gzipped_tsv(
        title_dataset,
        ["tconst", "titleType", "primaryTitle", "originalTitle", "isAdult", "startYear", "endYear", "runtimeMinutes", "genres"],
        [["tt1234567", "movie", "Localized Title", "Original Title", "0", "2024", r"\N", "120", "Drama"]],
    )
    _write_gzipped_tsv(
        name_dataset,
        ["nconst", "primaryName", "birthYear", "deathYear", "primaryProfession", "knownForTitles"],
        [],
    )

    monkeypatch.setattr(workbook_validator.settings, "imdb_title_basics_url", str(title_dataset))
    monkeypatch.setattr(workbook_validator.settings, "imdb_name_basics_url", str(name_dataset))
    monkeypatch.setattr(workbook_validator.settings, "imdb_dataset_dir", str(dataset_cache_dir))
    monkeypatch.setattr(workbook_validator.settings, "omdb_api_key", "")

    def fail_network(*args, **kwargs):
        raise AssertionError("network fallback should not be used when dataset lookup succeeds")

    monkeypatch.setattr(workbook_validator, "_network_get", fail_network)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "imdb_id",
            "companies",
            "brand_set",
        ]
    )
    sheet.append(
        [
            "Original Title",
            "Movies",
            "Feature",
            "Drama",
            "Drama",
            "tt1234567",
            "Studio",
            "Competitive View",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "imdb-dataset.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert not any(issue.cell == "F2" and issue.rule == "reference_lookup_match" for issue in artifact.issues)
    assert catalog["F2"].fill.fill_type != "solid"


def test_imdb_lookup_reuses_existing_stale_index_when_rebuild_is_disabled(tmp_path, monkeypatch):
    dataset_dir = tmp_path / "cache"
    dataset_dir.mkdir()
    db_path = dataset_dir / "imdb_basics.sqlite3"
    title_path = dataset_dir / "title.basics.tsv.gz"
    name_path = dataset_dir / "name.basics.tsv.gz"

    db_path.write_bytes(b"existing index")
    _write_gzipped_tsv(
        title_path,
        ["tconst", "titleType", "primaryTitle", "originalTitle", "isAdult", "startYear", "endYear", "runtimeMinutes", "genres"],
        [["tt1234567", "movie", "Localized Title", "Original Title", "0", "2024", r"\N", "120", "Drama"]],
    )
    _write_gzipped_tsv(
        name_path,
        ["nconst", "primaryName", "birthYear", "deathYear", "primaryProfession", "knownForTitles"],
        [],
    )

    future_timestamp = title_path.stat().st_mtime + 60
    title_path.touch()
    name_path.touch()
    monkeypatch.setattr(workbook_validator.settings, "imdb_dataset_dir", str(dataset_dir))
    monkeypatch.setattr(workbook_validator.settings, "imdb_rebuild_stale_index", False)
    monkeypatch.setattr(workbook_validator.settings, "imdb_title_basics_url", str(title_path))
    monkeypatch.setattr(workbook_validator.settings, "imdb_name_basics_url", str(name_path))
    db_path.touch()
    db_timestamp = db_path.stat().st_mtime - 120
    os.utime(db_path, (db_timestamp, db_timestamp))
    os.utime(title_path, (future_timestamp, future_timestamp))
    os.utime(name_path, (future_timestamp, future_timestamp))

    def fail_rebuild(*args, **kwargs):
        raise AssertionError("stale index should be reused during validation")

    monkeypatch.setattr(workbook_validator, "_build_imdb_dataset_index", fail_rebuild)

    resolved_path = workbook_validator._ensure_imdb_dataset_index()

    assert resolved_path == db_path
    assert resolved_path.read_bytes() == b"existing index"


def test_wikipedia_lookup_resolves_redirects_and_aliases_via_wikidata(monkeypatch):
    def fake_network_get(url, client=None, params=None, headers=None):
        request = httpx.Request("GET", url)
        if "en.wikipedia.org/w/api.php" in url:
            return httpx.Response(
                200,
                request=request,
                json={
                    "query": {
                        "redirects": [{"from": "Correct_Title_(film)", "to": "Correct Title"}],
                        "pages": [
                            {
                                "title": "Correct Title",
                                "canonicalurl": "https://en.wikipedia.org/wiki/Correct_Title",
                                "pageprops": {"wikibase_item": "Q123"},
                            }
                        ],
                    }
                },
            )
        if "www.wikidata.org/w/api.php" in url:
            return httpx.Response(
                200,
                request=request,
                json={
                    "entities": {
                        "Q123": {
                            "labels": {"en": {"value": "Correct Title"}},
                            "aliases": {"en": [{"value": "Correct Title Film"}]},
                            "sitelinks": {"enwiki": {"title": "Correct Title"}},
                            "claims": {"P31": [{"mainsnak": {"datavalue": {"value": {"id": "Q11424"}}}}]},
                        }
                    }
                },
            )
        raise AssertionError(f"unexpected url: {url}")

    monkeypatch.setattr(workbook_validator, "_network_get", fake_network_get)

    success, metadata, detail = workbook_validator._lookup_wikipedia_record(
        "http://en.wikipedia.org/wiki/Correct_Title_(film)",
        client=None,
    )

    assert success is True
    assert detail == ""
    assert metadata is not None
    assert metadata["title"] == "Correct Title"
    assert metadata["wikidata_id"] == "Q123"
    assert metadata["redirect_source"] == "Correct_Title_(film)"
    assert "Correct Title Film" in metadata["alternate_titles"]


def test_wikipedia_lookup_uses_persistent_cache_when_fresh(tmp_path, monkeypatch):
    cache_dir = tmp_path / "wiki-cache"
    monkeypatch.setattr(workbook_validator.settings, "wikipedia_cache_dir", str(cache_dir))
    monkeypatch.setattr(workbook_validator.settings, "wikipedia_refresh_hours", 24)

    def api_success(url, client=None, params=None, headers=None):
        request = httpx.Request("GET", url)
        if "en.wikipedia.org/w/api.php" in url:
            return httpx.Response(
                200,
                request=request,
                json={
                    "query": {
                        "pages": [
                            {
                                "title": "Cached Topic",
                                "canonicalurl": "https://en.wikipedia.org/wiki/Cached_Topic",
                                "pageprops": {"wikibase_item": "Q456"},
                            }
                        ]
                    }
                },
            )
        if "www.wikidata.org/w/api.php" in url:
            return httpx.Response(
                200,
                request=request,
                json={
                    "entities": {
                        "Q456": {
                            "labels": {"en": {"value": "Cached Topic"}},
                            "aliases": {"en": [{"value": "Cached Topic Alias"}]},
                            "sitelinks": {"enwiki": {"title": "Cached Topic"}},
                            "claims": {"P31": [{"mainsnak": {"datavalue": {"value": {"id": "Q5"}}}}]},
                        }
                    }
                },
            )
        raise AssertionError("unexpected request")

    monkeypatch.setattr(workbook_validator, "_network_get", api_success)

    success, metadata, detail = _lookup_wikipedia_record("Cached Topic", client=None)
    assert success is True
    assert detail == ""
    assert metadata is not None
    assert metadata["title"] == "Cached Topic"

    def no_network(*args, **kwargs):
        raise AssertionError("fresh Wikipedia cache should avoid live network requests")

    monkeypatch.setattr(workbook_validator, "_network_get", no_network)

    cached_success, cached_metadata, cached_detail = _lookup_wikipedia_record("Cached Topic", client=None)
    assert cached_success is True
    assert cached_detail == ""
    assert cached_metadata is not None
    assert cached_metadata["title"] == "Cached Topic"
    assert cached_metadata["wikidata_id"] == "Q456"


def test_wikipedia_lookup_returns_stale_success_cache_when_live_refresh_fails(tmp_path, monkeypatch):
    cache_dir = tmp_path / "wiki-cache"
    monkeypatch.setattr(workbook_validator.settings, "wikipedia_cache_dir", str(cache_dir))
    monkeypatch.setattr(workbook_validator.settings, "wikipedia_refresh_hours", 1)

    def api_success(url, client=None, params=None, headers=None):
        request = httpx.Request("GET", url)
        if "en.wikipedia.org/w/api.php" in url:
            return httpx.Response(
                200,
                request=request,
                json={
                    "query": {
                        "pages": [
                            {
                                "title": "Stable Topic",
                                "canonicalurl": "https://en.wikipedia.org/wiki/Stable_Topic",
                                "pageprops": {"wikibase_item": "Q789"},
                            }
                        ]
                    }
                },
            )
        if "www.wikidata.org/w/api.php" in url:
            return httpx.Response(
                200,
                request=request,
                json={
                    "entities": {
                        "Q789": {
                            "labels": {"en": {"value": "Stable Topic"}},
                            "aliases": {"en": [{"value": "Stable Topic Alias"}]},
                            "sitelinks": {"enwiki": {"title": "Stable Topic"}},
                            "claims": {"P31": [{"mainsnak": {"datavalue": {"value": {"id": "Q11424"}}}}]},
                        }
                    }
                },
            )
        raise AssertionError("unexpected request")

    monkeypatch.setattr(workbook_validator, "_network_get", api_success)
    initial_success, initial_metadata, initial_detail = _lookup_wikipedia_record("Stable Topic", client=None)
    assert initial_success is True
    assert initial_detail == ""
    assert initial_metadata is not None

    db_path = cache_dir / "wikipedia_cache.sqlite3"
    with sqlite3.connect(db_path) as connection:
        connection.execute("UPDATE wikimedia_cache SET checked_at = 1 WHERE lookup_key = ?", ("wikipedia:stable_topic",))
        connection.commit()

    def api_failure(url, client=None, params=None, headers=None):
        request = httpx.Request("GET", url)
        raise httpx.ConnectError("boom", request=request)

    monkeypatch.setattr(workbook_validator, "_network_get", api_failure)

    stale_success, stale_metadata, stale_detail = _lookup_wikipedia_record("Stable Topic", client=None)
    assert stale_success is True
    assert stale_detail == ""
    assert stale_metadata is not None
    assert stale_metadata["title"] == "Stable Topic"


def test_wikipedia_lookup_rejects_disambiguation_pages(monkeypatch):
    def fake_network_get(url, client=None, params=None, headers=None):
        return httpx.Response(
            200,
            request=httpx.Request("GET", url),
            json={
                "query": {
                    "pages": [
                        {
                            "title": "Avatar (disambiguation)",
                            "canonicalurl": "https://en.wikipedia.org/wiki/Avatar_(disambiguation)",
                            "pageprops": {"disambiguation": "", "wikibase_item": "Q4167410"},
                        }
                    ]
                }
            },
        )

    monkeypatch.setattr(workbook_validator, "_network_get", fake_network_get)

    success, metadata, detail = _lookup_wikipedia_record("Avatar (disambiguation)", client=None)

    assert success is False
    assert metadata is None
    assert detail == "Wikipedia page is a disambiguation page"


def test_wikidata_lookup_accepts_qid_and_uses_aliases(monkeypatch):
    def fake_network_get(url, client=None, params=None, headers=None):
        return httpx.Response(
            200,
            request=httpx.Request("GET", url),
            json={
                "entities": {
                    "Q42": {
                        "labels": {"en": {"value": "Douglas Adams"}},
                        "aliases": {"en": [{"value": "Douglas Noel Adams"}]},
                        "sitelinks": {"enwiki": {"title": "Douglas Adams"}},
                        "claims": {"P31": [{"mainsnak": {"datavalue": {"value": {"id": "Q5"}}}}]},
                    }
                }
            },
        )

    monkeypatch.setattr(workbook_validator, "_network_get", fake_network_get)

    success, metadata, detail = _lookup_wikidata_record("Q42", client=None)

    assert success is True
    assert detail == ""
    assert metadata is not None
    assert metadata["title"] == "Douglas Adams"
    assert metadata["wikidata_id"] == "Q42"
    assert "Douglas Noel Adams" in metadata["alternate_titles"]


def test_tv_shows_subcategory_must_include_program_type():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "companies",
            "brand_set",
            "facebook_page",
        ]
    )
    sheet.append(
        [
            "Show One",
            "TV Shows",
            "Series",
            "Drama",
            "Drama",
            "Studio",
            "Competitive View",
            "",
        ]
    )
    sheet.append(
        [
            "Show Two",
            "TV Shows",
            "Program Type - Scripted",
            "Drama",
            "Drama",
            "Studio",
            "Competitive View",
            "",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "tv-shows.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert any(issue.cell == "C2" and issue.rule == "contains" for issue in artifact.issues)
    assert catalog["C2"].fill.fill_type == "solid"
    assert catalog["C3"].fill.fill_type != "solid"


def test_validate_workbook_highlights_invalid_cells_and_adds_summary_sheet():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tasks"
    sheet.append(["Task ID", "Task Name", "Status", "Completion %", "Due Date"])
    sheet.append(["TASK-001", "Build validator", "Open", 50, date.today() + timedelta(days=1)])
    sheet.append(["TASK-001", "", "Blocked", 120, date.today() - timedelta(days=2)])

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(
        """
        {
          "rules": [
            {"sheet": "Tasks", "column": "Task ID", "check": "unique"},
            {"sheet": "Tasks", "column": "Task Name", "check": "required"},
            {"sheet": "Tasks", "column": "Status", "check": "in", "values": ["Open", "Done"]},
            {"sheet": "Tasks", "column": "Completion %", "check": "between", "min": 0, "max": 100},
            {"sheet": "Tasks", "column": "Due Date", "check": "date_not_past"}
          ]
        }
        """
    )

    artifact = validate_workbook(buffer.getvalue(), "tasks.xlsx", rules)

    assert artifact.issue_count == 6

    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    tasks = validated["Tasks"]
    summary = validated["Validation Summary"]

    assert tasks["A2"].fill.fill_type == "solid"
    assert tasks["A3"].fill.fill_type == "solid"
    assert tasks["B3"].fill.fill_type == "solid"
    assert tasks["C3"].fill.fill_type == "solid"
    assert tasks["D3"].fill.fill_type == "solid"
    assert tasks["E3"].fill.fill_type == "solid"
    assert summary.max_row == 7


def test_validate_workbook_supports_conditional_dar_and_url_rules():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "companies",
            "brand_set",
            "facebook_page",
        ]
    )
    sheet.append(
        [
            "Sample Title - DAR",
            "Movies",
            "Feature",
            "Drama",
            "Drama",
            "Another Company",
            "Competitive View",
            "https://www.facebook.com/p/example",
        ]
    )
    sheet.append(
        [
            "Regular Title",
            "Media",
            "",
            "",
            "",
            "",
            "",
            "https://www.facebook.com/example",
        ]
    )
    sheet.append(
        [
            "Talent Title - DAR",
            "Talent",
            "Talent Subtype - Athlete\nGender - Man",
            "",
            "",
            "Pristine Talent",
            "Pristine DAR Brands",
            "",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "catalog.xlsx", rules)

    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert artifact.issue_count >= 4
    assert catalog["F2"].fill.fill_type == "solid"
    assert catalog["G2"].fill.fill_type == "solid"
    assert catalog["H2"].fill.fill_type == "solid"
    assert catalog["C3"].fill.fill_type != "solid"
    assert catalog["D3"].fill.fill_type != "solid"
    assert catalog["G3"].fill.fill_type == "solid"
    assert catalog["F4"].fill.fill_type != "solid"


def test_tv_show_dar_titles_allow_pristine_tv_in_companies():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "companies",
            "brand_set",
            "facebook_page",
        ]
    )
    sheet.append(
        [
            "Example Show - DAR",
            "TV Shows",
            "Program Type - Scripted",
            "Drama",
            "Drama",
            "Pristine TV",
            "Pristine DAR Brands",
            "",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "tv-show-dar.xlsx", rules)
    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    catalog = validated["Catalog"]

    assert not any(issue.cell == "F2" for issue in artifact.issues)
    assert catalog["F2"].fill.fill_type != "solid"


def test_validate_workbook_flags_talent_subcategory_without_gender_and_type():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Talent"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "companies",
            "brand_set",
            "facebook_page",
        ]
    )
    sheet.append(
        [
            "Jane Doe",
            "Talent",
            "Talent Type - Actor",
            "",
            "",
            "Competitive Co",
            "Competitive View",
            "",
        ]
    )
    sheet.append(
        [
            "John Doe",
            "Talent",
            "Talent Subtype - Actor\nGender - Man",
            "",
            "",
            "Competitive Co",
            "Competitive View",
            "",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "talent.xlsx", rules)

    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    talent = validated["Talent"]

    assert talent["C2"].fill.fill_type == "solid"
    assert talent["C3"].fill.fill_type != "solid"
    assert any(issue.cell == "C2" and issue.rule == "talent_subcategory_format" for issue in artifact.issues)


def test_validate_workbook_allows_blank_talent_genre():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Talent"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "companies",
            "brand_set",
            "facebook_page",
        ]
    )
    sheet.append(
        [
            "Jane Doe",
            "Talent",
            "Talent Subtype - Actor\nGender - Woman",
            "",
            "",
            "Competitive Co",
            "Competitive View",
            "",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "talent-no-genre.xlsx", rules)

    validated = load_workbook(io.BytesIO(artifact.file_bytes))
    talent = validated["Talent"]

    assert talent["D2"].fill.fill_type != "solid"
    assert not any(issue.cell == "D2" and issue.rule == "not_blank_and_not_in" for issue in artifact.issues)


def test_it_internet_computing_category_is_allowed():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalog"
    sheet.append(
        [
            "title",
            "title_category",
            "title_sub_category",
            "genre",
            "primary_genre",
            "companies",
            "brand_set",
            "facebook_page",
        ]
    )
    sheet.append(
        [
            "Tech Title",
            "IT, Internet, Computing",
            "",
            "",
            "",
            "Company",
            "Competitive View",
            "",
        ]
    )

    buffer = io.BytesIO()
    workbook.save(buffer)

    rules = parse_validation_rules(build_sample_rules_json())
    artifact = validate_workbook(buffer.getvalue(), "combined-it-categories.xlsx", rules)

    title_category_issues = {issue.cell for issue in artifact.issues if issue.rule == "in"}
    assert "B2" not in title_category_issues
