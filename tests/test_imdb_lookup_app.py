import csv
import gzip
import io
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from imdb_lookup_app.config import settings
from imdb_lookup_app.main import app
from imdb_lookup_app.routes import cache


def test_index_page_renders():
    client = TestClient(app)
    response = client.get("/")
    assert response.status_code == 200
    assert "IMDb Bulk Lookup" in response.text
    assert "Quick Bulk Lookup" in response.text
    assert "CSV Or Excel Input" in response.text


def test_api_lookup_resolves_ids_and_reverse_matches(monkeypatch, tmp_path):
    _configure_lookup_datasets(monkeypatch, tmp_path)
    client = TestClient(app)

    response = client.post(
        "/api/lookup",
        json={
            "mode": "auto",
            "values": ["tt0000001", "Jane Example", "Shared Name"],
        },
    )

    assert response.status_code == 200
    payload = response.json()
    rows = payload["rows"]
    assert any(row["imdb_id"] == "tt0000001" and row["display_name"] == "First Title" for row in rows)
    assert any(row["imdb_id"] == "nm0000001" and row["display_name"] == "Jane Example" for row in rows)

    shared_rows = [row for row in rows if row["input_value"] == "Shared Name"]
    assert len(shared_rows) == 3
    assert all(row["status"] == "multiple_matches" for row in shared_rows)
    assert any(row["entity_kind"] == "title" for row in shared_rows)
    assert any(row["entity_kind"] == "person" for row in shared_rows)


def test_file_upload_returns_downloadable_csv_and_xlsx(monkeypatch, tmp_path):
    _configure_lookup_datasets(monkeypatch, tmp_path)
    client = TestClient(app)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Input"
    sheet.append(["input"])
    sheet.append(["tt0000001"])
    sheet.append(["Jane Example"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    response = client.post(
        "/lookup/file",
        data={"mode": "auto", "sheet_name": "Input", "input_column": "input"},
        files={
            "dataset_file": (
                "lookup.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    assert "Download CSV" in response.text
    assert "Download Excel" in response.text
    export_id = _extract_export_id(response.text, "csv")

    csv_response = client.get(f"/download/{export_id}/csv")
    assert csv_response.status_code == 200
    csv_rows = list(csv.DictReader(io.StringIO(csv_response.content.decode("utf-8-sig"))))
    assert any(row["imdb_id"] == "tt0000001" for row in csv_rows)
    assert any(row["display_name"] == "Jane Example" for row in csv_rows)

    xlsx_response = client.get(f"/download/{export_id}/xlsx")
    assert xlsx_response.status_code == 200
    workbook_result = load_workbook(io.BytesIO(xlsx_response.content))
    assert "Summary" in workbook_result.sheetnames
    assert "Results" in workbook_result.sheetnames
    result_sheet = workbook_result["Results"]
    assert result_sheet["H2"].value == "tt0000001"

    cache._items.clear()


def _extract_export_id(html: str, fmt: str) -> str:
    marker = "/download/"
    start = html.index(marker) + len(marker)
    end = html.index(f"/{fmt}", start)
    return html[start:end]


def _configure_lookup_datasets(monkeypatch, tmp_path: Path) -> None:
    title_dataset = tmp_path / "title.basics.tsv.gz"
    name_dataset = tmp_path / "name.basics.tsv.gz"
    cache_dir = tmp_path / "cache"

    _write_gzipped_tsv(
        title_dataset,
        ["tconst", "titleType", "primaryTitle", "originalTitle", "isAdult", "startYear", "endYear", "runtimeMinutes", "genres"],
        [
            ["tt0000001", "movie", "First Title", "Original First", "0", "1999", r"\N", "136", "Action"],
            ["tt0000002", "tvSeries", "Shared Name", "Shared Name", "0", "2010", "2014", "45", "Drama"],
            ["tt0000003", "movie", "Shared Name", "Shared Name", "0", "2024", r"\N", "110", "Drama"],
        ],
    )
    _write_gzipped_tsv(
        name_dataset,
        ["nconst", "primaryName", "birthYear", "deathYear", "primaryProfession", "knownForTitles"],
        [
            ["nm0000001", "Jane Example", "1980", r"\N", "actor,producer", "tt0000001,tt0000002"],
            ["nm0000002", "Shared Name", "1975", r"\N", "actor", "tt0000003"],
        ],
    )

    monkeypatch.setattr(settings, "imdb_lookup_title_basics_url", str(title_dataset))
    monkeypatch.setattr(settings, "imdb_lookup_name_basics_url", str(name_dataset))
    monkeypatch.setattr(settings, "imdb_lookup_dataset_dir", str(cache_dir))
    monkeypatch.setattr(settings, "imdb_lookup_refresh_hours", 24)
    monkeypatch.setattr(settings, "imdb_lookup_export_ttl_seconds", 900)
    cache._items.clear()


def _write_gzipped_tsv(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    with gzip.open(path, "wt", encoding="utf-8", newline="") as file_handle:
        writer = csv.writer(file_handle, delimiter="\t")
        writer.writerow(headers)
        writer.writerows(rows)


def test_api_fallback_when_db_unavailable(monkeypatch):
    from imdb_lookup_app.services.lookup import ImdbLookupService, ImdbLookupServiceError
    import httpx

    # Force database to be unavailable
    def mock_ensure_index(*args, **kwargs):
        raise ImdbLookupServiceError("Index is not available on Vercel.")
    monkeypatch.setattr(ImdbLookupService, "_ensure_imdb_dataset_index", mock_ensure_index)

    # Configure API keys
    monkeypatch.setattr(settings, "tmdb_api_key", "mock_tmdb_key")
    monkeypatch.setattr(settings, "omdb_api_key", "mock_omdb_key")

    # Mock client responses
    def mock_get(self, url, params=None, headers=None, **kwargs):
        url_str = str(url)
        if "api.themoviedb.org/3/find/tt1234567" in url_str:
            return httpx.Response(200, json={
                "movie_results": [{
                    "id": 101,
                    "title": "Mock Movie via TMDB Find",
                    "original_title": "Original Title",
                    "release_date": "2026-05-29"
                }],
                "tv_results": [],
                "tv_episode_results": [],
                "tv_season_results": [],
                "person_results": []
            })
        elif "api.themoviedb.org/3/find/nm1234567" in url_str:
            return httpx.Response(200, json={
                "movie_results": [],
                "tv_results": [],
                "tv_episode_results": [],
                "tv_season_results": [],
                "person_results": [{
                    "id": 201,
                    "name": "Mock Person via TMDB Find",
                    "known_for_department": "Acting",
                    "known_for": [{"title": "Known Work Title"}]
                }]
            })
        elif "omdbapi.com" in url_str:
            # Check if search or ID query
            p = params or {}
            if "s" in p:
                return httpx.Response(200, json={
                    "Response": "True",
                    "Search": [{
                        "Title": "OMDb Search Result",
                        "Year": "2025",
                        "imdbID": "tt2222222",
                        "Type": "movie"
                    }]
                })
            else:
                return httpx.Response(200, json={
                    "Response": "True",
                    "Title": "OMDb ID Result",
                    "Year": "2024",
                    "imdbID": p.get("i", ""),
                    "Type": "series"
                })
        elif "api.themoviedb.org/3/search/person" in url_str:
            return httpx.Response(200, json={
                "results": [{
                    "id": 6384,
                    "name": "Keanu Reeves",
                    "known_for_department": "Acting",
                    "known_for": [{"title": "The Matrix"}]
                }]
            })
        elif "api.themoviedb.org/3/person/6384/external_ids" in url_str:
            return httpx.Response(200, json={
                "imdb_id": "nm0000206"
            })
        return httpx.Response(404)

    monkeypatch.setattr(httpx.Client, "get", mock_get)

    client = TestClient(app)

    # 1. Test auto mode with ID (tt...) -> TMDB Find Movie
    resp1 = client.post("/api/lookup", json={"mode": "auto", "values": ["tt1234567"]})
    assert resp1.status_code == 200
    rows1 = resp1.json()["rows"]
    assert len(rows1) == 1
    assert rows1[0]["display_name"] == "Mock Movie via TMDB Find"
    assert rows1[0]["status"] == "matched"
    assert rows1[0]["notes"] == "Resolved via TMDB Find API."

    # 2. Test auto mode with ID (nm...) -> TMDB Find Person
    resp2 = client.post("/api/lookup", json={"mode": "auto", "values": ["nm1234567"]})
    assert resp2.status_code == 200
    rows2 = resp2.json()["rows"]
    assert len(rows2) == 1
    assert rows2[0]["display_name"] == "Mock Person via TMDB Find"
    assert rows2[0]["entity_kind"] == "person"
    assert rows2[0]["known_for_titles"] == "Known Work Title"

    # 3. Test title_to_id -> OMDb Search
    resp3 = client.post("/api/lookup", json={"mode": "title_to_id", "values": ["OMDb Search Result"]})
    assert resp3.status_code == 200
    rows3 = resp3.json()["rows"]
    assert len(rows3) == 1
    assert rows3[0]["imdb_id"] == "tt2222222"
    assert rows3[0]["notes"] == "Resolved via OMDb search API."

    # 4. Test person_to_id -> TMDB Person Search
    resp4 = client.post("/api/lookup", json={"mode": "person_to_id", "values": ["Keanu Reeves"]})
    assert resp4.status_code == 200
    rows4 = resp4.json()["rows"]
    assert len(rows4) == 1
    assert rows4[0]["imdb_id"] == "nm0000206"
    assert rows4[0]["notes"] == "Resolved via TMDB person search API."

