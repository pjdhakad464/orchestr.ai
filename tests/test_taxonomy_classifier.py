from __future__ import annotations

import pytest
from fastapi.testclient import TestClient
from app.main import app
from app.cache import TTLCache
from app.services.taxonomy_classifier import TaxonomyClassifier
from unittest.mock import AsyncMock, patch

@pytest.mark.asyncio
async def test_taxonomy_classifier_talent_classification():
    cache = TTLCache(60)
    classifier = TaxonomyClassifier(cache)

    # Mock search and page response
    mock_search_json = {
        "query": {
            "search": [
                {"title": "Taylor Swift"}
            ]
        }
    }
    mock_page_json = {
        "query": {
            "pages": [
                {
                    "title": "Taylor Swift",
                    "extract": "Taylor Swift is an American singer-songwriter. She is a woman.",
                    "categories": [
                        {"title": "Category:1989 births"},
                        {"title": "Category:American women singers"}
                    ],
                    "pageprops": {
                        "wikibase_item": "Q26876"
                    }
                }
            ]
        }
    }
    mock_wd_json = {
        "entities": {
            "Q26876": {
                "claims": {
                    "P31": [{"mainsnak": {"datavalue": {"value": {"entity-type": "item", "id": "Q5"}}}}],  # human
                    "P21": [{"mainsnak": {"datavalue": {"value": {"entity-type": "item", "id": "Q6581072"}}}}],  # female
                    "P106": [{"mainsnak": {"datavalue": {"value": {"entity-type": "item", "id": "Q177220"}}}}],  # singer
                }
            }
        }
    }

    # Patch httpx.AsyncClient.get
    with patch("httpx.AsyncClient.get") as mock_get:
        mock_get.side_effect = [
            AsyncMock(status_code=200, json=lambda: mock_search_json, raise_for_status=lambda: None),
            AsyncMock(status_code=200, json=lambda: mock_page_json, raise_for_status=lambda: None),
            AsyncMock(status_code=200, json=lambda: mock_wd_json, raise_for_status=lambda: None),
        ]

        result = await classifier.classify("Taylor Swift")
        assert result["category"] == "Talent"
        assert "Gender - Woman" in result["sub_category"]
        assert "Talent Type - Musician" in result["sub_category"]
        assert "Talent Subtype - Musician - Singer" in result["sub_category"]


@pytest.mark.asyncio
async def test_taxonomy_classifier_brand_classification():
    cache = TTLCache(60)
    classifier = TaxonomyClassifier(cache)

    mock_search_json = {
        "query": {
            "search": [
                {"title": "Nike, Inc."}
            ]
        }
    }
    mock_page_json = {
        "query": {
            "pages": [
                {
                    "title": "Nike, Inc.",
                    "extract": "Nike is an American multinational corporation that designs and manufactures footwear, apparel, and fashion accessories.",
                    "categories": [
                        {"title": "Category:Clothing brands of the United States"}
                    ],
                    "pageprops": {
                        "wikibase_item": "Q2539"
                    }
                }
            ]
        }
    }
    mock_wd_json = {
        "entities": {
            "Q2539": {
                "claims": {
                    "P31": [{"mainsnak": {"datavalue": {"value": {"entity-type": "item", "id": "Q4830453"}}}}],  # business
                }
            }
        }
    }

    with patch("httpx.AsyncClient.get") as mock_get:
        mock_get.side_effect = [
            AsyncMock(status_code=200, json=lambda: mock_search_json, raise_for_status=lambda: None),
            AsyncMock(status_code=200, json=lambda: mock_page_json, raise_for_status=lambda: None),
            AsyncMock(status_code=200, json=lambda: mock_wd_json, raise_for_status=lambda: None),
        ]

        result = await classifier.classify("Nike")
        assert result["category"] == "Fashion"
        assert any(sub in result["sub_category"] for sub in ["Product Category - Apparel", "Product Category - Footwear", "Women's Apparel"])


def test_classify_title_route_returns_html():
    client = TestClient(app)

    # Patch the classifier's classify method to return a mock result
    with patch("app.routes.taxonomy_classifier.classify", new_callable=AsyncMock) as mock_classify:
        mock_classify.return_value = {
            "category": "Talent",
            "sub_category": "Gender - Woman\nTalent Subtype - Actress\nTalent Type - Actress"
        }

        response = client.post("/classify-title", data={"title_name": "Meryl Streep"})
        assert response.status_code == 200
        assert "Taxonomy Classification Result" in response.text
        assert "Meryl Streep" in response.text
        assert "Talent" in response.text
        assert "Gender - Woman" in response.text
        assert "Talent Subtype - Actress" in response.text


def test_bulk_taxonomy_endpoint():
    import io
    import openpyxl
    import re
    client = TestClient(app)
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Title Name", "Instagram Link"])
    sheet.append(["Selena Gomez", "https://instagram.com/selenagomez"])
    sheet.append(["Nike", "@nike"])
    
    buffer = io.BytesIO()
    wb.save(buffer)
    file_bytes = buffer.getvalue()
    
    with patch("app.routes.taxonomy_classifier.classify", new_callable=AsyncMock) as mock_classify:
        mock_classify.side_effect = [
            {"category": "Talent", "sub_category": "Gender - Woman\nTalent Type - Musician"},
            {"category": "Fashion", "sub_category": "Women's Apparel"}
        ]
        
        response = client.post(
            "/bulk-classify-taxonomy",
            files={"workbook": ("test_titles.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"run_by": "Test Operator"}
        )
        
        assert response.status_code == 200
        assert "Taxonomy Identification Completed" in response.text
        assert "Selena Gomez" in response.text
        assert "Nike" in response.text
        assert "Talent" in response.text
        assert "Fashion" in response.text
        
        match = re.search(r'/taxonomy/download/([a-fA-F0-9\-]+)', response.text)
        assert match is not None
        job_id = match.group(1)
        
        download_response = client.get(f"/taxonomy/download/{job_id}")
        assert download_response.status_code == 200
        assert "spreadsheetml.sheet" in download_response.headers["content-type"]
        
        downloaded_wb = openpyxl.load_workbook(io.BytesIO(download_response.content))
        downloaded_sheet = downloaded_wb.active
        
        assert downloaded_sheet.cell(row=2, column=1).value == "Selena Gomez"
        assert downloaded_sheet.cell(row=2, column=2).value == "https://instagram.com/selenagomez"
        assert downloaded_sheet.cell(row=2, column=3).value == "Talent"
        assert downloaded_sheet.cell(row=2, column=4).value == "Gender - Woman\nTalent Type - Musician"
        
        assert downloaded_sheet.cell(row=3, column=1).value == "Nike"
        assert downloaded_sheet.cell(row=3, column=2).value == "@nike"
        assert downloaded_sheet.cell(row=3, column=3).value == "Fashion"
        assert downloaded_sheet.cell(row=3, column=4).value == "Women's Apparel"
