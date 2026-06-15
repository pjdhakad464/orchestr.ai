import io
import json
import os
import openpyxl
from pathlib import Path
import pytest
from unittest.mock import patch, MagicMock

import app.services.workbook_validator as workbook_validator
from app.services.workbook_validator import (
    parse_validation_rules,
    validate_workbook,
)

@pytest.fixture
def base_rules():
    rules_json = """
    {
      "rules": [
        {"sheet": "BDR", "column": "primary_genre", "check": "genre_taxonomy_audit", "when": [{"column": "title_category", "operator": "equals", "value": "Movies"}]},
        {"sheet": "BDR", "column": "released_on", "check": "date_cross_check", "when": [{"column": "title_category", "operator": "equals", "value": "Movies"}]},
        {"sheet": "BDR", "column": "network", "check": "network_platform_audit", "when": [{"column": "title_category", "operator": "equals", "value": "Movies"}]},
        {"sheet": "BDR", "column": "wikipedia_url", "check": "wikipedia_url_audit"},
        {"sheet": "BDR", "column": "imdb_id", "check": "imdb_url_audit"}
      ]
    }
    """
    return parse_validation_rules(rules_json)

def _build_test_workbook(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BDR"
    ws.append(["title", "title_category", "primary_genre", "released_on", "network", "wikipedia_url", "imdb_id"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def test_genre_taxonomy_audit_various_cases(base_rules, monkeypatch):
    # Mock entity metadata returned by TMDB
    mock_meta = {
        "type": "movie",
        "title": "Inception",
        "genres": ["Action", "Science Fiction"],
        "release_date": "2010-07-16",
        "networks": [],
        "companies": ["Warner Bros."]
    }
    monkeypatch.setattr(workbook_validator, "_get_entity_metadata", lambda row, cache: (True, mock_meta, ""))

    rows = [
        # Pass case
        ["Inception", "Movies", "Action", "2010-07-16", "Warner Bros.", "https://en.wikipedia.org/wiki/Inception", "tt1375666"],
        # Trailing newline format error
        ["Inception", "Movies", "Action\n", "2010-07-16", "Warner Bros.", "https://en.wikipedia.org/wiki/Inception", "tt1375666"],
        # Placeholder fail
        ["Inception", "Movies", "N/A", "2010-07-16", "Warner Bros.", "https://en.wikipedia.org/wiki/Inception", "tt1375666"],
        # Genre mismatch
        ["Inception", "Movies", "Comedy", "2010-07-16", "Warner Bros.", "https://en.wikipedia.org/wiki/Inception", "tt1375666"],
    ]

    workbook_bytes = _build_test_workbook(rows)
    artifact = validate_workbook(workbook_bytes, "test.xlsx", base_rules)

    genre_issues = [issue for issue in artifact.issues if issue.column == "C"]
    
    # Check issue count and details
    assert len(genre_issues) == 3
    
    # Trailing newline (row 3 in sheet, index 1 in data rows)
    row3_issue = next(i for i in genre_issues if i.row == 3)
    assert row3_issue.finding_category == "Formatting Error"
    assert "format_error" in row3_issue.message.lower()

    # Placeholder (row 4 in sheet)
    row4_issue = next(i for i in genre_issues if i.row == 4)
    assert row4_issue.finding_category == "Placeholder Found"
    assert "placeholder" in row4_issue.message.lower()

    # Genre Mismatch (row 5 in sheet)
    row5_issue = next(i for i in genre_issues if i.row == 5)
    assert row5_issue.finding_category == "Suspected Incorrect"
    assert "genre_mismatch" in row5_issue.message.lower()


def test_date_cross_check_various_cases(base_rules, monkeypatch):
    mock_meta = {
        "type": "movie",
        "title": "Inception",
        "genres": ["Action"],
        "release_date": "2010-07-16",
        "networks": [],
        "companies": ["Warner Bros."]
    }
    monkeypatch.setattr(workbook_validator, "_get_entity_metadata", lambda row, cache: (True, mock_meta, ""))

    rows = [
        # Pass
        ["Inception", "Movies", "Action", "2010-07-16", "Warner Bros.", "https://en.wikipedia.org/wiki/Inception", "tt1375666"],
        # Date mismatch
        ["Inception", "Movies", "Action", "2012-07-16", "Warner Bros.", "https://en.wikipedia.org/wiki/Inception", "tt1375666"],
        # Anomalous year (suspected typo)
        ["Inception", "Movies", "Action", "1899-07-16", "Warner Bros.", "https://en.wikipedia.org/wiki/Inception", "tt1375666"],
    ]

    workbook_bytes = _build_test_workbook(rows)
    artifact = validate_workbook(workbook_bytes, "test.xlsx", base_rules)

    date_issues = [issue for issue in artifact.issues if issue.column == "D"]
    assert len(date_issues) == 2

    # Date mismatch
    row3_issue = next(i for i in date_issues if i.row == 3)
    assert row3_issue.finding_category == "Suspected Incorrect"
    assert "date_mismatch" in row3_issue.message.lower()

    # Anomalous year formatting error
    row4_issue = next(i for i in date_issues if i.row == 4)
    assert row4_issue.finding_category == "Formatting Error"
    assert "anomalous year" in row4_issue.message.lower()


def test_network_platform_audit(base_rules, monkeypatch):
    mock_meta = {
        "type": "movie",
        "title": "Inception",
        "genres": ["Action"],
        "release_date": "2010-07-16",
        "networks": [],
        "companies": ["Warner Bros. Pictures", "Legendary Pictures"]
    }
    monkeypatch.setattr(workbook_validator, "_get_entity_metadata", lambda row, cache: (True, mock_meta, ""))

    rows = [
        # Match company loosely -> Pass
        ["Inception", "Movies", "Action", "2010-07-16", "Warner Bros.", "https://en.wikipedia.org/wiki/Inception", "tt1375666"],
        # Mismatch -> Fail
        ["Inception", "Movies", "Action", "2010-07-16", "Universal Pictures", "https://en.wikipedia.org/wiki/Inception", "tt1375666"],
    ]

    workbook_bytes = _build_test_workbook(rows)
    artifact = validate_workbook(workbook_bytes, "test.xlsx", base_rules)

    network_issues = [issue for issue in artifact.issues if issue.column == "E"]
    assert len(network_issues) == 1
    assert network_issues[0].row == 3
    assert network_issues[0].finding_category == "Suspected Incorrect"
    assert "network_mismatch" in network_issues[0].message.lower()


def test_wikipedia_url_audit(base_rules, monkeypatch):
    monkeypatch.setattr(workbook_validator, "_get_entity_metadata", lambda row, cache: (True, {}, ""))
    
    # Mock Wikipedia URL resolver
    def mock_wiki_resolver(url, client):
        if "Inception" in url:
            return True, {"title": "Inception", "url": url}, ""
        return True, {"title": "Avatar", "url": url}, ""
        
    monkeypatch.setattr(workbook_validator, "_lookup_wikipedia_record", mock_wiki_resolver)

    rows = [
        # Pass matching title
        ["Inception", "Movies", "Action", "2010-07-16", "Warner Bros.", "https://en.wikipedia.org/wiki/Inception", "tt1375666"],
        # Mismatching title
        ["Inception", "Movies", "Action", "2010-07-16", "Warner Bros.", "https://en.wikipedia.org/wiki/Avatar", "tt1375666"],
        # Missing URL
        ["Inception", "Movies", "Action", "2010-07-16", "Warner Bros.", "", "tt1375666"],
    ]

    workbook_bytes = _build_test_workbook(rows)
    artifact = validate_workbook(workbook_bytes, "test.xlsx", base_rules)

    wiki_issues = [issue for issue in artifact.issues if issue.column == "F"]
    assert len(wiki_issues) == 2

    # Mismatched title
    row3_issue = next(i for i in wiki_issues if i.row == 3)
    assert row3_issue.finding_category == "Suspected Incorrect"
    assert "wikipedia_mismatch" in row3_issue.message.lower()

    # Missing
    row4_issue = next(i for i in wiki_issues if i.row == 4)
    assert row4_issue.finding_category == "Missing Data"
    assert "missing_wikipedia" in row4_issue.message.lower()


def test_imdb_url_audit(base_rules, monkeypatch):
    monkeypatch.setattr(workbook_validator, "_get_entity_metadata", lambda row, cache: (True, {}, ""))

    # Mock IMDb lookup record
    def mock_imdb_resolver(imdb_id, client):
        if imdb_id == "tt1375666":
            return True, {"title": "Inception", "id": imdb_id}, ""
        return True, {"title": "Avatar", "id": imdb_id}, ""

    monkeypatch.setattr(workbook_validator, "_lookup_imdb_record", mock_imdb_resolver)

    rows = [
        # Pass
        ["Inception", "Movies", "Action", "2010-07-16", "Warner Bros.", "https://en.wikipedia.org/wiki/Inception", "tt1375666"],
        # Title mismatch
        ["Inception", "Movies", "Action", "2010-07-16", "Warner Bros.", "https://en.wikipedia.org/wiki/Inception", "tt0499549"],
        # Missing ID
        ["Inception", "Movies", "Action", "2010-07-16", "Warner Bros.", "https://en.wikipedia.org/wiki/Inception", ""],
    ]

    workbook_bytes = _build_test_workbook(rows)
    artifact = validate_workbook(workbook_bytes, "test.xlsx", base_rules)

    imdb_issues = [issue for issue in artifact.issues if issue.column == "G"]
    assert len(imdb_issues) == 2

    # Mismatched title
    row3_issue = next(i for i in imdb_issues if i.row == 3)
    assert row3_issue.finding_category == "Suspected Incorrect"
    assert "imdb_mismatch" in row3_issue.message.lower()

    # Missing ID
    row4_issue = next(i for i in imdb_issues if i.row == 4)
    assert row4_issue.finding_category == "Missing Data"
    assert "missing_imdb" in row4_issue.message.lower()


def test_validate_bdr_cli_script(monkeypatch, tmp_path):
    # Mock workbook_validator functions
    mock_meta = {
        "type": "movie",
        "title": "Inception",
        "genres": ["Action"],
        "release_date": "2010-07-16",
        "networks": [],
        "companies": ["Warner Bros."]
    }
    monkeypatch.setattr(workbook_validator, "_get_entity_metadata", lambda row, cache: (True, mock_meta, ""))
    monkeypatch.setattr(workbook_validator, "_lookup_wikipedia_record", lambda url, client: (True, {"title": "Inception", "url": url}, ""))
    monkeypatch.setattr(workbook_validator, "_lookup_imdb_record", lambda imdb_id, client: (True, {"title": "Inception", "id": imdb_id}, ""))

    # Create temporary input workbook
    rows = [
        ["Inception", "Movies", "Action", "2010-07-16", "Warner Bros.", "https://en.wikipedia.org/wiki/Inception", "tt1375666"],
        ["Avatar", "Movies", "Comedy", "2010-07-16", "Warner Bros.", "https://en.wikipedia.org/wiki/Inception", "tt1375666"], # Mismatches/errors
    ]
    workbook_bytes = _build_test_workbook(rows)
    input_file = tmp_path / "BDR_Input.xlsx"
    input_file.write_bytes(workbook_bytes)

    # Output file path
    output_file = tmp_path / "BDR_QA_Output.xlsx"

    # Mock command line arguments
    import sys
    test_args = [
        "scripts/validate_bdr.py",
        str(input_file),
        "--output", str(output_file),
        "--rules", str(Path(__file__).resolve().parent.parent / "data" / "bdr_qa_rules.json")
    ]

    with patch.object(sys, 'argv', test_args):
        import sys
        sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))
        import validate_bdr as validate_bdr_script
        validate_bdr_script.main()

    # Verify output file exists and has correct sheet structure
    assert output_file.exists()
    
    out_wb = openpyxl.load_workbook(output_file)
    assert "QA Findings" in out_wb.sheetnames
    assert "Clean Rows" in out_wb.sheetnames
    assert "Summary Metrics" in out_wb.sheetnames

    # Check metrics
    ws_metrics = out_wb["Summary Metrics"]
    # Metric rows start at row 4
    metric_values = {}
    for r in range(4, 7):
        metric = ws_metrics.cell(row=r, column=1).value
        val = ws_metrics.cell(row=r, column=2).value
        metric_values[metric] = val

    assert metric_values.get("Total Checked Rows") == 2
    assert metric_values.get("Clean Rows (No Issues)") == 1
    assert metric_values.get("Flagged Rows (With Issues)") == 1
