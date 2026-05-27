import asyncio
import csv
import io
from dataclasses import dataclass

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from title_url_lookup_app.main import app
from title_url_lookup_app.models import (
    BulkTitleLookupResponse,
    SiteLookupResult,
    TitleLookupQuery,
    TitleLookupResponse,
    TitleUrlCandidate,
)
from title_url_lookup_app.routes import cache
from title_url_lookup_app.search_models import SearchResult
from title_url_lookup_app.services.imdb_dataset import ImdbTitleMatch
from title_url_lookup_app.services.title_lookup import TitleUrlLookupService


class FakeSearchProvider:
    def __init__(self, responses: dict[str, list[SearchResult]]) -> None:
        self.responses = responses

    async def search(self, query: str, *, limit: int = 10) -> list[SearchResult]:
        return self.responses.get(query, [])[:limit]


@dataclass
class FakeImdbDatasetLookup:
    responses: dict[tuple[str, str, str], list[ImdbTitleMatch]]

    def lookup_title(self, query: TitleLookupQuery) -> list[ImdbTitleMatch]:
        return self.responses.get((query.title, query.year, query.title_type), [])


class StubLookupService:
    async def lookup_title(self, query: TitleLookupQuery) -> TitleLookupResponse:
        return TitleLookupResponse(
            query=query,
            results=[
                SiteLookupResult(
                    site_key="imdb",
                    site_label="IMDb",
                    status="found",
                    primary=TitleUrlCandidate(
                        url="https://imdb.com/title/tt14619456",
                        canonical_url="https://imdb.com/title/tt14619456",
                        result_title="Faces of Death (2026)",
                        score=91,
                        matched_on=["stubbed"],
                    ),
                )
            ],
            notes=["stubbed"],
        )

    async def lookup_titles(self, queries: list[TitleLookupQuery]) -> BulkTitleLookupResponse:
        entries = [await self.lookup_title(query) for query in queries]
        return BulkTitleLookupResponse(entries=entries, notes=["bulk stubbed"])


def test_service_picks_best_urls_for_faces_of_death():
    query = TitleLookupQuery(title="Faces of Death", year="2026", title_type="movie")
    ddg_provider = FakeSearchProvider(
        {
            '"Faces of Death" 2026 movie site:rottentomatoes.com/m': [
                SearchResult(
                    title="Faces of Death | Rotten Tomatoes",
                    url="https://www.rottentomatoes.com/m/faces_of_death",
                    snippet="1978 horror film",
                    source_domain="www.rottentomatoes.com",
                    position=1,
                ),
                SearchResult(
                    title="Faces of Death (2026) | Rotten Tomatoes",
                    url="https://www.rottentomatoes.com/m/faces_of_death_2026",
                    snippet="2026 horror movie directed by Daniel Goldhaber.",
                    source_domain="www.rottentomatoes.com",
                    position=2,
                ),
            ],
            '"Faces of Death" 2026 movie site:metacritic.com/movie': [
                SearchResult(
                    title="Faces of Death details - Metacritic",
                    url="https://www.metacritic.com/movie/faces-of-death/details/",
                    snippet="Release Date: Apr 10, 2026",
                    source_domain="www.metacritic.com",
                    position=1,
                )
            ],
        }
    )
    imdb_lookup = FakeImdbDatasetLookup(
        {
            ("Faces of Death", "2026", "movie"): [
                ImdbTitleMatch(
                    imdb_id="tt14619456",
                    url="https://www.imdb.com/title/tt14619456/",
                    display_title="Faces of Death",
                    original_title="Faces of Death",
                    title_type="movie",
                    start_year="2026",
                    end_year="",
                    score=188,
                    matched_on=["primary_title", "movie type matched", "start year 2026 matched"],
                ),
                ImdbTitleMatch(
                    imdb_id="tt0077533",
                    url="https://www.imdb.com/title/tt0077533/",
                    display_title="Faces of Death",
                    original_title="Faces of Death",
                    title_type="movie",
                    start_year="1978",
                    end_year="",
                    score=140,
                    matched_on=["primary_title", "movie type matched"],
                ),
            ]
        }
    )
    wiki_provider = FakeSearchProvider(
        {
            "Faces of Death 2026 film": [
                SearchResult(
                    title="Faces of Death (2026 film)",
                    url="https://en.wikipedia.org/wiki/Faces_of_Death_%282026_film%29",
                    snippet="2026 American horror film by Daniel Goldhaber.",
                    source_domain="en.wikipedia.org",
                    position=1,
                ),
                SearchResult(
                    title="Faces of Death",
                    url="https://en.wikipedia.org/wiki/Faces_of_Death",
                    snippet="1978 film",
                    source_domain="en.wikipedia.org",
                    position=2,
                ),
            ]
        }
    )

    service = TitleUrlLookupService(ddg_provider=ddg_provider, wiki_provider=wiki_provider, imdb_dataset_lookup=imdb_lookup)
    result = asyncio.run(service.lookup_title(query))

    urls = {item.site_key: item.primary.url for item in result.results if item.primary}
    assert urls["imdb"] == "https://imdb.com/title/tt14619456"
    assert urls["wikipedia"] == "https://en.wikipedia.org/wiki/Faces_of_Death_%282026_film%29"
    assert urls["rottentomatoes"] == "https://rottentomatoes.com/m/faces_of_death_2026"
    assert urls["metacritic"] == "https://metacritic.com/movie/faces-of-death"


def test_service_marks_close_matches_as_uncertain_without_year():
    query = TitleLookupQuery(title="Faces of Death", title_type="movie")
    ddg_provider = FakeSearchProvider({})
    imdb_lookup = FakeImdbDatasetLookup(
        {
            ("Faces of Death", "", "movie"): [
                ImdbTitleMatch(
                    imdb_id="tt14619456",
                    url="https://www.imdb.com/title/tt14619456/",
                    display_title="Faces of Death",
                    original_title="Faces of Death",
                    title_type="movie",
                    start_year="2026",
                    end_year="",
                    score=138,
                    matched_on=["primary_title", "movie type matched"],
                ),
                ImdbTitleMatch(
                    imdb_id="tt0077533",
                    url="https://www.imdb.com/title/tt0077533/",
                    display_title="Faces of Death",
                    original_title="Faces of Death",
                    title_type="movie",
                    start_year="1978",
                    end_year="",
                    score=134,
                    matched_on=["primary_title", "movie type matched"],
                ),
            ]
        }
    )
    wiki_provider = FakeSearchProvider({"Faces of Death film": []})

    service = TitleUrlLookupService(ddg_provider=ddg_provider, wiki_provider=wiki_provider, imdb_dataset_lookup=imdb_lookup)
    result = asyncio.run(service.lookup_title(query))

    imdb_result = next(item for item in result.results if item.site_key == "imdb")
    assert imdb_result.status == "uncertain"
    assert imdb_result.primary is not None
    assert imdb_result.alternates


def test_index_single_lookup_bulk_lookup_and_downloads(monkeypatch):
    from title_url_lookup_app import routes

    monkeypatch.setattr(routes, "lookup_service", StubLookupService())
    client = TestClient(app)

    index_response = client.get("/")
    assert index_response.status_code == 200
    assert "Standalone Title Link Engine" in index_response.text
    assert "Multiple Titles In One Run" in index_response.text
    assert "/api/bulk-lookup" in index_response.text

    form_response = client.post("/lookup", data={"title": "Faces of Death", "year": "2026", "title_type": "movie"})
    assert form_response.status_code == 200
    assert "stubbed" in form_response.text

    bulk_response = client.post(
        "/bulk-lookup",
        data={"entries": "Faces of Death | 2026 | movie\nThe Office | 2005 | tv", "title_type": "any"},
    )
    assert bulk_response.status_code == 200
    assert "bulk stubbed" in bulk_response.text
    assert "Download CSV" in bulk_response.text
    assert "Download Excel" in bulk_response.text

    export_id = _extract_export_id(bulk_response.text, "csv")
    csv_response = client.get(f"/download/{export_id}/csv")
    assert csv_response.status_code == 200
    csv_rows = list(csv.DictReader(io.StringIO(csv_response.content.decode("utf-8-sig"))))
    assert csv_rows[0]["input_title"] == "Faces of Death"
    assert csv_rows[0]["site"] == "IMDb"

    xlsx_response = client.get(f"/download/{export_id}/xlsx")
    assert xlsx_response.status_code == 200
    workbook = load_workbook(io.BytesIO(xlsx_response.content))
    assert "Summary" in workbook.sheetnames
    assert "Results" in workbook.sheetnames

    api_response = client.post(
        "/api/lookup",
        json={"title": "Faces of Death", "year": "2026", "title_type": "movie"},
    )
    assert api_response.status_code == 200
    assert api_response.json()["notes"] == ["stubbed"]

    bulk_api_response = client.post(
        "/api/bulk-lookup",
        json={
            "entries": [
                {"title": "Faces of Death", "year": "2026", "title_type": "movie"},
                {"title": "The Office", "year": "2005", "title_type": "tv"},
            ]
        },
    )
    assert bulk_api_response.status_code == 200
    assert bulk_api_response.json()["notes"] == ["bulk stubbed"]
    cache._items.clear()


def _extract_export_id(html: str, fmt: str) -> str:
    marker = "/download/"
    start = html.index(marker) + len(marker)
    end = html.index(f"/{fmt}", start)
    return html[start:end]
