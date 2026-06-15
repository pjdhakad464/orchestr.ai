import sys
import os
import argparse
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Adjust python path to import app services
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.services.workbook_validator import validate_loaded_workbook, parse_validation_rules
from app.models import ValidationRule
from app.config import settings

def main():
    parser = argparse.ArgumentParser(description="Audit BDR spreadsheet and generate QA findings report.")
    parser.add_argument("bdr_path", type=str, help="Path to the Brand Definition Report spreadsheet")
    parser.add_argument("--rules", type=str, default=None, help="Path to the JSON rules file (defaults to data/bdr_qa_rules.json)")
    parser.add_argument("--output", type=str, default=None, help="Path to save the output QA Excel report")
    
    args = parser.parse_args()
    
    bdr_file = Path(args.bdr_path)
    if not bdr_file.exists():
        print(f"Error: BDR spreadsheet file not found at {args.bdr_path}")
        sys.exit(1)
        
    # Load rules
    if args.rules:
        rules_path = Path(args.rules)
    else:
        rules_path = Path(__file__).resolve().parent.parent / "data" / "bdr_qa_rules.json"
        
    if not rules_path.exists():
        print(f"Error: Rules JSON file not found at {rules_path}")
        sys.exit(1)
        
    with open(rules_path, "r", encoding="utf-8") as f:
        rules_content = f.read()
        
    try:
        rules = parse_validation_rules(rules_content)
    except Exception as e:
        print(f"Error parsing rules JSON: {e}")
        sys.exit(1)
        
    print(f"Loaded {len(rules)} rules from {rules_path}")
    print(f"Loading workbook {bdr_file.name}...")
    
    try:
        wb = openpyxl.load_workbook(bdr_file)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        sys.exit(1)
        
    # Run validation
    print("Running BDR Metadata QA audits...")
    # Using 'full' review mode to run all rules
    artifact = validate_loaded_workbook(wb, bdr_file.name, rules, review_mode="full")
    issues = artifact.issues
    print(f"Audit completed. Found {len(issues)} issues.")
    
    # Process issues and group flagged rows
    flagged_rows_by_sheet = {}
    issues_by_code = {}
    
    for issue in issues:
        flagged_rows_by_sheet.setdefault(issue.sheet, set()).add(issue.row)
        issues_by_code[issue.rule] = issues_by_code.get(issue.rule, 0) + 1
        
    # Create the QA Report workbook
    qa_wb = openpyxl.Workbook()
    
    # Sheet 1: QA Findings
    ws_findings = qa_wb.active
    ws_findings.title = "QA Findings"
    findings_headers = ["Row #", "Sheet", "Title", "Field/Column", "Issue Code / Rule", "Current Value", "Finding Category", "Confidence", "Confidence Reason", "Message"]
    ws_findings.append(findings_headers)
    
    # Header styling
    header_fill = PatternFill(fill_type="solid", start_color="1F497D", end_color="1F497D")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num in range(1, len(findings_headers) + 1):
        cell = ws_findings.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        
    # We need to find the title for each flagged row to include it in the report.
    # Title column can be auto-detected per sheet.
    for issue in issues:
        sheet = wb[issue.sheet]
        # Let's find the title in the row.
        # We look for a column header matching 'title' or 'artist name' or 'brand/property tracked'
        title_col_idx = None
        for col_idx in range(1, sheet.max_column + 1):
            header_val = str(sheet.cell(row=1, column=col_idx).value or "").lower().strip()
            if header_val in {"title", "artist name", "brand/property tracked", "show/movie"}:
                title_col_idx = col_idx
                break
        
        # Fallback to Column B if not found
        if title_col_idx is None and sheet.max_column >= 2:
            title_col_idx = 2
            
        title_val = ""
        if title_col_idx is not None:
            title_val = str(sheet.cell(row=issue.row, column=title_col_idx).value or "")
            
        ws_findings.append([
            issue.row,
            issue.sheet,
            title_val,
            issue.column,
            issue.rule,
            issue.value,
            issue.finding_category,
            issue.confidence,
            issue.confidence_reason,
            issue.message
        ])
        
    # Style findings rows
    error_fill = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
    error_font = Font(color="9C0006")
    warning_fill = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")
    warning_font = Font(color="9C6500")
    
    for row_idx in range(2, ws_findings.max_row + 1):
        category = ws_findings.cell(row=row_idx, column=7).value
        # If finding category is Suspected Incorrect or Placeholder, style it
        row_fill = error_fill if category in {"Suspected Incorrect", "Placeholder Found", "Formatting Error"} else warning_fill
        row_font = error_font if category in {"Suspected Incorrect", "Placeholder Found", "Formatting Error"} else warning_font
        ws_findings.cell(row=row_idx, column=5).fill = row_fill
        ws_findings.cell(row=row_idx, column=5).font = row_font
        
    # Sheet 2: Clean Rows
    # For every sheet in the original workbook, copy headers and clean rows
    ws_clean = qa_wb.create_sheet(title="Clean Rows")
    total_clean_rows = 0
    total_original_rows = 0
    
    first_sheet = True
    for sheet_name in wb.sheetnames:
        if sheet_name == "Validation Summary":
            continue
        sheet = wb[sheet_name]
        flagged_rows = flagged_rows_by_sheet.get(sheet_name, set())
        
        # Write sheet name separator if there are multiple sheets
        if not first_sheet:
            ws_clean.append([])
            ws_clean.append([f"Sheet: {sheet_name}"])
            ws_clean.cell(row=ws_clean.max_row, column=1).font = Font(bold=True, size=12)
            
        # Copy headers
        headers = [cell.value for cell in sheet[1]]
        ws_clean.append(headers)
        header_row_idx = ws_clean.max_row
        for col_num in range(1, len(headers) + 1):
            cell = ws_clean.cell(row=header_row_idx, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            
        # Copy clean rows
        sheet_original_rows = 0
        sheet_clean_rows = 0
        for r_idx in range(2, sheet.max_row + 1):
            # Check if row is empty
            is_empty = all(sheet.cell(row=r_idx, column=c_idx).value is None for c_idx in range(1, sheet.max_column + 1))
            if is_empty:
                continue
                
            sheet_original_rows += 1
            if r_idx not in flagged_rows:
                row_vals = [sheet.cell(row=r_idx, column=c_idx).value for c_idx in range(1, sheet.max_column + 1)]
                ws_clean.append(row_vals)
                sheet_clean_rows += 1
                
        total_clean_rows += sheet_clean_rows
        total_original_rows += sheet_original_rows
        first_sheet = False
        
    # Sheet 3: Summary Metrics
    ws_summary = qa_wb.create_sheet(title="Summary Metrics")
    ws_summary.append(["QA Summary Metrics", ""])
    ws_summary.cell(row=1, column=1).font = Font(size=14, bold=True, color="1F497D")
    ws_summary.append([])
    
    ws_summary.append(["Metric", "Value"])
    ws_summary.cell(row=3, column=1).font = Font(bold=True)
    ws_summary.cell(row=3, column=2).font = Font(bold=True)
    
    ws_summary.append(["Total Checked Rows", total_original_rows])
    ws_summary.append(["Clean Rows (No Issues)", total_clean_rows])
    ws_summary.append(["Flagged Rows (With Issues)", total_original_rows - total_clean_rows])
    ws_summary.append([])
    
    ws_summary.append(["Issue Breakdown by Rule / Code", "Count"])
    ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True, color="1F497D")
    
    for rule_code, count in sorted(issues_by_code.items(), key=lambda x: x[1], reverse=True):
        ws_summary.append([rule_code, count])
        
    # Formatting adjustments (auto-fit columns)
    for ws in qa_wb.worksheets:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)
            
    # Save output
    if args.output:
        out_path = Path(args.output)
    else:
        out_path = Path(__file__).resolve().parent.parent / "validated_runs" / "BrandDefinitionReport_QA.xlsx"
        
    out_path.parent.mkdir(parents=True, exist_ok=True)
    qa_wb.save(out_path)
    print(f"QA Findings Report successfully saved to: {out_path}")
    print(f"Total Rows: {total_original_rows} | Clean: {total_clean_rows} | Flagged: {total_original_rows - total_clean_rows}")

if __name__ == "__main__":
    main()
