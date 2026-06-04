import asyncio
import os
import sys
import openpyxl
from datetime import date, datetime
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# Ensure settings are loaded with high refresh hours to avoid downloading dataset
os.environ["IMDB_DATASET_REFRESH_HOURS"] = "8760"

from metacritic_calendar_app.services.billboard import BillboardService
from metacritic_calendar_app.services.imdb_episode_counts import TvImdbEpisodeCountService
from imdb_lookup_app.services.lookup import ImdbLookupService

BILLBOARD_FILE = Path(__file__).resolve().parent.parent / "Billboard_Top_Artists.xlsx"
TV_FILE = Path(__file__).resolve().parent.parent / "TV-Seasons-and-Episodes.xlsx"

def format_date_dd_mm_yyyy(date_str: str) -> str:
    if not date_str or date_str == "-":
        return ""
    try:
        # If in YYYY-MM-DD, convert to DD-MM-YYYY
        if "-" in date_str:
            parts = date_str.split("-")
            if len(parts) == 3 and len(parts[0]) == 4:  # YYYY-MM-DD
                return f"{parts[2]}-{parts[1]}-{parts[0]}"
        return date_str
    except Exception:
        return date_str

def format_imdb_profession(prof: str) -> str:
    if not prof:
        return ""
    # Convert 'music_artist,actor,producer' -> 'Music Artist; Actor; Producer'
    cleaned = prof.replace("_", " ").title()
    parts = [p.strip() for p in cleaned.split(",")]
    return "; ".join(parts)

async def run_billboard_ingestion():
    print("\n--- Starting Billboard Artist 100 Ingestion ---")
    if not BILLBOARD_FILE.exists():
        print(f"Error: {BILLBOARD_FILE} does not exist at {BILLBOARD_FILE}!")
        return

    # Fetch top artists and resolve their details
    billboard_service = BillboardService()
    print("Fetching top artists from Billboard Artist 100...")
    snapshot = await billboard_service.get_top_artists_snapshot()
    print(f"Fetched {len(snapshot.items)} artist items.")

    # Load existing workbook
    wb = openpyxl.load_workbook(BILLBOARD_FILE)
    sheet = wb.active
    print(f"Loaded workbook {BILLBOARD_FILE}. Active sheet: {sheet.title}. Max row before: {sheet.max_row}")

    # Truncate sheet back to 8 rows to clear previous run data (keeping header + 7 original example rows)
    if sheet.max_row > 8:
        print(f"Truncating spreadsheet from {sheet.max_row} rows down to 8 rows to clear previous run data...")
        sheet.delete_rows(9, sheet.max_row - 8)

    imdb_service = ImdbLookupService()

    appended_count = 0
    for idx, item in enumerate(snapshot.items, 1):
        # Resolve IMDb Primary Profession from local lookup if we have imdb_id
        imdb_prof = ""
        if item.imdb_id:
            batch_res = imdb_service.lookup_values([item.imdb_id], mode="id_to_name")
            if batch_res.rows:
                imdb_prof = format_imdb_profession(batch_res.rows[0].primary_profession)

        imdb_url = f"https://www.imdb.com/name/{item.imdb_id}/" if item.imdb_id else ""
        bb_url = f"https://www.billboard.com/artist/{item.slug}/"

        row_data = [
            item.rank,
            item.name,
            item.imdb_id,
            imdb_url,
            imdb_prof,
            item.wikipedia_url,
            item.gender,
            item.profession,
            bb_url
        ]
        sheet.append(row_data)
        appended_count += 1

    wb.save(BILLBOARD_FILE)
    print(f"Saved {BILLBOARD_FILE}. Max row after: {sheet.max_row}. Appended {appended_count} rows.")

def run_tv_ingestion():
    print("\n--- Starting TV Releases Ingestion ---")
    if not TV_FILE.exists():
        print(f"Error: {TV_FILE} does not exist!")
        return

    # Fetch releases for rolling weekend of Monday, June 1, 2026
    anchor_date = date(2026, 6, 1)
    tv_service = TvImdbEpisodeCountService()
    print(f"Fetching TV releases for rolling weekend around {anchor_date}...")
    snapshot = tv_service.fetch_snapshot(date_window="daily_segment", today=anchor_date)
    print(f"Fetched {len(snapshot.items)} releases.")

    # Load existing workbook
    wb = openpyxl.load_workbook(TV_FILE)
    # Ensure sheet 'Export' exists or is active
    sheet = wb["Export"] if "Export" in wb.sheetnames else wb.active
    print(f"Loaded workbook {TV_FILE}. Sheet name: {sheet.title}. Max row before: {sheet.max_row}")

    # Truncate sheet back to 18 rows to clear previous run data (keeping header + 17 original example rows)
    if sheet.max_row > 18:
        print(f"Truncating spreadsheet from {sheet.max_row} rows down to 18 rows to clear previous run data...")
        sheet.delete_rows(19, sheet.max_row - 18)

    appended_count = 0
    for item in snapshot.items:
        row_data = [
            format_date_dd_mm_yyyy(item.release_date),
            item.title,
            item.network_distributor,
            item.imdb_id,
            item.metacritic_url,
            item.latest_season_number if item.latest_season_number is not None else "",
            item.latest_season_episode_count if item.latest_season_episode_count is not None else "",
            format_date_dd_mm_yyyy(item.latest_season_start_date),
            format_date_dd_mm_yyyy(item.latest_season_end_date)
        ]
        sheet.append(row_data)
        appended_count += 1

    wb.save(TV_FILE)
    print(f"Saved {TV_FILE}. Max row after: {sheet.max_row}. Appended {appended_count} rows.")

async def main():
    await run_billboard_ingestion()
    run_tv_ingestion()
    print("\n--- Monday Ingestion Complete! ---")

if __name__ == "__main__":
    asyncio.run(main())
