import re
import html
import time
import sqlite3
import asyncio
import difflib
from pathlib import Path
from urllib.parse import quote
import httpx
from pydantic import BaseModel, Field
from datetime import datetime

import os

from app.config import BASE_DIR, is_vercel

DB_PATH = Path("/tmp/wikipedia_cache.sqlite3") if is_vercel else BASE_DIR / "data" / "wikipedia_cache" / "wikipedia_cache.sqlite3"
REFERENCE_XLSX = BASE_DIR / "Billboard_Top_Artists.xlsx"
# A "new entry" on Billboard is rendered with a literal "-" in the LW column.
# We intentionally do NOT treat empty strings, em/en dashes, or "NEW"/"N/A"
# placeholders as new entries - an empty value usually means the LW regex
# failed to capture rather than a genuine new-entry signal.
NEW_ENTRY_TOKENS = {"-"}
FUZZY_MATCH_THRESHOLD = 0.85

class BillboardArtistItem(BaseModel):
    rank: int
    name: str
    slug: str
    gender: str = ""
    profession: str = ""
    imdb_id: str = ""
    imdb_url: str = ""
    imdb_primary_profession: str = ""
    wikipedia_url: str = ""
    billboard_url: str = ""
    last_week: str = ""
    peak_position: str = ""
    weeks_on_chart: str = ""
    is_new_entry: bool = False
    in_reference: bool = False
    reference_match: str = ""
    reference_match_score: float = 0.0

class BillboardArtistSnapshot(BaseModel):
    generated_at: datetime
    export_id: str | None = None
    items: list[BillboardArtistItem] = Field(default_factory=list)
    notes: list[str] = Field(default_factory=list)
    new_entry_count: int = 0

def ensure_cache_table():
    try:
        DB_PATH.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS billboard_artist_cache (
                    name TEXT PRIMARY KEY,
                    slug TEXT,
                    gender TEXT,
                    profession TEXT,
                    imdb_id TEXT,
                    wikipedia_url TEXT,
                    resolved_at INTEGER
                )
            """)
            conn.commit()
    except Exception as e:
        print(f"Failed to ensure cache table: {e}")

def get_cached_artist(name: str) -> dict | None:
    try:
        with sqlite3.connect(DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            row = conn.execute("SELECT * FROM billboard_artist_cache WHERE LOWER(name) = ?", (name.lower(),)).fetchone()
            return dict(row) if row else None
    except Exception:
        return None

def cache_artist(name: str, slug: str, gender: str, profession: str, imdb_id: str, wikipedia_url: str):
    try:
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute("""
                INSERT OR REPLACE INTO billboard_artist_cache (name, slug, gender, profession, imdb_id, wikipedia_url, resolved_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (name, slug, gender, profession, imdb_id, wikipedia_url, int(time.time())))
            conn.commit()
    except Exception as e:
        print(f"Failed to cache artist {name}: {e}")

class BillboardService:
    CHART_URL = "https://www.billboard.com/charts/artist-100/"
    WIKIDATA_API = "https://www.wikidata.org/w/api.php"
    
    def __init__(self, request_timeout_seconds: int = 15):
        self.timeout = request_timeout_seconds
        self.headers = {
            "User-Agent": "OfficialProfileFinder/0.1 (test@example.com) Python/httpx"
        }
        ensure_cache_table()

    async def fetch_billboard_artists(self) -> list[dict]:
        """Scrapes the Billboard Artist 100 chart.

        Returns a list of dicts with keys: rank, name, slug, last_week.
        ``last_week`` is the raw token from the LW column ("-" for new entries,
        otherwise the prior week's position as a string).
        """
        async with httpx.AsyncClient(headers=self.headers, timeout=self.timeout, follow_redirects=True) as client:
            r = await client.get(self.CHART_URL)
            if r.status_code != 200:
                raise RuntimeError(f"Billboard returned status code {r.status_code}")

            html_content = r.text
            rows = html_content.split('o-chart-results-list-row-container')

            seen = set()
            artists: list[dict] = []
            for rank, part in enumerate(rows[1:101], start=1):
                match = re.search(r'<h3[^>]+class="[^"]*c-title[^"]*"[^>]*>(.*?)</h3>', part, re.DOTALL)
                if not match:
                    continue
                h3_content = match.group(1).strip()
                a_match = re.search(r'<a[^>]*>(.*?)</a>', h3_content, re.DOTALL)
                name = a_match.group(1).strip() if a_match else h3_content
                name = re.sub(r'<[^>]+>', '', name).strip()
                unescaped_name = html.unescape(name).strip()
                if not unescaped_name:
                    continue

                slug = ""
                link_match = re.search(r'href="[^"]*/artist/([^/"]+)', h3_content)
                if link_match:
                    slug = link_match.group(1).strip()
                else:
                    slug = unescaped_name.lower().replace(" ", "-").replace("&", "and").replace(".", "")
                    slug = "".join(c for c in slug if c.isalnum() or c == "-")

                last_week = self._extract_chart_metric(part, ("LW", "LAST WEEK"))
                peak_position = self._extract_chart_metric(part, ("PEAK", "PEAK POS", "PEAK POSITION"))
                weeks_on_chart = self._extract_chart_metric(part, ("WEEKS", "WKS ON CHART", "WEEKS ON CHART"))

                if unescaped_name.lower() in seen:
                    continue
                seen.add(unescaped_name.lower())
                artists.append({
                    "rank": rank,
                    "name": unescaped_name,
                    "slug": slug,
                    "last_week": last_week,
                    "peak_position": peak_position,
                    "weeks_on_chart": weeks_on_chart,
                })
            return artists

    @staticmethod
    def _extract_chart_metric(row_html: str, label_aliases: tuple[str, ...]) -> str:
        """Extract a labeled chart metric (LW / PEAK / WEEKS) from a row.

        Billboard renders each metric as a ``c-span`` label (e.g. ``LW`` or
        ``PEAK``) followed by an ``<li>`` whose first ``c-label`` carries the
        value. The value is "-" for a new entry on LW, otherwise the numeric
        position or weeks-on-chart count.
        """
        # Build an alternation of escaped aliases with flexible whitespace,
        # so "PEAK POS" matches "PEAK<br>POS." or "PEAK POS." in the HTML.
        alts = "|".join(
            r"\s*".join(re.escape(token) for token in alias.split())
            for alias in label_aliases
        )
        pattern = (
            r'<span[^>]*class="c-span[^"]*"[^>]*>\s*(?:' + alts + r')\s*\.?\s*</span>'
            r'.*?<span[^>]*class="c-label[^"]*"[^>]*>\s*([^<]+?)\s*</span>'
        )
        m = re.search(pattern, row_html, re.DOTALL | re.IGNORECASE)
        if not m:
            return ""
        return html.unescape(m.group(1)).strip()

    @staticmethod
    def _is_new_entry(last_week: str) -> bool:
        return last_week.strip().lower() in NEW_ENTRY_TOKENS

    @staticmethod
    def _imdb_url(imdb_id: str) -> str:
        return f"https://www.imdb.com/name/{imdb_id}/" if imdb_id else ""

    @staticmethod
    def _billboard_url(slug: str) -> str:
        return f"https://www.billboard.com/artist/{slug}/" if slug else ""

    async def _make_wikidata_request(self, client: httpx.AsyncClient, params: dict) -> dict | None:
        max_retries = 5
        backoff = 1.5
        for attempt in range(max_retries):
            try:
                r = await client.get(self.WIKIDATA_API, params=params)
                if r.status_code == 429:
                    print(f"Wikidata rate limit (429) hit for {params.get('search') or params.get('ids')}. Retrying in {backoff}s...")
                    await asyncio.sleep(backoff)
                    backoff *= 2.0
                    continue
                if r.status_code != 200:
                    print(f"Wikidata request failed with status {r.status_code}")
                    return None
                return r.json()
            except (httpx.HTTPError, asyncio.TimeoutError) as exc:
                print(f"Wikidata request exception: {exc}. Retrying in {backoff}s...")
                await asyncio.sleep(backoff)
                backoff *= 2.0
        return None

    def _select_best_qid(self, search_hits: list[dict], artist_name: str) -> str | None:
        if not search_hits:
            return None
            
        best_qid = None
        best_score = -9999
        
        artist_keywords = [
            "singer", "rapper", "musician", "band", "music group", "musical group", 
            "composer", "songwriter", "vocalist", "disc jockey", "dj", "boy band", 
            "girl group", "orchestra", "duo", "trio", "quartet", "quintet", "music project", 
            "performing arts group", "rock group", "metal group", "pop group", "hip hop group", 
            "indie group"
        ]
        negative_keywords = [
            "album by", "song by", "single by", "track by", "discography", "tour by", 
            "given name", "family name", "disambiguation page", "surname", "chemical compound", 
            "film by", "novel by", "book by"
        ]
        other_positive_keywords = [
            "actor", "actress", "celebrity", "entertainer", "human", "person", 
            "music", "musical", "album", "song", "single"
        ]
        
        for hit in search_hits:
            qid = hit.get("id")
            label = hit.get("label", "").lower()
            desc = hit.get("description", "").lower()
            
            score = 0
            if label == artist_name.lower():
                score += 5
                
            for kw in artist_keywords:
                if kw in desc:
                    score += 10
                    break
                    
            for kw in other_positive_keywords:
                if kw in desc:
                    score += 2
                    break
                    
            for kw in negative_keywords:
                if kw in desc:
                    score -= 15
                    break
                    
            if score > best_score:
                best_score = score
                best_qid = qid
                
        return best_qid

    async def _resolve_imdb_id_via_serpapi(self, name: str) -> dict | None:
        """Resolves IMDb ID and Wikipedia URL using SerpApi search."""
        from app.config import settings
        if not settings.serpapi_api_key:
            return None
            
        import httpx
        import re
        
        type_filter = "site:imdb.com"
        queries = [
            f"{name} music artist {type_filter}",
            f"{name} IMDb",
            name
        ]
        
        res = {"imdb_id": "", "wikipedia_url": ""}
        
        async with httpx.AsyncClient(timeout=10, headers=self.headers) as client:
            for query in queries:
                params = {
                    "engine": settings.serpapi_engine or "google",
                    "q": query,
                    "api_key": settings.serpapi_api_key,
                    "num": "5",
                    "hl": "en",
                    "gl": "us",
                }
                try:
                    resp = await client.get("https://serpapi.com/search.json", params=params)
                    if resp.status_code == 200:
                        payload = resp.json()
                        
                        # Check knowledge graph profiles
                        kg = payload.get("knowledge_graph", {})
                        for key in ("profiles", "social_profiles"):
                            for item in kg.get(key) or []:
                                link = item.get("link", "")
                                match = re.search(r"/name/(nm\d+)", link)
                                if match:
                                    res["imdb_id"] = match.group(1)
                                    break
                            if res["imdb_id"]:
                                break
                                
                        # Scan organic results
                        for result in payload.get("organic_results", []):
                            link = result.get("link", "")
                            if not res["imdb_id"]:
                                match = re.search(r"/name/(nm\d+)", link)
                                if match:
                                    res["imdb_id"] = match.group(1)
                            if not res["wikipedia_url"] and "en.wikipedia.org/wiki/" in link:
                                res["wikipedia_url"] = link
                                
                        if res["imdb_id"]:
                            return res
                except Exception as e:
                    print(f"SerpApi query failed for '{query}': {e}")
        return None if not res["imdb_id"] else res

    async def resolve_artist_details(self, name: str, slug: str) -> dict:
        """Resolves gender, profession, IMDb ID and Wikipedia URL for an artist name, using cache when available."""
        cached = get_cached_artist(name)
        if cached:
            return cached

        # Resolve from Wikidata
        result = {
            "name": name,
            "slug": slug,
            "gender": "Unknown",
            "profession": "Unknown",
            "imdb_id": "",
            "wikipedia_url": ""
        }

        try:
            async with httpx.AsyncClient(headers=self.headers, timeout=self.timeout) as client:
                # 1. Search Wikidata
                search_params = {
                    "action": "wbsearchentities",
                    "format": "json",
                    "language": "en",
                    "type": "item",
                    "limit": 3,
                    "search": name
                }
                search_data = await self._make_wikidata_request(client, search_params)
                if not search_data:
                    # Try SerpApi fallback
                    serp_res = await self._resolve_imdb_id_via_serpapi(name)
                    if serp_res:
                        result["imdb_id"] = serp_res["imdb_id"]
                        result["wikipedia_url"] = serp_res["wikipedia_url"]
                    cache_artist(name, slug, result["gender"], result["profession"], result["imdb_id"], result["wikipedia_url"])
                    return result
                
                search_hits = search_data.get("search", [])
                if not search_hits:
                    # Try SerpApi fallback
                    serp_res = await self._resolve_imdb_id_via_serpapi(name)
                    if serp_res:
                        result["imdb_id"] = serp_res["imdb_id"]
                        result["wikipedia_url"] = serp_res["wikipedia_url"]
                    cache_artist(name, slug, result["gender"], result["profession"], result["imdb_id"], result["wikipedia_url"])
                    return result
                
                # Take best result using disambiguation scoring
                qid = self._select_best_qid(search_hits, name)
                if not qid:
                    qid = search_hits[0].get("id")
                
                # 2. Get claims and sitelinks
                entity_params = {
                    "action": "wbgetentities",
                    "format": "json",
                    "ids": qid,
                    "props": "claims|sitelinks",
                    "languages": "en"
                }
                entity_data = await self._make_wikidata_request(client, entity_params)
                if not entity_data:
                    # Try SerpApi fallback
                    serp_res = await self._resolve_imdb_id_via_serpapi(name)
                    if serp_res:
                        result["imdb_id"] = serp_res["imdb_id"]
                        result["wikipedia_url"] = serp_res["wikipedia_url"]
                    cache_artist(name, slug, result["gender"], result["profession"], result["imdb_id"], result["wikipedia_url"])
                    return result
                
                entity = entity_data.get("entities", {}).get(qid, {})
                
                # Sitelinks
                enwiki_title = entity.get("sitelinks", {}).get("enwiki", {}).get("title")
                if enwiki_title:
                    result["wikipedia_url"] = f"https://en.wikipedia.org/wiki/{quote(enwiki_title.replace(' ', '_'))}"
                
                claims = entity.get("claims", {})
                
                # IMDb ID (P345)
                imdb_claims = claims.get("P345", [])
                if imdb_claims:
                    result["imdb_id"] = imdb_claims[0].get("mainsnak", {}).get("datavalue", {}).get("value", "")
                
                # Gender QID (P21)
                gender_claims = claims.get("P21", [])
                gender_qid = ""
                if gender_claims:
                    gender_qid = gender_claims[0].get("mainsnak", {}).get("datavalue", {}).get("value", {}).get("id", "")
                
                # Profession QIDs (P106)
                profession_claims = claims.get("P106", [])
                prof_qids = []
                for claim in profession_claims[:3]:
                    pqid = claim.get("mainsnak", {}).get("datavalue", {}).get("value", {}).get("id", "")
                    if pqid:
                        prof_qids.append(pqid)
                
                # Resolve labels
                qids_to_resolve = [qid for qid in ([gender_qid] + prof_qids) if qid]
                if qids_to_resolve:
                    resolve_params = {
                        "action": "wbgetentities",
                        "format": "json",
                        "ids": "|".join(qids_to_resolve),
                        "props": "labels",
                        "languages": "en"
                    }
                    resolved_data = await self._make_wikidata_request(client, resolve_params)
                    if resolved_data:
                        resolved_entities = resolved_data.get("entities", {})
                        
                        if gender_qid:
                            result["gender"] = resolved_entities.get(gender_qid, {}).get("labels", {}).get("en", {}).get("value", "Unknown")
                        
                        prof_labels = []
                        for pqid in prof_qids:
                            pl = resolved_entities.get(pqid, {}).get("labels", {}).get("en", {}).get("value", "")
                            if pl:
                                prof_labels.append(pl)
                        if prof_labels:
                            result["profession"] = ", ".join(prof_labels)
                            
                # Fallback to SerpApi if IMDb ID is still missing
                if not result["imdb_id"]:
                    serp_res = await self._resolve_imdb_id_via_serpapi(name)
                    if serp_res:
                        result["imdb_id"] = serp_res["imdb_id"]
                        if not result["wikipedia_url"] and serp_res["wikipedia_url"]:
                            result["wikipedia_url"] = serp_res["wikipedia_url"]
            
            # Cache the result
            cache_artist(name, slug, result["gender"], result["profession"], result["imdb_id"], result["wikipedia_url"])
        except Exception as e:
            print(f"Error resolving details for {name}: {e}")
            
        return result

    @staticmethod
    def _load_reference_artists() -> dict[str, dict]:
        """Load the reference Billboard artist roster from the workbook at the repo root.

        Returns a mapping of normalized lowercase name -> reference row dict with
        keys: name, imdb_id, imdb_url, profession, wikipedia_url, gender,
        occupations, billboard_url.
        """
        reference: dict[str, dict] = {}
        if not REFERENCE_XLSX.exists():
            return reference
        try:
            from openpyxl import load_workbook
            wb = load_workbook(REFERENCE_XLSX, read_only=True, data_only=True)
            sheet = wb.active
            header_row = next(sheet.iter_rows(max_row=1, values_only=True), None)
            if not header_row:
                wb.close()
                return reference
            header = [str(c).strip().lower() if c is not None else "" for c in header_row]

            def idx_of(*candidates: str) -> int:
                for cand in candidates:
                    if cand in header:
                        return header.index(cand)
                return -1

            i_name = idx_of("artist name", "name")
            i_imdb = idx_of("imdb nmcode", "imdb id")
            i_imdb_url = idx_of("imdb url")
            i_prof = idx_of("imdb primary profession", "profession")
            i_wiki = idx_of("wikipedia url", "wikipedia")
            i_gender = idx_of("gender")
            i_occ = idx_of("occupations", "occupation")
            i_billboard = idx_of("billboard artist url", "billboard url")

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or i_name < 0 or i_name >= len(row):
                    continue
                name = str(row[i_name]).strip() if row[i_name] is not None else ""
                if not name:
                    continue

                def cell(i: int) -> str:
                    if i < 0 or i >= len(row) or row[i] is None:
                        return ""
                    return str(row[i]).strip()

                reference[name.lower()] = {
                    "name": name,
                    "imdb_id": cell(i_imdb),
                    "imdb_url": cell(i_imdb_url),
                    "profession": cell(i_prof),
                    "wikipedia_url": cell(i_wiki),
                    "gender": cell(i_gender),
                    "occupations": cell(i_occ),
                    "billboard_url": cell(i_billboard),
                }
            wb.close()
        except Exception as e:
            print(f"Failed to load Billboard reference workbook: {e}")
        return reference

    NEW_ENTRY_EXPORT_HEADERS = [
        "Rank", "Last Week", "Peak", "Weeks on Chart",
        "Artist Name", "IMDb nmcode", "IMDb URL",
        "IMDb Primary Profession", "Wikipedia URL", "Gender",
        "Occupations", "Billboard Artist URL",
    ]

    @classmethod
    def export_new_entries_xlsx(cls, snapshot: "BillboardArtistSnapshot") -> bytes:
        """Render the new-entries snapshot as an xlsx workbook.

        Columns (in order): Rank, Last Week, Peak, Weeks on Chart, Artist Name,
        IMDb nmcode, IMDb URL, IMDb Primary Profession, Wikipedia URL, Gender,
        Occupations, Billboard Artist URL.
        """
        from io import BytesIO
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Billboard New Entries"
        ws.append(cls.NEW_ENTRY_EXPORT_HEADERS)
        for item in snapshot.items:
            ws.append([
                item.rank,
                item.last_week,
                item.peak_position,
                item.weeks_on_chart,
                item.name,
                item.imdb_id,
                item.imdb_url,
                item.imdb_primary_profession,
                item.wikipedia_url,
                item.gender,
                item.profession,
                item.billboard_url,
            ])
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def _fuzzy_match(name: str, reference: dict[str, dict]) -> tuple[dict | None, float]:
        if not reference:
            return None, 0.0
        target = name.strip().lower()
        if target in reference:
            return reference[target], 1.0
        candidates = list(reference.keys())
        match = difflib.get_close_matches(target, candidates, n=1, cutoff=FUZZY_MATCH_THRESHOLD)
        if not match:
            return None, 0.0
        score = difflib.SequenceMatcher(None, target, match[0]).ratio()
        return reference[match[0]], score

    async def get_top_artists_snapshot(self) -> BillboardArtistSnapshot:
        """Fetch Billboard Artist 100 and resolve detail profiles for all artists in parallel."""
        raw_artists = await self.fetch_billboard_artists()
        reference = self._load_reference_artists()

        semaphore = asyncio.Semaphore(10)

        async def resolve_with_sem(entry):
            async with semaphore:
                return await self.resolve_artist_details(entry["name"], entry["slug"])

        resolved_results = await asyncio.gather(*[resolve_with_sem(e) for e in raw_artists])

        items: list[BillboardArtistItem] = []
        new_count = 0
        for entry, res in zip(raw_artists, resolved_results):
            ref_row, score = self._fuzzy_match(entry["name"], reference)
            is_new = self._is_new_entry(entry["last_week"])
            if is_new:
                new_count += 1
            imdb_id = res["imdb_id"] or (ref_row["imdb_id"] if ref_row else "")
            items.append(
                BillboardArtistItem(
                    rank=entry["rank"],
                    name=res["name"],
                    slug=res["slug"],
                    gender=res["gender"] or (ref_row["gender"] if ref_row else ""),
                    profession=res["profession"] or (ref_row["occupations"] if ref_row else ""),
                    imdb_id=imdb_id,
                    imdb_url=self._imdb_url(imdb_id),
                    imdb_primary_profession=(ref_row["profession"] if ref_row else ""),
                    wikipedia_url=res["wikipedia_url"] or (ref_row["wikipedia_url"] if ref_row else ""),
                    billboard_url=self._billboard_url(res["slug"]),
                    last_week=entry["last_week"],
                    peak_position=entry["peak_position"],
                    weeks_on_chart=entry["weeks_on_chart"],
                    is_new_entry=is_new,
                    in_reference=ref_row is not None,
                    reference_match=ref_row["name"] if ref_row else "",
                    reference_match_score=round(score, 3),
                )
            )

        return BillboardArtistSnapshot(
            generated_at=datetime.now().astimezone(),
            items=items,
            new_entry_count=new_count,
            notes=[
                "Source: Billboard Artist 100 Chart.",
                "Artist gender, professions, IMDb ID, and Wikipedia URLs resolved dynamically via Wikidata API.",
                f"{new_count} of {len(items)} entries are new this week (Last Week = '-').",
            ],
        )

    async def get_new_entries_snapshot(self) -> BillboardArtistSnapshot:
        """Fetch only new entries on the current Artist 100 chart with full metadata.

        A 'new entry' is a row whose Last Week column is '-' (or equivalent
        placeholder). Each new entry is fuzzy-matched against the local
        reference workbook so callers can tell which new chart entries are
        genuinely new to the company's tracked roster.
        """
        raw_artists = await self.fetch_billboard_artists()
        new_raw = [e for e in raw_artists if self._is_new_entry(e["last_week"])]
        reference = self._load_reference_artists()

        semaphore = asyncio.Semaphore(10)

        async def resolve_with_sem(entry):
            async with semaphore:
                return await self.resolve_artist_details(entry["name"], entry["slug"])

        resolved_results = await asyncio.gather(*[resolve_with_sem(e) for e in new_raw])

        items: list[BillboardArtistItem] = []
        for entry, res in zip(new_raw, resolved_results):
            ref_row, score = self._fuzzy_match(entry["name"], reference)
            gender = res["gender"] or (ref_row["gender"] if ref_row else "")
            occupations = res["profession"] or (ref_row["occupations"] if ref_row else "")
            imdb_primary = ref_row["profession"] if ref_row else ""
            imdb_id = res["imdb_id"] or (ref_row["imdb_id"] if ref_row else "")
            wikipedia_url = res["wikipedia_url"] or (ref_row["wikipedia_url"] if ref_row else "")
            items.append(
                BillboardArtistItem(
                    rank=entry["rank"],
                    name=res["name"],
                    slug=res["slug"],
                    gender=gender,
                    profession=occupations,
                    imdb_id=imdb_id,
                    imdb_url=self._imdb_url(imdb_id),
                    imdb_primary_profession=imdb_primary,
                    wikipedia_url=wikipedia_url,
                    billboard_url=self._billboard_url(res["slug"]),
                    last_week=entry["last_week"],
                    peak_position=entry["peak_position"],
                    weeks_on_chart=entry["weeks_on_chart"],
                    is_new_entry=True,
                    in_reference=ref_row is not None,
                    reference_match=ref_row["name"] if ref_row else "",
                    reference_match_score=round(score, 3),
                )
            )

        return BillboardArtistSnapshot(
            generated_at=datetime.now().astimezone(),
            items=items,
            new_entry_count=len(items),
            notes=[
                "Source: Billboard Artist 100 Chart - new entries only.",
                f"{len(items)} new entries detected this week (Last Week = '-').",
                f"{sum(1 for i in items if not i.in_reference)} of those are not yet in the reference roster.",
                "Gender/profession/IMDb/Wikipedia resolved via Wikidata with reference workbook fallback.",
            ],
        )
