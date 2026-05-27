from __future__ import annotations

import csv
import html
import io
import json
import re
from datetime import date, datetime
from typing import Any
from urllib.parse import urljoin

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from metacritic_calendar_app.models import (
    MetacriticTvClassificationItem,
    MetacriticTvClassificationSnapshot,
)
from metacritic_calendar_app.services.calendar import MetacriticCalendarError, MetacriticCalendarService
from metacritic_calendar_app.services.text import contains_rent_buy, repair_mojibake
from title_url_lookup_app.models import TitleLookupQuery
from title_url_lookup_app.services.imdb_dataset import ImdbDatasetLookupError, ImdbDatasetLookupService


TV_CLASSIFICATION_OUTPUT_COLUMNS = [
    "release_date",
    "title",
    "imdb_ttcode",
    "network",
    "daypart",
    "program_type",
    "language_type",
    "genre_1",
    "genre_2",
    "genre_3",
]

TV_CLASSIFICATION_EXPORT_HEADERS = [
    "Release Date",
    "Title",
    "IMDb ttcode",
    "Network",
    "Daypart",
    "Program Type",
    "Language Type",
    "Genre 1",
    "Genre 2",
    "Genre 3",
]

DAYPART_OTHER = "Daypart - Other"
DAYPART_PRIME_TIME = "Daypart - Prime Time (8:00pm-11:00pm (M-Sat) / 7:00pm-11:00pm (Sun))"
PROGRAM_TYPE_OTHER = "Program Type - Other"
PROGRAM_TYPE_SERIES = "Program Type - Series"
PROGRAM_TYPE_MINI_SERIES = "Program Type - Mini-Series"
PROGRAM_TYPE_MOVIE = "Program Type - Movie"
LANGUAGE_TYPE_ENGLISH = "Language Type - English"
LANGUAGE_TYPE_OTHER = "Language Type - Other"


class MetacriticTvClassificationReportService:
    BASE_URL = "https://www.metacritic.com"
    TABLE_ROW_RE = re.compile(r"<tr[^>]*>(.*?)</tr>", re.IGNORECASE | re.DOTALL)
    CELL_RE = re.compile(r"<td[^>]*>(.*?)</td>", re.IGNORECASE | re.DOTALL)
    LINK_RE = re.compile(r'<a [^>]*href="(?P<href>[^"]+)"[^>]*>(?P<label>.*?)</a>', re.IGNORECASE | re.DOTALL)
    STRONG_RE = re.compile(r"<strong[^>]*>(.*?)</strong>", re.IGNORECASE | re.DOTALL)
    IMG_ALT_RE = re.compile(r'<img\b[^>]*\balt="(?P<alt>[^"]*)"', re.IGNORECASE | re.DOTALL)
    SHORTCODE_API_RE = re.compile(r'api="([^"]+)"', re.IGNORECASE | re.DOTALL)
    SHORTCODE_RE = re.compile(r"<shortcode\b[^>]*>.*?</shortcode>", re.IGNORECASE | re.DOTALL)
    BR_RE = re.compile(r"<br\s*/?>", re.IGNORECASE)
    IMG_RE = re.compile(r"<img\b[^>]*>", re.IGNORECASE)
    TAG_RE = re.compile(r"<[^>]+>")
    WHITESPACE_RE = re.compile(r"\s+")
    TRAILER_RE = re.compile(r"\b(?:color\s+|b&w\s+)?trailer\b", re.IGNORECASE)
    AIR_TIME_RE = re.compile(
        r"\b(?P<hour>1[0-2]|[1-9])(?::(?P<minute>\d{2}))?\s*(?P<meridiem>[ap])\b",
        re.IGNORECASE,
    )
    NETWORK_TIME_TAIL_RE = re.compile(
        r"\b(?:1[0-2]|[1-9])(?::\d{2})?\s*[ap]\b.*$",
        re.IGNORECASE,
    )

    def __init__(
        self,
        calendar_service: MetacriticCalendarService | None = None,
        imdb_dataset_lookup: ImdbDatasetLookupService | None = None,
    ) -> None:
        self.calendar_service = calendar_service or MetacriticCalendarService()
        self.imdb_dataset_lookup = imdb_dataset_lookup or ImdbDatasetLookupService()

    def fetch_snapshot(self, today: date | None = None) -> MetacriticTvClassificationSnapshot:
        target = self.calendar_service.TARGETS["tv"]
        window_start = today or datetime.now().astimezone().date()
        window_end = self._add_months(window_start, 3)
        with httpx.Client(
            timeout=self.calendar_service.timeout_seconds,
            headers=self.calendar_service._build_headers(),
            follow_redirects=True,
        ) as client:
            payload = self.calendar_service._fetch_calendar_payload(client, target)

        parsed_items = self.parse_payload(payload)
        non_rent_buy_items, skipped_rent_buy = self._filter_rent_buy_items(parsed_items)
        items, skipped_outside_window, skipped_without_date = self._filter_items_for_window(
            non_rent_buy_items,
            window_start,
            window_end,
        )
        imdb_not_found_count, imdb_error = self._attach_imdb_ttcodes(items)
        notes = [
            "Source: Metacritic TV premiere calendar.",
            f"Date window: {window_start.isoformat()} to {window_end.isoformat()}.",
            "Includes Metacritic movie, series, mini-series, and other program-type rows in the selected window.",
        ]
        if skipped_outside_window:
            notes.append(f"{skipped_outside_window} parsed TV row(s) were outside the 3-month date window.")
        if skipped_without_date:
            notes.append(f"{skipped_without_date} parsed TV row(s) were skipped because the release date was unavailable.")
        if skipped_rent_buy:
            notes.append(f"{skipped_rent_buy} parsed TV row(s) were skipped because the availability is Rent/Buy.")
        if imdb_not_found_count:
            notes.append(f"{imdb_not_found_count} row(s) did not receive an IMDb ttcode.")
        if imdb_error:
            notes.append(f"IMDb ttcode lookup failed: {imdb_error}")
        if not items:
            notes.append("No TV classification rows were found in the 3-month date window.")

        return MetacriticTvClassificationSnapshot(
            generated_at=datetime.now().astimezone(),
            source_url=target.source_url,
            window_start=window_start,
            window_end=window_end,
            items=items,
            notes=notes,
        )

    def parse_payload(self, payload: dict[str, Any]) -> list[MetacriticTvClassificationItem]:
        item = payload.get("data", {}).get("item")
        if not isinstance(item, dict):
            raise MetacriticCalendarError("Metacritic TV payload did not include article data.")

        body_html = str(item.get("body") or "")
        if not body_html:
            raise MetacriticCalendarError("Metacritic TV payload did not include article body HTML.")

        base_year = self.calendar_service._resolve_base_year(item)
        current_year = base_year
        last_release_date: datetime | None = None
        parsed_items: list[MetacriticTvClassificationItem] = []

        headings = list(self.calendar_service.HEADING_RE.finditer(body_html))
        for index, match in enumerate(headings):
            tag = match.group(1).lower()
            heading_text = self.calendar_service._clean_html(match.group(2))
            body_start = match.end()
            body_end = headings[index + 1].start() if index + 1 < len(headings) else len(body_html)
            section_html = body_html[body_start:body_end]

            if tag == "h2":
                year_match = re.search(r"\b(20\d{2})\b", heading_text)
                if year_match:
                    current_year = int(year_match.group(1))
                    last_release_date = None
                continue

            release_date, last_release_date, current_year = self.calendar_service._derive_release_date(
                heading_text,
                fallback_text="",
                current_year=current_year,
                last_release_date=last_release_date,
            )

            for row_html in self.TABLE_ROW_RE.findall(section_html):
                report_item = self._parse_row(row_html, release_date)
                if report_item is not None:
                    parsed_items.append(report_item)

        parsed_items.sort(key=lambda row: (row.release_date or "9999-99-99", row.title.casefold()))
        return parsed_items

    def snapshot_to_csv_bytes(self, snapshot: MetacriticTvClassificationSnapshot) -> bytes:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(TV_CLASSIFICATION_EXPORT_HEADERS)
        for item in snapshot.items:
            writer.writerow(self._item_to_export_row(item))
        return output.getvalue().encode("utf-8-sig")

    def snapshot_to_xlsx_bytes(self, snapshot: MetacriticTvClassificationSnapshot) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        sheet.append(TV_CLASSIFICATION_EXPORT_HEADERS)
        for item in snapshot.items:
            sheet.append(self._item_to_export_row(item))

        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="185B63")
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = "@"
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 48)
        sheet.freeze_panes = "A2"

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def item_to_api_payload(self, item: MetacriticTvClassificationItem) -> dict[str, object]:
        return item.model_dump(include=set(TV_CLASSIFICATION_OUTPUT_COLUMNS))

    def _attach_imdb_ttcodes(self, items: list[MetacriticTvClassificationItem]) -> tuple[int, str]:
        lookup_cache: dict[tuple[str, str, str], str] = {}
        not_found_count = 0
        lookup_error = ""
        lookup_disabled = False
        for item in items:
            if lookup_disabled:
                not_found_count += 1
                continue
            title_type = "movie" if item.program_type == PROGRAM_TYPE_MOVIE else "tv"
            release_year = item.release_date[:4] if re.fullmatch(r"\d{4}-\d{2}-\d{2}", item.release_date) else ""
            cache_key = (item.title.casefold(), title_type, release_year)
            if cache_key not in lookup_cache:
                try:
                    matches = self.imdb_dataset_lookup.lookup_title(
                        TitleLookupQuery(title=item.title, title_type=title_type, year=release_year)
                    )
                except (ImdbDatasetLookupError, ValueError) as exc:
                    lookup_error = str(exc)
                    lookup_disabled = True
                    matches = []
                lookup_cache[cache_key] = matches[0].imdb_id if matches else ""

            item.imdb_ttcode = lookup_cache[cache_key]
            if not item.imdb_ttcode:
                not_found_count += 1

        return not_found_count, lookup_error

    def _filter_rent_buy_items(
        self,
        items: list[MetacriticTvClassificationItem],
    ) -> tuple[list[MetacriticTvClassificationItem], int]:
        filtered_items: list[MetacriticTvClassificationItem] = []
        skipped_rent_buy = 0
        for item in items:
            if contains_rent_buy(item.network):
                skipped_rent_buy += 1
                continue
            filtered_items.append(item)
        return filtered_items, skipped_rent_buy

    def _filter_items_for_window(
        self,
        items: list[MetacriticTvClassificationItem],
        window_start: date,
        window_end: date,
    ) -> tuple[list[MetacriticTvClassificationItem], int, int]:
        filtered_items: list[MetacriticTvClassificationItem] = []
        skipped_outside_window = 0
        skipped_without_date = 0
        for item in items:
            if not item.release_date:
                skipped_without_date += 1
                continue
            try:
                release_date = date.fromisoformat(item.release_date)
            except ValueError:
                skipped_without_date += 1
                continue
            if window_start <= release_date <= window_end:
                filtered_items.append(item)
            else:
                skipped_outside_window += 1
        return filtered_items, skipped_outside_window, skipped_without_date

    def _parse_row(self, row_html: str, release_date: str) -> MetacriticTvClassificationItem | None:
        cells = self.CELL_RE.findall(row_html)
        if len(cells) < 2:
            return None

        meta_cell = cells[0]
        data_cell = cells[1]
        network_cell = cells[2] if len(cells) > 2 else ""

        title, metacritic_url = self._extract_title_and_url(data_cell, meta_cell)
        if not title:
            return None

        genre_text = self._extract_genre_text(title, data_cell)
        language_type, genres = self._split_language_and_genres(genre_text)
        genre_1, genre_2, genre_3 = (genres + ["", "", ""])[:3]

        return MetacriticTvClassificationItem(
            release_date=release_date,
            title=title,
            metacritic_url=metacritic_url,
            network=self._extract_network(network_cell),
            daypart=self._extract_daypart(network_cell),
            program_type=self._extract_program_type(data_cell),
            language_type=language_type,
            genre_1=genre_1,
            genre_2=genre_2,
            genre_3=genre_3,
        )

    def _item_to_export_row(self, item: MetacriticTvClassificationItem) -> list[str]:
        return [
            item.release_date,
            item.title,
            item.imdb_ttcode,
            item.network,
            item.daypart,
            item.program_type,
            item.language_type,
            item.genre_1,
            item.genre_2,
            item.genre_3,
        ]

    def _extract_title_and_url(self, data_cell: str, meta_cell: str) -> tuple[str, str]:
        strong_match = self.STRONG_RE.search(data_cell)
        if strong_match:
            title = self._clean_html(strong_match.group(1))
            if title and title != "($)":
                return title, self._find_matching_title_url(data_cell, title)

        for match in self.LINK_RE.finditer(data_cell):
            href = html.unescape(match.group("href")).strip()
            label = self._clean_html(match.group("label"))
            if not label or self.TRAILER_RE.fullmatch(label):
                continue
            if self._is_external_media_url(href):
                continue
            if href.startswith("/") or "metacritic.com" in href:
                return label, urljoin(self.BASE_URL, href)

        api_payload = self._extract_shortcode_api(meta_cell)
        api_title = str(api_payload.get("title") or "").strip()
        api_url = str(api_payload.get("url") or "").strip()
        if api_title and not re.fullmatch(r"Season\s+\d+", api_title, re.IGNORECASE):
            return api_title, urljoin(self.BASE_URL, api_url) if api_url else ""

        lines = self._extract_lines(data_cell)
        if lines:
            fallback_title = self._strip_title_line_noise(lines[0], "")
            return fallback_title, self._find_matching_title_url(data_cell, fallback_title)

        return "", ""

    def _find_matching_title_url(self, data_cell: str, title: str) -> str:
        normalized_title = title.casefold().strip()
        for match in self.LINK_RE.finditer(data_cell):
            href = html.unescape(match.group("href")).strip()
            label = self._clean_html(match.group("label"))
            if label.casefold().strip() == normalized_title and (href.startswith("/") or "metacritic.com" in href):
                return urljoin(self.BASE_URL, href)
        return ""

    def _extract_genre_text(self, title: str, data_cell: str) -> str:
        for line in self._extract_lines(data_cell):
            candidate = self._strip_title_line_noise(line, title)
            if not candidate or candidate.casefold() == title.casefold():
                continue
            if ":" in candidate:
                candidate = candidate.split(":", 1)[0].strip()
            if candidate:
                return candidate
        return ""

    def _split_language_and_genres(self, genre_text: str) -> tuple[str, list[str]]:
        language_type = LANGUAGE_TYPE_ENGLISH
        genres: list[str] = []
        for raw_part in re.split(r"/", genre_text):
            part = self._normalize_genre(raw_part)
            if not part:
                continue
            if part.casefold() in {"foreign", "international"}:
                language_type = LANGUAGE_TYPE_OTHER
                continue
            genres.extend(self._expand_compound_genre(part))
        return language_type, genres[:3]

    def _extract_program_type(self, data_cell: str) -> str:
        alt_values = [html.unescape(match.group("alt")).casefold() for match in self.IMG_ALT_RE.finditer(data_cell)]
        if any("movie" in alt for alt in alt_values) or re.search(r'["/]movie/', data_cell, re.IGNORECASE):
            return PROGRAM_TYPE_MOVIE
        if any("limited series" in alt for alt in alt_values):
            return PROGRAM_TYPE_MINI_SERIES
        if any("new series" in alt for alt in alt_values):
            return PROGRAM_TYPE_SERIES
        return PROGRAM_TYPE_OTHER

    def _extract_network(self, network_cell: str) -> str:
        networks: list[str] = []
        for line in self._extract_lines(network_cell):
            cleaned = self._clean_network_line(line)
            if cleaned and cleaned not in networks:
                networks.append(cleaned)
        return "; ".join(networks)

    def _extract_daypart(self, network_cell: str) -> str:
        text = self._clean_html(network_cell)
        for match in self.AIR_TIME_RE.finditer(text):
            hour = int(match.group("hour"))
            meridiem = match.group("meridiem").lower()
            hour_24 = hour
            if meridiem == "p" and hour != 12:
                hour_24 += 12
            if meridiem == "a" and hour == 12:
                hour_24 = 0
            if 19 <= hour_24 < 23:
                return DAYPART_PRIME_TIME
        return DAYPART_OTHER

    def _clean_network_line(self, value: str) -> str:
        cleaned = value.strip()
        streaming_match = re.fullmatch(r"Streaming\s*\((?P<network>[^)]+)\)", cleaned, re.IGNORECASE)
        if streaming_match:
            cleaned = streaming_match.group("network")
        cleaned = self.NETWORK_TIME_TAIL_RE.sub("", cleaned)
        cleaned = re.sub(r"\btbd\b", "", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"\s*\[\s*app/VOD only\s*\]\s*", "", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"\s*\(\d{1,2}/\d{1,2}\)\s*$", "", cleaned)
        cleaned = self.WHITESPACE_RE.sub(" ", cleaned).strip(" -/,\u00a0")
        return cleaned

    def _strip_title_line_noise(self, value: str, title: str) -> str:
        cleaned = self.TRAILER_RE.sub("", value)
        cleaned = re.sub(r"^\(\$\)\s*", "", cleaned)
        if title and cleaned.casefold().startswith(title.casefold()):
            cleaned = cleaned[len(title):]
        cleaned = self.WHITESPACE_RE.sub(" ", cleaned).strip(" -:\u00a0")
        return cleaned

    def _normalize_genre(self, value: str) -> str:
        cleaned = self.WHITESPACE_RE.sub(" ", value).strip(" -:\u00a0")
        return cleaned

    def _expand_compound_genre(self, value: str) -> list[str]:
        replacements = {
            "Sci-Fi": "Sci Fi",
            "Rom-Com": "Rom Com",
        }
        normalized = replacements.get(value, value)
        if normalized.casefold() in {"action-adventure", "action adventure"}:
            return ["Action", "Adventure"]
        return [normalized]

    def _extract_shortcode_api(self, fragment: str) -> dict[str, Any]:
        match = self.SHORTCODE_API_RE.search(fragment)
        if not match:
            return {}
        try:
            return json.loads(html.unescape(match.group(1)))
        except json.JSONDecodeError:
            return {}

    def _extract_lines(self, fragment: str) -> list[str]:
        cleaned = self._clean_html(fragment, keep_line_breaks=True)
        return [line for line in cleaned.splitlines() if line]

    def _clean_html(self, value: str, *, keep_line_breaks: bool = False) -> str:
        if not value:
            return ""
        text = self.SHORTCODE_RE.sub("", value)
        text = self.IMG_RE.sub("", text)
        if keep_line_breaks:
            text = self.BR_RE.sub("\n", text)
        else:
            text = self.BR_RE.sub(" ", text)
        text = self.TAG_RE.sub("", text)
        text = html.unescape(text)
        text = repair_mojibake(text)
        if keep_line_breaks:
            lines = []
            for line in text.splitlines():
                normalized = self.WHITESPACE_RE.sub(" ", line).strip(" -\u00a0")
                if normalized:
                    lines.append(normalized)
            return "\n".join(lines)
        return self.WHITESPACE_RE.sub(" ", text).strip(" -\u00a0")

    def _is_external_media_url(self, href: str) -> bool:
        normalized = href.casefold()
        return "youtube.com" in normalized or "youtu.be" in normalized

    def _add_months(self, value: date, months: int) -> date:
        year = value.year + (value.month - 1 + months) // 12
        month = (value.month - 1 + months) % 12 + 1
        day = min(value.day, self._days_in_month(year, month))
        return date(year, month, day)

    def _days_in_month(self, year: int, month: int) -> int:
        if month == 12:
            return (date(year + 1, 1, 1) - date(year, month, 1)).days
        return (date(year, month + 1, 1) - date(year, month, 1)).days
