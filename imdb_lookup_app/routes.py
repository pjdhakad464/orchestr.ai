from __future__ import annotations

import csv
import io
import re
import uuid
from pathlib import Path
from typing import Annotated

from fastapi import APIRouter, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook

from app.cache import TTLCache
from imdb_lookup_app.config import settings
from imdb_lookup_app.models import DownloadArtifact, LookupBatchResult, LookupRequest, LookupRow
from imdb_lookup_app.services.lookup import ImdbLookupService, ImdbLookupServiceError


BASE_DIR = Path(__file__).resolve().parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
router = APIRouter()
cache = TTLCache(settings.imdb_lookup_export_ttl_seconds)
lookup_service = ImdbLookupService()

MODE_OPTIONS = [
    ("auto", "Auto detect"),
    ("id_to_name", "IMDb ID to title/person"),
    ("title_to_id", "Title to tt code"),
    ("person_to_id", "Person name to nm code"),
]
EXPORT_COLUMNS = [
    ("input_value", "Input"),
    ("normalized_input", "Normalized"),
    ("requested_mode", "Requested Mode"),
    ("resolved_lookup", "Resolved Lookup"),
    ("status", "Status"),
    ("match_rank", "Match Rank"),
    ("total_matches", "Total Matches"),
    ("imdb_id", "IMDb ID"),
    ("entity_kind", "Entity Kind"),
    ("display_name", "Display Name"),
    ("original_title", "Original Title"),
    ("title_type", "Title Type"),
    ("start_year", "Start Year"),
    ("end_year", "End Year"),
    ("birth_year", "Birth Year"),
    ("death_year", "Death Year"),
    ("primary_profession", "Primary Profession"),
    ("known_for_titles", "Known For Titles"),
    ("matched_on", "Matched On"),
    ("source_url", "IMDb URL"),
    ("notes", "Notes"),
]
PREFERRED_INPUT_HEADERS = [
    "input",
    "value",
    "lookup_value",
    "lookup",
    "identifier",
    "imdb_id",
    "id",
    "title",
    "name",
]


def static_asset_version(*relative_paths: str) -> str:
    versions: list[int] = []
    for relative_path in relative_paths:
        asset_path = BASE_DIR / "static" / relative_path
        try:
            versions.append(asset_path.stat().st_mtime_ns)
        except FileNotFoundError:
            continue
    if not versions:
        return "1"
    return str(max(versions))


def build_context(
    request: Request,
    *,
    result: LookupBatchResult | None = None,
    error_message: str = "",
    raw_values: str = "",
    selected_mode: str = "auto",
    source_hint: str = "",
    selected_sheet: str = "",
    selected_column: str = "",
) -> dict[str, object]:
    return {
        "request": request,
        "result": result,
        "error_message": error_message,
        "raw_values": raw_values,
        "selected_mode": selected_mode,
        "mode_options": MODE_OPTIONS,
        "source_hint": source_hint,
        "selected_sheet": selected_sheet,
        "selected_column": selected_column,
        "asset_version": static_asset_version("styles.css"),
    }


def parse_bulk_values(raw_text: str) -> list[str]:
    values: list[str] = []
    for token in re.split(r"[\r\n,;]+", raw_text):
        cleaned = token.strip()
        if cleaned:
            values.append(cleaned)
    return values


def store_download_artifact(result: LookupBatchResult, filename_hint: str) -> str:
    export_id = str(uuid.uuid4())
    filename_base = _slugify_filename(filename_hint or "imdb_lookup_results")
    artifact = DownloadArtifact(
        export_id=export_id,
        filename_base=filename_base,
        csv_bytes=rows_to_csv_bytes(result.rows),
        xlsx_bytes=rows_to_xlsx_bytes(result.rows, result.summary),
        rows=result.rows,
        summary=result.summary,
    )
    cache.set(f"export:{export_id}", artifact)
    return export_id


def rows_to_csv_bytes(rows: list[LookupRow]) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=[field for field, _ in EXPORT_COLUMNS])
    writer.writeheader()
    for row in rows:
        writer.writerow(row.model_dump())
    return buffer.getvalue().encode("utf-8-sig")


def rows_to_xlsx_bytes(rows: list[LookupRow], summary: list[str]) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Item", "Value"])
    for index, item in enumerate(summary, start=1):
        summary_sheet.append([f"Summary {index}", item])

    results_sheet = workbook.create_sheet("Results")
    results_sheet.append([label for _, label in EXPORT_COLUMNS])
    for row in rows:
        payload = row.model_dump()
        results_sheet.append([payload[field] for field, _ in EXPORT_COLUMNS])

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def load_values_from_upload(
    file_bytes: bytes,
    filename: str,
    sheet_name: str = "",
    input_column: str = "",
) -> tuple[list[str], str]:
    suffix = Path(filename).suffix.casefold()
    if suffix == ".csv":
        values, selected_column_name = _load_values_from_csv(file_bytes, input_column)
        source_hint = f"CSV source: {filename} | column: {selected_column_name}"
        return values, source_hint
    if suffix == ".xlsx":
        values, selected_sheet_name, selected_column_name = _load_values_from_workbook(file_bytes, sheet_name, input_column)
        source_hint = f"Workbook source: {filename} | sheet: {selected_sheet_name} | column: {selected_column_name}"
        return values, source_hint
    raise ImdbLookupServiceError("Upload a CSV or XLSX file.")


def _load_values_from_csv(file_bytes: bytes, input_column: str) -> tuple[list[str], str]:
    decoded = _decode_csv_bytes(file_bytes)
    rows = list(csv.reader(io.StringIO(decoded)))
    if not rows:
        raise ImdbLookupServiceError("The CSV file is empty.")

    selected_index, selected_name, include_first_row = _select_input_column(rows[0], input_column)
    data_rows = rows if include_first_row else rows[1:]
    values = _extract_values_from_rows(data_rows, selected_index)
    if not values:
        raise ImdbLookupServiceError("No lookup values were found in the selected CSV column.")
    return values, selected_name


def _load_values_from_workbook(
    file_bytes: bytes,
    sheet_name: str,
    input_column: str,
) -> tuple[list[str], str, str]:
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        if sheet_name.strip():
            if sheet_name not in workbook.sheetnames:
                raise ImdbLookupServiceError(f"Sheet '{sheet_name}' was not found in the workbook.")
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active

        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        if not rows:
            raise ImdbLookupServiceError("The workbook sheet is empty.")

        selected_index, selected_name, include_first_row = _select_input_column(rows[0], input_column)
        data_rows = rows if include_first_row else rows[1:]
        values = _extract_values_from_rows(data_rows, selected_index)
        if not values:
            raise ImdbLookupServiceError("No lookup values were found in the selected workbook column.")
        return values, worksheet.title, selected_name
    finally:
        workbook.close()


def _select_input_column(header_row: list[object], requested_column: str) -> tuple[int, str, bool]:
    normalized_headers = [_normalize_header_name(value) for value in header_row]
    if requested_column.strip():
        requested_normalized = _normalize_header_name(requested_column)
        if requested_normalized in normalized_headers:
            index = normalized_headers.index(requested_normalized)
            return index, _header_label(header_row[index], index), False
        raise ImdbLookupServiceError(f"Column '{requested_column}' was not found.")

    for preferred in PREFERRED_INPUT_HEADERS:
        if preferred in normalized_headers:
            index = normalized_headers.index(preferred)
            return index, _header_label(header_row[index], index), False

    non_empty_indexes = [index for index, value in enumerate(header_row) if str(value or "").strip()]
    if not non_empty_indexes:
        return 0, "Column A", True

    first_index = non_empty_indexes[0]
    first_value = str(header_row[first_index] or "").strip()
    include_first_row = bool(first_value) and (
        re.fullmatch(r"(tt|nm)\d{7,}", first_value, flags=re.IGNORECASE) is not None
        or _normalize_header_name(first_value) not in PREFERRED_INPUT_HEADERS
    )
    return first_index, _header_label(header_row[first_index], first_index), include_first_row


def _extract_values_from_rows(rows: list[list[object]], column_index: int) -> list[str]:
    values: list[str] = []
    for row in rows:
        if column_index >= len(row):
            continue
        value = str(row[column_index] or "").strip()
        if value:
            values.append(value)
    return values


def _decode_csv_bytes(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ImdbLookupServiceError("The CSV file could not be decoded. Save it as UTF-8 or Windows-1252.")


def _normalize_header_name(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().casefold()).strip("_")


def _header_label(value: object, index: int) -> str:
    text = str(value or "").strip()
    if text:
        return text
    return f"Column {chr(ord('A') + index)}"


def _slugify_filename(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_-]+", "_", value.strip()).strip("_")
    return cleaned or "imdb_lookup_results"


@router.get("/", response_class=HTMLResponse)
async def index(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(request, "index.html", build_context(request))


@router.post("/lookup/text", response_class=HTMLResponse)
async def lookup_text(
    request: Request,
    values: Annotated[str, Form()],
    mode: Annotated[str, Form()] = "auto",
) -> HTMLResponse:
    parsed_values = parse_bulk_values(values)
    if not parsed_values:
        return templates.TemplateResponse(
            request,
            "index.html",
            build_context(
                request,
                error_message="Paste at least one IMDb id, title, or person name.",
                raw_values=values,
                selected_mode=mode,
            ),
        )

    try:
        result = lookup_service.lookup_values(parsed_values, mode=mode)  # type: ignore[arg-type]
    except ImdbLookupServiceError as exc:
        return templates.TemplateResponse(
            request,
            "index.html",
            build_context(request, error_message=str(exc), raw_values=values, selected_mode=mode),
        )

    result.export_id = store_download_artifact(result, "imdb_text_lookup_results")
    return templates.TemplateResponse(
        request,
        "index.html",
        build_context(request, result=result, raw_values=values, selected_mode=mode),
    )


@router.post("/lookup/file", response_class=HTMLResponse)
async def lookup_file(
    request: Request,
    dataset_file: UploadFile | None = File(None),
    mode: Annotated[str, Form()] = "auto",
    sheet_name: Annotated[str | None, Form()] = None,
    input_column: Annotated[str | None, Form()] = None,
) -> HTMLResponse:
    uploaded_filename = Path(dataset_file.filename or "").name if dataset_file else ""
    selected_sheet = (sheet_name or "").strip()
    selected_column = (input_column or "").strip()
    if not uploaded_filename:
        return templates.TemplateResponse(
            request,
            "index.html",
            build_context(
                request,
                error_message="Choose a CSV or XLSX file to upload.",
                selected_mode=mode,
                selected_sheet=selected_sheet,
                selected_column=selected_column,
            ),
        )

    try:
        file_bytes = await dataset_file.read()
        values, source_hint = load_values_from_upload(
            file_bytes,
            uploaded_filename,
            sheet_name=selected_sheet,
            input_column=selected_column,
        )
        result = lookup_service.lookup_values(values, mode=mode)  # type: ignore[arg-type]
    except ImdbLookupServiceError as exc:
        return templates.TemplateResponse(
            request,
            "index.html",
            build_context(
                request,
                error_message=str(exc),
                selected_mode=mode,
                selected_sheet=selected_sheet,
                selected_column=selected_column,
            ),
        )

    result.summary.insert(0, source_hint)
    result.export_id = store_download_artifact(result, Path(uploaded_filename).stem + "_lookup_results")
    return templates.TemplateResponse(
        request,
        "index.html",
        build_context(
            request,
            result=result,
            selected_mode=mode,
            source_hint=source_hint,
            selected_sheet=selected_sheet,
            selected_column=selected_column,
        ),
    )


@router.post("/api/lookup")
async def api_lookup(payload: LookupRequest) -> JSONResponse:
    result = lookup_service.lookup_values(payload.values, mode=payload.mode)
    return JSONResponse(result.model_dump())


@router.get("/download/{export_id}/{fmt}")
async def download_export(export_id: str, fmt: str):
    artifact = cache.get(f"export:{export_id}")
    if not isinstance(artifact, DownloadArtifact):
        return HTMLResponse("Export expired. Run the lookup again.", status_code=404)

    if fmt == "csv":
        return StreamingResponse(
            io.BytesIO(artifact.csv_bytes),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{artifact.filename_base}.csv"'},
        )
    if fmt == "xlsx":
        return StreamingResponse(
            io.BytesIO(artifact.xlsx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{artifact.filename_base}.xlsx"'},
        )
    return HTMLResponse("Unsupported export format.", status_code=400)
