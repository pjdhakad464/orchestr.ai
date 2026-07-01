from __future__ import annotations

import openpyxl
from pathlib import Path
from typing import Any

from app.services.imdb_enricher import IMDbEnricher, TitleQuery
from app.services.duplicate_detector import DuplicateDetector
from app.services.excel_comparator import ExcelComparator
from app.services.health_scorer import MetadataHealthScorer
from app.services.anomaly_detector import AnomalyDetector
from app.services.workbook_validator import validate_loaded_workbook
from app.models import ValidationRuleSet

class PipelineSteps:
    """Contains standard step handlers for DataOps automation pipelines."""

    @staticmethod
    async def load_excel(inputs: dict[str, Any]) -> dict[str, Any]:
        """Loads an Excel workbook and returns list of rows (dicts)."""
        file_path = Path(inputs["file_path"])
        sheet_name = inputs.get("sheet_name")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        
        headers = [str(cell.value or "").strip() for cell in sheet[1]]
        rows = []
        for r_idx in range(2, sheet.max_row + 1):
            row_vals = [sheet.cell(row=r_idx, column=c_idx).value for c_idx in range(1, len(headers) + 1)]
            if not any(v is not None for v in row_vals):
                continue
            rows.append(dict(zip(headers, row_vals)))

        wb.close()
        return {"rows": rows, "headers": headers, "sheet_name": sheet.title}

    @staticmethod
    async def validate_workbook(inputs: dict[str, Any]) -> dict[str, Any]:
        """Validates the loaded workbook using workbook_validator."""
        file_path = Path(inputs["file_path"])
        rules_data = inputs.get("rules") # dict representing ValidationRuleSet

        wb = openpyxl.load_workbook(file_path)
        
        # Load rules
        if isinstance(rules_data, dict):
            ruleset = ValidationRuleSet.model_validate(rules_data)
        else:
            # Load default BDR QA rules if not specified
            import json
            from app.config import BASE_DIR
            rules_path = BASE_DIR / "data" / "bdr_qa_rules.json"
            if rules_path.exists():
                with open(rules_path, "r", encoding="utf-8") as f:
                    ruleset = ValidationRuleSet.model_validate(json.load(f))
            else:
                ruleset = ValidationRuleSet(rules=[])

        artifact = validate_loaded_workbook(wb, ruleset)
        wb.close()
        
        return {
            "issues": [issue.model_dump() for issue in artifact.issues],
            "issue_count": artifact.issue_count
        }

    @staticmethod
    async def enrich_imdb(inputs: dict[str, Any]) -> dict[str, Any]:
        """Performs bulk IMDb/TMDB enrichment for entertainment titles."""
        rows = inputs["rows"]
        title_col = inputs.get("title_column", "title")
        year_col = inputs.get("year_column", "released_on")
        type_col = inputs.get("type_column", "title_category")

        enricher = IMDbEnricher()
        queries = []
        for row in rows:
            title_val = str(row.get(title_col) or "").strip()
            if title_val:
                # Year parse
                yr_val = str(row.get(year_col) or "").strip()
                match = re.search(r"\b\d{4}\b", yr_val) if yr_val else None
                year = int(match.group()) if match else None
                
                # Content type
                ct_val = str(row.get(type_col) or "").strip().casefold()
                content_type = "movie" if "movie" in ct_val else ("tv" if "tv" in ct_val else "any")
                
                queries.append(TitleQuery(title=title_val, year=year, content_type=content_type))

        results = await enricher.bulk_enrich(queries)
        
        # Merge enriched data back into rows
        enriched_rows = []
        for idx, row in enumerate(rows):
            new_row = dict(row)
            if idx < len(results) and results[idx].metadata:
                meta = results[idx].metadata
                new_row["enriched_imdb_id"] = meta.imdb_id
                new_row["enriched_rating"] = meta.rating
                new_row["enriched_genres"] = ", ".join(meta.genres)
                new_row["enriched_networks"] = meta.networks
                new_row["enriched_start_year"] = meta.year
            enriched_rows.append(new_row)

        return {"rows": enriched_rows, "enrichment_count": sum(1 for r in results if r.metadata)}

    @staticmethod
    async def detect_duplicates(inputs: dict[str, Any]) -> dict[str, Any]:
        """Detects duplicates in row data."""
        rows = inputs["rows"]
        columns = inputs.get("columns", [])
        title_col = inputs.get("title_column")
        threshold = inputs.get("fuzzy_threshold", 0.85)

        detector = DuplicateDetector()
        report_exact = detector.detect_exact_duplicates(rows, columns) if columns else None
        report_fuzzy = detector.detect_fuzzy_duplicates(rows, title_col, threshold) if title_col else None

        return {
            "exact_duplicates": report_exact.model_dump() if report_exact else {},
            "fuzzy_duplicates": report_fuzzy.model_dump() if report_fuzzy else {}
        }

    @staticmethod
    async def compare_excel(inputs: dict[str, Any]) -> dict[str, Any]:
        """Compares two Excel spreadsheets side by side."""
        file_a = Path(inputs["file_a"])
        file_b = Path(inputs["file_b"])
        key_cols = inputs.get("key_columns")

        comparator = ExcelComparator()
        report = comparator.compare(file_a, file_b, key_cols)
        
        # Save diff report Excel if specified
        out_path = inputs.get("output_path")
        if out_path:
            comparator.export_to_xlsx(report, Path(out_path))

        return {"report": report.model_dump()}

    @staticmethod
    async def score_health(inputs: dict[str, Any]) -> dict[str, Any]:
        """Runs quality score computations across rows."""
        rows = inputs["rows"]
        weights = inputs.get("field_weights")

        scorer = MetadataHealthScorer()
        report = scorer.score_workbook(rows, weights)

        return {"report": report.model_dump()}

    @staticmethod
    async def detect_anomalies(inputs: dict[str, Any]) -> dict[str, Any]:
        """Flags statistical and category anomalies."""
        rows = inputs["rows"]
        date_cols = inputs.get("date_columns")
        cat_cols = inputs.get("categorical_columns")

        detector = AnomalyDetector()
        report = detector.detect_all(rows, date_cols, cat_cols)

        return {"report": report.model_dump()}

import re
