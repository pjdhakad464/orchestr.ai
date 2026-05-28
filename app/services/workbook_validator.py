from __future__ import annotations

import csv
from dataclasses import dataclass, field
import gzip
import inspect
import io
import json
import re
import shutil
import sqlite3
import time
import uuid
from datetime import date, datetime
from html import unescape
from pathlib import Path
from threading import Lock
from typing import Any
from urllib.parse import quote, unquote, urlparse

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import BASE_DIR, settings
from app.models import (
    ValidationCondition,
    ValidationRule,
    ValidationRuleSet,
    WorkbookValidationArtifact,
    WorkbookValidationIssue,
)
from app.platforms.facebook import has_blocked_facebook_path


ERROR_FILL = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
ERROR_FONT = Font(color="9C0006")
WARNING_FILL = PatternFill(fill_type="solid", start_color="FFFCE4D6", end_color="FFFCE4D6")
WARNING_FONT = Font(color="FFC65911")
SUMMARY_HEADERS = ["Sheet", "Row", "Column", "Cell", "Rule", "Finding Category", "Confidence", "Confidence Reason", "Message", "Value"]
COMMON_DATE_FORMATS = (
    "%d-%m-%Y",
    "%B-%d-%Y",
    "%b-%d-%Y",
    "%m-%d-%Y",
    "%m/%d/%Y",
    "%d/%m/%Y",
    "%Y/%m/%d",
    "%B %d, %Y",
    "%b %d, %Y",
    "%d %B %Y",
    "%d %b %Y",
)
IMDB_DATASET_INDEX_LOCK = Lock()
WIKIPEDIA_CACHE_LOCK = Lock()
IMDB_TITLE_BASICS_FILENAME = "title.basics.tsv.gz"
IMDB_NAME_BASICS_FILENAME = "name.basics.tsv.gz"
IMDB_INDEX_FILENAME = "imdb_title_lookup.sqlite3"
WIKIPEDIA_CACHE_FILENAME = "wikipedia_cache.sqlite3"
WIKIMEDIA_CACHE_TABLE = "wikimedia_cache"
WIKIDATA_DISAMBIGUATION_IDS = {"Q4167410", "Q22808320"}
WIKIDATA_TYPE_MAP = {
    "person": {"Q5"},
    "movie": {"Q11424", "Q24869", "Q24862", "Q506240"},
    "series": {"Q5398426", "Q15416", "Q3464665"},
    "episode": {"Q1983062"},
}


@dataclass(slots=True)
class CompiledValidationCondition:
    condition: ValidationCondition
    key: str


@dataclass(slots=True)
class WorksheetValidationContext:
    worksheet: Any
    header_row: int
    header_map: dict[str, int]
    rows_values: list[tuple[Any, ...]]
    active_rows: list[int]
    row_context_cache: dict[int, dict[str, Any]] = field(default_factory=dict)



class WorkbookValidationConfigError(ValueError):
    pass


def parse_validation_rules(raw_text: str) -> list[ValidationRule]:
    import json

    cleaned = raw_text.strip()
    if not cleaned:
        raise WorkbookValidationConfigError("Provide at least one validation rule in JSON format.")

    try:
        payload = json.loads(cleaned)
    except json.JSONDecodeError as exc:
        raise WorkbookValidationConfigError(f"Rules JSON is invalid: {exc.msg}.") from exc

    if isinstance(payload, list):
        payload = {"rules": payload}

    try:
        rule_set = ValidationRuleSet.model_validate(payload)
    except Exception as exc:  # pragma: no cover
        raise WorkbookValidationConfigError(f"Rules JSON could not be parsed: {exc}") from exc

    if not rule_set.rules:
        raise WorkbookValidationConfigError("Add at least one rule to the 'rules' list.")

    return rule_set.rules


def validate_workbook(
    file_bytes: bytes,
    filename: str,
    rules: list[ValidationRule],
    review_mode: str = "full",
    platform_filter: str | None = None,
) -> WorkbookValidationArtifact:
    workbook = load_validation_workbook(file_bytes, filename)
    return validate_loaded_workbook(workbook, filename, rules, review_mode, platform_filter)


def validate_loaded_workbook(
    workbook: Workbook,
    filename: str,
    rules: list[ValidationRule],
    review_mode: str = "full",
    platform_filter: str | None = None,
) -> WorkbookValidationArtifact:
    issues: list[WorkbookValidationIssue] = []
    social_cache: dict[str, tuple[bool, str]] = {}
    movie_release_cache: dict[str, tuple[bool, dict[str, Any] | None, str]] = {}
    reference_cache: dict[str, tuple[bool, dict[str, Any] | None, str]] = {}
    rottentomatoes_cache: dict[str, tuple[bool, dict[str, Any] | None, str]] = {}
    worksheet_contexts: dict[tuple[str, int], WorksheetValidationContext] = {}

    # Split review_mode by comma to check for multiple modes
    modes = [m.strip() for m in review_mode.split(",") if m.strip()]
    if not modes:
        modes = ["full"]

    # Run duplicate conflict scan if requested
    if "duplicate_conflict" in modes:
        for worksheet in workbook.worksheets:
            if worksheet.title == "Validation Summary":
                continue
            try:
                worksheet_context = _get_worksheet_validation_context(worksheet_contexts, worksheet, 1)
            except Exception:
                continue
            _perform_duplicate_conflict_scan(worksheet_context, worksheet, issues)
            
        if len(modes) == 1:
            _append_summary_sheet(workbook, issues)
            output = io.BytesIO()
            workbook.save(output)
            safe_name = f"{Path(filename).stem}_validated.xlsx"
            return WorkbookValidationArtifact(
                validation_id=str(uuid.uuid4()),
                filename=safe_name,
                file_bytes=output.getvalue(),
                issues=issues,
            )

    with _social_http_client() as social_client:
        for rule in rules:
            if not _rule_matches_review_mode(rule, review_mode, platform_filter):
                continue
            wildcard_sheet = (rule.sheet or "*").strip() in {"*", "Any"}
            for worksheet in _select_worksheets(workbook, rule.sheet):
                worksheet_context = _get_worksheet_validation_context(worksheet_contexts, worksheet, rule.header_row)
                header_map = worksheet_context.header_map
                column_key = rule.column.strip().casefold()
                column_index = header_map.get(column_key)
                if column_index is None:
                    if wildcard_sheet:
                        continue
                    raise WorkbookValidationConfigError(
                        f"Column '{rule.column}' was not found in sheet '{worksheet.title}' on header row {rule.header_row}."
                    )

                compiled_conditions = _compile_conditions(
                    worksheet_context,
                    rule.when,
                    wildcard_sheet=wildcard_sheet,
                )
                if compiled_conditions is None:
                    continue

                extra_required_columns = _required_columns_for_rule(rule)
                missing_extra_columns = [
                    column_name for column_name in extra_required_columns if _get_optional_column_index(header_map, column_name) is None
                ]
                if missing_extra_columns:
                    if wildcard_sheet:
                        continue
                    missing_label = ", ".join(missing_extra_columns)
                    raise WorkbookValidationConfigError(
                        f"Column '{missing_label}' was not found in sheet '{worksheet.title}' on header row {rule.header_row}."
                    )

                if rule.check == "unique":
                    issues.extend(
                        _validate_unique_rule_with_context(
                            worksheet_context,
                            column_index,
                            column_key,
                            rule,
                            compiled_conditions,
                        )
                    )
                    continue

                for row_number in worksheet_context.active_rows:
                    row_context = _get_row_context(worksheet_context, row_number)
                    if not _conditions_match_row_context(row_context, compiled_conditions):
                        continue

                    cell_value = row_context.get(column_key)
                    if "missing_only" in modes and not _is_blank(cell_value):
                        continue
                    if "existing_qa" in modes and _is_blank(cell_value):
                        continue

                    cell = worksheet.cell(row=row_number, column=column_index)
                    passed, message, category, confidence, reason = _evaluate_rule(
                        cell_value,
                        rule,
                        row_context=row_context,
                        social_cache=social_cache,
                        social_client=social_client,
                        movie_release_cache=movie_release_cache,
                        reference_cache=reference_cache,
                        rottentomatoes_cache=rottentomatoes_cache,
                    )
                    if passed:
                        continue

                    issues.append(
                        _mark_issue(
                            cell=cell, 
                            rule=rule, 
                            message=message, 
                            value=cell.value,
                            finding_category=category,
                            confidence=confidence,
                            confidence_reason=reason
                        )
                    )

    _append_summary_sheet(workbook, issues)
    output = io.BytesIO()
    workbook.save(output)
    safe_name = f"{Path(filename).stem}_validated.xlsx"
    return WorkbookValidationArtifact(
        validation_id=str(uuid.uuid4()),
        filename=safe_name,
        file_bytes=output.getvalue(),
        issues=issues,
    )


def load_validation_workbook(file_bytes: bytes, filename: str) -> Workbook:
    suffix = Path(filename).suffix.casefold()
    if suffix == ".xlsx":
        try:
            return load_workbook(io.BytesIO(file_bytes))
        except Exception as exc:  # pragma: no cover
            raise WorkbookValidationConfigError(f"The Excel workbook could not be opened: {exc}") from exc
    if suffix == ".csv":
        return _load_csv_workbook(file_bytes, filename)
    raise WorkbookValidationConfigError(
        "Upload an Excel workbook (.xlsx), a CSV file (.csv), or use a Google Sheets URL."
    )


def load_google_sheet_workbook(sheet_reference: str) -> tuple[Workbook, str]:
    spreadsheet_id = _extract_google_sheet_id(sheet_reference)
    if not spreadsheet_id:
        raise WorkbookValidationConfigError("Enter a valid Google Sheets URL or spreadsheet id.")

    if not settings.google_service_account_file:
        raise WorkbookValidationConfigError(
            "GOOGLE_SERVICE_ACCOUNT_FILE is not configured. Add it to .env before validating a Google Sheet."
        )

    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
    except ImportError as exc:  # pragma: no cover
        raise WorkbookValidationConfigError(
            "Google API packages are not installed. Run: python -m pip install -e .[dev]"
        ) from exc

    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    credentials = service_account.Credentials.from_service_account_file(
        settings.google_service_account_file,
        scopes=scopes,
    )
    service = build("sheets", "v4", credentials=credentials)

    try:
        spreadsheet = service.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            fields="properties.title,sheets.properties.title",
        ).execute()
    except Exception as exc:  # pragma: no cover
        raise WorkbookValidationConfigError(f"Google Sheet lookup failed: {exc}") from exc

    sheet_titles = [
        (sheet.get("properties") or {}).get("title", "").strip()
        for sheet in spreadsheet.get("sheets", [])
        if (sheet.get("properties") or {}).get("title")
    ]
    if not sheet_titles:
        raise WorkbookValidationConfigError("The Google Sheet does not contain any readable tabs.")

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for title in sheet_titles:
        try:
            values_response = service.spreadsheets().values().get(
                spreadsheetId=spreadsheet_id,
                range=title,
            ).execute()
        except Exception as exc:  # pragma: no cover
            raise WorkbookValidationConfigError(f"Google Sheet tab '{title}' could not be read: {exc}") from exc

        worksheet = workbook.create_sheet(title=_safe_sheet_name(title))
        for row in values_response.get("values", []):
            worksheet.append(list(row))
        if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet["A1"].value is None:
            worksheet["A1"] = ""

    source_name = f"{spreadsheet.get('properties', {}).get('title') or 'google_sheet'}.gsheet"
    return workbook, source_name


def build_sample_rules_json() -> str:
    return """{
  "rules": [
    {
      "sheet": "*",
      "column": "title",
      "check": "not_blank_and_not_in",
      "tokens": ["#NA", "N/A"],
      "message": "Title cannot be blank, #NA, or N/A."
    },
    {
      "sheet": "*",
      "column": "title_category",
      "check": "in",
      "values": [
        "Airlines",
        "Automotive",
        "Car Rental",
        "Consumer Electronics",
        "CPG",
        "Health & Beauty",
        "Education",
        "Energy",
        "Venues, Events & Attractions",
        "Fashion",
        "Financial Services",
        "Food Products",
        "Health, Wellness, Fitness",
        "Hospital & Health Care",
        "Hospitality",
        "Insurance",
        "Internet Services",
        "IT",
        "Internet",
        "Computing",
        "IT, Internet, Computing",
        "Legal",
        "Government Entities",
        "Marketing",
        "Advertising and Research",
        "Materials and Construction",
        "Media",
        "Movies",
        "Music and Entertainment",
        "TV Network",
        "Non-Profit/Charity/Philanthropy",
        "Pets",
        "Pet Foods & Pet Supplies",
        "Pharmaceuticals",
        "Radio",
        "Real Estate",
        "Restaurants",
        "Retail",
        "Beverages",
        "Sports Franchise",
        "Sports Organizations and Bodies",
        "Film Studio",
        "Supermarket",
        "Grocery",
        "Food & Convenience Stores",
        "Talent",
        "Tourism Boards",
        "Travel",
        "TV Shows",
        "Video Game",
        "Video Game Publishers",
        "Wireless and Telecom",
        "Publishers",
        "Podcasts",
        "Other",
        "Manufacturing & Infrastructure"
      ],
      "message": "Title category is blank or not in the approved list."
    },
    {
      "sheet": "*",
      "column": "title_sub_category",
      "check": "not_blank_and_not_in",
      "tokens": ["#NA", "N/A"],
      "when": [
        {
          "column": "title_category",
          "operator": "in",
          "values": ["Talent", "Movies", "TV Shows"]
        }
      ],
      "message": "Title sub category cannot be blank, #NA, or N/A."
    },
    {
      "sheet": "*",
      "column": "title_sub_category",
      "check": "talent_subcategory_format",
      "tokens": ["Gender -", "Talent Type -", "Talent Subtype -"],
      "when": [
        {
          "column": "title_category",
          "operator": "equals",
          "value": "Talent"
        }
      ],
      "message": "Talent title_sub_category must include gender and talent type or subtype."
    },
    {
      "sheet": "*",
      "column": "title_sub_category",
      "check": "contains",
      "value": "Program Type -",
      "when": [
        {
          "column": "title_category",
          "operator": "equals",
          "value": "TV Shows"
        }
      ],
      "message": "TV Shows title_sub_category must include Program Type."
    },
    {
      "sheet": "*",
      "column": "genre",
      "check": "not_blank_and_not_in",
      "tokens": ["#NA", "N/A"],
      "when": [
        {
          "column": "title_category",
          "operator": "in",
          "values": ["Movies", "TV Shows"]
        }
      ],
      "message": "Genre is required for Movies and TV Shows."
    },
    {
      "sheet": "*",
      "column": "primary_genre",
      "check": "not_blank_and_not_in",
      "tokens": ["#NA", "N/A"],
      "when": [
        {
          "column": "genre",
          "operator": "not_in",
          "values": ["", "#NA", "N/A"]
        }
      ],
      "message": "Primary genre is required when genre is populated."
    },
    {
      "sheet": "*",
      "column": "network",
      "check": "not_blank_and_not_in",
      "tokens": ["#NA", "N/A"],
      "when": [
        {
          "column": "title_category",
          "operator": "in",
          "values": ["Movies", "TV Shows"]
        }
      ],
      "message": "Network cannot be blank for Movies and TV Shows."
    },
    {
      "sheet": "*",
      "column": "released_on",
      "check": "movie_us_release_date_match",
      "when": [
        {
          "column": "title_category",
          "operator": "equals",
          "value": "Movies"
        }
      ]
    },
    {
      "sheet": "*",
      "column": "release_type",
      "check": "movie_release_type_match",
      "when": [
        {
          "column": "title_category",
          "operator": "equals",
          "value": "Movies"
        }
      ],
      "message": "Release type should match the TMDB USA recommendation."
    },
    {
      "sheet": "*",
      "column": "genre",
      "check": "movie_genre_match",
      "when": [
        {
          "column": "title_category",
          "operator": "equals",
          "value": "Movies"
        }
      ]
    },
    {
      "sheet": "*",
      "column": "companies",
      "check": "contains_any",
      "values": ["Pristine Brand", "Pristine Talent", "Pristine Film", "Pristine TV"],
      "when": [
        {
          "column": "title",
          "operator": "endswith",
          "value": " - DAR"
        },
        {
          "column": "title_category",
          "operator": "equals",
          "value": "TV Shows"
        }
      ],
      "message": "DAR TV Shows must include Pristine Brand, Pristine Talent, Pristine Film, or Pristine TV in companies."
    },
    {
      "sheet": "*",
      "column": "companies",
      "check": "contains_any",
      "values": ["Pristine Brand", "Pristine Talent", "Pristine Film"],
      "when": [
        {
          "column": "title",
          "operator": "endswith",
          "value": " - DAR"
        },
        {
          "column": "title_category",
          "operator": "not_equals",
          "value": "TV Shows"
        }
      ],
      "message": "DAR titles must include Pristine Brand, Pristine Talent, or Pristine Film in companies."
    },
    {
      "sheet": "*",
      "column": "companies",
      "check": "not_blank_and_not_in",
      "tokens": ["#NA", "N/A"],
      "when": [
        {
          "column": "title",
          "operator": "not_endswith",
          "value": " - DAR"
        }
      ],
      "message": "Companies cannot be blank for non-DAR titles."
    },
    {
      "sheet": "*",
      "column": "brand_set",
      "check": "contains",
      "value": "Pristine DAR Brands",
      "when": [
        {
          "column": "title",
          "operator": "endswith",
          "value": " - DAR"
        }
      ],
      "message": "DAR titles must include Pristine DAR Brands in brand_set."
    },
    {
      "sheet": "*",
      "column": "brand_set",
      "check": "contains",
      "value": "Competitive View",
      "when": [
        {
          "column": "title",
          "operator": "not_endswith",
          "value": " - DAR"
        }
      ],
      "message": "Non-DAR titles must include Competitive View in brand_set."
    },
    {
      "sheet": "*",
      "column": "brand_set",
      "check": "contains",
      "value": "[Data Feed] Film - Wide Release + Custom Requests",
      "when": [
        {
          "column": "title",
          "operator": "not_endswith",
          "value": " - DAR"
        },
        {
          "column": "title_sub_category",
          "operator": "contains",
          "value": "Release - Wide"
        }
      ],
      "message": "Wide release titles must include [Data Feed] Film - Wide Release + Custom Requests in brand_set."
    },
    {
      "sheet": "*",
      "column": "brand_set",
      "check": "contains",
      "value": "[Data Feed] Film - Wide Release + Custom Requests",
      "when": [
        {
          "column": "title",
          "operator": "not_endswith",
          "value": " - DAR"
        },
        {
          "column": "title_sub_category",
          "operator": "contains",
          "value": "Release Type - Wide"
        }
      ],
      "message": "Wide release titles must include [Data Feed] Film - Wide Release + Custom Requests in brand_set."
    },
    {
      "sheet": "*",
      "column": "brand_set",
      "check": "contains",
      "value": "LF // Film - Majors + Independents",
      "when": [
        {
          "column": "title_category",
          "operator": "equals",
          "value": "Movies"
        },
        {
          "column": "title",
          "operator": "endswith",
          "value": " - DAR"
        },
        {
          "column": "title_sub_category",
          "operator": "contains",
          "value": "Release - Wide"
        }
      ],
      "message": "DAR wide release movies must include LF // Film - Majors + Independents in brand_set."
    },
    {
      "sheet": "*",
      "column": "brand_set",
      "check": "contains",
      "value": "LF // Film - Majors + Independents",
      "when": [
        {
          "column": "title_category",
          "operator": "equals",
          "value": "Movies"
        },
        {
          "column": "title",
          "operator": "endswith",
          "value": " - DAR"
        },
        {
          "column": "title_sub_category",
          "operator": "contains",
          "value": "Release Type - Wide"
        }
      ],
      "message": "DAR wide release movies must include LF // Film - Majors + Independents in brand_set."
    },
    {
      "sheet": "*",
      "column": "brand_set",
      "check": "contains",
      "value": "LF // Film - Majors + Independents",
      "when": [
        {
          "column": "title_category",
          "operator": "equals",
          "value": "Movies"
        },
        {
          "column": "title",
          "operator": "endswith",
          "value": " - DAR"
        },
        {
          "column": "title_sub_category",
          "operator": "contains",
          "value": "Release - Limited"
        }
      ],
      "message": "DAR limited release movies must include LF // Film - Majors + Independents in brand_set."
    },
    {
      "sheet": "*",
      "column": "brand_set",
      "check": "contains",
      "value": "LF // Film - Majors + Independents",
      "when": [
        {
          "column": "title_category",
          "operator": "equals",
          "value": "Movies"
        },
        {
          "column": "title",
          "operator": "endswith",
          "value": " - DAR"
        },
        {
          "column": "title_sub_category",
          "operator": "contains",
          "value": "Release Type - Limited"
        }
      ],
      "message": "DAR limited release movies must include LF // Film - Majors + Independents in brand_set."
    },
    {
      "sheet": "*",
      "column": "brand_set",
      "check": "regex",
      "pattern": "^\\\\s*Competitive View\\\\s*$",
      "when": [
        {
          "column": "title_category",
          "operator": "equals",
          "value": "Movies"
        },
        {
          "column": "title",
          "operator": "not_endswith",
          "value": " - DAR"
        },
        {
          "column": "title_sub_category",
          "operator": "contains",
          "value": "Release - Limited"
        }
      ],
      "message": "Non-DAR limited release movies must have only Competitive View in brand_set."
    },
    {
      "sheet": "*",
      "column": "brand_set",
      "check": "regex",
      "pattern": "^\\\\s*Competitive View\\\\s*$",
      "when": [
        {
          "column": "title_category",
          "operator": "equals",
          "value": "Movies"
        },
        {
          "column": "title",
          "operator": "not_endswith",
          "value": " - DAR"
        },
        {
          "column": "title_sub_category",
          "operator": "contains",
          "value": "Release Type - Limited"
        }
      ],
      "message": "Non-DAR limited release movies must have only Competitive View in brand_set."
    },
    {
      "sheet": "*",
      "column": "brand_set",
      "check": "not_blank_and_not_in",
      "tokens": ["#NA", "N/A"],
      "message": "Brand set cannot be blank, #NA, or N/A."
    },
    {
      "sheet": "*",
      "column": "facebook_page",
      "check": "url_not_contains_if_present",
      "tokens": ["/p/", "/page/", "/pages/", "/php/", "profile.php"],
      "message": "Facebook URL cannot contain /p/, /page/, /pages/, /php/, or profile.php."
    },
    {
      "sheet": "*",
      "column": "twitter_search_terms",
      "check": "not_blank_and_not_in",
      "tokens": ["#NA", "N/A"],
      "when": [
        {
          "column": "title_category",
          "operator": "in",
          "values": ["Movies", "TV Shows"]
        }
      ],
      "message": "twitter_search_terms cannot be blank for Movies and TV Shows."
    },
    {
      "sheet": "*",
      "column": "twitter_search_term_keywords",
      "check": "not_blank_and_not_in",
      "tokens": ["#NA", "N/A"],
      "when": [
        {
          "column": "title_category",
          "operator": "in",
          "values": ["Movies", "TV Shows"]
        }
      ],
      "message": "twitter_search_term_keywords cannot be blank for Movies and TV Shows."
    },
    {
      "sheet": "*",
      "column": "twitter_search_term_keywords",
      "check": "not_equals",
      "value": "#VALUE!",
      "message": "twitter_search_term_keywords cannot contain #VALUE!."
    },
    {
      "sheet": "*",
      "column": "twitter_handle",
      "check": "social_reference_format",
      "platform": "twitter",
      "message": "Twitter/X field must be a handle or profile URL, not a tweet/status link."
    },
    {
      "sheet": "*",
      "column": "instagram_user",
      "check": "social_reference_format",
      "platform": "instagram",
      "message": "Instagram field must be a username or profile URL, not a post, reel, story, or TV link."
    },
    {
      "sheet": "*",
      "column": "youtube_channel_username",
      "check": "social_reference_format",
      "platform": "youtube",
      "message": "YouTube field must be a channel handle or channel URL, not a video, shorts, or playlist link."
    },
    {
      "sheet": "*",
      "column": "youtube_channel_username",
      "check": "social_reference_reachable",
      "platform": "youtube",
      "message": "Each YouTube channel listed here must resolve to a valid channel and appear related to the title."
    },
    {
      "sheet": "*",
      "column": "youtube_channel_username",
      "check": "url_not_contains_if_present",
      "tokens": ["%20", "%7"],
      "message": "There is a syntax error in the URL. Please correct it by removing '%20' or '%7' and replacing it with a space."
    },
    {
      "sheet": "*",
      "column": "youtube_channel_company",
      "check": "url_not_contains_if_present",
      "tokens": ["%20"],
      "message": "There is a syntax error in the URL. Please correct it by removing '%20' and replacing it with a space."
    },
    {
      "sheet": "*",
      "column": "tiktok_user",
      "check": "social_reference_format",
      "platform": "tiktok",
      "message": "TikTok field must be a username or profile URL, not a video link."
    },
    {
      "sheet": "*",
      "column": "wikidata_id",
      "check": "social_reference_format",
      "platform": "wikidata",
      "message": "Wikidata field must be a valid Q-id or Wikidata item URL."
    },
    {
      "sheet": "*",
      "column": "wikidata_id",
      "check": "reference_lookup_match",
      "platform": "wikidata"
    },
    {
      "sheet": "*",
      "column": "imdb_id",
      "check": "social_reference_format",
      "platform": "imdb",
      "message": "IMDb field must be a valid tt... or nm... id, or an IMDb title/name URL."
    },
    {
      "sheet": "*",
      "column": "imdb_id",
      "check": "reference_lookup_match",
      "platform": "imdb"
    }
  ]
}"""


def _select_worksheets(workbook: Workbook, sheet_name: str) -> list:
    cleaned = (sheet_name or "*").strip()
    if cleaned in {"*", "Any"}:
        return [workbook[sheet] for sheet in workbook.sheetnames]
    if cleaned not in workbook.sheetnames:
        raise WorkbookValidationConfigError(f"Sheet '{cleaned}' was not found in the workbook.")
    return [workbook[cleaned]]


def _build_header_map(worksheet, header_row: int) -> dict[str, int]:
    header_map: dict[str, int] = {}
    aliases = {
        "brand/property tracked": "title",
        "brand/property or talent": "title_category",
        "brand/talent": "title",
        "facebook url": "facebook_page",
        "facebook page url": "facebook_page",
        "facebook page": "facebook_page",
        "twitter/x url": "twitter_handle",
        "twitter handle": "twitter_handle",
        "twitter url": "twitter_handle",
        "twitter/x handle": "twitter_handle",
        "instagram url": "instagram_user",
        "instagram user": "instagram_user",
        "instagram handle": "instagram_user",
        "youtube url": "youtube_channel_username",
        "youtube channel": "youtube_channel_username",
        "youtube channel username": "youtube_channel_username",
        "tiktok url": "tiktok_user",
        "tiktok user": "tiktok_user",
        "tiktok handle": "tiktok_user",
        "wikipedia url": "wikipedia_url",
        "wikipedia page": "wikipedia_url",
        "wikidata id": "wikidata_id",
        "imdb id": "imdb_id",
    }
    for cell in worksheet[header_row]:
        value = str(cell.value).strip() if cell.value is not None else ""
        if value:
            norm_val = value.casefold()
            header_map[norm_val] = cell.column
            if norm_val in aliases:
                header_map[aliases[norm_val]] = cell.column
    return header_map


def _get_worksheet_validation_context(
    cache: dict[tuple[str, int], WorksheetValidationContext],
    worksheet,
    header_row: int,
) -> WorksheetValidationContext:
    cache_key = (worksheet.title, header_row)
    cached = cache.get(cache_key)
    if cached is not None:
        return cached

    rows_values = list(worksheet.iter_rows(values_only=True))

    context = WorksheetValidationContext(
        worksheet=worksheet,
        header_row=header_row,
        header_map=_build_header_map(worksheet, header_row),
        rows_values=rows_values,
        active_rows=[],
    )
    context.active_rows = _build_active_rows(context)
    cache[cache_key] = context
    return context


def _build_active_rows(context: WorksheetValidationContext) -> list[int]:
    active_rows: list[int] = []
    header_row = context.header_row
    for idx in range(header_row, len(context.rows_values)):
        row_values = context.rows_values[idx]
        is_empty = True
        for val in row_values:
            if val is not None and not (isinstance(val, str) and not val.strip()):
                is_empty = False
                break
        if not is_empty:
            active_rows.append(idx + 1)
    return active_rows



def _compile_conditions(
    worksheet_context: WorksheetValidationContext,
    conditions: list[ValidationCondition],
    *,
    wildcard_sheet: bool,
) -> list[CompiledValidationCondition] | None:
    compiled: list[CompiledValidationCondition] = []
    for condition in conditions:
        key = condition.column.strip().casefold()
        if key not in worksheet_context.header_map:
            if wildcard_sheet:
                return None
            raise WorkbookValidationConfigError(
                f"Column '{condition.column}' was not found in sheet '{worksheet_context.worksheet.title}' "
                f"on header row {worksheet_context.header_row}."
            )
        compiled.append(CompiledValidationCondition(condition=condition, key=key))
    return compiled


def _get_required_column_index(header_map: dict[str, int], column_name: str, sheet_name: str, header_row: int) -> int:
    column_index = header_map.get(column_name.strip().casefold())
    if column_index is None:
        raise WorkbookValidationConfigError(
            f"Column '{column_name}' was not found in sheet '{sheet_name}' on header row {header_row}."
        )
    return column_index


def _get_optional_column_index(header_map: dict[str, int], column_name: str) -> int | None:
    return header_map.get(column_name.strip().casefold())


def _required_columns_for_rule(rule: ValidationRule) -> list[str]:
    if rule.check in {
        "rottentomatoes_url_match",
        "movie_us_release_date_match",
        "movie_release_type_match",
        "movie_genre_match",
        "reference_lookup_match",
    }:
        required_columns = ["title"]
        if rule.check == "rottentomatoes_url_match":
            required_columns.append("released_on")
        return required_columns
    if (
        rule.platform == "youtube"
        and rule.column == "youtube_channel_username"
        and rule.check in {"social_reference_format", "social_reference_reachable"}
    ):
        return ["title"]
    return []


def _validate_unique_rule(worksheet, header_map: dict[str, int], column_index: int, rule: ValidationRule) -> list[WorkbookValidationIssue]:
    value_rows: dict[str, list[int]] = {}
    for row_number in range(rule.header_row + 1, worksheet.max_row + 1):
        if _row_is_empty(worksheet, row_number):
            continue
        if not _conditions_match(worksheet, row_number, header_map, rule.when, rule.header_row):
            continue
        cell = worksheet.cell(row=row_number, column=column_index)
        normalized = _normalize_value(cell.value, ignore_case=rule.ignore_case)
        if not normalized:
            continue
        value_rows.setdefault(normalized, []).append(row_number)

    issues: list[WorkbookValidationIssue] = []
    for duplicate_rows in value_rows.values():
        if len(duplicate_rows) < 2:
            continue
        for row_number in duplicate_rows:
            cell = worksheet.cell(row=row_number, column=column_index)
            issues.append(
                _mark_issue(
                    cell=cell,
                    rule=rule,
                    message=rule.message or f"{rule.column} must be unique.",
                    value=cell.value,
                )
            )
    return issues


def _validate_unique_rule_with_context(
    worksheet_context: WorksheetValidationContext,
    column_index: int,
    column_key: str,
    rule: ValidationRule,
    compiled_conditions: list[CompiledValidationCondition],
) -> list[WorkbookValidationIssue]:
    value_rows: dict[str, list[int]] = {}
    for row_number in worksheet_context.active_rows:
        row_context = _get_row_context(worksheet_context, row_number)
        if not _conditions_match_row_context(row_context, compiled_conditions):
            continue

        normalized = _normalize_value(row_context.get(column_key), ignore_case=rule.ignore_case)
        if not normalized:
            continue
        value_rows.setdefault(normalized, []).append(row_number)

    issues: list[WorkbookValidationIssue] = []
    for duplicate_rows in value_rows.values():
        if len(duplicate_rows) < 2:
            continue
        for row_number in duplicate_rows:
            cell = worksheet_context.worksheet.cell(row=row_number, column=column_index)
            issues.append(
                _mark_issue(
                    cell=cell,
                    rule=rule,
                    message=rule.message or f"{rule.column} must be unique.",
                    value=cell.value,
                    finding_category="Duplicate",
                    confidence="High",
                    confidence_reason="Column value must be unique, but identical value exists in other row(s)."
                )
            )
    return issues


def _conditions_match(
    worksheet,
    row_number: int,
    header_map: dict[str, int],
    conditions: list[ValidationCondition],
    header_row: int,
) -> bool:
    for condition in conditions:
        column_index = _get_required_column_index(header_map, condition.column, worksheet.title, header_row)
        value = worksheet.cell(row=row_number, column=column_index).value
        if not _evaluate_condition(value, condition):
            return False
    return True


def _conditions_match_row_context(
    row_context: dict[str, Any],
    compiled_conditions: list[CompiledValidationCondition],
) -> bool:
    for compiled in compiled_conditions:
        if not _evaluate_condition(row_context.get(compiled.key), compiled.condition):
            return False
    return True


def _evaluate_condition(value: Any, condition: ValidationCondition) -> bool:
    if condition.operator == "not_blank":
        return not _is_blank(value)
    if condition.operator == "blank":
        return _is_blank(value)

    if condition.operator == "equals":
        return _normalize_value(value, condition.ignore_case) == _normalize_value(condition.value, condition.ignore_case)
    if condition.operator == "not_equals":
        return _normalize_value(value, condition.ignore_case) != _normalize_value(condition.value, condition.ignore_case)
    if condition.operator == "in":
        actual = _normalize_value(value, condition.ignore_case)
        expected = {_normalize_value(item, condition.ignore_case) for item in condition.values}
        return actual in expected
    if condition.operator == "not_in":
        actual = _normalize_value(value, condition.ignore_case)
        expected = {_normalize_value(item, condition.ignore_case) for item in condition.values}
        return actual not in expected
    if condition.operator == "endswith":
        return _normalize_value(value, condition.ignore_case).endswith(
            _normalize_value(condition.value, condition.ignore_case)
        )
    if condition.operator == "not_endswith":
        return not _normalize_value(value, condition.ignore_case).endswith(
            _normalize_value(condition.value, condition.ignore_case)
        )
    if condition.operator == "contains":
        return _normalize_value(condition.value, condition.ignore_case) in _normalize_value(value, condition.ignore_case)

    raise WorkbookValidationConfigError(f"Unsupported condition type: {condition.operator}")


def _classify_facebook_page(url: str, title: str, page_text: str) -> tuple[str, str, str]:
    normalized = re.sub(r"\s+", " ", page_text).casefold() if page_text else ""
    url_lower = url.lower() if url else ""
    title_lower = title.lower() if title else ""
    
    # Extract handle/ID from URL
    handle = ""
    parsed = urlparse(url)
    segments = [s for s in parsed.path.split("/") if s]
    if segments:
        handle = segments[0].lower()
    if handle == "profile.php" and parsed.query:
        from urllib.parse import parse_qs
        handle = parse_qs(parsed.query).get("id", [""])[0]
    
    # 1. Suspicious / Impersonation Mismatch
    title_words = [w for w in re.findall(r"\b[a-z]{3,}\b", title_lower)]
    if handle and title_words and not any(w in handle for w in title_words):
        return "Suspicious/Impersonation", "Low", f"Facebook handle '{handle}' does not match title words."
        
    # 2. Fan Page
    if "fan page" in normalized or "fanpage" in normalized or "unofficial" in normalized or "backup" in normalized or "parody" in normalized:
        return "Fan Page", "High", "Page description indicates fan page, parody, or unofficial archive."
        
    # 3. Community Page
    if "community group" in normalized or "community page" in normalized or "public group" in normalized:
        return "Community Page", "High", "Page is labeled as a community group or public group."
        
    # 4. Auto-generated
    if "interest page" in normalized or "topic page" in normalized or "auto-generated" in normalized:
        return "Auto-generated", "High", "Interest page generated automatically by Facebook."
        
    # 5. Official Regional
    regional_words = {"france", "germany", "spain", "italy", "japan", "brazil", "mexico", "india", "uk", "canada", "australia", "regional", "latam", "europe", "asia", "localized"}
    if any(word in title_lower or word in handle for word in regional_words):
        return "Official Regional", "Medium", "Local branch or regional keyword found in title/handle."
        
    # 6. Official Verified
    if "verified account" in normalized or "verified profile" in normalized or "official page" in normalized:
        return "Official", "High", "Verified official page badge or metadata confirmed."
        
    # 7. Unofficial
    if "unofficial" in normalized or "fan-made" in normalized:
        return "Unofficial", "Medium", "Unverified or unofficial fan page attributes."
        
    # Default is Unable to Verify if page is empty or geoblocked, or Official with Medium confidence if active
    if not page_text or len(page_text) < 100:
        return "Unable to Verify", "Low", "Empty page content or page is private/geoblocked."
        
    return "Official", "Medium", "Active profile matches title name, but lacks explicit verification badge."


def _evaluate_rule(
    value: Any,
    rule: ValidationRule,
    row_context: dict[str, Any] | None = None,
    social_cache: dict[str, tuple[bool, str]] | None = None,
    social_client: httpx.Client | None = None,
    movie_release_cache: dict[str, tuple[bool, dict[str, Any] | None, str]] | None = None,
    reference_cache: dict[str, tuple[bool, dict[str, Any] | None, str]] | None = None,
    rottentomatoes_cache: dict[str, tuple[bool, dict[str, Any] | None, str]] | None = None,
) -> tuple[bool, str, str, str, str]:
    passed, message = _evaluate_rule_core(
        value,
        rule,
        row_context,
        social_cache,
        social_client,
        movie_release_cache,
        reference_cache,
        rottentomatoes_cache,
    )
    if passed:
        return True, "", "Verified Correct", "High", ""

    category = "Suspected Incorrect"
    confidence = "High"
    reason = ""

    message_lower = message.lower()
    
    # Check for quality issues / Fan pages
    if "fan page" in message_lower or "fanpage" in message_lower:
        category = "Fan Page Detected"
        confidence = "High"
        reason = "Page content or attributes indicate this is a fan page, not an official account."
    elif "friends instead of followers" in message_lower:
        category = "Suspected Incorrect"
        confidence = "High"
        reason = "Profile shows friends count instead of followers, suggesting a personal timeline rather than an official page."
    elif "blocked url pattern" in message_lower or "cannot contain" in message_lower:
        category = "Suspected Incorrect"
        confidence = "High"
        reason = "The URL format contains path segments reserved for personal timelines or page setup templates."
    elif "http 404" in message_lower or "404" in message_lower:
        category = "No Official Presence Found"
        confidence = "High"
        reason = "The social link returns HTTP 404 Not Found, indicating the account does not exist or has been deleted."
    elif "http 403" in message_lower or "forbidden" in message_lower or "403" in message_lower:
        category = "Unable to Verify"
        confidence = "Low"
        reason = "Access to the profile is restricted or geoblocked (HTTP 403 Forbidden)."
    elif "timeout" in message_lower or "connect" in message_lower:
        category = "Unable to Verify"
        confidence = "Low"
        reason = "The server timed out or failed to connect while verifying the account."
    elif "does not appear related to title" in message_lower:
        category = "Suspected Incorrect"
        confidence = "Medium"
        reason = "The profile is reachable but its content does not match the title name."
    elif "unable to verify" in message_lower:
        category = "Unable to Verify"
        confidence = "Low"
        reason = "Validation check failed to verify the profile authenticity due to restricted access or missing metadata."
    elif "is required" in message_lower:
        category = "Needs Manual Review"
        confidence = "High"
        reason = "Field is required according to rules but is currently blank."
    elif "must match one of the allowed values" in message_lower or "approved list" in message_lower:
        category = "Suspected Incorrect"
        confidence = "High"
        reason = "Cell value is not in the list of approved classification tags."
    
    return False, message, category, confidence, reason


def _evaluate_rule_core(
    value: Any,
    rule: ValidationRule,
    row_context: dict[str, Any] | None = None,
    social_cache: dict[str, tuple[bool, str]] | None = None,
    social_client: httpx.Client | None = None,
    movie_release_cache: dict[str, tuple[bool, dict[str, Any] | None, str]] | None = None,
    reference_cache: dict[str, tuple[bool, dict[str, Any] | None, str]] | None = None,
    rottentomatoes_cache: dict[str, tuple[bool, dict[str, Any] | None, str]] | None = None,
) -> tuple[bool, str]:
    if rule.check == "required":
        passed = not _is_blank(value)
        return passed, rule.message or f"{rule.column} is required."

    if rule.check == "not_blank_and_not_in":
        actual = _normalize_value(value, rule.ignore_case)
        invalid = {_normalize_value(item, rule.ignore_case) for item in rule.tokens}
        passed = bool(actual) and actual not in invalid
        return passed, rule.message or f"{rule.column} cannot be blank or use placeholder values."

    if rule.check == "equals":
        actual = _normalize_value(value, rule.ignore_case)
        expected = _normalize_value(rule.value, rule.ignore_case)
        return actual == expected, rule.message or f"{rule.column} must equal {rule.value!r}."

    if rule.check == "not_equals":
        actual = _normalize_value(value, rule.ignore_case)
        expected = _normalize_value(rule.value, rule.ignore_case)
        return actual != expected, rule.message or f"{rule.column} must not equal {rule.value!r}."

    if rule.check == "in":
        actual = _normalize_value(value, rule.ignore_case)
        expected = {_normalize_value(item, rule.ignore_case) for item in rule.values}
        passed = bool(actual) and actual in expected
        return passed, rule.message or f"{rule.column} must match one of the allowed values."

    if rule.check == "regex":
        text = _normalize_value(value, ignore_case=False)
        passed = bool(text and re.fullmatch(rule.pattern or "", text))
        return passed, rule.message or f"{rule.column} does not match the expected format."

    if rule.check == "min":
        numeric = _as_number(value)
        passed = numeric is not None and numeric >= float(rule.value)
        return passed, rule.message or f"{rule.column} must be at least {rule.value}."

    if rule.check == "max":
        numeric = _as_number(value)
        passed = numeric is not None and numeric <= float(rule.value)
        return passed, rule.message or f"{rule.column} must be at most {rule.value}."

    if rule.check == "between":
        numeric = _as_number(value)
        passed = numeric is not None and rule.min_value is not None and rule.max_value is not None
        passed = bool(passed and rule.min_value <= numeric <= rule.max_value)
        return passed, rule.message or f"{rule.column} must be between {rule.min_value} and {rule.max_value}."

    if rule.check == "unique":
        raise WorkbookValidationConfigError("Unique rules must be evaluated separately.")

    if rule.check == "date_not_past":
        parsed_date = _as_date(value)
        passed = parsed_date is not None and parsed_date >= date.today()
        return passed, rule.message or f"{rule.column} cannot be earlier than today."

    if rule.check == "date_not_future":
        parsed_date = _as_date(value)
        passed = parsed_date is not None and parsed_date <= date.today()
        return passed, rule.message or f"{rule.column} cannot be later than today."

    if rule.check == "contains":
        actual = _normalize_value(value, rule.ignore_case)
        needle = _normalize_value(rule.value, rule.ignore_case)
        return bool(actual) and needle in actual, rule.message or f"{rule.column} must contain {rule.value!r}."

    if rule.check == "contains_any":
        actual = _normalize_value(value, rule.ignore_case)
        needles = [_normalize_value(item, rule.ignore_case) for item in rule.values]
        return bool(actual) and any(item in actual for item in needles), rule.message or (
            f"{rule.column} must contain one of the required values."
        )

    if rule.check == "url_not_contains_if_present":
        raw_value = _normalize_value(value, ignore_case=False)
        actual = _normalize_value(value, rule.ignore_case)
        blocked_tokens = [_normalize_value(item, rule.ignore_case) for item in rule.tokens]
        if rule.column == "facebook_page" and raw_value:
            passed = not has_blocked_facebook_path(raw_value)
        else:
            passed = not actual or all(token not in actual for token in blocked_tokens)
        if not passed and rule.column == "youtube_channel_username":
            if "%7" in actual:
                return False, "Remove %7 and add filter that we called pipe."
            if "%20" in actual:
                return (
                    False,
                    "There is a syntax error in the URL. Please correct it by removing '%20' and replacing it with a space.",
                )
        return passed, rule.message or f"{rule.column} contains a blocked URL pattern."

    if rule.check == "talent_subcategory_format":
        actual = _normalize_value(value, rule.ignore_case)
        if not actual:
            return False, rule.message or f"{rule.column} must include gender and talent type or subtype."
        required_tokens = [_normalize_value(item, rule.ignore_case) for item in rule.tokens]
        has_gender = required_tokens[0] in actual
        has_type_or_subtype = any(token in actual for token in required_tokens[1:])
        passed = has_gender and has_type_or_subtype
        return passed, rule.message or f"{rule.column} must include gender and talent type or subtype."

    if rule.check == "rottentomatoes_url_match":
        passed, detail = _validate_rottentomatoes_url(
            row_context or {},
            value,
            rottentomatoes_cache if rottentomatoes_cache is not None else {},
        )
        if passed:
            return True, rule.message or ""
        if _should_skip_rottentomatoes_validation(detail):
            return True, ""
        if detail:
            return False, detail
        return False, rule.message or f"{rule.column} does not match the Rotten Tomatoes URL."

    if rule.check == "movie_us_release_date_match":
        passed, detail = _validate_movie_us_release_date(
            row_context or {},
            value,
            movie_release_cache if movie_release_cache is not None else {},
        )
        if passed:
            return True, rule.message or ""
        if _should_skip_tmdb_validation(detail):
            return True, ""
        if detail:
            return False, detail
        return False, rule.message or f"{rule.column} does not match the movie USA release date."

    if rule.check == "movie_release_type_match":
        passed, detail = _validate_movie_release_type(
            row_context or {},
            value,
            movie_release_cache if movie_release_cache is not None else {},
        )
        if passed:
            return True, rule.message or ""
        if _should_skip_tmdb_validation(detail):
            return True, ""
        if detail:
            return False, f"{rule.message} ({detail})" if rule.message else detail
        return False, rule.message or f"{rule.column} does not match the movie release type recommendation."

    if rule.check == "movie_genre_match":
        passed, detail = _validate_movie_genre(
            row_context or {},
            value,
            movie_release_cache if movie_release_cache is not None else {},
        )
        if passed:
            return True, rule.message or ""
        if _should_skip_tmdb_validation(detail):
            return True, ""
        if detail:
            return False, detail
        return False, rule.message or f"{rule.column} does not match the movie genre recommendation."

    if _should_skip_wikipedia_rule(rule):
        return True, ""

    if rule.check == "social_reference_format":
        passed, detail = _validate_social_reference_format(value, rule, row_context or {})
        if passed:
            return True, rule.message or ""
        if detail:
            return False, f"{rule.message} ({detail})" if rule.message else detail
        return False, rule.message or f"{rule.column} has an invalid format."

    if rule.check == "social_reference_reachable":
        passed, detail = _validate_social_reference(
            value,
            rule,
            social_cache if social_cache is not None else {},
            social_client,
            row_context or {},
        )
        if passed:
            return True, rule.message or ""
        if detail:
            return False, f"{rule.message} ({detail})" if rule.message else detail
        return False, rule.message or f"{rule.column} is not reachable."

    if rule.check == "reference_lookup_match":
        passed, detail = _validate_reference_lookup(
            value,
            rule,
            row_context or {},
            reference_cache if reference_cache is not None else {},
            social_client,
        )
        if passed:
            return True, rule.message or ""
        if detail:
            return False, detail
        return False, rule.message or f"{rule.column} could not be verified."

    raise WorkbookValidationConfigError(f"Unsupported rule type: {rule.check}")


def _should_skip_wikipedia_rule(rule: ValidationRule) -> bool:
    return rule.platform == "wikipedia" and rule.check in {
        "social_reference_format",
        "social_reference_reachable",
        "reference_lookup_match",
    }


def _row_is_empty(worksheet, row_number: int) -> bool:
    for cell in worksheet[row_number]:
        if not _is_blank(cell.value):
            return False
    return True


def _mark_issue(
    cell, 
    rule: ValidationRule | None, 
    message: str, 
    value: Any,
    finding_category: str = "Needs Manual Review",
    confidence: str = "High",
    confidence_reason: str = ""
) -> WorkbookValidationIssue:
    issue_fill, issue_font = ERROR_FILL, ERROR_FONT
    if rule:
        issue_fill, issue_font = _issue_style_for_rule(rule, message)
    cell.fill = issue_fill
    cell.font = issue_font

    comment_body = f"{rule.check if rule else 'duplicate_conflict'}: {message}"
    if cell.comment and comment_body not in cell.comment.text:
        cell.comment = Comment(f"{cell.comment.text}\n{comment_body}", "Validator")
    elif cell.comment is None:
        cell.comment = Comment(comment_body, "Validator")

    return WorkbookValidationIssue(
        sheet=cell.parent.title,
        row=cell.row,
        column=get_column_letter(cell.column),
        cell=cell.coordinate,
        rule=rule.check if rule else "duplicate_conflict",
        message=message,
        value="" if value is None else str(value),
        finding_category=finding_category,
        confidence=confidence,
        confidence_reason=confidence_reason,
    )


def _issue_style_for_rule(rule: ValidationRule, message: str) -> tuple[PatternFill, Font]:
    normalized_message = _normalize_value(message, ignore_case=False)
    if (
        rule.check == "url_not_contains_if_present"
        and rule.column in {"youtube_channel_username", "youtube_channel_company"}
        and ("%20" in normalized_message or "%7" in normalized_message)
    ):
        return WARNING_FILL, WARNING_FONT
    return ERROR_FILL, ERROR_FONT


def _append_summary_sheet(workbook: Workbook, issues: list[WorkbookValidationIssue]) -> None:
    if "Validation Summary" in workbook.sheetnames:
        del workbook["Validation Summary"]

    summary = workbook.create_sheet("Validation Summary")
    summary.append(SUMMARY_HEADERS)

    for issue in issues:
        summary.append(
            [
                issue.sheet,
                issue.row,
                issue.column,
                issue.cell,
                issue.rule,
                issue.finding_category,
                issue.confidence,
                issue.confidence_reason,
                issue.message,
                issue.value,
            ]
        )

    if not issues:
        summary.append(["Workbook", 0, "-", "-", "passed", "Verified Correct", "High", "", "No validation issues found.", "-"])

    for column in summary.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        summary.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 14), 40)


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def _normalize_value(value: Any, ignore_case: bool = True) -> str:
    if value is None:
        text = ""
    elif isinstance(value, datetime):
        text = value.isoformat()
    else:
        text = str(value).strip()
    return text.casefold() if ignore_case else text


def _reference_titles_from_metadata(metadata: dict[str, Any]) -> list[str]:
    titles: list[str] = []
    for value in [metadata.get("title"), *(metadata.get("alternate_titles") or [])]:
        normalized = _normalize_value(value, ignore_case=False)
        if normalized and normalized not in titles:
            titles.append(normalized)
    return titles


def _reference_type_matches_title_category(title_category: str, media_type: str) -> bool:
    if not media_type:
        return True
    if title_category == "movies":
        return media_type in {"movie", "short", "video", "tvmovie"}
    if title_category == "tv shows":
        return media_type in {"series", "episode", "tvseries", "tvminiseries", "tvepisode"}
    if title_category == "talent":
        return media_type == "person"
    return True


def _normalize_movie_lookup_title(value: Any) -> str:
    title = _normalize_value(value, ignore_case=False)
    if not title:
        return ""
    title = re.sub(r"\s+-\s+dar$", "", title, flags=re.IGNORECASE)
    title = re.sub(r"\s+", " ", title).strip()
    return title


def _rottentomatoes_media_type(row_context: dict[str, Any]) -> str:
    title_category = _normalize_value(row_context.get("title_category"))
    if title_category == "movies":
        return "movie"
    if title_category == "tv shows":
        return "tvSeries"
    return ""


def _normalize_rottentomatoes_url(value: Any) -> str | None:
    raw_value = unescape(_normalize_value(value, ignore_case=False))
    if not raw_value:
        return None

    cleaned = raw_value.strip()
    if cleaned.startswith("/"):
        cleaned = f"https://www.rottentomatoes.com{cleaned}"
    elif cleaned.casefold().startswith("www.rottentomatoes.com/"):
        cleaned = f"https://{cleaned}"
    elif cleaned.casefold().startswith("rottentomatoes.com/"):
        cleaned = f"https://www.{cleaned}"
    elif not _looks_like_url(cleaned):
        return None

    parsed = urlparse(cleaned)
    if "rottentomatoes.com" not in parsed.netloc.casefold():
        return None

    normalized_path = re.sub(r"/+$", "", parsed.path or "")
    if not normalized_path:
        return None
    if not normalized_path.casefold().startswith("/m/"):
        return None
    return f"https://www.rottentomatoes.com{normalized_path}"


def _strip_html_tags(raw_html: str) -> str:
    text = re.sub(r"<[^>]+>", " ", raw_html)
    return re.sub(r"\s+", " ", unescape(text)).strip()


def _extract_html_attribute(raw_html: str, attribute_name: str) -> str:
    pattern = re.compile(rf'\b{re.escape(attribute_name)}="([^"]*)"', flags=re.IGNORECASE)
    match = pattern.search(raw_html)
    if not match:
        return ""
    return unescape(match.group(1)).strip()


def _extract_rottentomatoes_page_identity(html: str) -> tuple[str, int | None]:
    for block in re.findall(r'<script type="application/ld\+json">\s*(.*?)\s*</script>', html, flags=re.DOTALL):
        try:
            payload = json.loads(block)
        except json.JSONDecodeError:
            continue
        for item in _iter_json_ld_objects(payload):
            title = _normalize_value(item.get("name"), ignore_case=False)
            if not title:
                continue
            published = _normalize_value(item.get("datePublished"), ignore_case=False)
            year_match = re.search(r"\b(\d{4})\b", published or "")
            return title, int(year_match.group(1)) if year_match else None

    match = re.search(r"<title>(.*?)</title>", html, flags=re.IGNORECASE | re.DOTALL)
    if not match:
        return "", None

    title = unescape(match.group(1)).strip()
    title = re.sub(r"\s*[\-|]\s*Rotten Tomatoes.*$", "", title, flags=re.IGNORECASE).strip()
    return title, None


def _iter_json_ld_objects(payload: Any) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    if isinstance(payload, dict):
        items.append(payload)
        graph_items = payload.get("@graph")
        if isinstance(graph_items, list):
            for item in graph_items:
                if isinstance(item, dict):
                    items.append(item)
    elif isinstance(payload, list):
        for item in payload:
            if isinstance(item, dict):
                items.extend(_iter_json_ld_objects(item))
    return items


def _parse_int(value: Any) -> int | None:
    text = _normalize_value(value, ignore_case=False)
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _parse_rottentomatoes_search_results(page_html: str, media_type: str) -> list[dict[str, Any]]:
    section_pattern = re.compile(
        rf'<search-page-result\b[^>]*\btype="{re.escape(media_type)}"[^>]*>(?P<section>.*?)</search-page-result>',
        flags=re.IGNORECASE | re.DOTALL,
    )
    section_match = section_pattern.search(page_html)
    if not section_match:
        return []

    row_pattern = re.compile(
        r"<search-page-media-row\b(?P<attributes>[^>]*)>(?P<body>.*?)</search-page-media-row>",
        flags=re.IGNORECASE | re.DOTALL,
    )
    title_pattern = re.compile(
        r'<a\b(?P<attributes>[^>]*)data-qa="info-name"(?P<rest>[^>]*)>(?P<content>.*?)</a>',
        flags=re.IGNORECASE | re.DOTALL,
    )

    results: list[dict[str, Any]] = []
    for row_match in row_pattern.finditer(section_match.group("section")):
        row_attributes = row_match.group("attributes")
        body = row_match.group("body")
        title_match = title_pattern.search(body)
        if not title_match:
            continue

        anchor_attributes = f"{title_match.group('attributes')} {title_match.group('rest')}"
        title = _strip_html_tags(title_match.group("content"))
        url = _normalize_rottentomatoes_url(_extract_html_attribute(anchor_attributes, "href"))
        if not title or not url:
            continue

        results.append(
            {
                "title": title,
                "url": url,
                "type": media_type,
                "release_year": _parse_int(
                    _extract_html_attribute(row_attributes, "release-year")
                    or _extract_html_attribute(row_attributes, "releaseyear")
                ),
                "start_year": _parse_int(
                    _extract_html_attribute(row_attributes, "start-year")
                    or _extract_html_attribute(row_attributes, "startyear")
                ),
                "end_year": _parse_int(
                    _extract_html_attribute(row_attributes, "end-year")
                    or _extract_html_attribute(row_attributes, "endyear")
                ),
            }
        )

    return results


def _rottentomatoes_year_matches(result: dict[str, Any], release_year: int) -> bool:
    release_result_year = _parse_int(result.get("release_year"))
    start_year = _parse_int(result.get("start_year"))
    end_year = _parse_int(result.get("end_year"))

    if release_result_year is not None:
        return release_result_year == release_year
    if start_year is not None and end_year is not None:
        return start_year <= release_year <= end_year
    if start_year is not None:
        return start_year == release_year
    return False


def _rottentomatoes_year_distance(result: dict[str, Any], release_year: int | None) -> int:
    if release_year is None:
        return 0
    if _rottentomatoes_year_matches(result, release_year):
        return 0

    candidate_years = [
        year
        for year in [
            _parse_int(result.get("release_year")),
            _parse_int(result.get("start_year")),
            _parse_int(result.get("end_year")),
        ]
        if year is not None
    ]
    if not candidate_years:
        return 9999
    return min(abs(candidate_year - release_year) for candidate_year in candidate_years)


def _select_rottentomatoes_search_result(
    query_title: str,
    results: list[dict[str, Any]],
    release_year: int | None = None,
) -> dict[str, Any] | None:
    matched_results = [result for result in results if _titles_loosely_match(query_title, str(result.get("title") or ""))]
    if not matched_results:
        return None

    if release_year is not None:
        year_matches = [result for result in matched_results if _rottentomatoes_year_matches(result, release_year)]
        if not year_matches:
            return None
        matched_results = year_matches

    normalized_query_title = _normalize_lookup_title(query_title)
    matched_results.sort(
        key=lambda result: (
            0 if _normalize_lookup_title(str(result.get("title") or "")) == normalized_query_title else 1,
            len(_normalize_lookup_title(str(result.get("title") or ""))),
        )
    )
    return matched_results[0]


def _lookup_rottentomatoes_title(
    title: str,
    media_type: str,
    release_year: int | None = None,
    client: httpx.Client | None = None,
) -> tuple[bool, dict[str, Any] | None, str]:
    try:
        response = _network_get(
            "https://www.rottentomatoes.com/search",
            client=client,
            params={"search": title},
        )
    except httpx.HTTPStatusError as exc:
        return False, None, f"Rotten Tomatoes lookup failed with HTTP {exc.response.status_code}"
    except httpx.HTTPError as exc:
        return False, None, f"Rotten Tomatoes lookup failed: {exc.__class__.__name__}"

    results = _parse_rottentomatoes_search_results(response.text, media_type)
    if not results:
        return False, None, "Rotten Tomatoes search results were empty"

    selected_result = _select_rottentomatoes_search_result(title, results, release_year=release_year)
    if selected_result is None:
        if release_year is not None:
            return False, None, f"Rotten Tomatoes title not found for release year {release_year}"
        return False, None, "Rotten Tomatoes title not found"

    verification_success, verified_result, verification_detail = _verify_rottentomatoes_result(
        title,
        selected_result,
        release_year=release_year,
        client=client,
    )
    if not verification_success:
        return False, None, verification_detail

    return True, verified_result, ""


def _verify_rottentomatoes_result(
    query_title: str,
    selected_result: dict[str, Any],
    release_year: int | None,
    client: httpx.Client | None = None,
) -> tuple[bool, dict[str, Any] | None, str]:
    normalized_url = _normalize_rottentomatoes_url(selected_result.get("url"))
    if not normalized_url:
        return False, None, "Rotten Tomatoes did not return a valid movie URL"

    try:
        response = _network_get(normalized_url, client=client)
    except httpx.HTTPStatusError as exc:
        return False, None, f"Rotten Tomatoes lookup failed with HTTP {exc.response.status_code}"
    except httpx.HTTPError as exc:
        return False, None, f"Rotten Tomatoes lookup failed: {exc.__class__.__name__}"

    page_url = _normalize_rottentomatoes_url(str(response.url))
    if not page_url:
        return False, None, "Rotten Tomatoes page did not resolve to a movie URL"

    page_title, page_year = _extract_rottentomatoes_page_identity(response.text)
    if not page_title:
        return False, None, "Rotten Tomatoes page title could not be read"

    if _normalize_lookup_title(page_title) != _normalize_lookup_title(query_title):
        return False, None, f"Rotten Tomatoes page title mismatch: {page_title}"

    verified_year = page_year or _parse_int(selected_result.get("release_year")) or _parse_int(selected_result.get("start_year"))
    if release_year is not None and verified_year is not None and verified_year != release_year:
        return False, None, f"Rotten Tomatoes page year mismatch: {verified_year}"

    merged_result = dict(selected_result)
    merged_result["title"] = page_title
    merged_result["url"] = page_url
    if verified_year is not None:
        merged_result["release_year"] = verified_year
    return True, merged_result, ""


def _get_rottentomatoes_metadata(
    row_context: dict[str, Any],
    cache: dict[str, tuple[bool, dict[str, Any] | None, str]],
) -> tuple[bool, dict[str, Any] | None, str]:
    title = _normalize_movie_lookup_title(row_context.get("title"))
    if not title:
        return False, None, "title is missing, so the Rotten Tomatoes URL cannot be checked"

    media_type = _rottentomatoes_media_type(row_context)
    if not media_type:
        return False, None, "title_category is not supported for Rotten Tomatoes lookup"

    release_year = _movie_lookup_release_year(row_context)
    if release_year is None:
        return False, None, "released_on must include a valid year so the Rotten Tomatoes URL can be checked"

    lookup_key = f"{media_type}|{title.casefold()}|{release_year or ''}"
    if lookup_key not in cache:
        cache[lookup_key] = _lookup_rottentomatoes_title(title, media_type=media_type, release_year=release_year)

    return cache[lookup_key]


def _validate_rottentomatoes_url(
    row_context: dict[str, Any],
    rottentomatoes_value: Any,
    cache: dict[str, tuple[bool, dict[str, Any] | None, str]],
) -> tuple[bool, str]:
    success, metadata, detail = _get_rottentomatoes_metadata(row_context, cache)
    if not success:
        return False, detail

    expected_url = _normalize_rottentomatoes_url((metadata or {}).get("url"))
    if not expected_url:
        return False, "Rotten Tomatoes did not return a URL"

    actual_url = _normalize_rottentomatoes_url(rottentomatoes_value)
    if actual_url == expected_url:
        return True, ""

    return False, f"Rotten Tomatoes URL: {expected_url}"


def _should_skip_rottentomatoes_validation(detail: str) -> bool:
    normalized_detail = _normalize_value(detail)
    if not normalized_detail:
        return False
    return normalized_detail.startswith("rotten tomatoes lookup failed:")


def _should_skip_tmdb_validation(detail: str) -> bool:
    normalized_detail = _normalize_value(detail)
    if not normalized_detail:
        return False
    return normalized_detail.startswith("tmdb lookup failed:") or normalized_detail == "tmdb is not configured"


def _build_row_context(context: WorksheetValidationContext, row_number: int) -> dict[str, Any]:
    row_idx = row_number - 1
    row_values = context.rows_values[row_idx]
    row_ctx: dict[str, Any] = {}
    for header_name, column_index in context.header_map.items():
        val_idx = column_index - 1
        if val_idx < len(row_values):
            row_ctx[header_name] = row_values[val_idx]
        else:
            row_ctx[header_name] = None
    return row_ctx


def _get_row_context(worksheet_context: WorksheetValidationContext, row_number: int) -> dict[str, Any]:
    cached = worksheet_context.row_context_cache.get(row_number)
    if cached is not None:
        return cached

    context = _build_row_context(worksheet_context, row_number)
    worksheet_context.row_context_cache[row_number] = context
    return context



def _as_number(value: Any) -> float | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def _as_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        parsers = [date.fromisoformat]
        parsers.extend(lambda item, fmt=fmt: datetime.strptime(item, fmt).date() for fmt in COMMON_DATE_FORMATS)
        for parser in parsers:
            try:
                return parser(text)
            except ValueError:
                continue
    return None


def _validate_social_reference(
    value: Any,
    rule: ValidationRule,
    cache: dict[str, tuple[bool, str]],
    client: httpx.Client | None,
    row_context: dict[str, Any] | None = None,
) -> tuple[bool, str]:
    raw_value = _normalize_value(value, ignore_case=False)
    if not raw_value:
        return True, ""
    if _contains_control_characters(raw_value):
        return False, "value contains a non-printable control character"

    if rule.platform is None:
        raise WorkbookValidationConfigError("Social reference rule is missing a platform.")

    references = _split_social_reference_values(raw_value)
    title = _normalize_value((row_context or {}).get("title"), ignore_case=False)
    for index, reference in enumerate(references, start=1):
        normalized_url = _normalize_social_reference(rule.platform, reference)
        if not normalized_url:
            return False, _social_reference_entry_detail(index, len(references), "could not normalize the value into a valid URL")

        cache_key = f"{rule.platform}:{normalized_url}"
        if cache_key not in cache:
            cache[cache_key] = _fetch_social_reference(normalized_url, rule.platform, client, title)

        success, detail = cache[cache_key]
        if not success:
            return False, _social_reference_entry_detail(index, len(references), detail)
        if _should_match_youtube_reference_to_title(rule) and not _youtube_reference_matches_title(reference, title):
            return False, _social_reference_entry_detail(
                index,
                len(references),
                f"does not appear related to title '{title}'",
            )

    return True, ""


def _validate_social_reference_format(
    value: Any,
    rule: ValidationRule,
    row_context: dict[str, Any] | None = None,
) -> tuple[bool, str]:
    raw_value = _normalize_value(value, ignore_case=False)
    if not raw_value:
        return True, ""
    if _contains_control_characters(raw_value):
        return False, "value contains a non-printable control character"
    if rule.platform is None:
        raise WorkbookValidationConfigError("Social reference format rule is missing a platform.")

    references = _split_social_reference_values(raw_value)
    title = _normalize_value((row_context or {}).get("title"), ignore_case=False)
    for index, reference in enumerate(references, start=1):
        passed, detail = _matches_social_reference_format(rule.platform, reference)
        if not passed:
            return False, _social_reference_entry_detail(index, len(references), detail)
        if _should_match_youtube_reference_to_title(rule) and not _youtube_reference_matches_title(reference, title):
            return False, _social_reference_entry_detail(
                index,
                len(references),
                f"does not appear related to title '{title}'",
            )
    return True, ""


def _validate_reference_lookup(
    value: Any,
    rule: ValidationRule,
    row_context: dict[str, Any],
    cache: dict[str, tuple[bool, dict[str, Any] | None, str]],
    client: httpx.Client | None,
) -> tuple[bool, str]:
    raw_value = _normalize_value(value, ignore_case=False)
    if not raw_value:
        return True, ""
    if _contains_control_characters(raw_value):
        return False, "value contains a line break or other non-printable character"
    if rule.platform not in {"wikipedia", "wikidata", "imdb"}:
        raise WorkbookValidationConfigError("Reference lookup is only supported for Wikipedia, Wikidata, and IMDb.")

    title = _normalize_value(row_context.get("title"), ignore_case=False)
    if not title:
        return False, "title is missing, so the reference cannot be verified"

    cache_key = f"{rule.platform}:{raw_value.strip()}"
    if cache_key not in cache:
        cache[cache_key] = _lookup_reference_record(rule.platform, raw_value, client)

    success, metadata, detail = cache[cache_key]
    if not success:
        return False, detail
    if metadata is None:
        return False, "lookup did not return any metadata"

    reference_titles = _reference_titles_from_metadata(metadata)
    if not reference_titles:
        return False, f"{rule.platform.title()} lookup did not return a title"
    if not any(_titles_loosely_match(title, reference_title) for reference_title in reference_titles):
        label = "IMDb title" if rule.platform == "imdb" else "Wikimedia title"
        return False, f"{label}: {metadata.get('title')}"

    if rule.platform in {"wikipedia", "wikidata"}:
        expected_wikidata_id = _extract_wikidata_identifier(_normalize_value(row_context.get("wikidata_id"), ignore_case=False))
        actual_wikidata_id = _normalize_value(metadata.get("wikidata_id"), ignore_case=False)
        if expected_wikidata_id and actual_wikidata_id and expected_wikidata_id.casefold() != actual_wikidata_id.casefold():
            return False, f"Wikidata id: {actual_wikidata_id}"

    title_category = _normalize_value(row_context.get("title_category"))
    media_type = _normalize_value(metadata.get("type"))
    if rule.platform == "imdb":
        if not _reference_type_matches_title_category(title_category, media_type):
            return False, f"IMDb type: {metadata.get('type')}"
    elif rule.platform in {"wikipedia", "wikidata"}:
        if not _reference_type_matches_title_category(title_category, media_type):
            return False, f"Wikimedia type: {metadata.get('type')}"

    return True, ""


def _matches_social_reference_format(platform: str, raw_value: str) -> tuple[bool, str]:
    cleaned = raw_value.strip()
    if not cleaned:
        return True, ""

    if platform == "twitter":
        if _looks_like_url(cleaned):
            parsed = urlparse(cleaned)
            path = parsed.path.casefold()
            if "/status/" in path or path.startswith("/i/"):
                return False, "tweet/status URLs are not allowed"
            parts = [part for part in parsed.path.split("/") if part]
            return (len(parts) == 1, "profile URLs must point to a user profile")
        handle = cleaned.lstrip("@").strip("/")
        return (bool(re.fullmatch(r"[A-Za-z0-9_]{1,15}", handle)), "handle must be 1-15 letters, numbers, or underscores")

    if platform == "instagram":
        if _looks_like_url(cleaned):
            parsed = urlparse(cleaned)
            path = parsed.path.casefold()
            if any(token in path for token in ["/p/", "/reel/", "/stories/", "/tv/"]):
                return False, "post, reel, story, and TV URLs are not allowed"
            parts = [part for part in parsed.path.split("/") if part]
            return (len(parts) == 1, "profile URLs must point to a username profile")
        handle = cleaned.lstrip("@").strip("/")
        return (bool(re.fullmatch(r"[A-Za-z0-9._]{1,30}", handle)), "username contains invalid Instagram characters")

    if platform == "youtube":
        if _looks_like_url(cleaned):
            parsed = urlparse(cleaned)
            path = parsed.path.casefold()
            if parsed.netloc.casefold().endswith("youtu.be") or path.startswith("/watch") or path.startswith("/shorts/") or path.startswith("/playlist"):
                return False, "video, shorts, and playlist URLs are not allowed"
            valid_channel_paths = [
                path.startswith("/@"),
                path.startswith("/channel/"),
                path.startswith("/user/"),
            ]
            return (any(valid_channel_paths), "URL must point to a YouTube channel")
        handle = cleaned.strip()
        if handle.startswith("@"):
            return (bool(re.fullmatch(r"@[A-Za-z0-9._-]{1,100}", handle)), "channel handle contains invalid characters")
        return (bool(re.fullmatch(r"[A-Za-z0-9._-]{1,100}", handle)), "channel username contains invalid characters")

    if platform == "tiktok":
        if _looks_like_url(cleaned):
            parsed = urlparse(cleaned)
            path = parsed.path.casefold()
            if "/video/" in path:
                return False, "video URLs are not allowed"
            return (path.startswith("/@"), "URL must point to a TikTok profile")
        handle = cleaned.lstrip("@").strip("/")
        return (bool(re.fullmatch(r"[A-Za-z0-9._]{1,50}", handle)), "username contains invalid TikTok characters")

    if platform == "wikipedia":
        if _looks_like_url(cleaned):
            parsed = urlparse(cleaned)
            path = parsed.path.casefold()
            if "special:search" in path or not path.startswith("/wiki/"):
                return False, "URL must point to a Wikipedia article"
            return True, ""
        return (bool(cleaned.strip()), "page slug cannot be blank")

    if platform == "wikidata":
        if _looks_like_url(cleaned):
            parsed = urlparse(cleaned)
            path = parsed.path.casefold()
            return (
                "wikidata.org" in parsed.netloc.casefold() and bool(re.match(r"^/wiki/q\d+/?$", path)),
                "URL must point to a Wikidata item",
            )
        return (bool(re.fullmatch(r"Q\d+", cleaned, flags=re.IGNORECASE)), "value must be a Wikidata Q-id")

    if platform == "imdb":
        if _looks_like_url(cleaned):
            parsed = urlparse(cleaned)
            path = parsed.path.casefold()
            return (
                bool(re.match(r"^/(title/tt\d{7,}|name/nm\d{7,})/?$", path)),
                "URL must point to an IMDb title or name page",
            )
        return (
            bool(re.fullmatch(r"(tt|nm)\d{7,}", cleaned)),
            "value must be a tt... or nm... IMDb id",
        )

    if platform == "facebook":
        if _looks_like_url(cleaned) and has_blocked_facebook_path(cleaned):
            return False, "post and generic Facebook page URLs are not allowed"
        return True, ""

    raise WorkbookValidationConfigError(f"Unsupported social platform: {platform}")


def _split_social_reference_values(raw_value: str) -> list[str]:
    text = raw_value.strip()
    if not text:
        return []

    url_matches = [match.rstrip(".,);]") for match in re.findall(r"https?://[^\s,;|]+", text)]
    if len(url_matches) > 1:
        return [match for match in url_matches if match]

    return [item.strip(" \"'") for item in re.split(r"[\n,;|]+", text) if item.strip(" \"'")]


def _social_reference_entry_detail(index: int, total: int, detail: str) -> str:
    prefix = f"entry {index}" if total > 1 else "value"
    return f"{prefix}: {detail}" if detail else prefix


def _should_match_youtube_reference_to_title(rule: ValidationRule) -> bool:
    return rule.platform == "youtube" and rule.column == "youtube_channel_username"


def _youtube_reference_matches_title(raw_value: str, title: str) -> bool:
    normalized_title = _normalize_value(title, ignore_case=False)
    if not normalized_title:
        return True

    label = _extract_youtube_reference_label(raw_value)
    if not label:
        return True

    return _titles_loosely_match(normalized_title, label)


def _extract_youtube_reference_label(raw_value: str) -> str | None:
    normalized_url = _normalize_social_reference("youtube", raw_value)
    if not normalized_url:
        return None

    parsed = urlparse(normalized_url)
    parts = [unquote(part).strip() for part in parsed.path.split("/") if part.strip()]
    if not parts:
        return None

    if parts[0].startswith("@"):
        return parts[0][1:] or None
    if parts[0] in {"c", "user"} and len(parts) > 1:
        return parts[1] or None
    return None


def _normalize_social_reference(platform: str, raw_value: str) -> str | None:
    cleaned = raw_value.strip()
    if not cleaned:
        return None

    if platform == "facebook":
        return cleaned if _looks_like_url(cleaned) else f"https://www.facebook.com/{cleaned.lstrip('@/')}"

    if platform == "twitter":
        if _looks_like_url(cleaned):
            return cleaned
        handle = cleaned.lstrip("@").strip("/")
        if not re.fullmatch(r"[A-Za-z0-9_]{1,15}", handle):
            return None
        return f"https://x.com/{handle}"

    if platform == "instagram":
        if _looks_like_url(cleaned):
            return cleaned
        handle = cleaned.lstrip("@").strip("/")
        if not re.fullmatch(r"[A-Za-z0-9._]{1,30}", handle):
            return None
        return f"https://www.instagram.com/{handle}/"

    if platform == "youtube":
        if _looks_like_url(cleaned):
            return cleaned
        handle = cleaned.strip()
        if handle.startswith("@"):
            safe = handle.strip("/")
            return f"https://www.youtube.com/{safe}"
        if not re.fullmatch(r"[A-Za-z0-9._-]{1,100}", handle):
            return None
        return f"https://www.youtube.com/@{handle}"

    if platform == "tiktok":
        if _looks_like_url(cleaned):
            return cleaned
        handle = cleaned.lstrip("@").strip("/")
        if not re.fullmatch(r"[A-Za-z0-9._]{1,50}", handle):
            return None
        return f"https://www.tiktok.com/@{handle}"

    if platform == "wikipedia":
        if _looks_like_url(cleaned):
            return cleaned
        return f"https://en.wikipedia.org/wiki/{quote(cleaned.replace(' ', '_'))}"

    if platform == "wikidata":
        if _looks_like_url(cleaned):
            return cleaned
        if not re.fullmatch(r"Q\d+", cleaned, flags=re.IGNORECASE):
            return None
        return f"https://www.wikidata.org/wiki/{cleaned.upper()}"

    if platform == "imdb":
        if _looks_like_url(cleaned):
            return cleaned
        if re.fullmatch(r"tt\d{7,}", cleaned):
            return f"https://www.imdb.com/title/{cleaned}/"
        if re.fullmatch(r"nm\d{7,}", cleaned):
            return f"https://www.imdb.com/name/{cleaned}/"
        return None

    raise WorkbookValidationConfigError(f"Unsupported social platform: {platform}")


def _lookup_reference_record(
    platform: str,
    raw_value: str,
    client: httpx.Client | None,
) -> tuple[bool, dict[str, Any] | None, str]:
    if platform == "wikipedia":
        return _lookup_wikipedia_record(raw_value, client)
    if platform == "wikidata":
        return _lookup_wikidata_record(raw_value, client)
    if platform == "imdb":
        return _lookup_imdb_record(raw_value, client)
    raise WorkbookValidationConfigError(f"Unsupported reference lookup platform: {platform}")


def _lookup_wikipedia_record(raw_value: str, client: httpx.Client | None) -> tuple[bool, dict[str, Any] | None, str]:
    slug = _wikipedia_slug_from_reference(raw_value)
    if not slug:
        return False, None, "Wikipedia article could not be normalized"
    normalized_url = _normalize_social_reference("wikipedia", raw_value)
    if not normalized_url:
        return False, None, "Wikipedia article could not be normalized"

    cache_keys = [_wikimedia_cache_key("wikipedia", slug)]
    cached_result = _load_wikimedia_cached_record(cache_keys)
    if cached_result is not None and _is_wikipedia_cache_fresh(cached_result["checked_at"]):
        return _wikimedia_cached_row_to_result(cached_result)

    live_success, live_metadata, live_detail = _lookup_wikipedia_live_record(slug, normalized_url, client)
    if live_success:
        _save_wikimedia_cached_record(
            _wikimedia_cache_keys_from_metadata("wikipedia", slug, live_metadata),
            normalized_url,
            live_success,
            live_metadata,
            live_detail,
        )
        return live_success, live_metadata, live_detail

    if cached_result is not None:
        cached_success, cached_metadata, cached_detail = _wikimedia_cached_row_to_result(cached_result)
        if cached_success:
            return cached_success, cached_metadata, cached_detail

    _save_wikimedia_cached_record(cache_keys, normalized_url, live_success, live_metadata, live_detail)
    return live_success, live_metadata, live_detail


def _lookup_wikipedia_live_record(
    slug: str,
    normalized_url: str,
    client: httpx.Client | None,
) -> tuple[bool, dict[str, Any] | None, str]:
    try:
        response = _network_get(
            "https://en.wikipedia.org/w/api.php",
            client=client,
            params={
                "action": "query",
                "titles": slug,
                "redirects": "1",
                "prop": "info|pageprops",
                "inprop": "url",
                "ppprop": "wikibase_item|disambiguation",
                "format": "json",
                "formatversion": "2",
            },
            headers={"Api-User-Agent": settings.wikimedia_contact or "local-app"},
        )
    except httpx.HTTPStatusError as exc:
        if exc.response.status_code == 404:
            return False, None, "Wikipedia page not found"
        return False, None, f"Wikipedia lookup failed with HTTP {exc.response.status_code}"
    except httpx.HTTPError as exc:
        return False, None, f"Wikipedia lookup failed: {exc.__class__.__name__}"

    payload = response.json()
    pages = ((payload.get("query") or {}).get("pages") or [])
    if not pages:
        return False, None, "Wikipedia page not found"
    page = pages[0]
    if page.get("missing") is True:
        return False, None, "Wikipedia page not found"
    if _is_wikipedia_disambiguation_page(page):
        return False, None, "Wikipedia page is a disambiguation page"

    redirects = (payload.get("query") or {}).get("redirects") or []
    redirect_source = _normalize_value(redirects[0].get("from"), ignore_case=False) if redirects else ""
    pageprops = page.get("pageprops") or {}
    wikidata_id = _normalize_value(pageprops.get("wikibase_item"), ignore_case=False).upper()
    if not wikidata_id:
        return False, None, "Wikipedia page is not linked to a Wikidata item"

    canonical_title = _normalize_value(page.get("title"), ignore_case=False) or slug.replace("_", " ")
    canonical_url = (
        _normalize_value(page.get("canonicalurl"), ignore_case=False)
        or _normalize_value(page.get("fullurl"), ignore_case=False)
        or normalized_url
    )
    entity_success, entity_metadata, entity_detail = _lookup_wikidata_live_record(
        wikidata_id,
        f"https://www.wikidata.org/wiki/{wikidata_id}",
        client,
    )
    if not entity_success:
        return False, None, entity_detail
    if entity_metadata is None:
        return False, None, "Wikidata lookup did not return any metadata"

    metadata = _merge_wikipedia_and_wikidata_metadata(
        canonical_title=canonical_title,
        canonical_url=canonical_url,
        redirect_source=redirect_source,
        wikidata_metadata=entity_metadata,
    )
    return True, metadata, ""


def _lookup_wikidata_record(raw_value: str, client: httpx.Client | None) -> tuple[bool, dict[str, Any] | None, str]:
    wikidata_id = _extract_wikidata_identifier(raw_value)
    if not wikidata_id:
        return False, None, "Wikidata item could not be normalized"
    normalized_url = _normalize_social_reference("wikidata", raw_value)
    if not normalized_url:
        return False, None, "Wikidata item could not be normalized"

    cache_keys = [_wikimedia_cache_key("wikidata", wikidata_id)]
    cached_result = _load_wikimedia_cached_record(cache_keys)
    if cached_result is not None and _is_wikipedia_cache_fresh(cached_result["checked_at"]):
        return _wikimedia_cached_row_to_result(cached_result)

    live_success, live_metadata, live_detail = _lookup_wikidata_live_record(wikidata_id, normalized_url, client)
    if live_success:
        _save_wikimedia_cached_record(
            _wikimedia_cache_keys_from_metadata("wikidata", wikidata_id, live_metadata),
            normalized_url,
            live_success,
            live_metadata,
            live_detail,
        )
        return live_success, live_metadata, live_detail

    if cached_result is not None:
        cached_success, cached_metadata, cached_detail = _wikimedia_cached_row_to_result(cached_result)
        if cached_success:
            return cached_success, cached_metadata, cached_detail

    _save_wikimedia_cached_record(cache_keys, normalized_url, live_success, live_metadata, live_detail)
    return live_success, live_metadata, live_detail


def _lookup_wikidata_live_record(
    wikidata_id: str,
    normalized_url: str,
    client: httpx.Client | None,
) -> tuple[bool, dict[str, Any] | None, str]:
    try:
        response = _network_get(
            "https://www.wikidata.org/w/api.php",
            client=client,
            params={
                "action": "wbgetentities",
                "format": "json",
                "ids": wikidata_id,
                "props": "labels|aliases|sitelinks|claims",
                "sitefilter": "enwiki",
            },
            headers={"Api-User-Agent": settings.wikimedia_contact or "local-app"},
        )
    except httpx.HTTPStatusError as exc:
        if exc.response.status_code == 404:
            return False, None, "Wikidata item not found"
        return False, None, f"Wikidata lookup failed with HTTP {exc.response.status_code}"
    except httpx.HTTPError as exc:
        return False, None, f"Wikidata lookup failed: {exc.__class__.__name__}"

    payload = response.json()
    entity = (payload.get("entities") or {}).get(wikidata_id)
    if not entity or entity.get("missing"):
        return False, None, "Wikidata item not found"

    instance_of_ids = _extract_wikidata_instance_of_ids(entity)
    if _is_wikidata_disambiguation(entity, instance_of_ids):
        return False, None, "Wikidata item is a disambiguation page"

    metadata = _build_wikidata_metadata(
        wikidata_id=wikidata_id,
        entity=entity,
        normalized_url=normalized_url,
        instance_of_ids=instance_of_ids,
    )
    return True, metadata, ""


def _load_wikimedia_cached_record(lookup_keys: list[str]) -> sqlite3.Row | None:
    db_path = _wikipedia_cache_db_path()
    if not db_path.exists():
        return None
    with sqlite3.connect(db_path) as connection:
        connection.row_factory = sqlite3.Row
        try:
            for lookup_key in lookup_keys:
                row = connection.execute(
                    f"""
                    SELECT lookup_key, lookup_value, success, title, alternate_titles, canonical_url,
                           wikipedia_url, wikidata_url, wikidata_id, entity_type, redirect_source, detail, checked_at
                    FROM {WIKIMEDIA_CACHE_TABLE}
                    WHERE lookup_key = ?
                    """,
                    (lookup_key,),
                ).fetchone()
                if row is not None:
                    return row
        except sqlite3.OperationalError:
            return None
    return None


def _save_wikimedia_cached_record(
    lookup_keys: list[str],
    lookup_value: str,
    success: bool,
    metadata: dict[str, Any] | None,
    detail: str,
) -> None:
    db_path = _wikipedia_cache_db_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    normalized_keys = [key for key in dict.fromkeys(lookup_keys) if key]
    if not normalized_keys:
        return

    with WIKIPEDIA_CACHE_LOCK:
        with sqlite3.connect(db_path) as connection:
            connection.execute(
                f"""
                CREATE TABLE IF NOT EXISTS {WIKIMEDIA_CACHE_TABLE} (
                    lookup_key TEXT PRIMARY KEY,
                    lookup_value TEXT NOT NULL,
                    success INTEGER NOT NULL,
                    title TEXT NOT NULL,
                    alternate_titles TEXT NOT NULL,
                    canonical_url TEXT NOT NULL,
                    wikipedia_url TEXT NOT NULL,
                    wikidata_url TEXT NOT NULL,
                    wikidata_id TEXT NOT NULL,
                    entity_type TEXT NOT NULL,
                    redirect_source TEXT NOT NULL,
                    detail TEXT NOT NULL,
                    checked_at INTEGER NOT NULL
                )
                """
            )
            payload = (
                lookup_value,
                1 if success else 0,
                _normalize_value((metadata or {}).get("title"), ignore_case=False),
                json.dumps(_reference_titles_from_metadata(metadata or {})),
                _normalize_value((metadata or {}).get("url"), ignore_case=False),
                _normalize_value((metadata or {}).get("wikipedia_url"), ignore_case=False),
                _normalize_value((metadata or {}).get("wikidata_url"), ignore_case=False),
                _normalize_value((metadata or {}).get("wikidata_id"), ignore_case=False),
                _normalize_value((metadata or {}).get("type"), ignore_case=False),
                _normalize_value((metadata or {}).get("redirect_source"), ignore_case=False),
                _normalize_value(detail, ignore_case=False),
                int(time.time()),
            )
            for lookup_key in normalized_keys:
                connection.execute(
                    f"""
                    INSERT OR REPLACE INTO {WIKIMEDIA_CACHE_TABLE} (
                        lookup_key, lookup_value, success, title, alternate_titles, canonical_url,
                        wikipedia_url, wikidata_url, wikidata_id, entity_type, redirect_source, detail, checked_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (lookup_key, *payload),
                )
            connection.commit()


def _wikimedia_cached_row_to_result(row: sqlite3.Row) -> tuple[bool, dict[str, Any] | None, str]:
    success = bool(row["success"])
    if not success:
        return False, None, _normalize_value(row["detail"], ignore_case=False)
    try:
        alternate_titles = json.loads(row["alternate_titles"] or "[]")
    except json.JSONDecodeError:
        alternate_titles = []
    return (
        True,
        {
            "title": _normalize_value(row["title"], ignore_case=False),
            "alternate_titles": [item for item in alternate_titles if _normalize_value(item, ignore_case=False)],
            "url": _normalize_value(row["canonical_url"], ignore_case=False),
            "wikipedia_url": _normalize_value(row["wikipedia_url"], ignore_case=False),
            "wikidata_url": _normalize_value(row["wikidata_url"], ignore_case=False),
            "wikidata_id": _normalize_value(row["wikidata_id"], ignore_case=False),
            "type": _normalize_value(row["entity_type"], ignore_case=False),
            "redirect_source": _normalize_value(row["redirect_source"], ignore_case=False),
        },
        "",
    )


def _wikimedia_cache_keys_from_metadata(platform: str, raw_value: str, metadata: dict[str, Any] | None) -> list[str]:
    keys = {_wikimedia_cache_key(platform, raw_value)}
    wikipedia_title = _normalize_value((metadata or {}).get("wikipedia_title"), ignore_case=False)
    if wikipedia_title:
        keys.add(_wikimedia_cache_key("wikipedia", wikipedia_title))
    wikidata_id = _normalize_value((metadata or {}).get("wikidata_id"), ignore_case=False)
    if wikidata_id:
        keys.add(_wikimedia_cache_key("wikidata", wikidata_id))
    return [key for key in keys if key]


def _wikimedia_cache_key(platform: str, raw_value: str) -> str:
    cleaned = _normalize_value(raw_value, ignore_case=False)
    if platform == "wikipedia":
        cleaned = cleaned.replace(" ", "_").strip("/")
    if platform == "wikidata":
        cleaned = cleaned.upper()
    return f"{platform}:{cleaned.casefold()}"


def _build_wikidata_metadata(
    wikidata_id: str,
    entity: dict[str, Any],
    normalized_url: str,
    instance_of_ids: set[str],
) -> dict[str, Any]:
    wikipedia_title = _normalize_value(((entity.get("sitelinks") or {}).get("enwiki") or {}).get("title"), ignore_case=False)
    labels = _extract_wikidata_text_values(entity.get("labels"))
    aliases = _extract_wikidata_alias_values(entity.get("aliases"))
    title = wikipedia_title or (labels[0] if labels else wikidata_id)
    alternate_titles = _deduplicate_preserving_order([*labels, *aliases, wikipedia_title])
    wikipedia_url = ""
    if wikipedia_title:
        wikipedia_url = f"https://en.wikipedia.org/wiki/{quote(wikipedia_title.replace(' ', '_'))}"

    return {
        "title": title,
        "alternate_titles": alternate_titles,
        "type": _normalize_wikidata_entity_type(instance_of_ids),
        "wikidata_id": wikidata_id,
        "url": normalized_url or f"https://www.wikidata.org/wiki/{wikidata_id}",
        "wikidata_url": normalized_url or f"https://www.wikidata.org/wiki/{wikidata_id}",
        "wikipedia_url": wikipedia_url,
        "wikipedia_title": wikipedia_title,
        "redirect_source": "",
    }


def _merge_wikipedia_and_wikidata_metadata(
    canonical_title: str,
    canonical_url: str,
    redirect_source: str,
    wikidata_metadata: dict[str, Any],
) -> dict[str, Any]:
    alternate_titles = _deduplicate_preserving_order(
        [
            canonical_title,
            redirect_source,
            wikidata_metadata.get("title"),
            *(wikidata_metadata.get("alternate_titles") or []),
        ]
    )
    return {
        **wikidata_metadata,
        "title": canonical_title or _normalize_value(wikidata_metadata.get("title"), ignore_case=False),
        "alternate_titles": alternate_titles,
        "url": canonical_url or _normalize_value(wikidata_metadata.get("url"), ignore_case=False),
        "wikipedia_url": canonical_url or _normalize_value(wikidata_metadata.get("wikipedia_url"), ignore_case=False),
        "wikipedia_title": canonical_title,
        "redirect_source": redirect_source,
    }

def _is_wikipedia_cache_fresh(checked_at: int | None) -> bool:
    if not checked_at:
        return False
    refresh_age_hours = max(int(settings.wikipedia_refresh_hours or 24), 1)
    age_seconds = time.time() - int(checked_at)
    return age_seconds < refresh_age_hours * 3600


def _wikipedia_cache_db_path() -> Path:
    configured = _normalize_value(settings.wikipedia_cache_dir, ignore_case=False)
    if configured:
        return Path(configured) / WIKIPEDIA_CACHE_FILENAME
    return BASE_DIR / "data" / "wikipedia_cache" / WIKIPEDIA_CACHE_FILENAME


def _lookup_imdb_dataset_record(imdb_id: str) -> tuple[bool, dict[str, Any] | None, str]:
    try:
        db_path = _ensure_imdb_dataset_index()
    except Exception as exc:
        return False, None, f"IMDb dataset lookup failed: {exc}"

    with sqlite3.connect(db_path) as connection:
        connection.row_factory = sqlite3.Row
        if imdb_id.startswith("tt"):
            row = connection.execute(
                """
                SELECT tconst, primary_title, original_title, title_type
                FROM title_basics
                WHERE tconst = ?
                """,
                (imdb_id,),
            ).fetchone()
            if row is None:
                return False, None, "IMDb id not found in official dataset"

            primary_title = _normalize_value(row["primary_title"], ignore_case=False)
            original_title = _normalize_value(row["original_title"], ignore_case=False)
            title = primary_title or original_title
            alternate_titles = [candidate for candidate in [original_title, primary_title] if candidate and candidate != title]
            return (
                True,
                {
                    "title": title,
                    "alternate_titles": alternate_titles,
                    "type": _normalize_imdb_dataset_title_type(row["title_type"]),
                    "id": row["tconst"],
                    "url": f"https://www.imdb.com/title/{row['tconst']}/",
                    "source": "imdb-datasets",
                },
                "",
            )

        row = connection.execute(
            """
            SELECT nconst, primary_name
            FROM name_basics
            WHERE nconst = ?
            """,
            (imdb_id,),
        ).fetchone()
        if row is None:
            return False, None, "IMDb id not found in official dataset"

        return (
            True,
            {
                "title": _normalize_value(row["primary_name"], ignore_case=False),
                "type": "person",
                "id": row["nconst"],
                "url": f"https://www.imdb.com/name/{row['nconst']}/",
                "source": "imdb-datasets",
            },
            "",
        )


def _ensure_imdb_dataset_index() -> Path:
    dataset_dir = _imdb_dataset_dir()
    db_path = dataset_dir / IMDB_INDEX_FILENAME

    if db_path.exists() and not settings.imdb_rebuild_stale_index:
        return db_path

    dataset_dir.mkdir(parents=True, exist_ok=True)

    title_path = _ensure_imdb_dataset_file(
        settings.imdb_title_basics_url,
        dataset_dir / IMDB_TITLE_BASICS_FILENAME,
    )
    name_path = _ensure_imdb_dataset_file(
        settings.imdb_name_basics_url,
        dataset_dir / IMDB_NAME_BASICS_FILENAME,
    )

    with IMDB_DATASET_INDEX_LOCK:
        if _imdb_index_is_current(db_path, [title_path, name_path]):
            return db_path
        if db_path.exists() and not settings.imdb_rebuild_stale_index:
            return db_path
        _build_imdb_dataset_index(db_path, title_path=title_path, name_path=name_path)
        return db_path


def _imdb_dataset_dir() -> Path:
    configured = _normalize_value(settings.imdb_dataset_dir, ignore_case=False)
    if configured:
        return Path(configured)
    return BASE_DIR / "data" / "imdb_datasets"


def _ensure_imdb_dataset_file(source: str, destination: Path) -> Path:
    source_value = _normalize_value(source, ignore_case=False)
    if not source_value:
        raise WorkbookValidationConfigError("IMDb dataset source is not configured.")

    if _looks_like_url(source_value):
        if destination.exists():
            refresh_age_hours = max(int(settings.imdb_dataset_refresh_hours or 24), 1)
            age_seconds = time.time() - destination.stat().st_mtime
            if age_seconds < refresh_age_hours * 3600:
                return destination
        try:
            _download_imdb_dataset_file(source_value, destination)
        except Exception:
            if destination.exists():
                return destination
            raise
        return destination

    source_path = Path(source_value)
    if not source_path.exists():
        raise WorkbookValidationConfigError(f"IMDb dataset file was not found: {source_path}")
    if not destination.exists() or source_path.stat().st_mtime > destination.stat().st_mtime or source_path.stat().st_size != destination.stat().st_size:
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, destination)
    return destination


def _download_imdb_dataset_file(url: str, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = destination.with_suffix(destination.suffix + ".tmp")
    timeout = httpx.Timeout(connect=10.0, read=60.0, write=60.0, pool=10.0)
    with httpx.Client(timeout=timeout, follow_redirects=True) as client:
        with client.stream("GET", url) as response:
            response.raise_for_status()
            with temporary_path.open("wb") as file_handle:
                for chunk in response.iter_bytes():
                    if chunk:
                        file_handle.write(chunk)
    temporary_path.replace(destination)


def _imdb_index_is_current(db_path: Path, source_paths: list[Path]) -> bool:
    if not db_path.exists():
        return False
    db_mtime = db_path.stat().st_mtime
    return all(source_path.exists() and source_path.stat().st_mtime <= db_mtime for source_path in source_paths)


def _build_imdb_dataset_index(db_path: Path, title_path: Path, name_path: Path) -> None:
    temporary_db = db_path.with_suffix(db_path.suffix + ".tmp")
    if temporary_db.exists():
        temporary_db.unlink()

    connection = sqlite3.connect(temporary_db)
    try:
        connection.execute("PRAGMA journal_mode = OFF")
        connection.execute("PRAGMA synchronous = OFF")
        connection.execute("PRAGMA temp_store = MEMORY")
        connection.execute(
            """
            CREATE TABLE title_basics (
                tconst TEXT PRIMARY KEY,
                primary_title TEXT NOT NULL,
                original_title TEXT NOT NULL,
                title_type TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE name_basics (
                nconst TEXT PRIMARY KEY,
                primary_name TEXT NOT NULL
            )
            """
        )
        _load_title_basics_index(connection, title_path)
        _load_name_basics_index(connection, name_path)
        connection.execute("CREATE INDEX idx_title_basics_type ON title_basics(title_type)")
        connection.commit()
    finally:
        connection.close()

    temporary_db.replace(db_path)


def _load_title_basics_index(connection: sqlite3.Connection, title_path: Path) -> None:
    with gzip.open(title_path, "rt", encoding="utf-8", newline="") as file_handle:
        reader = csv.DictReader(file_handle, delimiter="\t")
        rows: list[tuple[str, str, str, str]] = []
        for row in reader:
            rows.append(
                (
                    _clean_imdb_dataset_value(row.get("tconst")),
                    _clean_imdb_dataset_value(row.get("primaryTitle")),
                    _clean_imdb_dataset_value(row.get("originalTitle")),
                    _clean_imdb_dataset_value(row.get("titleType")),
                )
            )
            if len(rows) >= 10000:
                connection.executemany(
                    "INSERT OR REPLACE INTO title_basics (tconst, primary_title, original_title, title_type) VALUES (?, ?, ?, ?)",
                    rows,
                )
                rows.clear()
        if rows:
            connection.executemany(
                "INSERT OR REPLACE INTO title_basics (tconst, primary_title, original_title, title_type) VALUES (?, ?, ?, ?)",
                rows,
            )


def _load_name_basics_index(connection: sqlite3.Connection, name_path: Path) -> None:
    with gzip.open(name_path, "rt", encoding="utf-8", newline="") as file_handle:
        reader = csv.DictReader(file_handle, delimiter="\t")
        rows: list[tuple[str, str]] = []
        for row in reader:
            rows.append(
                (
                    _clean_imdb_dataset_value(row.get("nconst")),
                    _clean_imdb_dataset_value(row.get("primaryName")),
                )
            )
            if len(rows) >= 10000:
                connection.executemany(
                    "INSERT OR REPLACE INTO name_basics (nconst, primary_name) VALUES (?, ?)",
                    rows,
                )
                rows.clear()
        if rows:
            connection.executemany(
                "INSERT OR REPLACE INTO name_basics (nconst, primary_name) VALUES (?, ?)",
                rows,
            )


def _clean_imdb_dataset_value(value: str | None) -> str:
    cleaned = _normalize_value(value, ignore_case=False)
    return "" if cleaned == r"\N" else cleaned


def _normalize_imdb_dataset_title_type(value: str | None) -> str:
    normalized = _normalize_value(value)
    if normalized in {"movie", "short", "video", "tvmovie"}:
        return "movie" if normalized == "movie" else normalized
    if normalized in {"tvseries", "tvminiseries"}:
        return "series"
    if normalized == "tvepisode":
        return "episode"
    return normalized


def _lookup_imdb_record(raw_value: str, client: httpx.Client | None) -> tuple[bool, dict[str, Any] | None, str]:
    imdb_id = _extract_imdb_identifier(raw_value)
    if not imdb_id:
        return False, None, "IMDb id could not be normalized"

    dataset_success, dataset_metadata, dataset_detail = _lookup_imdb_dataset_record(imdb_id)
    if dataset_success:
        return dataset_success, dataset_metadata, dataset_detail

    fallback_details: list[str] = []
    if dataset_detail:
        fallback_details.append(dataset_detail)

    if imdb_id.startswith("tt") and settings.omdb_api_key:
        omdb_success, omdb_metadata, omdb_detail = _lookup_omdb_title_record(imdb_id, client)
        if omdb_success:
            return omdb_success, omdb_metadata, omdb_detail
        if omdb_detail:
            fallback_details.append(omdb_detail)

    normalized_url = _normalize_social_reference("imdb", raw_value)
    if not normalized_url:
        return False, None, "IMDb URL could not be normalized"

    try:
        response = _network_get(normalized_url, client=client)
    except httpx.HTTPStatusError as exc:
        return False, None, f"IMDb lookup failed with HTTP {exc.response.status_code}"
    except httpx.HTTPError as exc:
        return False, None, f"IMDb lookup failed: {exc.__class__.__name__}"

    title = _extract_imdb_html_title(response.text)
    if not title:
        return False, None, "IMDb page title could not be read"

    record_type = "person" if imdb_id.startswith("nm") else ""
    return True, {"title": title, "type": record_type, "id": imdb_id, "url": str(response.url)}, ""


def _lookup_omdb_title_record(imdb_id: str, client: httpx.Client | None) -> tuple[bool, dict[str, Any] | None, str]:
    try:
        response = _network_get(
            "https://www.omdbapi.com/",
            client=client,
            params={"apikey": settings.omdb_api_key, "i": imdb_id},
        )
    except httpx.HTTPError as exc:
        return False, None, f"OMDb lookup failed: {exc.__class__.__name__}"

    payload = response.json()
    if payload.get("Response") == "False":
        return False, None, payload.get("Error") or "OMDb did not find the IMDb id"

    return (
        True,
        {
            "title": payload.get("Title") or "",
            "type": payload.get("Type") or "",
            "id": payload.get("imdbID") or imdb_id,
            "url": f"https://www.imdb.com/title/{payload.get('imdbID') or imdb_id}/",
        },
        "",
    )


def _network_get(
    url: str,
    client: httpx.Client | None = None,
    params: dict[str, Any] | None = None,
    headers: dict[str, str] | None = None,
) -> httpx.Response:
    if client is None:
        with _build_social_http_client() as http_client:
            response = http_client.get(url, params=params, headers=headers)
    else:
        response = client.get(url, params=params, headers=headers)
    response.raise_for_status()
    return response


def _wikipedia_slug_from_reference(raw_value: str) -> str | None:
    normalized_url = _normalize_social_reference("wikipedia", raw_value)
    if not normalized_url:
        return None
    parsed = urlparse(normalized_url)
    path = parsed.path
    if not path.startswith("/wiki/"):
        return None
    slug = unquote(path[len("/wiki/") :]).strip()
    return slug or None


def _extract_wikidata_identifier(raw_value: str) -> str | None:
    cleaned = raw_value.strip()
    if re.fullmatch(r"Q\d+", cleaned, flags=re.IGNORECASE):
        return cleaned.upper()
    if not _looks_like_url(cleaned):
        return None
    parsed = urlparse(cleaned)
    match = re.search(r"/wiki/(Q\d+)", parsed.path, flags=re.IGNORECASE)
    if not match:
        return None
    return match.group(1).upper()


def _extract_imdb_identifier(raw_value: str) -> str | None:
    cleaned = raw_value.strip()
    if re.fullmatch(r"(tt|nm)\d{7,}", cleaned):
        return cleaned
    if not _looks_like_url(cleaned):
        return None
    parsed = urlparse(cleaned)
    match = re.search(r"/(title|name)/((tt|nm)\d{7,})", parsed.path, flags=re.IGNORECASE)
    if not match:
        return None
    return match.group(2)


def _extract_imdb_html_title(html: str) -> str:
    for block in re.findall(r'<script type="application/ld\+json">\s*(.*?)\s*</script>', html, flags=re.DOTALL):
        try:
            payload = json.loads(block)
        except json.JSONDecodeError:
            continue
        if isinstance(payload, dict) and payload.get("name"):
            return str(payload["name"]).strip()

    match = re.search(r"<title>(.*?)</title>", html, flags=re.IGNORECASE | re.DOTALL)
    if not match:
        return ""
    title = unescape(match.group(1)).strip()
    title = re.sub(r"\s*-\s*IMDb.*$", "", title, flags=re.IGNORECASE)
    return title.strip()


def _extract_wikipedia_html_title(html: str) -> str:
    match = re.search(r"<title>(.*?)</title>", html, flags=re.IGNORECASE | re.DOTALL)
    if not match:
        return ""
    title = unescape(match.group(1)).strip()
    title = re.sub(r"\s*-\s*Wikipedia.*$", "", title, flags=re.IGNORECASE)
    return title.strip()


def _titles_loosely_match(expected: str, actual: str) -> bool:
    normalized_expected = _normalize_lookup_title(expected)
    normalized_actual = _normalize_lookup_title(actual)
    if not normalized_expected or not normalized_actual:
        return False
    if normalized_expected == normalized_actual:
        return True
    return normalized_expected in normalized_actual or normalized_actual in normalized_expected


def _normalize_lookup_title(value: str) -> str:
    text = value.strip()
    text = re.sub(r"\s+-\s+dar$", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\([^)]*\)", "", text)
    text = text.replace("&", " and ")
    text = re.sub(r"[^A-Za-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip().casefold()
    return text


def _extract_wikidata_text_values(payload: Any) -> list[str]:
    if not isinstance(payload, dict):
        return []
    values: list[str] = []
    for item in payload.values():
        value = _normalize_value((item or {}).get("value"), ignore_case=False)
        if value and value not in values:
            values.append(value)
    return values


def _extract_wikidata_alias_values(payload: Any) -> list[str]:
    if not isinstance(payload, dict):
        return []
    values: list[str] = []
    for language_entries in payload.values():
        if not isinstance(language_entries, list):
            continue
        for item in language_entries:
            value = _normalize_value((item or {}).get("value"), ignore_case=False)
            if value and value not in values:
                values.append(value)
    return values


def _extract_wikidata_instance_of_ids(entity: dict[str, Any]) -> set[str]:
    claims = (entity.get("claims") or {}).get("P31") or []
    values: set[str] = set()
    for claim in claims:
        claim_value = (((claim or {}).get("mainsnak") or {}).get("datavalue") or {}).get("value") or {}
        entity_id = _normalize_value(claim_value.get("id"), ignore_case=False).upper()
        if entity_id:
            values.add(entity_id)
    return values


def _is_wikipedia_disambiguation_page(page: dict[str, Any]) -> bool:
    pageprops = page.get("pageprops") or {}
    if "disambiguation" in pageprops:
        return True
    title = _normalize_value(page.get("title"), ignore_case=False)
    return bool(title and title.casefold().endswith("(disambiguation)"))


def _is_wikidata_disambiguation(entity: dict[str, Any], instance_of_ids: set[str]) -> bool:
    if instance_of_ids & WIKIDATA_DISAMBIGUATION_IDS:
        return True
    enwiki_title = _normalize_value(((entity.get("sitelinks") or {}).get("enwiki") or {}).get("title"), ignore_case=False)
    return bool(enwiki_title and enwiki_title.casefold().endswith("(disambiguation)"))


def _normalize_wikidata_entity_type(instance_of_ids: set[str]) -> str:
    for normalized_type, qids in WIKIDATA_TYPE_MAP.items():
        if instance_of_ids & qids:
            return normalized_type
    return ""


def _deduplicate_preserving_order(values: list[Any]) -> list[str]:
    deduplicated: list[str] = []
    for value in values:
        normalized = _normalize_value(value, ignore_case=False)
        if normalized and normalized not in deduplicated:
            deduplicated.append(normalized)
    return deduplicated


def _fetch_social_reference(url: str, platform: str, client: httpx.Client | None = None, title: str = "") -> tuple[bool, str]:
    try:
        if client is None:
            with _build_social_http_client() as http_client:
                response = http_client.get(url)
        else:
            response = client.get(url)
    except httpx.InvalidURL as exc:
        return False, str(exc)
    except httpx.HTTPError as exc:
        return False, str(exc)

    if response.status_code >= 400:
        return False, f"http {response.status_code}"

    final_url = str(response.url)
    passed, detail = _validate_final_social_url(platform, url, final_url)
    if not passed:
        return passed, detail

    content_issue = _detect_social_page_quality_issue(platform, response.text, url, title)
    if content_issue:
        return False, content_issue

    return True, ""


def _validate_final_social_url(platform: str, requested_url: str, final_url: str) -> tuple[bool, str]:
    parsed = urlparse(final_url)
    host = parsed.netloc.casefold()
    path = parsed.path.casefold()
    requested = urlparse(requested_url)
    requested_path = requested.path.casefold()

    if platform == "facebook":
        if "facebook.com" not in host:
            return False, "redirected away from facebook.com"
        return True, ""

    if platform == "twitter":
        if "x.com" not in host and "twitter.com" not in host:
            return False, "redirected away from X/Twitter"
        return True, ""

    if platform == "instagram":
        if "instagram.com" not in host:
            return False, "redirected away from instagram.com"
        return True, ""

    if platform == "youtube":
        if "youtube.com" not in host and "youtu.be" not in host:
            return False, "redirected away from YouTube"
        return True, ""

    if platform == "tiktok":
        if "tiktok.com" not in host:
            return False, "redirected away from tiktok.com"
        return True, ""

    if platform == "wikipedia":
        if "wikipedia.org" not in host:
            return False, "redirected away from wikipedia.org"
        if "/wiki/" not in path or "special:search" in path:
            return False, "did not resolve to a Wikipedia article"
        return True, ""

    if platform == "wikidata":
        if "wikidata.org" not in host:
            return False, "redirected away from wikidata.org"
        if not re.match(r"^/wiki/q\d+/?$", path):
            return False, "did not resolve to a Wikidata item"
        return True, ""

    if platform == "imdb":
        if "imdb.com" not in host:
            return False, "redirected away from imdb.com"
        if not (path.startswith("/title/") or path.startswith("/name/")):
            return False, "did not resolve to an IMDb title or name page"
        if requested_path and requested_path not in path:
            return False, "resolved to a different IMDb page"
        return True, ""

    return False, f"unsupported platform {platform}"


def _looks_like_url(value: str) -> bool:
    return value.startswith("http://") or value.startswith("https://")


def _contains_control_characters(value: str) -> bool:
    allowed_control_characters = {"\n", "\r", "\t"}
    return any((ord(char) < 32 and char not in allowed_control_characters) or ord(char) == 127 for char in value)


def _validate_movie_us_release_date(
    row_context: dict[str, Any],
    released_on_value: Any,
    cache: dict[str, tuple[bool, dict[str, Any] | None, str]],
) -> tuple[bool, str]:
    actual_date = _as_date(released_on_value)
    if actual_date is None:
        return False, "released_on is blank or not a valid date"

    success, metadata, detail = _get_movie_metadata(row_context, cache)
    if not success:
        return False, detail

    expected_date = _normalize_tmdb_date((metadata or {}).get("release_date")) or _normalize_tmdb_date(
        (metadata or {}).get("us_release_date")
    )
    if not expected_date:
        return False, "TMDB did not return a release date"

    if actual_date.isoformat() != expected_date:
        return False, f"{_movie_release_date_label(metadata)}: {expected_date}"

    return True, ""


def _validate_movie_release_type(
    row_context: dict[str, Any],
    release_type_value: Any,
    cache: dict[str, tuple[bool, dict[str, Any] | None, str]],
) -> tuple[bool, str]:
    success, metadata, detail = _get_movie_metadata(row_context, cache)
    if not success:
        return False, detail

    expected_release_type = _normalize_value((metadata or {}).get("release_type"))
    if not expected_release_type:
        return False, "TMDB did not return a release type recommendation"

    actual_release_type = _normalize_value(release_type_value)
    if actual_release_type != expected_release_type:
        return False, f"TMDB recommends {expected_release_type.title()}"

    return True, ""


def _validate_movie_genre(
    row_context: dict[str, Any],
    genre_value: Any,
    cache: dict[str, tuple[bool, dict[str, Any] | None, str]],
) -> tuple[bool, str]:
    success, metadata, detail = _get_movie_metadata(row_context, cache)
    if not success:
        return False, detail

    expected_genres = (metadata or {}).get("genres", [])
    if not expected_genres:
        return False, "TMDB did not return any genres"

    actual_genres = {_normalize_genre_name(item) for item in _split_multi_value(genre_value)}
    recommended_genres = {_normalize_genre_name(item) for item in expected_genres}
    if actual_genres and actual_genres == recommended_genres:
        return True, ""

    recommended_text = ", ".join(expected_genres)
    return False, f"TMDB genres: {recommended_text}"


def _get_movie_metadata(
    row_context: dict[str, Any],
    cache: dict[str, tuple[bool, dict[str, Any] | None, str]],
) -> tuple[bool, dict[str, Any] | None, str]:
    title = _normalize_movie_lookup_title(row_context.get("title"))
    if not title:
        return False, None, "title is missing, so the movie metadata cannot be checked"

    release_year = _movie_lookup_release_year(row_context)
    lookup_key = f"{title.casefold()}|{release_year or ''}"
    if lookup_key not in cache:
        lookup_signature = inspect.signature(_lookup_movie_metadata)
        if "release_year" in lookup_signature.parameters:
            cache[lookup_key] = _lookup_movie_metadata(title, release_year=release_year)
        else:
            cache[lookup_key] = _lookup_movie_metadata(title)

    return cache[lookup_key]


def _lookup_movie_us_release_date(title: str) -> tuple[bool, str | None, str]:
    success, metadata, detail = _lookup_movie_metadata(title)
    if not success:
        return False, None, detail

    expected_date = _normalize_tmdb_date((metadata or {}).get("release_date")) or _normalize_tmdb_date(
        (metadata or {}).get("us_release_date")
    )
    if expected_date:
        return True, expected_date, ""
    return False, None, "TMDB did not return a release date"


def _lookup_movie_metadata(title: str, release_year: int | None = None) -> tuple[bool, dict[str, Any] | None, str]:
    if not (settings.tmdb_api_key or settings.tmdb_read_access_token):
        return False, None, "TMDB is not configured"

    try:
        with _build_tmdb_http_client() as client:
            selected_movie = _search_tmdb_movie(client, title, release_year=release_year)
            if not selected_movie:
                return False, None, "movie not found in TMDB"

            movie_id = selected_movie.get("id")
            if not movie_id:
                return False, None, "movie id missing in TMDB search results"

            details_payload = _tmdb_get(client, f"/movie/{movie_id}")
            release_payload = _tmdb_get(client, f"/movie/{movie_id}/release_dates")
    except httpx.HTTPError as exc:
        return False, None, f"TMDB lookup failed: {exc.__class__.__name__}"

    us_entries: list[dict[str, Any]] = []
    for result in release_payload.get("results", []):
        if result.get("iso_3166_1") == "US":
            us_entries.extend(result.get("release_dates", []))

    us_release_date = _preferred_us_release_date(us_entries)
    fallback_release_date = _normalize_tmdb_date(details_payload.get("release_date"))
    genres = [genre.get("name") for genre in details_payload.get("genres", []) if genre.get("name")]
    metadata = {
        "us_release_date": us_release_date,
        "release_date": us_release_date or fallback_release_date,
        "release_date_source": "US" if us_release_date else ("global" if fallback_release_date else ""),
        "release_type": _infer_movie_release_type(us_entries),
        "genres": genres,
    }
    return True, metadata, ""


def _movie_lookup_release_year(row_context: dict[str, Any]) -> int | None:
    released_on = _as_date(row_context.get("released_on"))
    if released_on is not None:
        return released_on.year
    return None


def _movie_release_date_label(metadata: dict[str, Any] | None) -> str:
    source = _normalize_value((metadata or {}).get("release_date_source"))
    if source == "global":
        return "TMDB release date"
    return "TMDB USA release date"


def _search_tmdb_movie(client: httpx.Client, title: str, release_year: int | None = None) -> dict[str, Any] | None:
    fallback_result: dict[str, Any] | None = None
    attempted: set[tuple[str, int | None]] = set()

    search_years: list[int | None] = [release_year] if release_year is not None else [None]
    if release_year is not None:
        search_years.append(None)

    for year in search_years:
        for query in _movie_lookup_queries(title):
            query_key = (query.casefold(), year)
            if query_key in attempted:
                continue
            attempted.add(query_key)

            params: dict[str, Any] = {"query": query}
            if year is not None:
                params["primary_release_year"] = year

            search_payload = _tmdb_get(client, "/search/movie", params)
            results = search_payload.get("results", [])
            if not results:
                continue

            fallback_result = fallback_result or results[0]
            matched = _select_tmdb_movie_search_result(title, results, release_year=release_year)
            if matched is not None:
                return matched

    return fallback_result


def _movie_lookup_queries(title: str) -> list[str]:
    queries: list[str] = []

    def add(candidate: str) -> None:
        cleaned = re.sub(r"\s+", " ", candidate).strip()
        if cleaned and cleaned not in queries:
            queries.append(cleaned)

    add(title)
    add(re.sub(r"\s+\([^)]{1,20}\)$", "", title))
    add(title.replace("&", " and "))
    add(title.replace(" and ", " & "))
    return queries


def _select_tmdb_movie_search_result(
    query_title: str,
    results: list[dict[str, Any]],
    release_year: int | None = None,
) -> dict[str, Any] | None:
    matched_results: list[dict[str, Any]] = []
    for item in results:
        candidate_titles = [item.get("title"), item.get("original_title")]
        if any(_titles_loosely_match(query_title, str(candidate)) for candidate in candidate_titles if candidate):
            matched_results.append(item)

    if not matched_results:
        return None

    if release_year is not None:
        year_matches = [item for item in matched_results if _tmdb_result_release_year(item) == release_year]
        if year_matches:
            matched_results = year_matches

    matched_results.sort(
        key=lambda item: (
            0 if _tmdb_result_exact_title_match(query_title, item) else 1,
            -float(item.get("popularity") or 0),
        )
    )
    return matched_results[0]


def _tmdb_result_exact_title_match(query_title: str, result: dict[str, Any]) -> bool:
    normalized_query = _normalize_lookup_title(query_title)
    for candidate in (result.get("title"), result.get("original_title")):
        if candidate and _normalize_lookup_title(str(candidate)) == normalized_query:
            return True
    return False


def _tmdb_result_release_year(result: dict[str, Any]) -> int | None:
    normalized_date = _normalize_tmdb_date(result.get("release_date"))
    if not normalized_date:
        return None
    try:
        return int(normalized_date[:4])
    except ValueError:
        return None


def _normalize_tmdb_date(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    normalized = text[:10]
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", normalized):
        return None
    return normalized


def _normalize_genre_name(value: Any) -> str:
    normalized = _normalize_value(value)
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized).strip()
    aliases = {
        "scifi": "science fiction",
        "sci fi": "science fiction",
        "science fiction": "science fiction",
        "tv movie": "tv movie",
        "television movie": "tv movie",
    }
    return aliases.get(normalized, normalized)


def _preferred_us_release_date(entries: list[dict[str, Any]]) -> str | None:
    if not entries:
        return None

    type_priority = [3, 2, 1, 4, 5, 6]
    parsed_entries: list[tuple[int, str]] = []
    for entry in entries:
        normalized = _normalize_tmdb_date(entry.get("release_date"))
        if not normalized:
            continue
        release_type = entry.get("type")
        priority = type_priority.index(release_type) if release_type in type_priority else len(type_priority)
        parsed_entries.append((priority, normalized))

    if not parsed_entries:
        return None

    parsed_entries.sort(key=lambda item: (item[0], item[1]))
    return parsed_entries[0][1]


def _infer_movie_release_type(entries: list[dict[str, Any]]) -> str:
    release_types = {entry.get("type") for entry in entries if entry.get("type")}
    if 3 in release_types:
        return "Wide"
    if 2 in release_types:
        return "Limited"
    if 4 in release_types:
        return "Digital"
    if 6 in release_types:
        return "TV"
    if 5 in release_types:
        return "Physical"
    if 1 in release_types:
        return "Premiere"
    return ""


def _split_multi_value(value: Any) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    return [item.strip() for item in re.split(r"[,;/|\n]+", text) if item.strip()]


def _load_csv_workbook(file_bytes: bytes, filename: str) -> Workbook:
    decoded = _decode_csv_bytes(file_bytes)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _safe_sheet_name(Path(filename).stem or "Sheet1")
    reader = csv.reader(io.StringIO(decoded))
    for row in reader:
        sheet.append(row)
    if sheet.max_row == 1 and sheet.max_column == 1 and sheet["A1"].value is None:
        sheet["A1"] = ""
    return workbook


def _decode_csv_bytes(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise WorkbookValidationConfigError("The CSV file could not be decoded. Save it as UTF-8 or Windows-1252 and try again.")


def _extract_google_sheet_id(sheet_reference: str) -> str:
    cleaned = sheet_reference.strip()
    if not cleaned:
        return ""
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", cleaned)
    if match:
        return match.group(1)
    if re.fullmatch(r"[a-zA-Z0-9-_]{20,}", cleaned):
        return cleaned
    return ""


def _safe_sheet_name(raw_name: str) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]", " ", raw_name).strip()
    return (cleaned or "Sheet1")[:31]


def _build_tmdb_http_client() -> httpx.Client:
    headers = {"Accept": "application/json"}
    if settings.tmdb_read_access_token:
        headers["Authorization"] = f"Bearer {settings.tmdb_read_access_token}"
    timeout = httpx.Timeout(connect=2.0, read=5.0, write=5.0, pool=2.0)
    return httpx.Client(timeout=timeout, headers=headers)


def _tmdb_get(client: httpx.Client, path: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
    query = dict(params or {})
    if settings.tmdb_api_key and "Authorization" not in client.headers:
        query["api_key"] = settings.tmdb_api_key
    response = client.get(f"https://api.themoviedb.org/3{path}", params=query)
    response.raise_for_status()
    return response.json()


def _detect_social_page_quality_issue(platform: str, page_text: str, url: str = "", title: str = "") -> str:
    if platform != "facebook":
        return ""

    normalized = re.sub(r"\s+", " ", page_text).casefold() if page_text else ""
    if not normalized:
        return "Unable to Verify (page is empty or geoblocked)"

    if re.search(r"\bfriends\b", normalized) and not re.search(r"\bfollowers?\b", normalized):
        return "page appears to show friends instead of followers (personal timeline)"

    classification, confidence, reason = _classify_facebook_page(url, title, page_text)
    if classification not in {"Official", "Official Regional"}:
        return f"Facebook page classified as {classification}: {reason}"

    return ""


def _social_http_client():
    return _ReusableSocialClient()


def _build_social_http_client() -> httpx.Client:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
        )
    }
    timeout = httpx.Timeout(connect=2.0, read=4.0, write=4.0, pool=2.0)
    limits = httpx.Limits(max_connections=10, max_keepalive_connections=5)
    return httpx.Client(follow_redirects=True, timeout=timeout, headers=headers, limits=limits)


class _ReusableSocialClient:
    def __enter__(self) -> httpx.Client:
        self.client = _build_social_http_client()
        return self.client

    def __exit__(self, exc_type, exc, tb) -> None:
        self.client.close()


def _rule_matches_review_mode(rule: ValidationRule, review_mode: str, platform_filter: str | None) -> bool:
    if not review_mode:
        return True
        
    modes = [m.strip() for m in review_mode.split(",") if m.strip()]
    if not modes or "full" in modes:
        return True
        
    col = (rule.column or "").strip().lower()
    
    target_modes = [m for m in modes if m in {"social_only", "categorization", "platform_specific"}]
    if not target_modes:
        return True
        
    matched = False
    for mode in target_modes:
        if mode == "social_only":
            social_cols = {
                "facebook_page", "twitter_handle", "instagram_user", "tiktok_user", 
                "youtube_channel_username", "wikipedia_url", "wikidata_id", "imdb_id", 
                "linkedin_page", "website", "official_website"
            }
            if col in social_cols:
                if platform_filter and platform_filter != "all":
                    platforms = [p.strip().lower() for p in platform_filter.split(",") if p.strip()]
                    if col == "facebook_page" and "facebook" in platforms:
                        matched = True
                    elif col == "instagram_user" and "instagram" in platforms:
                        matched = True
                    elif col == "twitter_handle" and "twitter" in platforms:
                        matched = True
                    elif col == "tiktok_user" and "tiktok" in platforms:
                        matched = True
                    elif col == "youtube_channel_username" and "youtube" in platforms:
                        matched = True
                    elif col == "wikipedia_url" and "wikipedia" in platforms:
                        matched = True
                    elif col == "imdb_id" and "imdb" in platforms:
                        matched = True
                    elif col == "linkedin_page" and "linkedin" in platforms:
                        matched = True
                    elif col in {"website", "official_website"} and "website" in platforms:
                        matched = True
                else:
                    matched = True
        elif mode == "categorization":
            cat_cols = {
                "title_category", "title_sub_category", "talent_type", "talent_subtype", 
                "genre", "primary_genre"
            }
            if col in cat_cols:
                matched = True
        elif mode == "platform_specific":
            if not platform_filter or platform_filter == "all":
                matched = True
            else:
                platforms = [p.strip().lower() for p in platform_filter.split(",") if p.strip()]
                for pf in platforms:
                    if pf == "facebook" and col == "facebook_page":
                        matched = True
                    elif pf == "instagram" and col == "instagram_user":
                        matched = True
                    elif pf == "twitter" and col == "twitter_handle":
                        matched = True
                    elif pf == "tiktok" and col == "tiktok_user":
                        matched = True
                    elif pf == "youtube" and col == "youtube_channel_username":
                        matched = True
                    elif pf == "wikipedia" and col == "wikipedia_url":
                        matched = True
                    elif pf == "imdb" and col == "imdb_id":
                        matched = True
                    elif pf == "linkedin" and col == "linkedin_page":
                        matched = True
                    elif pf == "website" and col in {"website", "official_website"}:
                        matched = True
    return matched



def _perform_duplicate_conflict_scan(worksheet_context, worksheet, issues):
    title_rows = {}
    handle_groups = {
        "facebook_page": {},
        "twitter_handle": {},
        "instagram_user": {},
        "youtube_channel_username": {},
        "tiktok_user": {},
        "wikidata_id": {},
        "imdb_id": {},
    }
    
    header_map = worksheet_context.header_map
    title_index = header_map.get("title")
    
    for row_number in worksheet_context.active_rows:
        row_ctx = _get_row_context(worksheet_context, row_number)
        title_val = row_ctx.get("title")
        if not _is_blank(title_val):
            norm_title = _normalize_value(title_val).strip().lower()
            title_rows.setdefault(norm_title, []).append((row_number, title_val))
            
        for handle_col, groups in handle_groups.items():
            col_idx = header_map.get(handle_col)
            if col_idx is not None:
                val = row_ctx.get(handle_col)
                if not _is_blank(val):
                    norm_val = _normalize_value(val).strip().lower()
                    groups.setdefault(norm_val, []).append((row_number, title_val or f"Row {row_number}", val))

    # Detect duplicate titles
    if title_index is not None:
        for norm_title, rows in title_rows.items():
            if len(rows) > 1:
                for row_num, orig_title in rows:
                    cell = worksheet.cell(row=row_num, column=title_index)
                    other_rows = ", ".join(str(r[0]) for r in rows if r[0] != row_num)
                    message = f"Duplicate entity: Title '{orig_title}' also found on row(s): {other_rows}."
                    issues.append(
                        _mark_issue(
                            cell=cell, 
                            rule=None, 
                            message=message, 
                            value=orig_title,
                            finding_category="Duplicate",
                            confidence="High",
                            confidence_reason="Identical entity name matches exactly on multiple rows in sheet."
                        )
                    )

    # Detect conflicting handles
    for handle_col, groups in handle_groups.items():
        col_idx = header_map.get(handle_col)
        for norm_val, entries in groups.items():
            unique_titles = {e[1].strip().lower(): e[1] for e in entries}
            if len(unique_titles) > 1:
                for row_num, title_val, orig_val in entries:
                    cell = worksheet.cell(row=row_num, column=col_idx)
                    others = ", ".join(f"row {e[0]} ('{e[1]}')" for e in entries if e[0] != row_num)
                    message = f"Conflicting handle: '{orig_val}' is shared with other entities: {others}."
                    issues.append(
                        _mark_issue(
                            cell=cell,
                            rule=None,
                            message=message,
                            value=orig_val,
                            finding_category="Duplicate",
                            confidence="High",
                            confidence_reason="Same social handle or external ID is assigned to multiple distinct title names."
                        )
                    )
