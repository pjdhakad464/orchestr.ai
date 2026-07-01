from __future__ import annotations

import openpyxl
from openpyxl.styles import PatternFill, Font
from pathlib import Path
from typing import Any
from pydantic import BaseModel, Field

class CellDiff(BaseModel):
    sheet: str
    row_num: int
    column: str
    value_a: str
    value_b: str
    diff_type: str  # added, removed, modified, matching

class RowDiff(BaseModel):
    row_key: str
    row_num_a: int | None = None
    row_num_b: int | None = None
    diffs: list[CellDiff] = Field(default_factory=list)
    status: str  # added, removed, modified, matching

class ComparisonReport(BaseModel):
    file_a: str
    file_b: str
    total_rows_a: int = 0
    total_rows_b: int = 0
    added_rows: int = 0
    removed_rows: int = 0
    modified_rows: int = 0
    unchanged_rows: int = 0
    column_diffs: list[str] = Field(default_factory=list)
    row_diffs: list[RowDiff] = Field(default_factory=list)

class ExcelComparator:
    def __init__(self) -> None:
        pass

    def compare(
        self,
        file_a_path: Path,
        file_b_path: Path,
        key_columns: list[str] | None = None,
        sheet_name: str | None = None
    ) -> ComparisonReport:
        """Compares two Excel workbooks sheet by sheet or on a specific sheet."""
        wb_a = openpyxl.load_workbook(file_a_path, data_only=True)
        wb_b = openpyxl.load_workbook(file_b_path, data_only=True)

        # Select sheet
        if sheet_name:
            sheets_a = [sheet_name] if sheet_name in wb_a.sheetnames else []
            sheets_b = [sheet_name] if sheet_name in wb_b.sheetnames else []
        else:
            sheets_a = wb_a.sheetnames
            sheets_b = wb_b.sheetnames

        common_sheets = [s for s in sheets_a if s in sheets_b]
        if not common_sheets:
            return ComparisonReport(
                file_a=file_a_path.name,
                file_b=file_b_path.name,
                column_diffs=["No matching sheets found between workbooks."]
            )

        # Let's focus on the first common sheet for row-level details
        sheet_to_compare = common_sheets[0]
        ws_a = wb_a[sheet_to_compare]
        ws_b = wb_b[sheet_to_compare]

        headers_a = [str(cell.value or "").strip() for cell in ws_a[1]]
        headers_b = [str(cell.value or "").strip() for cell in ws_b[1]]

        # Determine key column
        if not key_columns:
            # Fallback to auto-detection (Column B or first column)
            if len(headers_a) > 1:
                key_columns = [headers_a[1]] if headers_a[1] else [headers_a[0]]
            else:
                key_columns = [headers_a[0]]

        key_indices_a = [headers_a.index(k) for k in key_columns if k in headers_a]
        key_indices_b = [headers_b.index(k) for k in key_columns if k in headers_b]

        if not key_indices_a or not key_indices_b:
            return ComparisonReport(
                file_a=file_a_path.name,
                file_b=file_b_path.name,
                column_diffs=["Key columns not found in both workbooks."]
            )

        # Load row dicts
        rows_a = self._load_rows(ws_a, headers_a)
        rows_b = self._load_rows(ws_b, headers_b)

        report = ComparisonReport(
            file_a=file_a_path.name,
            file_b=file_b_path.name,
            total_rows_a=len(rows_a),
            total_rows_b=len(rows_b)
        )

        # Match rows based on key columns
        map_a = self._build_key_map(rows_a, key_indices_a, headers_a)
        map_b = self._build_key_map(rows_b, key_indices_b, headers_b)

        all_keys = set(map_a.keys()).union(map_b.keys())

        for key in all_keys:
            idx_a = map_a.get(key)
            idx_b = map_b.get(key)

            if idx_a is not None and idx_b is not None:
                # Row exists in both
                row_data_a = rows_a[idx_a]
                row_data_b = rows_b[idx_b]
                diffs = []
                
                # Check each common header
                common_headers = [h for h in headers_a if h in headers_b]
                for h in common_headers:
                    val_a = str(row_data_a.get(h) or "").strip()
                    val_b = str(row_data_b.get(h) or "").strip()
                    if val_a != val_b:
                        diffs.append(CellDiff(
                            sheet=sheet_to_compare,
                            row_num=idx_b + 2,
                            column=h,
                            value_a=val_a,
                            value_b=val_b,
                            diff_type="modified"
                        ))
                
                if diffs:
                    report.modified_rows += 1
                    report.row_diffs.append(RowDiff(
                        row_key=key,
                        row_num_a=idx_a + 2,
                        row_num_b=idx_b + 2,
                        diffs=diffs,
                        status="modified"
                    ))
                else:
                    report.unchanged_rows += 1
                    
            elif idx_a is not None:
                # Removed in B
                report.removed_rows += 1
                row_data_a = rows_a[idx_a]
                report.row_diffs.append(RowDiff(
                    row_key=key,
                    row_num_a=idx_a + 2,
                    status="removed"
                ))
            else:
                # Added in B
                report.added_rows += 1
                row_data_b = rows_b[idx_b]
                report.row_diffs.append(RowDiff(
                    row_key=key,
                    row_num_b=idx_b + 2,
                    status="added"
                ))

        wb_a.close()
        wb_b.close()
        return report

    def _load_rows(self, ws: openpyxl.worksheet.worksheet.Worksheet, headers: list[str]) -> list[dict[str, Any]]:
        rows = []
        for r_idx in range(2, ws.max_row + 1):
            row_vals = [ws.cell(row=r_idx, column=c_idx).value for c_idx in range(1, len(headers) + 1)]
            # Skip empty rows
            if not any(v is not None for v in row_vals):
                continue
            rows.append(dict(zip(headers, row_vals)))
        return rows

    def _build_key_map(self, rows: list[dict[str, Any]], key_indices: list[int], headers: list[str]) -> dict[str, int]:
        key_map = {}
        for idx, row in enumerate(rows):
            key_parts = []
            for k_idx in key_indices:
                col_name = headers[k_idx]
                key_parts.append(str(row.get(col_name) or "").strip().casefold())
            key_str = "||".join(key_parts)
            if key_str:
                key_map[key_str] = idx
        return key_map

    def export_to_xlsx(self, report: ComparisonReport, output_path: Path) -> None:
        """Exports a detailed comparison Excel sheet with color highlights."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Comparison Report"

        # Styles
        header_font = Font(name="Outfit", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        
        red_fill = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
        green_fill = PatternFill(start_color="D6FFD6", end_color="D6FFD6", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # Set headers
        headers = ["Row Key", "Status", "Row # (Original)", "Row # (New)", "Column", "Value (Original)", "Value (New)"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill

        r_idx = 2
        for rd in report.row_diffs:
            if rd.status == "added":
                ws.cell(row=r_idx, column=1, value=rd.row_key).fill = green_fill
                ws.cell(row=r_idx, column=2, value="Added").fill = green_fill
                ws.cell(row=r_idx, column=4, value=rd.row_num_b).fill = green_fill
                r_idx += 1
            elif rd.status == "removed":
                ws.cell(row=r_idx, column=1, value=rd.row_key).fill = red_fill
                ws.cell(row=r_idx, column=2, value="Removed").fill = red_fill
                ws.cell(row=r_idx, column=3, value=rd.row_num_a).fill = red_fill
                r_idx += 1
            elif rd.status == "modified":
                for diff in rd.diffs:
                    ws.cell(row=r_idx, column=1, value=rd.row_key).fill = yellow_fill
                    ws.cell(row=r_idx, column=2, value="Modified").fill = yellow_fill
                    ws.cell(row=r_idx, column=3, value=rd.row_num_a).fill = yellow_fill
                    ws.cell(row=r_idx, column=4, value=rd.row_num_b).fill = yellow_fill
                    ws.cell(row=r_idx, column=5, value=diff.column).fill = yellow_fill
                    ws.cell(row=r_idx, column=6, value=diff.value_a).fill = yellow_fill
                    ws.cell(row=r_idx, column=7, value=diff.value_b).fill = yellow_fill
                    r_idx += 1

        # Summary Tab
        ws_sum = wb.create_sheet("Summary Metrics")
        ws_sum.cell(row=1, column=1, value="Metric").font = header_font
        ws_sum.cell(row=1, column=1).fill = header_fill
        ws_sum.cell(row=1, column=2, value="Count").font = header_font
        ws_sum.cell(row=1, column=2).fill = header_fill

        metrics = [
            ("Original File", report.file_a),
            ("New File", report.file_b),
            ("Total Rows (Original)", report.total_rows_a),
            ("Total Rows (New)", report.total_rows_b),
            ("Added Rows", report.added_rows),
            ("Removed Rows", report.removed_rows),
            ("Modified Rows", report.modified_rows),
            ("Unchanged Rows", report.unchanged_rows)
        ]

        for idx, (m, val) in enumerate(metrics, 2):
            ws_sum.cell(row=idx, column=1, value=m)
            ws_sum.cell(row=idx, column=2, value=val)

        wb.save(output_path)
        wb.close()
