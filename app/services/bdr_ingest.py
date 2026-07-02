"""BDR (Brand Definition Report) Ingest Builder.

Applies a client change-list to an uploaded BDR workbook and returns an
ingest-ready workbook: every changed cell is colour-coded, carries a
"DataOps" comment, TST is rebuilt per the DAR rule, BDR hard rules are
enforced, and a Legend sheet is appended.

Scope: this is the deterministic core. It never fabricates data — metadata
that requires open-ended web research (IMDb nm id, Wikipedia URL, unprovided
social handles) is used only when supplied in the change-list; otherwise the
field is left blank with a NOT FOUND / NEEDS MANUAL REVIEW comment.
"""

from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass, field as _dcfield
from datetime import date
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ─── Colour convention (fill hex / font hex) ───────────────────────────────
FILL_ADDED = PatternFill("solid", fgColor="C6EFCE")
FONT_ADDED = Font(color="276221")
FILL_REMOVED = PatternFill("solid", fgColor="FFC7CE")
FONT_REMOVED = Font(color="9C0006")
FILL_UPDATED = PatternFill("solid", fgColor="FFEB9C")
FONT_UPDATED = Font(color="9C5700")
FILL_NEWROW = PatternFill("solid", fgColor="BDD7EE")
FONT_NEWROW = Font(color="1F4E79")

SHEET_NAME = "BrandIngest"
LEGEND_SHEET = "Legend"

# Friendly change-list keys → real BDR column names.
FIELD_ALIASES = {
    "title": "title",
    "category": "title_category",
    "title_category": "title_category",
    "subcategory": "title_sub_category",
    "sub_category": "title_sub_category",
    "title_sub_category": "title_sub_category",
    "companies": "companies",
    "brand_set": "brand_set",
    "active": "active",
    "facebook": "facebook_page",
    "facebook_page": "facebook_page",
    "fb": "facebook_page",
    "twitter": "twitter_handle",
    "twitter_handle": "twitter_handle",
    "x": "twitter_handle",
    "instagram": "instagram_user",
    "instagram_user": "instagram_user",
    "ig": "instagram_user",
    "youtube": "youtube_channel_username",
    "youtube_channel_username": "youtube_channel_username",
    "yt": "youtube_channel_username",
    "tiktok": "tiktok_user",
    "tiktok_user": "tiktok_user",
    "linkedin": "linkedin_page",
    "linkedin_page": "linkedin_page",
    "threads": "threads_page",
    "wikipedia": "wikipedia_page",
    "wikipedia_page": "wikipedia_page",
    "wiki": "wikipedia_page",
    "imdb": "imdb_id",
    "imdb_id": "imdb_id",
}

BARE_HANDLE_FIELDS = {"instagram_user", "tiktok_user"}
FULL_URL_FIELDS = {"facebook_page", "twitter_handle", "youtube_channel_username", "wikipedia_page", "imdb_id", "linkedin_page", "threads_page"}
PLATFORM_LABELS = {
    "facebook_page": "Facebook", "twitter_handle": "Twitter", "instagram_user": "Instagram",
    "youtube_channel_username": "YouTube", "tiktok_user": "TikTok", "wikipedia_page": "Wikipedia",
    "imdb_id": "IMDb", "linkedin_page": "LinkedIn", "threads_page": "Threads",
}
FAN_PATTERNS = ("fan", "fanpage", "unofficial", "daily", "updates", "real_", "the_real", "_real", "stans", "army", "hq", "fandom")


# ─── Instruction model ─────────────────────────────────────────────────────
@dataclass
class Instruction:
    kind: str  # ADD | REMOVE | RENAME | NEW | AMBIGUOUS
    target: str = ""            # title to match (ADD/REMOVE/RENAME old)
    field: str = ""             # column (ADD/REMOVE)
    value: str = ""             # new value (ADD) / new title (RENAME)
    attrs: dict[str, str] = _dcfield(default_factory=dict)  # NEW row k=v
    raw: str = ""
    note: str = ""


@dataclass
class Summary:
    rows_modified: int = 0
    handles_added: int = 0
    handles_removed: int = 0
    titles_renamed: int = 0
    new_rows: list[str] = _dcfield(default_factory=list)
    tst_rebuilt: int = 0
    fb_slashes_stripped: int = 0
    flags: list[str] = _dcfield(default_factory=list)   # QA / NOT FOUND / disambig
    unmatched: list[str] = _dcfield(default_factory=list)


# ─── TST builder (exact DAR rule) ──────────────────────────────────────────
def build_tst(twitter_handle_url: str | None, ig_handle: str | None, title: str) -> str:
    lines: list[str] = []
    if twitter_handle_url:
        slug = twitter_handle_url.rstrip("/").split("/")[-1].lstrip("@")
        if slug:
            lines.append(f"@{slug}|DAR|DAR")
    if ig_handle:
        tag = ig_handle.replace(".", "").replace("_", "").lower()
    else:
        base = (title or "").replace(" - DAR", "")
        nfkd = unicodedata.normalize("NFKD", base)
        ascii_base = nfkd.encode("ascii", "ignore").decode("ascii")
        tag = re.sub(r"[^a-z0-9]", "", ascii_base.lower())
    if tag:
        lines.append(f"#{tag}|DAR|DAR")
    return "\n".join(lines)


# ─── Change-list parser ────────────────────────────────────────────────────
def parse_change_list(text: str) -> list[Instruction]:
    """Parse the change list. Primary format is pipe-delimited (robust):

        ADD | <title> | <field> | <value>
        REMOVE | <title> | <field>
        RENAME | <old title> | <new title>
        NEW | <Name> | <category> | key=value; key=value; type=Actor,Director

    Free-text lines matching the common patterns (Instagram: <url>,
    "please remove the tiktok", "rename ... as ...") are best-effort parsed;
    anything else is flagged AMBIGUOUS rather than guessed.
    """
    instructions: list[Instruction] = []
    for raw_line in (text or "").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue

        if "|" in line and line.split("|", 1)[0].strip().upper() in {"ADD", "REMOVE", "RENAME", "NEW"}:
            parts = [p.strip() for p in line.split("|")]
            verb = parts[0].upper()
            if verb == "ADD" and len(parts) >= 4:
                instructions.append(Instruction("ADD", target=parts[1], field=_resolve_field(parts[2]), value="|".join(parts[3:]).strip(), raw=line))
            elif verb == "REMOVE" and len(parts) >= 3:
                instructions.append(Instruction("REMOVE", target=parts[1], field=_resolve_field(parts[2]), raw=line))
            elif verb == "RENAME" and len(parts) >= 3:
                instructions.append(Instruction("RENAME", target=parts[1], value=parts[2], raw=line))
            elif verb == "NEW" and len(parts) >= 2:
                attrs = _parse_new_attrs(parts[3] if len(parts) >= 4 else "")
                instructions.append(Instruction("NEW", target=parts[1], value=(parts[2] if len(parts) >= 3 else "Talent"), attrs=attrs, raw=line))
            else:
                instructions.append(Instruction("AMBIGUOUS", raw=line, note="Malformed pipe instruction"))
            continue

        parsed = _parse_freeform(line)
        instructions.append(parsed)
    return instructions


def _resolve_field(token: str) -> str:
    return FIELD_ALIASES.get(token.strip().lower().replace(" ", "_"), token.strip().lower())


def _parse_new_attrs(blob: str) -> dict[str, str]:
    attrs: dict[str, str] = {}
    for pair in re.split(r"[;\n]", blob):
        if "=" in pair:
            k, v = pair.split("=", 1)
            attrs[k.strip().lower()] = v.strip()
    return attrs


def _parse_freeform(line: str) -> Instruction:
    low = line.lower()
    # RENAME: "rename <old> as/to <new>"
    m = re.search(r"rename\s+(?:the\s+brand\s+)?['\"]?(.+?)['\"]?\s+(?:as|to)\s+['\"]?(.+?)['\"]?$", line, re.IGNORECASE)
    if m:
        return Instruction("RENAME", target=m.group(1).strip(), value=m.group(2).strip(), raw=line)
    # REMOVE: "remove the <platform> [from <title>]"
    if "remove" in low or "delete" in low:
        plat = _detect_platform(low)
        title = _extract_title_after(line, ("from", "for", "on"))
        if plat:
            return Instruction("REMOVE", target=title, field=plat, raw=line, note="" if title else "title not specified")
        return Instruction("AMBIGUOUS", raw=line, note="remove: platform not identified")
    # ADD: "<platform>: <value>" or "add <platform> <value> to <title>"
    plat = _detect_platform(low)
    urlm = re.search(r"(https?://\S+)", line)
    handlem = re.search(r"@([A-Za-z0-9_.]+)", line)
    if plat and (urlm or handlem):
        val = urlm.group(1) if urlm else handlem.group(1)
        title = _extract_title_after(line, ("to", "for", "on"))
        return Instruction("ADD", target=title, field=plat, value=val, raw=line, note="" if title else "title not specified")
    return Instruction("AMBIGUOUS", raw=line, note="Could not classify instruction")


def _detect_platform(low: str) -> str:
    for kw, col in (("facebook", "facebook_page"), ("instagram", "instagram_user"), ("tiktok", "tiktok_user"),
                    ("youtube", "youtube_channel_username"), ("twitter", "twitter_handle"), (" x ", "twitter_handle"),
                    ("wikipedia", "wikipedia_page"), ("imdb", "imdb_id"), ("linkedin", "linkedin_page"),
                    ("threads", "threads_page")):
        if kw in low:
            return col
    return ""


def _extract_title_after(line: str, preps: tuple[str, ...]) -> str:
    for prep in preps:
        m = re.search(rf"\b{prep}\s+['\"]?(.+?)['\"]?$", line, re.IGNORECASE)
        if m:
            return m.group(1).strip().rstrip(".")
    return ""


# ─── Value formatting per BDR rules ────────────────────────────────────────
def format_value(field_name: str, value: str) -> str:
    v = (value or "").strip()
    if not v:
        return v
    if field_name in BARE_HANDLE_FIELDS:
        v = v.lstrip("@")
        if v.startswith("http"):
            v = v.rstrip("/").split("/")[-1]
        return v
    if field_name == "facebook_page":
        return v.rstrip("/")
    if field_name == "imdb_id" and v.startswith("nm"):
        return f"https://www.imdb.com/name/{v}"
    return v


# ─── Workbook helpers ──────────────────────────────────────────────────────
def _comment(text: str) -> Comment:
    c = Comment(text, "DataOps")
    c.width = 340
    c.height = 80
    return c


def _paint(cell, fill, font, comment_text: str) -> None:
    cell.fill = fill
    cell.font = font
    cell.comment = _comment(comment_text)


def _colmap(ws) -> dict[str, int]:
    out: dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            out[str(cell.value).strip().lower()] = idx
    return out


def _find_row(ws, colmap: dict[str, int], title: str) -> int | None:
    tcol = colmap.get("title")
    if not tcol or not title:
        return None
    target = title.strip().lower()
    target_nodar = target.replace(" - dar", "")
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=tcol).value
        if val is None:
            continue
        cur = str(val).strip().lower()
        if cur == target or cur.replace(" - dar", "") == target_nodar:
            return r
    return None


# ─── Main entry point ──────────────────────────────────────────────────────
def build_ingest(workbook_bytes: bytes, change_list_text: str, today: str | None = None) -> tuple[bytes, Summary]:
    today = today or date.today().isoformat()
    wb = load_workbook(io.BytesIO(workbook_bytes))
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    colmap = _colmap(ws)
    summary = Summary()
    instructions = parse_change_list(change_list_text)
    # Apply existing-row handle changes (REMOVE/ADD) BEFORE renames so a handle
    # instruction that references the current title still matches; renames then
    # run, and new rows last. AMBIGUOUS lines are just flagged.
    _priority = {"REMOVE": 0, "ADD": 1, "RENAME": 2, "NEW": 3, "AMBIGUOUS": 4}
    instructions = sorted(instructions, key=lambda i: _priority.get(i.kind, 5))
    touched_rows: set[int] = set()

    def set_last_reviewed(row: int) -> None:
        c = colmap.get("last_reviewed")
        if c:
            cell = ws.cell(row=row, column=c)
            if cell.value != today:
                cell.value = today
                _paint(cell, FILL_UPDATED, FONT_UPDATED, f"UPDATED: last_reviewed set to {today}")

    for ins in instructions:
        if ins.kind == "AMBIGUOUS":
            summary.flags.append(f"AMBIGUOUS: {ins.raw} ({ins.note})")
            continue

        if ins.kind == "RENAME":
            row = _find_row(ws, colmap, ins.target)
            if not row:
                summary.unmatched.append(f"RENAME target not found: '{ins.target}'")
                continue
            tcol = colmap["title"]
            new_title = ins.value
            _paint(ws.cell(row=row, column=tcol), FILL_UPDATED, FONT_UPDATED,
                   f"RENAMED: '{ins.target}' → '{new_title}' per client instruction")
            ws.cell(row=row, column=tcol).value = new_title
            summary.titles_renamed += 1
            touched_rows.add(row)
            set_last_reviewed(row)
            continue

        if ins.kind == "REMOVE":
            row = _find_row(ws, colmap, ins.target)
            if not row or ins.field not in colmap:
                summary.unmatched.append(f"REMOVE not applied ({ins.field}) for '{ins.target}'")
                continue
            _clear_handle(ws, colmap, row, ins.field, summary)
            touched_rows.add(row)
            set_last_reviewed(row)
            continue

        if ins.kind == "ADD":
            row = _find_row(ws, colmap, ins.target)
            if not row or ins.field not in colmap:
                summary.unmatched.append(f"ADD not applied ({ins.field}) for '{ins.target}'")
                continue
            val = format_value(ins.field, ins.value)
            cell = ws.cell(row=row, column=colmap[ins.field])
            label = PLATFORM_LABELS.get(ins.field, ins.field)
            if _fan_risk(val):
                cell.value = val
                _paint(cell, FILL_ADDED, FONT_ADDED,
                       f"ADDED: {label} ({val}) — FAN/PARODY RISK: verify this is the official account")
                summary.flags.append(f"FAN/PARODY RISK: {label} '{val}' on '{ins.target}'")
            else:
                cell.value = val
                _paint(cell, FILL_ADDED, FONT_ADDED, f"ADDED: {label} ({val}) added per client instruction")
            summary.handles_added += 1
            touched_rows.add(row)
            set_last_reviewed(row)
            continue

        if ins.kind == "NEW":
            _build_new_row(ws, colmap, ins, today, summary)

    # Rebuild TST on every touched (existing) row that had a handle change.
    for row in sorted(touched_rows):
        _rebuild_tst(ws, colmap, row, summary)

    _enforce_hard_rules(ws, colmap, touched_rows, summary)
    summary.rows_modified = len(touched_rows)
    _write_legend(wb, summary)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), summary


def _clear_handle(ws, colmap, row, field_name, summary: Summary) -> None:
    cell = ws.cell(row=row, column=colmap[field_name])
    old = cell.value
    label = PLATFORM_LABELS.get(field_name, field_name)
    cell.value = None
    _paint(cell, FILL_REMOVED, FONT_REMOVED, f"REMOVED: {label} ({old}) removed per client instruction")
    summary.handles_removed += 1
    # Paired fields
    if field_name == "facebook_page" and "facebook_verified" in colmap:
        vc = ws.cell(row=row, column=colmap["facebook_verified"])
        vc.value = None
        _paint(vc, FILL_REMOVED, FONT_REMOVED, "REMOVED: facebook_verified cleared (facebook_page removed)")
    if field_name == "twitter_handle" and "twitter_verified" in colmap:
        vc = ws.cell(row=row, column=colmap["twitter_verified"])
        vc.value = None
        _paint(vc, FILL_REMOVED, FONT_REMOVED, "REMOVED: twitter_verified cleared (twitter_handle removed)")


def _rebuild_tst(ws, colmap, row, summary: Summary) -> None:
    if "twitter_search_terms" not in colmap:
        return
    tw = ws.cell(row=row, column=colmap["twitter_handle"]).value if "twitter_handle" in colmap else None
    ig = ws.cell(row=row, column=colmap["instagram_user"]).value if "instagram_user" in colmap else None
    title = ws.cell(row=row, column=colmap["title"]).value if "title" in colmap else ""
    new_tst = build_tst(str(tw) if tw else None, str(ig) if ig else None, str(title) if title else "")
    cell = ws.cell(row=row, column=colmap["twitter_search_terms"])
    if (cell.value or "") != new_tst:
        cell.value = new_tst
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        _paint(cell, FILL_UPDATED, FONT_UPDATED, "UPDATED: twitter_search_terms rebuilt per DAR rule")
        summary.tst_rebuilt += 1


def _build_new_row(ws, colmap, ins: Instruction, today: str, summary: Summary) -> None:
    row = ws.max_row + 1
    name = ins.target.strip()
    a = ins.attrs
    category = (ins.value or a.get("category") or "Talent").strip()
    title = name if name.endswith(" - DAR") else f"{name} - DAR"

    def put(field_name: str, value: str, comment: str) -> None:
        if field_name not in colmap or not value:
            return
        cell = ws.cell(row=row, column=colmap[field_name])
        cell.value = value
        cell.font = FONT_NEWROW
        cell.fill = FILL_NEWROW
        cell.comment = _comment(comment)

    # Deterministic scaffolding
    put("title", title, f"NEW ROW: {name}")
    put("title_category", category, f"NEW ROW: category = {category}")
    put("companies", "Pristine Brand", "NEW ROW: Pristine Brand (DAR)")
    put("brand_set", "LF // Talent\nPristine DAR Brands", "NEW ROW: default DAR brand_set")
    put("active", "true", "NEW ROW: active")
    if "last_reviewed" in colmap:
        put("last_reviewed", today, f"NEW ROW: last_reviewed {today}")

    # title_sub_category (multi-line) from provided gender/type; else flag
    sub = _build_subcategory(a)
    if sub:
        put("title_sub_category", sub, "NEW ROW: title_sub_category from client-provided gender/type")
    else:
        summary.flags.append(f"NEEDS MANUAL REVIEW: title_sub_category for '{title}' (gender/type not provided)")

    # Provided socials / metadata (format-normalised); flag research-required blanks
    provided_any = False
    for key, val in a.items():
        col = FIELD_ALIASES.get(key)
        if col and col not in ("title_category", "title_sub_category") and val:
            fv = format_value(col, val)
            label = PLATFORM_LABELS.get(col, col)
            if _fan_risk(fv):
                summary.flags.append(f"FAN/PARODY RISK: {label} '{fv}' on new row '{title}'")
            put(col, fv, f"NEW ROW: {label} provided by client")
            provided_any = True

    # Research-required fields left blank + flag (anti-hallucination)
    for col, label in (("imdb_id", "IMDb"), ("wikipedia_page", "Wikipedia")):
        if col in colmap and not ws.cell(row=row, column=colmap[col]).value:
            summary.flags.append(f"NOT FOUND: {label} for '{title}' — verify + populate manually (never inferred)")

    # TST for the new row
    if "twitter_search_terms" in colmap:
        tw = ws.cell(row=row, column=colmap["twitter_handle"]).value if "twitter_handle" in colmap else None
        ig = ws.cell(row=row, column=colmap["instagram_user"]).value if "instagram_user" in colmap else None
        tst = build_tst(str(tw) if tw else None, str(ig) if ig else None, title)
        cell = ws.cell(row=row, column=colmap["twitter_search_terms"])
        cell.value = tst
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = FONT_NEWROW
        cell.fill = FILL_NEWROW
        cell.comment = _comment("NEW ROW: twitter_search_terms built per DAR rule")
        summary.tst_rebuilt += 1

    # Paint any remaining populated cells in the row blue (leave blanks unfilled)
    summary.new_rows.append(title)


def _build_subcategory(attrs: dict[str, str]) -> str:
    lines: list[str] = []
    gender = attrs.get("gender")
    if gender:
        lines.append(f"Gender - {gender.strip().title()}")
    types_raw = attrs.get("type") or attrs.get("talent_type") or ""
    for t in re.split(r"[,;]", types_raw):
        t = t.strip()
        if t:
            lines.append(f"Talent Type - {t.title()}")
    return "\n".join(lines) if (gender and len(lines) >= 2) else ("\n".join(lines) if lines else "")


def _fan_risk(value: str) -> bool:
    low = (value or "").lower()
    return any(p in low for p in FAN_PATTERNS)


def _enforce_hard_rules(ws, colmap, touched_rows: set[int], summary: Summary) -> None:
    fb = colmap.get("facebook_page")
    if fb:
        for r in list(touched_rows):
            cell = ws.cell(row=r, column=fb)
            if isinstance(cell.value, str) and cell.value.endswith("/"):
                cell.value = cell.value.rstrip("/")
                summary.fb_slashes_stripped += 1


def _write_legend(wb, summary: Summary) -> None:
    if LEGEND_SHEET in wb.sheetnames:
        del wb[LEGEND_SHEET]
    ws = wb.create_sheet(LEGEND_SHEET)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80
    rows = [
        ("Colour Code Legend", ""),
        ("Green cell", "Handle / URL / value ADDED"),
        ("Red cell", "Handle / URL / value REMOVED (cell cleared)"),
        ("Yellow cell", "Value UPDATED, renamed, or rebuilt"),
        ("Blue row", "Entire NEW ROW ingested"),
        ("", ""),
        ("Analysis Notes — This Session", ""),
        (f"Rows modified: {summary.rows_modified}", ""),
        (f"Handles added: {summary.handles_added}", ""),
        (f"Handles removed: {summary.handles_removed}", ""),
        (f"Titles renamed: {summary.titles_renamed}", ""),
        (f"New rows: {len(summary.new_rows)}", "; ".join(summary.new_rows)),
        (f"TST rebuilt: {summary.tst_rebuilt}", ""),
        (f"Facebook trailing slashes stripped: {summary.fb_slashes_stripped}", ""),
    ]
    for label, val in rows:
        ws.append([label, val])
    ws["A1"].font = Font(bold=True, size=12)
    ws["A7"].font = Font(bold=True, size=12)
    swatches = {2: FILL_ADDED, 3: FILL_REMOVED, 4: FILL_UPDATED, 5: FILL_NEWROW}
    for r, fill in swatches.items():
        ws.cell(row=r, column=1).fill = fill
    for flag in summary.flags + summary.unmatched:
        ws.append(["FLAG", flag])
