"""Apply-BDR diff report.

Input: an ApplyBrandDefinitionReport workbook whose BrandIngest sheet holds
TWO rows per brand — record_type "INGESTED" (what was applied) and "FROM DB"
(the database state). Output: the SAME sheet, layout untouched, with the
INGESTED row of each pair highlighted wherever it differs from the DB row
(green = value added, red = value removed, yellow = value changed; every
highlight carries a DataOps comment with the DB value), plus:

- a "URL Manager" sheet comparing each row's url_managers entries against the
  platform URLs maintained in the row's own columns (mismatch / missing /
  orphan findings, mirrored as yellow highlights on the url_managers cell);
- a "Legend" sheet with the colour key and session findings.

Multi-line fields (brand_set, twitter_search_terms, url_managers, ...) are
compared as unordered line-sets so pure reordering never flags.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field as _dcfield

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

FILL_ADDED = PatternFill("solid", fgColor="C6EFCE")
FONT_ADDED = Font(color="276221")
FILL_REMOVED = PatternFill("solid", fgColor="FFC7CE")
FONT_REMOVED = Font(color="9C0006")
FILL_UPDATED = PatternFill("solid", fgColor="FFEB9C")
FONT_UPDATED = Font(color="9C5700")

SHEET_NAME = "BrandIngest"
URLM_SHEET = "URL Manager"
LEGEND_SHEET = "Legend"

MULTILINE_COLUMNS = {
    "brand_set", "composite_brand_set", "twitter_search_terms",
    "instagram_business_hashtags", "twitter_search_term_keywords",
    "url_managers", "title_sub_category", "companies", "genre",
}
SKIP_COLUMNS = {"record_type"}

# url_managers platform token -> the column that maintains that URL/handle
URLM_PLATFORM_COLUMNS = {
    "facebook": "facebook_page",
    "twitter": "twitter_handle",
    "instagram": "instagram_user",
    "youtube": "youtube_channel_username",
    "tiktok": "tiktok_user",
    "linkedin": "linkedin_page",
    "threads": "threads_page",
    "pinterest": "pinterest_user_username",
}


@dataclass
class ApplySummary:
    brands: int = 0
    pairs_diffed: int = 0
    cells_added: int = 0
    cells_removed: int = 0
    cells_changed: int = 0
    urlm_findings: list[str] = _dcfield(default_factory=list)
    notes: list[str] = _dcfield(default_factory=list)


def _comment(text: str) -> Comment:
    c = Comment(text, "DataOps")
    c.width = 340
    c.height = 90
    return c


def _norm_slug(value: str) -> str:
    """Normalise a URL or handle down to a comparable slug."""
    v = (value or "").strip().rstrip("/")
    v = re.sub(r"^https?://(www\.)?", "", v, flags=re.IGNORECASE)
    seg = v.split("/")[-1] if "/" in v else v
    return seg.lstrip("@").casefold()


def _norm_url(value: str) -> str:
    v = (value or "").strip().rstrip("/")
    return re.sub(r"^https?://(www\.)?", "", v, flags=re.IGNORECASE).casefold()


def _lineset(value: str) -> frozenset[str]:
    return frozenset(ln.strip() for ln in (value or "").splitlines() if ln.strip())


def build_apply_diff_report(workbook_bytes: bytes) -> tuple[bytes, ApplySummary]:
    wb = load_workbook(io.BytesIO(workbook_bytes))
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    summary = ApplySummary()

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    colmap = {h.lower(): i + 1 for i, h in enumerate(headers) if h}
    rt_col = colmap.get("record_type")
    id_col = colmap.get("brand_id")
    title_col = colmap.get("title")
    if not rt_col or not id_col:
        raise ValueError("This does not look like an Apply BDR: 'record_type' / 'brand_id' columns missing.")

    # Group row numbers by brand_id.
    groups: dict[str, dict[str, int]] = {}
    for r in range(2, ws.max_row + 1):
        bid = ws.cell(row=r, column=id_col).value
        rt = str(ws.cell(row=r, column=rt_col).value or "").strip().upper()
        if bid is None or not rt:
            continue
        groups.setdefault(str(bid), {})[rt] = r
    summary.brands = len(groups)

    def cellval(r: int, c: int) -> str:
        v = ws.cell(row=r, column=c).value
        return "" if v is None else str(v)

    # ── Pairwise diff: highlight the INGESTED cell ──────────────────────────
    for bid, rows in groups.items():
        ing, db = rows.get("INGESTED"), rows.get("FROM DB")
        if ing is None or db is None:
            summary.notes.append(f"brand_id {bid}: missing {'FROM DB' if db is None else 'INGESTED'} row — pair skipped")
            continue
        summary.pairs_diffed += 1
        title = cellval(ing, title_col) if title_col else bid

        for h in headers:
            key = h.lower()
            if not key or key in SKIP_COLUMNS:
                continue
            c = colmap[key]
            v_ing, v_db = cellval(ing, c).strip(), cellval(db, c).strip()
            if key in MULTILINE_COLUMNS:
                same = _lineset(v_ing) == _lineset(v_db)
            else:
                same = v_ing == v_db
            if same:
                continue
            cell = ws.cell(row=ing, column=c)
            db_disp = v_db if len(v_db) <= 240 else v_db[:240] + "…"
            if v_ing and not v_db:
                cell.fill, cell.font = FILL_ADDED, FONT_ADDED
                cell.comment = _comment(f"ADDED vs DB: '{h}' was blank in the database.")
                summary.cells_added += 1
            elif v_db and not v_ing:
                cell.fill, cell.font = FILL_REMOVED, FONT_REMOVED
                cell.comment = _comment(f"REMOVED vs DB: '{h}' had a value in the database:\n{db_disp}")
                summary.cells_removed += 1
            else:
                cell.fill, cell.font = FILL_UPDATED, FONT_UPDATED
                cell.comment = _comment(f"CHANGED vs DB: '{h}' database value:\n{db_disp}")
                summary.cells_changed += 1

    # ── URL Manager cross-check (per row, both INGESTED and FROM DB) ───────
    urlm_rows: list[tuple[str, str, str, str, str, str, str]] = []
    urlm_col = colmap.get("url_managers")
    if urlm_col:
        for bid, rows in groups.items():
            for rt, r in rows.items():
                title = cellval(r, title_col) if title_col else bid
                raw = cellval(r, urlm_col)
                entries: dict[str, tuple[str, str]] = {}
                for ln in raw.splitlines():
                    parts = [p.strip() for p in ln.split("|")]
                    if len(parts) >= 2 and parts[0]:
                        entries[parts[0].casefold()] = (parts[1], parts[2] if len(parts) > 2 else "")

                issues_here: list[str] = []
                for plat, colname in URLM_PLATFORM_COLUMNS.items():
                    col_idx = colmap.get(colname)
                    maintained = cellval(r, col_idx).strip() if col_idx else ""
                    um = entries.get(plat)
                    if um is None and not maintained:
                        continue
                    if um is None and maintained:
                        status = "MISSING in url_managers"
                        issues_here.append(f"{plat}: {status}")
                        urlm_rows.append((bid, title, rt, plat, "", maintained, status))
                        continue
                    um_val, um_company = um
                    if not maintained:
                        status = "ORPHAN entry (no maintained URL in row)"
                        issues_here.append(f"{plat}: {status}")
                        urlm_rows.append((bid, title, rt, plat, um_val, "", status))
                        continue
                    # A maintained cell may hold several handles/URLs on separate
                    # lines (e.g. primary + secondary twitter) — match any line.
                    maintained_lines = [ln for ln in maintained.splitlines() if ln.strip()] or [maintained]
                    matched = any(
                        _norm_slug(um_val) == _norm_slug(ml) or _norm_url(um_val) == _norm_url(ml)
                        for ml in maintained_lines
                    )
                    if matched:
                        urlm_rows.append((bid, title, rt, plat, um_val, maintained, "OK"))
                    else:
                        status = "MISMATCH vs maintained URL"
                        issues_here.append(f"{plat}: '{um_val}' vs '{maintained}'")
                        urlm_rows.append((bid, title, rt, plat, um_val, maintained, status))

                if issues_here:
                    cell = ws.cell(row=r, column=urlm_col)
                    # Don't clobber a pairwise-diff highlight colour; comment carries detail.
                    if cell.fill is None or cell.fill.fgColor is None or cell.fill.patternType is None:
                        cell.fill, cell.font = FILL_UPDATED, FONT_UPDATED
                    existing = (cell.comment.text + "\n\n") if cell.comment else ""
                    cell.comment = _comment(existing + "URL MANAGER: " + "; ".join(issues_here))
                    summary.urlm_findings.append(f"{title} [{rt}] — " + "; ".join(issues_here))

    _write_urlm_sheet(wb, urlm_rows)
    _write_legend(wb, summary)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), summary


def _write_urlm_sheet(wb, rows) -> None:
    if URLM_SHEET in wb.sheetnames:
        del wb[URLM_SHEET]
    ws = wb.create_sheet(URLM_SHEET)
    headers = ["brand_id", "title", "record_type", "platform", "url_managers value", "maintained value (row column)", "status"]
    ws.append(headers)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i).font = Font(bold=True)
    widths = [10, 42, 12, 12, 48, 48, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    for row in rows:
        ws.append(list(row))
        status = row[-1]
        cell = ws.cell(row=ws.max_row, column=len(headers))
        if status == "OK":
            cell.fill, cell.font = FILL_ADDED, FONT_ADDED
        elif status.startswith("MISSING"):
            cell.fill, cell.font = FILL_REMOVED, FONT_REMOVED
        else:
            if status != "OK":
                cell.fill, cell.font = FILL_UPDATED, FONT_UPDATED
    ws.freeze_panes = "A2"


def _write_legend(wb, summary: ApplySummary) -> None:
    if LEGEND_SHEET in wb.sheetnames:
        del wb[LEGEND_SHEET]
    ws = wb.create_sheet(LEGEND_SHEET)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 96
    rows = [
        ("Colour Code Legend", "(highlights are applied to the INGESTED row of each brand pair)"),
        ("Green cell", "Value ADDED vs database (DB was blank)"),
        ("Red cell", "Value REMOVED vs database (DB had a value, ingested is blank)"),
        ("Yellow cell", "Value CHANGED vs database (comment holds the DB value)"),
        ("", ""),
        ("Analysis Notes — This Session", ""),
        (f"Brands: {summary.brands}", ""),
        (f"Pairs diffed: {summary.pairs_diffed}", ""),
        (f"Cells added: {summary.cells_added}", ""),
        (f"Cells removed: {summary.cells_removed}", ""),
        (f"Cells changed: {summary.cells_changed}", ""),
        (f"URL Manager findings: {len(summary.urlm_findings)}", "see 'URL Manager' sheet for the full comparison"),
    ]
    for label, val in rows:
        ws.append([label, val])
    ws["A1"].font = Font(bold=True, size=12)
    ws["A6"].font = Font(bold=True, size=12)
    swatches = {2: FILL_ADDED, 3: FILL_REMOVED, 4: FILL_UPDATED}
    for r, fill in swatches.items():
        ws.cell(row=r, column=1).fill = fill
    for f in summary.urlm_findings:
        ws.append(["URL FLAG", f])
    for n in summary.notes:
        ws.append(["NOTE", n])
