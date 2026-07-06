from __future__ import annotations

import io
import uuid
import re
from pathlib import Path
from typing import Annotated

from fastapi import APIRouter, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from starlette.concurrency import run_in_threadpool

from app.cache import TTLCache
from app.config import settings
from app.models import EntityQuery, ExportPayload, SearchSession, ValidationHistoryEntry, WorkbookValidationArtifact
from app.services.entity_resolver import EntityResolver
from app.services.exporter import (
    build_ui_matrix_rows,
    build_export_payload_from_bulk,
    build_export_payload_from_search,
    export_to_google_sheets,
    rows_to_csv_bytes,
    rows_to_xlsx_bytes,
)
from app.services.factory import build_entity_resolver, build_media_resolver
from app.services.media_resolver import MediaResolver
from app.services.search_provider import SearchProviderUnavailableError
from app.services.tmdb_client import TmdbUnavailableError
from app.services.validation_history import list_validation_runs, load_saved_validation_file, record_validation_run
from app.services.workbook_validator import (
    WorkbookValidationConfigError,
    build_sample_rules_json,
    load_google_sheet_workbook,
    parse_validation_rules,
    validate_loaded_workbook,
    validate_workbook,
)
from app.services.taxonomy_classifier import TaxonomyClassifier
from app.landing import LANDING
from app.hub import CATEGORIES, TOOLS


BASE_DIR = Path(__file__).resolve().parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
router = APIRouter()
cache = TTLCache(settings.cache_ttl_seconds)
resolver = build_entity_resolver(cache)
media_resolver = build_media_resolver(cache)
taxonomy_classifier = TaxonomyClassifier(cache)


def render_partial(request: Request, template_name: str, **context) -> HTMLResponse:
    return templates.TemplateResponse(request, template_name, context)


def store_export_payload(payload: ExportPayload) -> str:
    cache.set(f"export:{payload.export_id}", payload)
    return payload.export_id


def store_validation_artifact(artifact: WorkbookValidationArtifact) -> str:
    cache.set(f"validation:{artifact.validation_id}", artifact)
    return artifact.validation_id


def parse_bulk_queries(raw_text: str, default_entity_type: str | None, default_country: str | None) -> list[EntityQuery]:
    queries: list[EntityQuery] = []
    for line in raw_text.splitlines():
        cleaned = line.strip()
        if not cleaned:
            continue
        parts = [part.strip() for part in cleaned.split("|")]
        name = parts[0]
        entity_type = parts[1] if len(parts) > 1 and parts[1] else default_entity_type
        country = parts[2] if len(parts) > 2 and parts[2] else default_country
        queries.append(EntityQuery(name=name, entity_type=entity_type or None, country=country or None))
    return queries


def static_asset_version(*relative_paths: str) -> str:
    versions: list[int] = []
    for relative_path in relative_paths:
        asset_path = BASE_DIR / "static" / relative_path
        try:
            versions.append(asset_path.stat().st_mtime_ns)
        except FileNotFoundError:
            continue
    if not versions:
        return "1"
    return str(max(versions))


def build_template_context(request: Request) -> dict[str, object]:
    return {
        "request": request,
        "company_entity_types": EntityResolver.COMPANY_ENTITY_TYPES,
        "media_types": MediaResolver.MEDIA_TYPES,
        "talent_types": EntityResolver.TALENT_TYPES,
        "talent_professions": EntityResolver.TALENT_PROFESSIONS,
        "platforms": EntityResolver.SUPPORTED_PLATFORMS,
        "validator_rules_example": build_sample_rules_json(),
        "validation_history": list_validation_runs(settings.validation_history_limit),
        "asset_version": static_asset_version("styles.css", "ai-tech-bg.svg", "htmx-fallback.js"),
    }


def build_validation_history_context(latest_history: ValidationHistoryEntry | None = None) -> dict[str, object]:
    history_runs = list_validation_runs(settings.validation_history_limit)
    return {
        "history_runs": history_runs,
        "latest_history": latest_history or (history_runs[0] if history_runs else None),
    }


@router.get("/", response_class=HTMLResponse)
async def index(request: Request) -> HTMLResponse:
    """Premium marketing landing (module: app/landing). The Data Ops Validator
    tool remains fully available at /excel-validator."""
    return templates.TemplateResponse(
        request, "landing/index.html", {"request": request, "landing": LANDING})


@router.get("/excel-validator", response_class=HTMLResponse)
async def excel_validator(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(request, "validator.html", build_template_context(request))


@router.get("/excel-validator/guide", response_class=HTMLResponse)
async def excel_validator_guide(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(request, "validator_guide.html", build_template_context(request))


@router.post("/bdr-ingest")
async def bdr_ingest(
    workbook: Annotated[UploadFile, File()],
    change_list: Annotated[str, Form()] = "",
):
    """BDR Ingest Builder: apply a change list to a BDR workbook and return an
    ingest-ready, colour-coded workbook (Legend sheet carries the summary)."""
    from app.services.bdr_ingest import build_ingest

    raw = await workbook.read()
    if not raw:
        return HTMLResponse("No workbook uploaded.", status_code=400)
    try:
        out_bytes, summary = await run_in_threadpool(build_ingest, raw, change_list)
    except Exception as exc:  # surface a clean error instead of a 500 stack
        import traceback
        traceback.print_exc()
        return HTMLResponse(f"BDR ingest failed: {exc}", status_code=500)

    base = (workbook.filename or "BDR").rsplit(".", 1)[0]
    filename = f"{base}_FINAL.xlsx"
    return StreamingResponse(
        io.BytesIO(out_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-BDR-Rows-Modified": str(summary.rows_modified),
            "X-BDR-New-Rows": str(len(summary.new_rows)),
            "X-BDR-Flags": str(len(summary.flags)),
        },
    )


@router.post("/bdr-apply-report")
async def bdr_apply_report(workbook: Annotated[UploadFile, File()]):
    """Apply-BDR diff QA: pair each brand's INGESTED row against its FROM DB
    row, highlight differences on the INGESTED row, cross-check url_managers
    against the row's maintained platform URLs, and return the annotated
    workbook (URL Manager + Legend sheets appended; layout unchanged)."""
    from app.services.apply_bdr_report import build_apply_diff_report

    raw = await workbook.read()
    if not raw:
        return HTMLResponse("No workbook uploaded.", status_code=400)
    try:
        out_bytes, summary = await run_in_threadpool(build_apply_diff_report, raw)
    except ValueError as exc:
        return HTMLResponse(str(exc), status_code=400)
    except Exception as exc:
        import traceback
        traceback.print_exc()
        return HTMLResponse(f"Apply-report QA failed: {exc}", status_code=500)

    base = (workbook.filename or "ApplyBDR").rsplit(".", 1)[0]
    filename = f"{base}_QA.xlsx"
    return StreamingResponse(
        io.BytesIO(out_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Apply-Pairs": str(summary.pairs_diffed),
            "X-Apply-Changed": str(summary.cells_changed),
            "X-Apply-URL-Findings": str(len(summary.urlm_findings)),
        },
    )


@router.post("/search", response_class=HTMLResponse)
async def search(
    request: Request,
    name: Annotated[str, Form()],
    entity_type: Annotated[str | None, Form()] = None,
    country: Annotated[str | None, Form()] = None,
) -> HTMLResponse:
    cleaned_name = name.strip()
    if not cleaned_name:
        return render_partial(request, "_error.html", message="Enter a name to search.")
    if entity_type in {"movie", "tv_show", "tv_network", "celebrity", "influencer"}:
        return render_partial(
            request,
            "_error.html",
            message="Use Media Finder for movies/TV and Talents for influencers or celebrities.",
        )

    query = EntityQuery(name=cleaned_name, entity_type=entity_type or None, country=country or None)
    try:
        response = await resolver.search(query)
    except SearchProviderUnavailableError as exc:
        return render_partial(request, "_error.html", message=str(exc))

    if response.disambiguation_required:
        return render_partial(request, "_disambiguation.html", response=response, resolve_path="/resolve-entity")

    payload = build_export_payload_from_search(response)
    response.export_id = store_export_payload(payload)
    return render_partial(request, "_results.html", response=response, table_rows=build_ui_matrix_rows(payload.rows))


@router.post("/search/media", response_class=HTMLResponse)
async def search_media(
    request: Request,
    name: Annotated[str, Form()],
    media_type: Annotated[str, Form()] = "movie",
    country: Annotated[str | None, Form()] = None,
) -> HTMLResponse:
    cleaned_name = name.strip()
    if not cleaned_name:
        return render_partial(request, "_error.html", message="Enter a movie, TV show, or network name to search.")

    query = EntityQuery(name=cleaned_name, entity_type=media_type or "movie", country=(country or "").strip() or None)
    try:
        response = await media_resolver.search(query)
    except TmdbUnavailableError as exc:
        return render_partial(request, "_error.html", message=str(exc))

    if response.disambiguation_required:
        session_id = response.session_id or str(uuid.uuid4())
        cache.set(
            f"media-session:{session_id}",
            SearchSession(session_id=session_id, query=query, candidates=response.entity_candidates),
        )
        response.session_id = session_id
        return render_partial(request, "_disambiguation.html", response=response, resolve_path="/resolve-media")

    payload = build_export_payload_from_search(response)
    response.export_id = store_export_payload(payload)
    return render_partial(request, "_results.html", response=response, table_rows=build_ui_matrix_rows(payload.rows))


@router.post("/search/talent", response_class=HTMLResponse)
async def search_talent(
    request: Request,
    name: Annotated[str, Form()],
    talent_type: Annotated[str, Form()] = "celebrity",
    profession: Annotated[str | None, Form()] = None,
    date_of_birth: Annotated[str | None, Form()] = None,
    country: Annotated[str | None, Form()] = None,
) -> HTMLResponse:
    cleaned_name = name.strip()
    if not cleaned_name:
        return render_partial(request, "_error.html", message="Enter a talent name to search.")

    query = EntityQuery(
        name=cleaned_name,
        entity_type=talent_type or "celebrity",
        profession=(profession or "").strip() or None,
        date_of_birth=(date_of_birth or "").strip() or None,
        country=(country or "").strip() or None,
    )
    try:
        response = await resolver.search(query)
    except SearchProviderUnavailableError as exc:
        return render_partial(request, "_error.html", message=str(exc))

    if response.disambiguation_required:
        return render_partial(request, "_disambiguation.html", response=response, resolve_path="/resolve-entity")

    payload = build_export_payload_from_search(response)
    response.export_id = store_export_payload(payload)
    return render_partial(request, "_results.html", response=response, table_rows=build_ui_matrix_rows(payload.rows))


@router.post("/resolve-entity", response_class=HTMLResponse)
async def resolve_entity(
    request: Request,
    session_id: Annotated[str, Form()],
    candidate_id: Annotated[str, Form()],
) -> HTMLResponse:
    try:
        response = await resolver.resolve_from_session(session_id=session_id, candidate_id=candidate_id)
    except SearchProviderUnavailableError as exc:
        return render_partial(request, "_error.html", message=str(exc))

    payload = build_export_payload_from_search(response)
    response.export_id = store_export_payload(payload)
    return render_partial(request, "_results.html", response=response, table_rows=build_ui_matrix_rows(payload.rows))


@router.post("/resolve-media", response_class=HTMLResponse)
async def resolve_media(
    request: Request,
    session_id: Annotated[str, Form()],
    candidate_id: Annotated[str, Form()],
) -> HTMLResponse:
    session = cache.get(f"media-session:{session_id}")
    if not isinstance(session, SearchSession):
        return render_partial(request, "_error.html", message="This media search session expired. Please search again.")

    selected = next((item for item in session.candidates if item.candidate_id == candidate_id), None)
    if selected is None:
        return render_partial(request, "_error.html", message="The selected media candidate was not found.")

    try:
        response = await media_resolver.resolve_candidate(session.query, selected)
    except TmdbUnavailableError as exc:
        return render_partial(request, "_error.html", message=str(exc))

    payload = build_export_payload_from_search(response)
    response.export_id = store_export_payload(payload)
    return render_partial(request, "_results.html", response=response, table_rows=build_ui_matrix_rows(payload.rows))


@router.post("/bulk-search", response_class=HTMLResponse)
async def bulk_search(
    request: Request,
    names: Annotated[str, Form()],
    entity_type: Annotated[str | None, Form()] = None,
    country: Annotated[str | None, Form()] = None,
) -> HTMLResponse:
    if entity_type in {"movie", "tv_show", "tv_network", "celebrity", "influencer"}:
        return render_partial(
            request,
            "_error.html",
            message="Bulk Lookup currently supports the company flow only. Use Media Finder for movies/TV.",
        )
    queries = parse_bulk_queries(names, entity_type or None, country or None)
    if not queries:
        return render_partial(request, "_error.html", message="Enter at least one line for bulk search.")
    if len(queries) > 500:
        return render_partial(request, "_error.html", message="Bulk search is limited to 500 lines at a time.")

    try:
        bulk_response = await resolver.bulk_search(queries)
    except SearchProviderUnavailableError as exc:
        return render_partial(request, "_error.html", message=str(exc))

    payload = build_export_payload_from_bulk(bulk_response)
    bulk_response.export_id = store_export_payload(payload)
    return render_partial(
        request,
        "_bulk_results.html",
        response=bulk_response,
        table_rows=build_ui_matrix_rows(payload.rows),
    )


@router.post("/validate-excel", response_class=HTMLResponse)
async def validate_excel(
    request: Request,
    workbook: UploadFile | None = File(None),
    rules_json: Annotated[str | None, Form()] = None,
    rules_file: UploadFile | None = File(None),
    google_sheet_url: Annotated[str | None, Form()] = None,
    run_by: Annotated[str | None, Form()] = None,
    review_mode: Annotated[str, Form()] = "full",
    platform_filter: Annotated[str | None, Form()] = None,
) -> HTMLResponse:
    run_by_name = (run_by or "").strip()
    if not run_by_name:
        return render_partial(request, "_error.html", message="Enter a name in the Run by field before validating.")

    if rules_file and rules_file.filename:
        rules_source = (await rules_file.read()).decode("utf-8")
    else:
        rules_source = (rules_json or "").strip()

    google_sheet_reference = (google_sheet_url or "").strip()
    uploaded_filename = Path(workbook.filename or "").name if workbook else ""
    workbook_name = uploaded_filename or "workbook.xlsx"
    if not uploaded_filename and not google_sheet_reference:
        return render_partial(
            request,
            "_error.html",
            message="Upload an Excel workbook (.xlsx), a CSV file (.csv), or paste a Google Sheets URL.",
        )

    try:
        rules = parse_validation_rules(rules_source)
        if uploaded_filename:
            suffix = Path(uploaded_filename).suffix.lower()
            if suffix not in {".xlsx", ".csv"}:
                return render_partial(
                    request,
                    "_error.html",
                    message="Upload an Excel workbook (.xlsx), a CSV file (.csv), or use a Google Sheets URL.",
                )
            artifact = await run_in_threadpool(validate_workbook, await workbook.read(), uploaded_filename, rules, review_mode, platform_filter)
            workbook_name = uploaded_filename
        else:
            google_workbook, source_name = await run_in_threadpool(load_google_sheet_workbook, google_sheet_reference)
            artifact = await run_in_threadpool(validate_loaded_workbook, google_workbook, source_name, rules, review_mode, platform_filter)
            workbook_name = source_name
    except WorkbookValidationConfigError as exc:
        return render_partial(request, "_error.html", message=str(exc))
    except UnicodeDecodeError:
        return render_partial(request, "_error.html", message="Rules file must be a UTF-8 JSON file.")
    except Exception as exc:
        return render_partial(request, "_error.html", message=f"Workbook validation failed: {exc}")

    validation_id = store_validation_artifact(artifact)
    latest_history = await run_in_threadpool(
        record_validation_run,
        artifact,
        original_filename=workbook_name,
        run_by=run_by_name,
        client_ip=request.client.host if request.client else "",
    )
    return render_partial(
        request,
        "_validation_results.html",
        artifact=artifact,
        validation_id=validation_id,
        **build_validation_history_context(latest_history),
    )


@router.get("/validate-excel/download/{validation_id}")
async def download_validated_workbook(validation_id: str):
    artifact = cache.get(f"validation:{validation_id}")
    if isinstance(artifact, WorkbookValidationArtifact):
        file_bytes = artifact.file_bytes
        filename = artifact.filename
    else:
        saved_file = load_saved_validation_file(validation_id)
        if saved_file is None:
            return HTMLResponse("Validated workbook expired. Upload the file again.", status_code=404)
        file_bytes, filename = saved_file

    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/{export_id}/{fmt}")
async def export_results(export_id: str, fmt: str):
    payload = cache.get(f"export:{export_id}")
    if not isinstance(payload, ExportPayload):
        return HTMLResponse("Export expired. Run the search again.", status_code=404)

    safe_name = payload.title.replace(" ", "_").replace("/", "_")
    if fmt == "csv":
        return StreamingResponse(
            io.BytesIO(rows_to_csv_bytes(payload)),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.csv"'},
        )
    if fmt == "xlsx":
        try:
            data = rows_to_xlsx_bytes(payload)
        except RuntimeError as exc:
            return HTMLResponse(str(exc), status_code=500)
        return StreamingResponse(
            io.BytesIO(data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
        )
    if fmt == "google-sheets":
        try:
            sheet_url = export_to_google_sheets(payload)
        except RuntimeError as exc:
            return HTMLResponse(str(exc), status_code=500)
        return RedirectResponse(sheet_url)

    return HTMLResponse("Unsupported export format.", status_code=400)


@router.post("/classify-title", response_class=HTMLResponse)
async def classify_title(
    request: Request,
    title_name: Annotated[str, Form()],
) -> HTMLResponse:
    cleaned_name = title_name.strip()
    if not cleaned_name:
        return render_partial(request, "_error.html", message="Enter a title name to classify.")

    try:
        result = await taxonomy_classifier.classify(cleaned_name)
    except Exception as exc:
        return render_partial(request, "_error.html", message=f"Taxonomy classification failed: {exc}")

    return render_partial(
        request,
        "_classification_results.html",
        title=cleaned_name,
        category=result["category"],
        sub_category=result["sub_category"],
    )


def _detect_columns(headers: list[str]) -> tuple[int, int | None]:
    title_idx = 0
    ig_idx = None
    
    for idx, h in enumerate(headers):
        h_lower = str(h).strip().lower()
        if h_lower in {"title", "name", "entity_name", "title_name", "entity", "brand", "brand_name", "talent", "talent_name"}:
            title_idx = idx
            break
    else:
        for idx, h in enumerate(headers):
            h_lower = str(h).strip().lower()
            if "name" in h_lower or "title" in h_lower or "brand" in h_lower:
                title_idx = idx
                break
                
    for idx, h in enumerate(headers):
        h_lower = str(h).strip().lower()
        if h_lower in {"instagram", "ig", "instagram_user", "instagram_handle", "ig_handle", "instagram_page", "ig_page", "instagram_url", "ig_url"}:
            ig_idx = idx
            break
    else:
        for idx, h in enumerate(headers):
            h_lower = str(h).strip().lower()
            if "instagram" in h_lower or "ig" in h_lower or "handle" in h_lower or "page" in h_lower or "link" in h_lower:
                ig_idx = idx
                break
                
    return title_idx, ig_idx


def process_taxonomy_bulk_file(
    file_bytes: bytes,
    filename: str,
    run_by: str,
    classifier: TaxonomyClassifier,
) -> tuple[str, int, list[dict[str, Any]], bytes]:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    import csv
    
    is_csv = filename.lower().endswith(".csv")
    input_rows = []
    headers = []
    
    if is_csv:
        content = file_bytes.decode("utf-8-sig", errors="ignore")
        reader = csv.reader(io.StringIO(content))
        all_lines = list(reader)
        if all_lines:
            headers = [h.strip() for h in all_lines[0]]
            input_rows = all_lines[1:]
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = wb.active
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col).value
            headers.append(str(val).strip() if val is not None else f"Column {col}")
        for row in range(2, sheet.max_row + 1):
            row_vals = []
            for col in range(1, len(headers) + 1):
                row_vals.append(sheet.cell(row=row, column=col).value)
            if any(v is not None for v in row_vals):
                input_rows.append(row_vals)

    total_rows = len(input_rows)
    if total_rows == 0:
        raise RuntimeError("The uploaded workbook or CSV has no data rows.")

    title_idx, ig_idx = _detect_columns(headers)

    import asyncio
    async def run_classification():
        semaphore = asyncio.Semaphore(15)
        results = {}
        
        async def process_row(idx, row):
            title_val = str(row[title_idx]).strip() if title_idx < len(row) and row[title_idx] is not None else ""
            ig_val = ""
            if ig_idx is not None and ig_idx < len(row) and row[ig_idx] is not None:
                ig_val = str(row[ig_idx]).strip()
                
            if not title_val:
                results[idx] = {"category": "UNVERIFIED", "sub_category": "Title name cannot be blank."}
                return
                
            async with semaphore:
                try:
                    res = await classifier.classify(title_val, ig_val)
                    results[idx] = res
                except Exception as e:
                    results[idx] = {"category": "UNVERIFIED", "sub_category": f"Error: {e}"}
                    
        tasks = [process_row(idx, row) for idx, row in enumerate(input_rows)]
        await asyncio.gather(*tasks)
        return results

    classification_results = asyncio.run(run_classification())

    out_wb = openpyxl.Workbook()
    out_sheet = out_wb.active
    out_sheet.title = "Taxonomy Classification"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    badge_font = Font(name="Segoe UI", size=10, bold=True, color="1E40AF")
    badge_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    cat_idx = None
    sub_idx = None
    for idx, h in enumerate(headers):
        h_lower = str(h).strip().lower()
        if h_lower in {"title_category", "title category", "category"}:
            cat_idx = idx
        elif h_lower in {"title_sub_category", "title_subcategory", "title sub category", "title sub-category", "sub_category", "subcategory", "sub-category"}:
            sub_idx = idx

    inplace = (cat_idx is not None and sub_idx is not None)
    if inplace:
        new_headers = list(headers)
    else:
        new_headers = headers + ["title_category", "title_sub_category"]

    for col_idx, h in enumerate(new_headers, 1):
        cell = out_sheet.cell(row=1, column=col_idx)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    preview_rows = []
    
    for row_idx, row in enumerate(input_rows, 2):
        res = classification_results.get(row_idx - 2, {"category": "UNVERIFIED", "sub_category": ""})
        category_val = res.get("category", "UNVERIFIED")
        subcategory_val = res.get("sub_category", "")
        
        for col_idx, val in enumerate(row, 1):
            cell = out_sheet.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.border = thin_border
            
        if inplace:
            cat_col = cat_idx + 1
            sub_col = sub_idx + 1
        else:
            cat_col = len(row) + 1
            sub_col = len(row) + 2

        cat_cell = out_sheet.cell(row=row_idx, column=cat_col)
        cat_cell.value = category_val
        cat_cell.font = badge_font
        cat_cell.fill = badge_fill
        cat_cell.border = thin_border
        cat_cell.alignment = Alignment(horizontal="center", vertical="center")

        sub_cell = out_sheet.cell(row=row_idx, column=sub_col)
        sub_cell.value = subcategory_val
        sub_cell.font = data_font
        sub_cell.border = thin_border
        sub_cell.alignment = Alignment(wrap_text=True)

        if row_idx - 2 < 15:
            title_val = str(row[title_idx]).strip() if title_idx < len(row) and row[title_idx] is not None else ""
            ig_val = ""
            if ig_idx is not None and ig_idx < len(row) and row[ig_idx] is not None:
                ig_val = str(row[ig_idx]).strip()
            preview_rows.append({
                "row_num": row_idx,
                "title": title_val,
                "instagram": ig_val,
                "category": category_val,
                "sub_category": subcategory_val
            })

    for col in out_sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        out_sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    out_buf = io.BytesIO()
    out_wb.save(out_buf)
    out_bytes = out_buf.getvalue()

    job_id = str(uuid.uuid4())
    return job_id, total_rows, preview_rows, out_bytes


@router.post("/bulk-classify-taxonomy", response_class=HTMLResponse)
async def bulk_classify_taxonomy(
    request: Request,
    workbook: UploadFile = File(...),
    run_by: Annotated[str, Form()] = "",
) -> HTMLResponse:
    run_by_name = run_by.strip()
    if not run_by_name:
        return render_partial(request, "_error.html", message="Enter your name or team name.")
        
    uploaded_filename = workbook.filename or "uploaded_taxonomy_sheet.xlsx"
    if not (uploaded_filename.lower().endswith(".xlsx") or uploaded_filename.lower().endswith(".csv")):
        return render_partial(request, "_error.html", message="Upload an Excel workbook (.xlsx) or a CSV file (.csv).")

    file_bytes = await workbook.read()
    
    try:
        job_id, total_rows, preview_rows, out_bytes = await run_in_threadpool(
            process_taxonomy_bulk_file,
            file_bytes,
            uploaded_filename,
            run_by_name,
            taxonomy_classifier
        )
        cache.set(f"taxonomy_export:{job_id}", (uploaded_filename, out_bytes))
    except Exception as exc:
        return render_partial(request, "_error.html", message=f"Bulk taxonomy identification failed: {exc}")

    return render_partial(
        request,
        "_taxonomy_results.html",
        filename=uploaded_filename,
        total_rows=total_rows,
        preview_rows=preview_rows,
        job_id=job_id,
    )


@router.get("/taxonomy/download-template")
async def download_taxonomy_template():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Taxonomy Template"
    
    headers = [
        "title", 
        "title_category", 
        "title_sub_category", 
        "facebook_page", 
        "twitter_handle", 
        "instagram_user", 
        "youtube_channel_username", 
        "tiktok_user", 
        "wikipedia_page", 
        "imdb_id"
    ]
    
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    examples = [
        ["Christopher Nolan", "", "", "", "", "christophernolanofficial", "", "", "https://en.wikipedia.org/wiki/Christopher_Nolan", "nm0634289"],
        ["Stranger Things", "", "", "", "", "strangerthingstv", "", "", "https://en.wikipedia.org/wiki/Stranger_Things", "tt5074352"],
        ["Minecraft", "", "", "", "", "minecraft", "", "", "https://en.wikipedia.org/wiki/Minecraft", "tt3560702"],
    ]
    
    for row_idx, row in enumerate(examples, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = Font(name="Segoe UI", size=10)
            
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    out_buf = io.BytesIO()
    wb.save(out_buf)
    
    return StreamingResponse(
        io.BytesIO(out_buf.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="taxonomy_template.xlsx"'},
    )


@router.get("/taxonomy/download/{job_id}")
async def download_taxonomy_file(job_id: str):
    cached = cache.get(f"taxonomy_export:{job_id}")
    if not cached:
        return HTMLResponse("Download expired or not found. Please classify again.", status_code=404)
    filename, file_bytes = cached
    
    safe_name = re.sub(r"[^\w\-_.]", "_", filename)
    stem = Path(safe_name).stem
    classified_filename = f"{stem}_classified.xlsx"
    
    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{classified_filename}"'},
    )


@router.get("/tools", response_class=HTMLResponse)
async def all_tools(request: Request) -> HTMLResponse:
    """Media Tools Hub — searchable, filterable tool catalog (module: app/hub)."""
    context = build_template_context(request)
    context["tools"] = TOOLS
    context["categories"] = CATEGORIES
    return templates.TemplateResponse(request, "hub.html", context)
